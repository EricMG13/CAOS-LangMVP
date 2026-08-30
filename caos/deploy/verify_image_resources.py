import hashlib
import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

root = Path(__file__).resolve().parent
source_server = root.parent / "server"
if (source_server / "caos").is_dir():
    sys.path.insert(0, str(source_server))

from caos.methodology.bundle import DeployVBundle  # noqa: E402
from caos.models import CpModelBundle  # noqa: E402


bundle_root = root / "caos" / "methodology" / "vendor" / "deploy_v"
if not bundle_root.is_dir():
    bundle_root = root.parent / "server" / "caos" / "methodology" / "vendor" / "deploy_v"
report = DeployVBundle(bundle_root).verify()
# `raise`, not `assert`: this is the image's bundle gate, and every other check
# in this file raises. An assert would vanish under `python -O` and pass an
# image whose vendored authority was never counted.
if report["checked"] != 307 or report["mismatches"]:
    raise RuntimeError(f"Deploy V bundle gate failed: {report}")
arguments = sys.argv[1:]
runtime = None
if arguments:
    if len(arguments) != 2 or arguments[0] != "--runtime" or arguments[1] not in {"app", "worker"}:
        raise SystemExit("usage: verify_image_resources.py [--runtime app|worker]")
    runtime = arguments[1]
office = shutil.which("soffice") or shutil.which("libreoffice")
if runtime == "app" and office is not None:
    raise RuntimeError("LibreOffice must not be installed in the API image")
if runtime == "worker" and office is None:
    raise RuntimeError("LibreOffice is required in the export worker image")
smoke: dict[str, object] = {}
if runtime == "worker":
    fixture = Path("/image-fixtures/cp_model")
    paths = {
        module_id: fixture / f"{module_id.lower().replace('-', '')}.md"
        for module_id in ("CP-1", "CP-1A", "CP-1B", "CP-2", "CP-2B")
    }
    if not all(path.is_file() for path in paths.values()):
        raise RuntimeError("canonical CP-MODEL worker fixture is incomplete")
    with tempfile.TemporaryDirectory(prefix="caos-image-model-") as temporary:
        result = CpModelBundle(bundle_root).export(paths, Path(temporary) / "output")
        with result.output.open("rb") as workbook_file:
            if hashlib.file_digest(workbook_file, "sha256").hexdigest() != result.sha256:
                raise RuntimeError("worker fixture export hash mismatch")
        workbook = load_workbook(result.output, read_only=True, data_only=False)
        try:
            if not {"Credit Snapshot", "Model", "KPIs"}.issubset(workbook.sheetnames):
                raise RuntimeError("worker fixture export is missing visible model tabs")
        finally:
            workbook.close()
        smoke = {
            "workbook": result.output.name,
            "formulas_validated": result.formulas_validated,
            "semantic_checks": result.semantic_checks,
            "calculation_engine": result.calculation_engine,
        }
print({**report, "runtime": runtime, "libreoffice": office, "smoke": smoke})
