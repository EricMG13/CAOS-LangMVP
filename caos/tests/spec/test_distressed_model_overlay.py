from __future__ import annotations

import copy
import io

import pytest
from openpyxl import load_workbook

from caos.contracts import ModelPreviewRequest, ModelSignOffRequest, digest
from caos.models.service import ModelService
from caos.storage.runs import run_snapshots
from spec_helpers import seed_case_with_source


@pytest.fixture()
def models(settings, store, engine):
    return ModelService(store=store, vault_dir=settings.storage_dir, engine=engine)


def _calculation(engine, calculator_id: str) -> dict:
    inputs = {
        "funding_gap": {
            "horizon_years": 2,
            "cash": 100,
            "currency": "USD",
            "forecast_fcf": 50,
            "instruments": [{
                "instrument": "Near-term notes",
                "amount": 300,
                "years_to_maturity": 1,
                "currency": "USD",
            }],
        },
        "recovery_waterfall": {
            "enterprise_value": 130,
            "claims": [
                *[
                    {"claim_id": f"FILLER-{index:03d}", "class": "Senior Secured", "amount": 1}
                    for index in range(80)
                ],
                {"claim_id": "SUN", "class": "Senior Unsecured", "amount": 200},
            ],
        },
    }
    return engine._calculation_runtime.execute("CP-4C", calculator_id, inputs[calculator_id])


def _attach_cp4c_calculations(engine, run_id: str, calculator_ids=None) -> None:
    calculator_ids = calculator_ids or ("funding_gap", "recovery_waterfall")
    artifact = next(
        item
        for item in engine.runs.artifacts_for_run(run_id)
        if item["module_id"] == "CP-4C"
    )
    payload = copy.deepcopy(artifact["payload"])
    payload["calculations"] = [
        _calculation(engine, calculator_id) for calculator_id in calculator_ids
    ]
    engine.runs.update_artifact_for_tests(
        run_id,
        "CP-4C",
        payload=payload,
        digest=digest(payload),
    )


async def _ready_full_credit(models, engine, store):
    case, _source = seed_case_with_source(store)
    run = await engine.run_scripted_for_tests(case["id"])
    snapshot = await engine.accept(run["id"], actor="analyst")
    queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == snapshot["id"]
    )
    build = models.run_build_for_tests(queued["id"])
    assert build["status"] == "READY"
    return case, snapshot, build


async def _accepted_distressed(engine, case_id: str, calculator_ids=None):
    run = await engine.run_scripted_for_tests(
        case_id,
        pathway="DISTRESSED_RESTRUCTURING",
    )
    _attach_cp4c_calculations(engine, run["id"], calculator_ids)
    return run, await engine.accept(run["id"], actor="analyst")


async def test_distressed_acceptance_queues_auditable_overlay_through_intermediate_snapshot(
    models,
    engine,
    store,
):
    case, base_snapshot, base_build = await _ready_full_credit(models, engine, store)
    earnings = await engine.run_scripted_for_tests(
        case["id"],
        pathway="EARNINGS_UPDATE",
    )
    earnings_snapshot = await engine.accept(earnings["id"], actor="analyst")

    _run, distressed_snapshot = await _accepted_distressed(engine, case["id"])
    assert distressed_snapshot["previous_snapshot_id"] == earnings_snapshot["id"]
    queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == distressed_snapshot["id"]
    )
    assert queued["status"] == "QUEUED"

    overlay = models.run_build_for_tests(queued["id"])
    assert overlay["status"] == "READY"
    assert overlay["input_fingerprint"] != base_build["input_fingerprint"]
    assert overlay["payload"]["tabs"] == base_build["payload"]["tabs"]
    assert overlay["payload_digest"] == digest(overlay["payload"])
    effect, = overlay["payload"]["pathway_effects"]
    assert effect["effect_id"] == "DISTRESSED_SCENARIO_RECOVERY"
    assert effect["base_model"] == {
        "build_id": base_build["id"],
        "run_id": base_snapshot["run_id"],
        "snapshot_id": base_snapshot["id"],
        "snapshot_digest": base_snapshot["digest"],
        "source_set_id": base_snapshot["source_set_id"],
        "source_set_version": base_snapshot["source_set_version"],
        "input_fingerprint": base_build["input_fingerprint"],
        "payload_digest": base_build["payload_digest"],
        "assumptions_digest": base_build["assumptions_digest"],
        "outputs_digest": base_build["outputs_digest"],
        "qa_digest": digest(base_build["qa"]),
        "methodology_build_id": base_build["methodology_build_id"],
        "calculation_runtime": base_build["calculation_runtime"],
    }
    assert {item["calculator_id"] for item in effect["calculations"]} == {
        "funding_gap",
        "recovery_waterfall",
    }
    assert effect["calculation_records_digest"] == digest(effect["calculations"])
    readiness = models.readiness(case["id"])
    assert readiness["status"] == "READY"
    assert readiness["snapshot_id"] == distressed_snapshot["id"]
    registry = models.assumption_registry(case["id"], overlay["id"])
    assert registry["snapshot_id"] == distressed_snapshot["id"]
    assert models.default_outputs(case["id"], overlay["id"])


async def test_distressed_ancestry_follows_acceptance_order_when_runs_were_created_out_of_order(
    models,
    engine,
    store,
):
    case, _source = seed_case_with_source(store)
    earnings_run = await engine.run_scripted_for_tests(
        case["id"], pathway="EARNINGS_UPDATE"
    )
    full_credit_run = await engine.run_scripted_for_tests(case["id"])

    full_credit_snapshot = await engine.accept(full_credit_run["id"], actor="analyst")
    base_queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == full_credit_snapshot["id"]
    )
    assert models.run_build_for_tests(base_queued["id"])["status"] == "READY"
    earnings_snapshot = await engine.accept(earnings_run["id"], actor="analyst")
    _run, distressed_snapshot = await _accepted_distressed(engine, case["id"])

    assert distressed_snapshot["previous_snapshot_id"] == earnings_snapshot["id"]
    assert earnings_snapshot["previous_snapshot_id"] == full_credit_snapshot["id"]
    queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == distressed_snapshot["id"]
    )
    assert models.run_build_for_tests(queued["id"])["status"] == "READY"
    assert models.readiness(case["id"])["status"] == "READY"


async def test_distressed_model_export_includes_the_auditable_pathway_effect(
    models,
    engine,
    store,
):
    case, _base_snapshot, _base_build = await _ready_full_credit(models, engine, store)
    _run, distressed_snapshot = await _accepted_distressed(engine, case["id"])
    queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == distressed_snapshot["id"]
    )
    overlay = models.run_build_for_tests(queued["id"])

    models.queue_export(overlay["id"], "analyst")
    models.run_export_for_tests(overlay["id"])
    content, _sha256 = models.download(case["id"], overlay["id"])
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    try:
        assert "Pathway Effects" in workbook.sheetnames
        sheet = workbook["Pathway Effects"]
        rows = {path: value for path, value in sheet.iter_rows(min_row=2, values_only=True)}
        assert rows["pathway_effects.0.effect_id"] == '"DISTRESSED_SCENARIO_RECOVERY"'
        funding = next(
            path.removesuffix(".calculator_id")
            for path, value in rows.items()
            if value == '"funding_gap"' and path.endswith(".calculator_id")
        )
        recovery = next(
            path.removesuffix(".calculator_id")
            for path, value in rows.items()
            if value == '"recovery_waterfall"' and path.endswith(".calculator_id")
        )
        assert rows[
            f"{funding}.canonical_output.as_of_balance_sheet_date.gap.funding_gap"
        ] == 150
        assert rows[f"{recovery}.canonical_output.base.claims.80.claim_id"] == '"SUN"'
        assert rows[f"{recovery}.canonical_output.base.claims.80.recovery_pct"] == 0.25
        assert all(cell.data_type != "f" for row in sheet.iter_rows() for cell in row)
    finally:
        workbook.close()


async def test_signed_distressed_export_matches_the_reviewed_worksheet(
    models,
    engine,
    store,
):
    case, _base_snapshot, _base_build = await _ready_full_credit(models, engine, store)
    _run, distressed_snapshot = await _accepted_distressed(engine, case["id"])
    queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == distressed_snapshot["id"]
    )
    overlay = models.run_build_for_tests(queued["id"])
    registry = models.assumption_registry(case["id"], overlay["id"])
    assumptions = copy.deepcopy(registry["defaults"])
    changed = next(row for row in assumptions if row["status"] == "READY" and row["case"] == "BASE")
    changed["value"] = float(changed["value"]) + 0.001
    request = ModelPreviewRequest.model_validate({
        "build_id": overlay["id"],
        "parent_revision_id": None,
        "registry_version": registry["version"],
        "registry_digest": registry["digest"],
        "assumptions": assumptions,
        "draft_generation": 1,
    })
    preview = models.preview(case["id"], request)
    signed = models.sign_off(case["id"], ModelSignOffRequest.model_validate({
        **request.model_dump(),
        "preview_digest": preview["preview_digest"],
        "expected_head_revision_id": None,
        "note": "Reviewed Distressed overlay",
    }))

    models.run_export_for_tests(signed["id"])
    content, _sha256 = models.download(case["id"], signed["id"])
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    try:
        expected_titles = [tab["title"] for tab in signed["worksheet"]["tabs"]]
        assert workbook.sheetnames == expected_titles
        assert {"Assumptions", "Pathway Effects"} <= set(expected_titles)
        assumptions_sheet = workbook["Assumptions"]
        signed_row = next(
            row for row in assumptions_sheet.iter_rows(min_row=2, values_only=True)
            if row[0] == changed["assumption_id"]
            and row[1] == changed["case"]
            and row[2] == changed["period_id"]
        )
        assert signed_row[5] == changed["value"]
        effect_rows = dict(workbook["Pathway Effects"].iter_rows(min_row=2, values_only=True))
        assert effect_rows["pathway_effects.0.effect_id"] == '"DISTRESSED_SCENARIO_RECOVERY"'
    finally:
        workbook.close()


async def test_distressed_without_ready_full_credit_refuses_with_typed_blocker(
    models,
    engine,
    store,
):
    case, _source = seed_case_with_source(store)
    _run, snapshot = await _accepted_distressed(engine, case["id"])

    assert models.list_builds(case["id"]) == []
    assert models.readiness(case["id"])["blockers"] == [{
        "code": "DISTRESSED_BASE_MODEL_REQUIRED",
        "detail": (
            "Accept and complete a Full Credit model before building a "
            "Distressed Restructuring overlay."
        ),
    }]
    with pytest.raises(ValueError, match="^DISTRESSED_BASE_MODEL_REQUIRED$"):
        models.queue_build(case["id"], "analyst")
    assert snapshot["id"] == store.get_case(case["id"])["accepted_snapshot_id"]


async def test_distressed_requires_both_cp4c_calculator_records(models, engine, store):
    case, base_snapshot, _build = await _ready_full_credit(models, engine, store)
    run = await engine.run_scripted_for_tests(
        case["id"], pathway="DISTRESSED_RESTRUCTURING",
    )
    _attach_cp4c_calculations(engine, run["id"], ("funding_gap",))

    with pytest.raises(Exception, match="RUN_NOT_READY"):
        await engine.accept(run["id"], actor="analyst")
    assert engine.runs.snapshot_for_run(run["id"]) is None
    assert store.get_case(case["id"])["accepted_snapshot_id"] == base_snapshot["id"]


async def test_distressed_rejects_a_forged_cp4c_calculator_binding(models, engine, store):
    case, base_snapshot, _build = await _ready_full_credit(models, engine, store)
    run = await engine.run_scripted_for_tests(
        case["id"],
        pathway="DISTRESSED_RESTRUCTURING",
    )
    _attach_cp4c_calculations(engine, run["id"])
    artifact = next(
        item
        for item in engine.runs.artifacts_for_run(run["id"])
        if item["module_id"] == "CP-4C"
    )
    payload = copy.deepcopy(artifact["payload"])
    payload["calculations"][0]["script_digest"] = "0" * 64
    engine.runs.update_artifact_for_tests(
        run["id"],
        "CP-4C",
        payload=payload,
        digest=digest(payload),
    )
    with pytest.raises(Exception, match="RUN_NOT_READY"):
        await engine.accept(run["id"], actor="analyst")
    assert engine.runs.snapshot_for_run(run["id"]) is None
    assert store.get_case(case["id"])["accepted_snapshot_id"] == base_snapshot["id"]


async def test_distressed_recomputes_calculation_outputs_before_model_use(
    models,
    engine,
    store,
    monkeypatch,
):
    case, base_snapshot, _build = await _ready_full_credit(models, engine, store)
    run = await engine.run_scripted_for_tests(
        case["id"],
        pathway="DISTRESSED_RESTRUCTURING",
    )
    _attach_cp4c_calculations(engine, run["id"])
    artifact = next(
        item
        for item in engine.runs.artifacts_for_run(run["id"])
        if item["module_id"] == "CP-4C"
    )
    payload = copy.deepcopy(artifact["payload"])
    funding_gap = next(
        record
        for record in payload["calculations"]
        if record["calculator_id"] == "funding_gap"
    )
    # Forge the real number the model would consume, and keep the record
    # internally consistent (its output digest matches the forged output), so
    # only re-executing the pinned calculator can tell the record is false.
    gap = funding_gap["canonical_output"]["as_of_balance_sheet_date"]["gap"]
    assert gap["funding_gap"] != 999
    gap["funding_gap"] = 999
    funding_gap["output_digest"] = digest(funding_gap["canonical_output"])
    engine.runs.update_artifact_for_tests(
        run["id"],
        "CP-4C",
        payload=payload,
        digest=digest(payload),
    )

    replay_failures: list[tuple[str, str]] = []
    real_replay = engine._replay_artifact_calculations

    def observed_replay(artifact, module_id):
        try:
            return real_replay(artifact, module_id)
        except Exception as exc:  # noqa: BLE001 - recorded, then re-raised unchanged
            replay_failures.append((module_id, str(exc)))
            raise

    monkeypatch.setattr(engine, "_replay_artifact_calculations", observed_replay)
    with pytest.raises(Exception, match="RUN_NOT_READY"):
        await engine.accept(run["id"], actor="analyst")
    assert replay_failures and replay_failures[0][0] == "CP-4C", (
        "the recomputation, not a downstream lineage check, must be what refuses the forged record"
    )
    assert "calculation record differs" in replay_failures[0][1]
    assert engine.runs.snapshot_for_run(run["id"]) is None
    assert store.get_case(case["id"])["accepted_snapshot_id"] == base_snapshot["id"]


async def test_chain_or_base_tampering_invalidates_a_queued_distressed_build(
    models,
    engine,
    store,
):
    case, base_snapshot, base_build = await _ready_full_credit(models, engine, store)
    _run, distressed_snapshot = await _accepted_distressed(engine, case["id"])
    queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == distressed_snapshot["id"]
    )

    with store.engine.begin() as connection:
        connection.execute(
            run_snapshots.update()
            .where(run_snapshots.c.id == distressed_snapshot["id"])
            .values(previous_snapshot_id=distressed_snapshot["id"])
        )
    assert models.readiness(case["id"])["blockers"][0]["code"] == (
        "CANONICAL_MODEL_INPUTS_INVALID"
    )

    with store.engine.begin() as connection:
        connection.execute(
            run_snapshots.update()
            .where(run_snapshots.c.id == distressed_snapshot["id"])
            .values(previous_snapshot_id=base_snapshot["id"])
        )
    tampered = copy.deepcopy(base_build["payload"])
    tampered["tampered"] = True
    models.builds.update_build(base_build["id"], payload=tampered)
    failed = models.run_build_for_tests(queued["id"])
    assert failed["status"] == "FAILED"
    assert failed["error"]["code"] == "MODEL_INPUT_INVALID"


async def test_distressed_base_pointer_is_bound_to_the_snapshot_digest(
    models,
    engine,
    store,
):
    case, first_snapshot, _first_build = await _ready_full_credit(models, engine, store)
    second_run = await engine.run_scripted_for_tests(case["id"])
    second_snapshot = await engine.accept(second_run["id"], actor="analyst")
    second_queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == second_snapshot["id"]
    )
    assert models.run_build_for_tests(second_queued["id"])["status"] == "READY"
    _run, distressed_snapshot = await _accepted_distressed(engine, case["id"])
    distressed_queued = next(
        build for build in models.list_builds(case["id"])
        if build["snapshot_id"] == distressed_snapshot["id"]
    )

    with store.engine.begin() as connection:
        connection.execute(
            run_snapshots.update()
            .where(run_snapshots.c.id == distressed_snapshot["id"])
            .values(previous_snapshot_id=first_snapshot["id"])
        )

    assert models.readiness(case["id"])["status"] == "CANONICAL_MODEL_INPUTS_INVALID"
    failed = models.run_build_for_tests(distressed_queued["id"])
    assert failed["status"] == "FAILED"
    assert failed["error"]["code"] == "MODEL_INPUT_INVALID"


async def test_full_credit_snapshot_identity_is_revalidated_before_model_use(
    models,
    engine,
    store,
):
    case, snapshot, build = await _ready_full_credit(models, engine, store)
    with store.engine.begin() as connection:
        connection.execute(
            run_snapshots.update()
            .where(run_snapshots.c.id == snapshot["id"])
            .values(accepted_at="1999-01-01T00:00:00+00:00")
        )

    assert models.readiness(case["id"])["status"] == "CANONICAL_MODEL_INPUTS_INVALID"
    with pytest.raises(ValueError, match="MODEL_REVISION_INVALID"):
        models.validated_build(case["id"], build["id"])


async def test_withdrawn_snapshot_source_revokes_model_authority(
    models,
    engine,
    store,
):
    case, _snapshot, build = await _ready_full_credit(models, engine, store)
    source, = store.list_sources(case["id"])
    store.withdraw(case["id"], source["id"], "analyst")

    assert models.readiness(case["id"])["status"] == "CANONICAL_MODEL_INPUTS_INVALID"
    with pytest.raises(ValueError, match="MODEL_REVISION_INVALID"):
        models.validated_build(case["id"], build["id"])


async def test_intermediate_snapshot_withdrawal_does_not_revoke_an_intact_full_credit_base(
    models,
    engine,
    store,
):
    """Full Credit at v1, an Earnings run at v2 after uploading a document,
    that document withdrawn, then Distressed at v3: the base Full Credit model
    never pinned the withdrawn document and stays usable."""
    import hashlib

    case, base_snapshot, base_build = await _ready_full_credit(models, engine, store)

    body = b"an earnings update document"
    extra = store.ingest({
        "case_id": case["id"], "filename": "update.txt", "media_type": "text/plain",
        "bytes": len(body), "sha256": hashlib.sha256(body).hexdigest(), "vault_path": None,
        "blocks": [{"block_id": "b00001", "locator": {"line": 1}, "text": body.decode(),
                    "extractor_version": "builtin-v1", "confidence": "MEDIUM", "untrusted_data": True}],
        "withdrawn": False,
    }, "analyst")
    earnings = await engine.run_scripted_for_tests(case["id"], pathway="EARNINGS_UPDATE")
    earnings_snapshot = await engine.accept(earnings["id"], actor="analyst")
    assert earnings_snapshot["source_set_version"] == base_snapshot["source_set_version"] + 1

    assert store.withdraw(case["id"], extra["id"], "analyst") is not None
    _run, distressed_snapshot = await _accepted_distressed(engine, case["id"])
    assert distressed_snapshot["previous_snapshot_id"] == earnings_snapshot["id"]

    assert models.builds.get_build(base_build["id"])["status"] == "READY"
    readiness = models.readiness(case["id"])
    assert all(blocker["code"] != "DISTRESSED_BASE_MODEL_REQUIRED" for blocker in readiness.get("blockers", [])), readiness
    queued = next(
        (build for build in models.list_builds(case["id"]) if build["snapshot_id"] == distressed_snapshot["id"]),
        None,
    ) or models.queue_build(case["id"], "analyst")
    build = models.run_build_for_tests(queued["id"])
    assert build["status"] == "READY", build.get("error")
    assert build["payload"]["pathway_effects"][0]["base_model"]["build_id"] == base_build["id"]
