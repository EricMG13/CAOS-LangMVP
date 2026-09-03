"""Cross-format publication goldens and semantic parity (Task 10; Phase 4 items
9–10 and 13; ETR PUB-012, PUB-018, PUB-026).

Six states — normal, dense, long-text, multilingual, held (changes requested)
and filed — each frozen through the worker path and exported to Markdown, PDF
and XLSX. For every state the test proves that the three exports and the
browser's publication document carry the same headings, table values, source
citations, origin labels, limitations, model identity and opinion, and pins the
identifier-normalised Markdown bytes, XLSX cell dump and PDF text structure to
approved goldens under caos/tests/fixtures/deliverables/publication/.

Regenerate the goldens deliberately with CAOS_REGENERATE_GOLDENS=1 after
inspecting every affected PDF page and XLSX sheet (the report records that
inspection); a silent drift is a failure, never an auto-update.
"""

from __future__ import annotations

import io
import json
import os
import re
from pathlib import Path

import pytest

from test_deliverables_spec import (
    add_approver,
    draft_request,
    file_request,
    freeze_now,
    make_service,
    optional_block,
    required_blocks,
    revision_selection,
    seed_model,
    seed_ready_case,
)

GOLDEN_DIR = Path(__file__).resolve().parents[1] / "fixtures" / "deliverables" / "publication"
REGENERATE = os.environ.get("CAOS_REGENERATE_GOLDENS") == "1"
STATES = ("normal", "dense", "long_text", "multilingual", "held", "filed")

_ID = re.compile(r"\b(?:src|snap|dlv|dth|opn|frz|rcpt|dlbld|dlrevn|set|case|run|aud|intake|mdl)-[0-9a-f]{6,}")
_SHA = re.compile(r"\b[0-9a-f]{64}\b")
_TS = re.compile(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:\+00:00|Z)")
_BUILD = re.compile(r"deploy-v-[0-9a-f]+")


def normalise(text: str) -> str:
    text = _SHA.sub("<sha256>", text)
    text = _ID.sub("<id>", text)
    text = _TS.sub("<timestamp>", text)
    return _BUILD.sub("<methodology-build>", text)


MULTILINGUAL = (
    "Judgment: the issuer’s liquidity is adequate. 债务重组条款仍在谈判中。 Ο δανειστής διατηρεί επαρκή ρευστότητα. "
    "Türkçe: nakit akışı yeterli görünüyor. Naïve façade — “curly” quotes, ﬁ/ﬂ ligature pairs, and a right-to-left mark‏ follow."
)


def _outputs(years: int, metrics: int) -> dict:
    names = ["total_leverage", "accessible_liquidity", "net_debt", "interest_cover", "fcf", "revenue", "adjusted_ebitda_calc", "capex"][:metrics]
    return {
        case_name: {
            f"FY{2024 + index}": {metric: round(3.1 + index * 0.13 + offset * 0.5 + (0 if case_name == "BASE" else 0.7), 2) for offset, metric in enumerate(names)}
            for index in range(years)
        }
        for case_name in ("BASE", "DOWNSIDE")
    }


def _build(service, store, state: str):
    case, source, _authority = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    years, metrics = (7, 8) if state == "dense" else (2, 2)
    model = seed_model(service, case, outputs=_outputs(years, metrics))
    if state == "long_text":
        sentence = "We judge that the committee should read the complete rationale because every clause of it carries weight. "
        narrative = (sentence * 190)[:19_800].rstrip()
    elif state == "multilingual":
        narrative = MULTILINGUAL
    elif state == "dense":
        narrative = "In our judgment coverage stays adequate across the seven forecast years under both cases."
    else:
        narrative = "In our judgment the credit remains adequately covered through the forecast horizon."
    blocks = required_blocks(template, source, narrative_text=narrative)
    extras = [optional_block(template, "GENERATED_METRIC"), optional_block(template, "MODEL_APPENDIX")]
    if state == "dense":
        extras = [
            optional_block(template, "GENERATED_METRIC", metric_ids=["total_leverage", "accessible_liquidity", "net_debt"]),
            optional_block(template, "GENERATED_METRIC", n=2, metric_ids=["interest_cover", "fcf"]),
            optional_block(template, "MODEL_APPENDIX"),
            optional_block(template, "LIMITATIONS", text="Covenant definitions unavailable; capex guidance not disclosed.", citations=[]),
        ]
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, blocks=blocks + extras, model_selection=revision_selection(model)),
        actor="analyst",
    )
    frozen = freeze_now(service, case["id"], revision)
    receipt = None
    if state == "held":
        from caos.contracts import RequestDeliverableChangesRequest

        add_approver(store, case)
        outcome = service.request_changes(case["id"], frozen["deliverable_id"], RequestDeliverableChangesRequest(
            preview_digest=frozen["preview_digest"], input_fingerprint=frozen["input_fingerprint"],
            comment="Clarify the downside bridge before filing.",
        ), actor="approver-user")
        frozen = outcome["frozen"]
    if state == "filed":
        add_approver(store, case)
        frozen = service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user")
        receipt = service.filing_receipt(case["id"], frozen["deliverable_id"])
    exports = {fmt: service.export(frozen["deliverable_id"], fmt)[0] for fmt in ("md", "pdf", "xlsx")}
    return case, source, model, frozen, exports, receipt


def _pdf_text(content: bytes) -> str:
    from pypdf import PdfReader

    return "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(content)).pages)


def _pdf_pages(content: bytes) -> int:
    from pypdf import PdfReader

    return len(PdfReader(io.BytesIO(content)).pages)


def _xlsx_dump(content: bytes) -> dict[str, list[list]]:
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(content))
    return {sheet.title: [[cell for cell in row] for row in sheet.iter_rows(values_only=True)] for sheet in book.worksheets}


def _xlsx_text(dump: dict[str, list[list]]) -> str:
    return "\n".join(str(cell) for rows in dump.values() for row in rows for cell in row if cell is not None)


def _publication_facts(payload: dict) -> dict:
    """Everything the four formats must agree on, read from the one frozen payload."""
    from caos.publishing.renderers import _section_rows, _walk_sections

    publication = payload["publication"]
    headings, values, origins = [], set(), set()
    for page in publication["pages"]:
        for section, _depth in _walk_sections(page["sections"]):
            headings.append(section["title"])
            origins.add("Analyst judgment" if section["origin"]["kind"] == "ANALYST" else f"Locked · {section['origin']['kind'].lower()}")
            if section["kind"] in {"table", "chart", "profile"}:
                for row in _section_rows(section)[1 if section["kind"] != "profile" else 0:]:
                    values.update(str(cell) for cell in row if str(cell).strip())
    return {
        "headings": headings,
        "values": values,
        "origins": origins,
        "citations": {row["source_id"] for row in payload["evidence"]},
        "limitations": publication["disclosures"]["limitations"],
        "model_identity": publication["masthead"]["model_identity"],
        "opinion": payload["opinion"]["opinion"],
        "opinion_owner": payload["opinion"]["signed_by"],
        "approval_state": publication["masthead"]["approval_state"],
    }


def _subsequence(needles: list[str], haystack: list[str]) -> bool:
    position = 0
    for needle in needles:
        try:
            position = haystack.index(needle, position) + 1
        except ValueError:
            return False
    return True


@pytest.fixture()
def service(tmp_path, store):
    return make_service(store, tmp_path / "deliverable-vault")


@pytest.mark.parametrize("state", STATES)
def test_every_format_carries_the_same_facts_and_matches_its_golden(service, store, state, tmp_path):
    case, source, model, frozen, exports, receipt = _build(service, store, state)
    payload = frozen["payload"]
    facts = _publication_facts(payload)
    markdown = exports["md"].decode("utf-8")
    pdf_text = _pdf_text(exports["pdf"])
    dump = _xlsx_dump(exports["xlsx"])
    xlsx_text = _xlsx_text(dump)

    # Headings in document order in every format (Markdown H3+, XLSX Report
    # sheet Section column, PDF text with headings on their own line).
    origin_suffix = re.compile(r" · (?:Analyst judgment|Locked · ).*$")
    markdown_headings = [origin_suffix.sub("", line.lstrip("#").strip()) for line in markdown.splitlines() if line.startswith("### ") or line.startswith("#### ")]
    assert _subsequence(facts["headings"], markdown_headings), (facts["headings"], markdown_headings)
    report_rows = dump["Report"][1:]
    xlsx_headings = list(dict.fromkeys(row[1] for row in report_rows if row[1]))
    assert _subsequence(facts["headings"], xlsx_headings), xlsx_headings
    pdf_lines = [line.strip() for line in pdf_text.splitlines()]
    for heading in facts["headings"]:
        assert any(line.startswith(heading) for line in pdf_lines), f"{heading!r} missing from the PDF"

    # Every table, chart and profile value; every citation; every origin label;
    # limitations; model identity; the opinion and its owner; the approval state.
    for value in facts["values"]:
        assert value in markdown and value in xlsx_text, value
        # PDF wraps long cells inside their column; compare on whitespace-free text.
        assert re.sub(r"\s+", "", value) in re.sub(r"\s+", "", pdf_text), value
    for source_id in facts["citations"]:
        assert source_id in markdown and source_id in xlsx_text and source_id in pdf_text
    for origin in facts["origins"]:
        assert origin in markdown and origin in pdf_text
        assert origin.split(" · ")[0] in xlsx_text
    for limitation in facts["limitations"]:
        assert limitation in markdown and limitation in xlsx_text
        assert re.sub(r"\s+", "", limitation) in re.sub(r"\s+", "", pdf_text)
    assert facts["model_identity"] in markdown and facts["model_identity"] in xlsx_text
    assert re.sub(r"\s+", "", facts["model_identity"]) in re.sub(r"\s+", "", pdf_text)
    assert facts["opinion"] in markdown and facts["opinion"] in xlsx_text
    assert re.sub(r"\s+", "", facts["opinion"]) in re.sub(r"\s+", "", pdf_text)
    assert facts["opinion_owner"] in markdown and facts["opinion_owner"] in xlsx_text and facts["opinion_owner"] in pdf_text
    assert facts["approval_state"] == "PENDING APPROVAL"
    assert markdown.count("PENDING APPROVAL") >= 2 and "PENDING APPROVAL" in pdf_text and "PENDING APPROVAL" in xlsx_text
    assert "approver-user" not in markdown and "approver-user" not in pdf_text and "approver-user" not in xlsx_text, \
        "the approver is never in the bytes, in any state"
    if state == "multilingual":
        for fragment in ("债务重组条款仍在谈判中", "Ο δανειστής διατηρεί επαρκή ρευστότητα", "Türkçe: nakit akışı", "Naïve façade", "“curly”"):
            assert fragment in markdown and fragment in xlsx_text
            assert re.sub(r"\s+", "", fragment) in re.sub(r"\s+", "", pdf_text), f"{fragment!r} was silently replaced in the PDF"
        assert "�" not in pdf_text and "�" not in markdown
    if state == "long_text":
        body = next(block["text"] for block in payload["content"]["blocks"] if block["kind"] == "NARRATIVE")
        assert body in markdown and body in xlsx_text, "no fixed-line truncation of a long narrative"
        assert re.sub(r"\s+", "", body) in re.sub(r"\s+", "", pdf_text), "the PDF carries the whole narrative"
    if state == "held":
        assert frozen["status"] == "CHANGES_REQUESTED"
    if state == "filed":
        assert frozen["status"] == "FILED" and receipt["approved_by"] == "approver-user"
        assert receipt["exports"] == {fmt: meta["sha256"] for fmt, meta in frozen["exports"].items()}
    assert _pdf_pages(exports["pdf"]) >= (3 if state in {"dense", "long_text"} else 1)

    # Goldens: identifier-normalised Markdown bytes, XLSX cell dump, PDF structure.
    observed = {
        "md": normalise(markdown),
        "xlsx": normalise(json.dumps(dump, ensure_ascii=False, sort_keys=True, default=str, indent=1)),
        "pdf": json.dumps({
            "pages": _pdf_pages(exports["pdf"]),
            "headings": facts["headings"],
            "first_lines": [normalise(line) for line in pdf_lines[:6]],
        }, ensure_ascii=False, indent=1),
    }
    GOLDEN_DIR.mkdir(parents=True, exist_ok=True)
    for fmt, text in observed.items():
        path = GOLDEN_DIR / f"{state}.{fmt}.golden"
        if REGENERATE or not path.exists():
            path.write_text(text, encoding="utf-8")
        expected = path.read_text(encoding="utf-8")
        assert text == expected, f"{state}.{fmt} drifted from its approved golden; inspect the render and regenerate deliberately"


def test_held_and_filed_bytes_are_the_frozen_bytes(service, store):
    """Neither a change request nor filing rerenders: the exports a frozen
    record carries are the exports its held and filed states serve."""
    case, source, model, frozen, exports, _receipt = _build(service, store, "normal")
    add_approver(store, case)
    from caos.contracts import RequestDeliverableChangesRequest

    held = service.request_changes(case["id"], frozen["deliverable_id"], RequestDeliverableChangesRequest(
        preview_digest=frozen["preview_digest"], input_fingerprint=frozen["input_fingerprint"], comment="Hold for review.",
    ), actor="approver-user")["frozen"]
    assert held["status"] == "CHANGES_REQUESTED" and held["exports"] == frozen["exports"]
    for fmt, content in exports.items():
        assert service.export(frozen["deliverable_id"], fmt)[0] == content
    second_case, _s, _m, second, second_exports, receipt = _build(service, store, "filed")
    assert second["exports"] == service.frozen_record(second_case["id"], second["deliverable_id"])["exports"]
    for fmt, content in second_exports.items():
        assert service.export(second["deliverable_id"], fmt)[0] == content
    assert receipt["signed_by"] == "analyst" and receipt["approved_by"] == "approver-user"
