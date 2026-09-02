"""Deliverables specification (invariant 5; DECISIONS §2, §10.5/10.7, §12.21-23).

All tests must FAIL today: every unbuilt import (caos.deliverables.*, caos.api via the
client/app fixtures) happens inside test bodies, fixtures, or helpers called from them.
A passing test in this file is a defect.

Sources: TEST_INVENTORY.md contractual rows for test_deliverables.py (21),
test_deliverable_exports.py (8), the two deliverable rows of test_ledger_contracts.py,
and the eight approval-gate rows of test_cp_dr_planning.py re-hosted onto the deliverable
filing interrupt per DECISIONS §11.9. Full row map at the end of the file.

Specification decisions this file pins (beyond the porting briefs):
- DeliverableService(store=..., vault_dir=..., engine=...) with workspace / save_draft /
  freeze / approve_filing / request_changes / export / templates. Freeze parks a
  filing-gate thread (freeze -> render -> interrupt() -> file | request-changes, §2).
- workspace(case_id, pathway) -> {"template", "draft" (current revision | None), "frozen" (list)}.
- Revision record: {draft_id, revision_id, version, digest, content}; frozen record:
  {deliverable_id, thread_id, status, preview_digest, input_fingerprint, build_id,
  payload, exports: {fmt: {sha256, size}}}.
- Typed errors are matched by their code substring in str(exc) (same convention as
  test_runs_spec.py); audit actions: deliverable.draft.saved / .frozen / .filed /
  .changes_requested.
- Required template slots are limited to HEADING / NARRATIVE / EVIDENCE_REGISTER /
  LIMITATIONS; model-dependent kinds are optional-only; optional slot ids are
  f"{slot_stem}-{n}". Every template has at least two required slots.
- service.export works from FROZEN (the gate's own render); the HTTP export route serves
  FILED only (pre-filing 409) and never re-renders (§12.23).
- The gate-level canonical approval hash is "sha256:" + preview_digest and matches
  ^sha256:[0-9a-f]{64}$ (re-hosted planning-gate contract); the HTTP request models keep
  their bare-64-hex fields.
- For-tests hooks invented on the future service (clearly suffixed): _for_tests hooks for
  seeding accepted authority / signed revisions / application builds, scenario previews,
  tampering frozen payloads and stored exports, deleting exports, terminating filing
  threads, reading audit events and thread state, and a renderer_for_tests constructor
  kwarg plus resume_filing_with_hash_for_tests.

Merged/adapted rows are documented in the ROW MAPPING block at the end.
"""

from __future__ import annotations

import copy
import hashlib
import io
import json
import re
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pytest

from spec_helpers import seed_case_with_source

SIX_PATHWAYS = (
    "FULL_CREDIT",
    "EARNINGS_UPDATE",
    "COVENANT_REFINANCING",
    "RELATIVE_VALUE",
    "DISTRESSED_RESTRUCTURING",
    "DEEP_RESEARCH",
)

ANALYST_H = {"x-caos-role": "ANALYST", "x-forwarded-user": "analyst"}
APPROVER_H = {"x-caos-role": "APPROVER", "x-forwarded-user": "approver-user"}
GOLDEN_DIR = Path(__file__).resolve().parents[1] / "fixtures" / "deliverables" / "golden"

SHOCK = {"assumption_id": "revenue_growth", "case": "DOWNSIDE", "period_id": "FY2026", "value": -0.05}
_DEFAULT_MODEL_SELECTION = object()
_TEST_MODEL_SELECTION = "_default_model_selection_for_tests"


# --- helpers (unbuilt imports stay inside; called only from test bodies) ----------


def make_service(store, vault_dir, **overrides):
    from caos.deliverables.service import DeliverableService

    return DeliverableService(store=store, vault_dir=vault_dir, engine=overrides.pop("engine", None), **overrides)


@pytest.fixture()
def service(tmp_path, store):
    return make_service(store, tmp_path / "deliverable-vault")


def _never_render(payload, fmt):
    raise AssertionError("filing never re-renders (§12.23)")


def required_blocks(template, source, *, narrative_text="Leverage is manageable."):
    """Blocks satisfying every required template block, in template order.

    Spec: required blocks use only non-model kinds — anything else is a template defect.
    """
    blocks = []
    for item in template["blocks"]:
        base = {"block_id": item["block_id"], "slot_id": item["slot_id"]}
        kind = item["kind"]
        if kind == "NARRATIVE":
            blocks.append({**base, "kind": "NARRATIVE", "text": narrative_text, "content_mode": "ANALYST_JUDGMENT", "citations": []})
        elif kind == "EVIDENCE_REGISTER":
            blocks.append({**base, "kind": "EVIDENCE_REGISTER", "citations": [
                {"source_id": source["id"], "block_ids": ["b00001"], "claim": "Pinned evidence line supports the opinion."},
            ]})
        elif kind == "LIMITATIONS":
            blocks.append({**base, "kind": "LIMITATIONS", "text": "Scope-limited review.", "citations": []})
        else:
            raise AssertionError(f"required blocks must be non-model kinds, got {kind}")
    return blocks


def draft_request(
    template,
    source=None,
    *,
    blocks=None,
    expected_version=0,
    model_selection=_DEFAULT_MODEL_SELECTION,
    extra_blocks=(),
):
    from caos.contracts import DeliverableDraftRequest

    if blocks is None:
        blocks = required_blocks(template, source) + list(extra_blocks)
    if model_selection is _DEFAULT_MODEL_SELECTION:
        model_selection = template.get(_TEST_MODEL_SELECTION)
    return DeliverableDraftRequest(
        expected_version=expected_version,
        template_id=template["template_id"],
        template_version=template["template_version"],
        model_selection=model_selection,
        blocks=blocks,
    )


def freeze_request(revision):
    from caos.contracts import FreezeDeliverableRequest

    return FreezeDeliverableRequest(draft_id=revision["draft_id"], draft_version=revision["version"], draft_digest=revision["digest"])


def file_request(frozen, **overrides):
    from caos.contracts import FileDeliverableRequest

    fields = {"preview_digest": frozen["preview_digest"], "input_fingerprint": frozen["input_fingerprint"]}
    fields.update(overrides)
    return FileDeliverableRequest(**fields)


def optional_slot(template, kind, n=1):
    policy = next(p for p in template["optional_blocks"] if p["kind"] == kind)
    return f"{policy['slot_stem']}.{n:02d}"


def optional_block(template, kind, n=1, **fields):
    slot = optional_slot(template, kind, n)
    base = {"block_id": f"blk-{slot}", "slot_id": slot, "kind": kind}
    defaults = {
        "GENERATED_METRIC": {"metric_ids": ["total_leverage"]},
        "GENERATED_TABLE": {"table_id": "debt_schedule", "field_ids": ["instrument", "amount"]},
        "GENERATED_CHART": {"recipe": {"chart_kind": "line", "fields": ["total_leverage"]}},
        "MODEL_APPENDIX": {},
        "LIMITATIONS": {"text": "Covenant definitions unavailable.", "citations": []},
    }
    base.update(defaults[kind])
    base.update(fields)
    return base


def scenario_exhibit_block(template, preview, n=1, **overrides):
    slot = optional_slot(template, "SCENARIO_EXHIBIT", n)
    block = {
        "block_id": f"blk-{slot}",
        "slot_id": slot,
        "kind": "SCENARIO_EXHIBIT",
        "title": "Downside revenue shock",
        "shocks": preview["shocks"],
        "scenario": preview["scenario"],
        "scenario_digest": preview["scenario_digest"],
    }
    block.update(overrides)
    return block


def revision_selection(model):
    return {"kind": "ANALYST_REVISION", "build_id": model["build_id"], "revision_id": model["revision_id"]}


def seed_ready_case(service, store):
    """Case + source + the pinned upstream accepted authority freeze requires."""
    case, source = seed_case_with_source(store)
    authority = service.seed_accepted_authority_for_tests(case["id"])
    return case, source, authority


def seed_model(service, case, *, outputs=None):
    """Signed analyst revision + its READY application build; returns identity + payloads:
    {build_id, revision_id, outputs, assumptions, build_payload, build_qa}."""
    return service.seed_signed_revision_for_tests(case["id"], outputs=outputs or {"total_leverage": 4.2})


def bind_default_model_for_tests(service, case, template):
    model = seed_model(service, case)
    template[_TEST_MODEL_SELECTION] = revision_selection(model)
    return model


def save_min_draft(service, store, pathway="FULL_CREDIT", **kwargs):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()[pathway]
    if template["model_requirement"] == "REQUIRED":
        bind_default_model_for_tests(service, case, template)
    revision = service.save_draft(case["id"], pathway, draft_request(template, source, **kwargs), actor="analyst")
    return case, source, template, revision


def freeze_min(service, store, pathway="FULL_CREDIT"):
    case, source, template, revision = save_min_draft(service, store, pathway)
    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    return case, source, revision, frozen


def add_approver(store, case):
    store.add_member(case["id"], "analyst", "approver-user", "APPROVER", actor_role="ADMIN")


def frozen_deliverable(service, store, pathway="FULL_CREDIT"):
    case, source, revision, frozen = freeze_min(service, store, pathway)
    add_approver(store, case)
    return case, revision, frozen


def http_seed(settings, store, pathway="FULL_CREDIT"):
    """Service twin over the app's own store + vault so seeding hooks feed HTTP routes."""
    svc = make_service(store, settings.storage_dir)
    case, source = seed_case_with_source(store)
    svc.seed_accepted_authority_for_tests(case["id"])
    template = svc.templates()[pathway]
    if template["model_requirement"] == "REQUIRED":
        bind_default_model_for_tests(svc, case, template)
    return svc, case, source, template


def ingest_second_source(store, case_id, body=b"late-arriving evidence"):
    return store.ingest({
        "case_id": case_id,
        "filename": "late.txt",
        "media_type": "text/plain",
        "bytes": len(body),
        "sha256": hashlib.sha256(body).hexdigest(),
        "vault_path": None,
        "blocks": [{"block_id": "b00002", "locator": {"line": 1}, "text": body.decode(), "extractor_version": "builtin-v1", "confidence": "MEDIUM", "untrusted_data": True}],
        "withdrawn": False,
    }, "analyst")


# --- template registry ------------------------------------------------------------


def test_template_registry_serves_six_pathways_with_stable_identity_and_evidence_register(service):
    templates = service.templates()
    assert set(templates) == set(SIX_PATHWAYS)
    assert service.templates() == templates, "registry identity is stable across reads"
    for pathway, template in templates.items():
        assert template["template_id"] and template["template_version"] and template["title"], pathway
        assert template["model_requirement"] in {"REQUIRED", "OPTIONAL"}, pathway
        block_ids = [b["block_id"] for b in template["blocks"]]
        assert len(block_ids) == len(set(block_ids)), f"{pathway}: block ids must be unique"
        assert all(b["title"] and b["required"] for b in template["blocks"]), f"{pathway}: required blocks are titled"
        assert [b for b in template["blocks"] if b["kind"] == "EVIDENCE_REGISTER"], f"{pathway}: mandatory evidence register"


def test_template_owns_optional_block_kind_stem_cap_order_and_model_dependence(service):
    template = service.templates()["RELATIVE_VALUE"]
    policy = template["optional_blocks"]
    for entry in policy:
        assert {"kind", "slot_stem", "max_items", "order", "model_dependent"} <= set(entry)
    by_kind = {entry["kind"]: entry for entry in policy}
    for kind in ("GENERATED_METRIC", "GENERATED_TABLE", "GENERATED_CHART", "SCENARIO_EXHIBIT", "MODEL_APPENDIX"):
        assert by_kind[kind]["model_dependent"] is True, kind
    assert by_kind["LIMITATIONS"]["model_dependent"] is False
    assert set(template["allowed_appendices"]) == set(by_kind), "allowed appendices derive from the optional policy"


# --- draft save + append-only revisions -------------------------------------------


def test_saved_revision_digest_is_digest_of_content_and_order_violation_appends_nothing(service, store):
    from caos.contracts import digest

    case, source, template, revision = save_min_draft(service, store)
    assert revision["version"] == 1
    assert revision["digest"] == digest(revision["content"]), "revision is content-addressed"
    blocks = required_blocks(template, source)
    assert len(blocks) >= 2, "spec premise: every template has at least two required slots"
    with pytest.raises(Exception, match="DELIVERABLE_TEMPLATE_ORDER_INVALID"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, blocks=list(reversed(blocks)), expected_version=1), actor="analyst")
    assert service.workspace(case["id"], "FULL_CREDIT")["draft"]["version"] == 1, "nothing appended on order violation"


def test_full_credit_draft_composes_one_governed_decision_document(service, store):
    case, source, _ = seed_ready_case(service, store)
    model = seed_model(service, case, outputs={
        "BASE": {"FY2027": {"total_leverage": 3.8, "accessible_liquidity": 210.0}},
        "DOWNSIDE": {"FY2027": {"total_leverage": 5.1, "accessible_liquidity": 95.0}},
    })
    template = service.templates()["FULL_CREDIT"]

    revision = service.save_draft(
        case["id"],
        "FULL_CREDIT",
        draft_request(template, source, model_selection=revision_selection(model)),
        actor="analyst",
    )

    content = revision["content"]
    sections = content["document_sections"]
    assert content["document_schema_version"] == "caos.deliverable.document.v1"
    assert [section["section_id"] for section in sections] == [
        "credit_snapshot",
        "recommendation",
        "thesis_variant",
        "business_industry",
        "capital_structure",
        "base_downside_model",
        "liquidity_covenants",
        "risks_catalysts_falsifiers",
        "monitoring",
        "evidence_register",
    ]
    assert sections[0]["origin"]["kind"] == "ARTIFACT"
    assert sections[1]["editable"] is True
    assert sections[5]["editable"] is False
    model_table = sections[5]["items"][1][0]
    assert model_table["origin"]["kind"] == "MODEL"
    assert ["BASE / FY2027 / total_leverage", "3.8"] in model_table["rows"]


def _document_structure(sections):
    def fields(section):
        if section["kind"] == "profile":
            return [row["label"] for row in section["rows"]]
        if section["kind"] == "table":
            return section["columns"]
        if section["kind"] == "chart":
            return section["accessible_columns"]
        if section["kind"] == "columns":
            labels = []
            for column in section["items"]:
                for item in column:
                    labels.append(item["title"])
                    labels.extend(f"{item['title']}: {label}" for label in fields(item))
            return labels
        return []

    return [{
        "section_id": section["section_id"],
        "kind": section["kind"],
        "title": section["title"],
        "page": section["page"],
        "editable": section["editable"],
        "origin": section["origin"]["kind"],
        "fields": fields(section),
    } for section in sections]


@pytest.mark.parametrize("pathway", SIX_PATHWAYS)
def test_document_recipe_matches_approved_structure(service, store, pathway):
    case, source, _ = seed_ready_case(service, store)
    model = seed_model(service, case, outputs={
        "BASE": {"FY2027": {"total_leverage": 3.8, "accessible_liquidity": 210.0}},
        "DOWNSIDE": {"FY2027": {"total_leverage": 5.1, "accessible_liquidity": 95.0}},
    })
    template = service.templates()[pathway]
    revision = service.save_draft(
        case["id"], pathway,
        draft_request(template, source, model_selection=revision_selection(model)),
        actor="analyst",
    )

    expected = json.loads((GOLDEN_DIR / f"{pathway.lower()}.json").read_text())
    assert _document_structure(revision["content"]["document_sections"]) == expected


def test_optional_generated_blocks_become_locked_document_appendices(service, store):
    case, source, _ = seed_ready_case(service, store)
    model = seed_model(service, case, outputs={
        "BASE": {"FY2027": {"total_leverage": 3.8}},
        "DOWNSIDE": {"FY2027": {"total_leverage": 5.1}},
        "debt_schedule": [
            {"instrument": "RCF", "amount": 120.0, "maturity": "FY2028", "margin": "S+350"},
        ],
    })
    template = service.templates()["FULL_CREDIT"]
    extras = [
        optional_block(template, "GENERATED_METRIC"),
        optional_block(template, "GENERATED_TABLE"),
        optional_block(template, "GENERATED_CHART"),
        optional_block(template, "MODEL_APPENDIX"),
        optional_block(template, "LIMITATIONS"),
    ]
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, source, model_selection=revision_selection(model), extra_blocks=extras),
        actor="analyst",
    )

    appendices = revision["content"]["document_sections"][-6:]
    assert [(section["kind"], section["title"], section["editable"], section["origin"]["kind"]) for section in appendices] == [
        ("table", "Generated Metrics · 01", False, "MODEL"),
        ("table", "Debt Schedule · 01", False, "MODEL"),
        ("chart", "Generated Chart · 01", False, "MODEL"),
        ("table", "Model Appendix", False, "MODEL"),
        ("text", "Limitations · 01", True, "ANALYST"),
        ("table", "Evidence Register", False, "ARTIFACT"),
    ]
    assert ["total_leverage / BASE / FY2027", "3.8"] in appendices[0]["rows"]
    assert appendices[1]["rows"] == [["RCF", "120"]]
    assert ["total_leverage / DOWNSIDE / FY2027", "5.1"] in appendices[2]["accessible_rows"]
    assert appendices[4]["body"] == "Covenant definitions unavailable."


def test_annual_model_generated_table_uses_only_selected_server_outputs(service, store):
    case, source, _ = seed_ready_case(service, store)
    model = seed_model(service, case, outputs={
        "BASE": {"BASE::FY2027": {"revenue": 517.4, "fcf": 53.6, "total_leverage": 1.7}},
        "DOWNSIDE": {"DOWNSIDE::FY2027": {"revenue": 423.9, "fcf": 10.3, "total_leverage": 3.2}},
    })
    template = service.templates()["FULL_CREDIT"]

    revision = service.save_draft(
        case["id"],
        "FULL_CREDIT",
        draft_request(
            template,
            source,
            model_selection=revision_selection(model),
            extra_blocks=[optional_block(
                template,
                "GENERATED_TABLE",
                table_id="annual_model",
                field_ids=["revenue", "total_leverage"],
            )],
        ),
        actor="analyst",
    )

    table = revision["content"]["document_sections"][-2]
    assert table["columns"] == ["Case", "Period", "Revenue", "Total Leverage"]
    assert table["rows"] == [
        ["BASE", "FY2027", "517.4", "1.7"],
        ["DOWNSIDE", "FY2027", "423.9", "3.2"],
    ]
    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    assert frozen["payload"]["content"]["document_sections"] == revision["content"]["document_sections"]


def test_pathway_calculation_rows_paginate_without_hiding_late_outputs(service, store):
    from caos.deliverables.document import compose_document

    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["DISTRESSED_RESTRUCTURING"]
    sections = compose_document(
        pathway="DISTRESSED_RESTRUCTURING",
        template=template,
        blocks=required_blocks(template, source),
        artifacts={"__authority__": {"id": "snap-pathway-pagination"}},
        model={
            "build_id": "bld-pathway-pagination",
            "outputs": {},
            "pathway_effects": [{
                "calculations": [
                    {
                        "calculator_id": "recovery_waterfall",
                        "canonical_output": {
                            "claims": list(range(501)),
                            "recovery_pct": 0.25,
                        },
                    },
                    {
                        "calculator_id": "liquidity_runway",
                        "canonical_output": {"funding_gap": 150},
                    },
                ],
            }],
        },
    )

    scenario = next(section for section in sections if section["section_id"] == "base_downside_and_scenario_exhibits")
    tables = scenario["items"][1][1:]
    rows = [row for table in tables for row in table["rows"]]
    assert len(rows) == 503
    assert all(len(table["rows"]) <= 500 for table in tables)
    assert ["recovery_waterfall / recovery_pct", "0.25"] in rows
    assert ["liquidity_runway / funding_gap", "150"] in rows
    assert len({table["section_id"] for table in tables}) == len(tables)


def test_optional_block_cannot_collide_with_a_canonical_document_section(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    bind_default_model_for_tests(service, case, template)
    collision = optional_block(template, "LIMITATIONS", block_id="credit_snapshot")

    with pytest.raises(Exception, match="DELIVERABLE_SECTION_ID_DUPLICATE"):
        service.save_draft(
            case["id"],
            "FULL_CREDIT",
            draft_request(template, source, extra_blocks=[collision]),
            actor="analyst",
        )
    assert service.workspace(case["id"], "FULL_CREDIT")["draft"] is None


@pytest.mark.parametrize("pathway", ["RELATIVE_VALUE", "DEEP_RESEARCH"])
def test_model_optional_pathways_compose_without_model_authority(service, store, pathway):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()[pathway]

    revision = service.save_draft(
        case["id"], pathway, draft_request(template, source), actor="analyst",
    )

    assert revision["content"]["model_identity"] is None
    assert all(section["origin"]["kind"] != "MODEL" for section in revision["content"]["document_sections"])


async def test_no_model_relative_value_freeze_pins_accepted_run_methodology(
    settings, store, engine, tmp_path,
):
    service = make_service(store, tmp_path / "deliverable-vault", engine=engine)
    case, source = seed_case_with_source(store)
    run = await engine.run_scripted_for_tests(case["id"], pathway="RELATIVE_VALUE")
    await engine.accept(run["id"], actor="analyst")
    template = service.templates()["RELATIVE_VALUE"]
    revision = service.save_draft(
        case["id"], "RELATIVE_VALUE", draft_request(template, source), actor="analyst",
    )

    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")

    assert frozen["payload"]["model"] is None
    assert frozen["payload"]["authority"]["build_id"] == "unbuilt"
    assert frozen["payload"]["methodology"]["build_id"] == run["plan"]["build_id"]
    assert frozen["payload"]["methodology"]["build_id"] == engine.bundle.build_id


def test_draft_request_rejects_client_supplied_document_sections(client, settings, store):
    _svc, case, source, template = http_seed(settings, store)
    body = draft_request(template, source).model_dump(mode="json")
    body["document_sections"] = [{"section_id": "forged", "body": "client content"}]

    response = client.put(
        f"/api/cases/{case['id']}/deliverables/FULL_CREDIT/draft",
        json=body,
        headers=ANALYST_H,
    )

    assert response.status_code == 422
    assert _svc.workspace(case["id"], "FULL_CREDIT")["draft"] is None


def test_freeze_recomposes_and_rejects_document_authority_drift(service, store):
    case, _source, revision, _frozen = freeze_min(service, store)
    service.supersede_accepted_authority_for_tests(case["id"])

    with pytest.raises(Exception, match="DELIVERABLE_COMPOSITION_MISMATCH"):
        service.freeze(case["id"], freeze_request(revision), actor="analyst")

    assert len(service.workspace(case["id"], "FULL_CREDIT")["frozen"]) == 1


def test_composition_verifies_accepted_artifact_bytes_before_saving(settings, store, engine, tmp_path):
    import asyncio

    service = make_service(store, tmp_path / "deliverable-vault", engine=engine)
    case, source = seed_case_with_source(store)

    async def accepted_run():
        run = await engine.run_scripted_for_tests(case["id"])
        await engine.accept(run["id"], actor="analyst")
        return run

    run = asyncio.run(accepted_run())
    template = service.templates()["FULL_CREDIT"]
    bind_default_model_for_tests(service, case, template)
    first = service.save_draft(
        case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst",
    )
    profile = first["content"]["document_sections"][0]["items"][0][0]
    assert "CP-1" in {row["label"] for row in profile["rows"]}

    engine.forge_node_artifact_for_tests(run["id"], module_id="CP-1", digest="0" * 64)
    with pytest.raises(Exception, match="DELIVERABLE_COMPOSITION_REQUIRED_VALUE_MISSING:artifact.CP-1"):
        service.save_draft(
            case["id"],
            "FULL_CREDIT",
            draft_request(template, source, expected_version=1),
            actor="analyst",
        )
    assert service.workspace(case["id"], "FULL_CREDIT")["draft"]["version"] == 1


def test_stale_expected_version_is_atomic_cas_conflict_carrying_current_head(service, store):
    case, source, template, revision = save_min_draft(service, store)

    def racing_save(text):
        blocks = required_blocks(template, source, narrative_text=text)
        try:
            return ("ok", service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, blocks=blocks, expected_version=1), actor="analyst"))
        except Exception as exc:  # noqa: BLE001 — the loser's typed conflict
            return ("conflict", exc)

    with ThreadPoolExecutor(max_workers=2) as pool:
        outcomes = list(pool.map(racing_save, ["Racer one narrative.", "Racer two narrative."]))
    assert sorted(kind for kind, _ in outcomes) == ["conflict", "ok"], "exactly one CAS winner"
    conflict = next(value for kind, value in outcomes if kind == "conflict")
    assert "VERSION_CONFLICT" in str(conflict)
    assert conflict.current["version"] == 2, "typed conflict carries the current head for rebase"
    saved_events = [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.draft.saved"]
    assert len(saved_events) == 2, "one audit entry per appended revision, none for the loser"


def test_restore_appends_new_revision_and_history_stays_byte_identical(service, store):
    case, source, template, r1 = save_min_draft(service, store)
    service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, blocks=required_blocks(template, source, narrative_text="Revised view."), expected_version=1),
        actor="analyst",
    )
    before = copy.deepcopy(service.revision_history(case["id"], "FULL_CREDIT"))
    r3 = service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, blocks=r1["content"]["blocks"], expected_version=2), actor="analyst")
    assert r3["version"] == 3
    assert r3["content"]["blocks"] == r1["content"]["blocks"], "restore carries the old content verbatim"
    after = service.revision_history(case["id"], "FULL_CREDIT")
    assert [r["version"] for r in after] == [1, 2, 3], "versions strictly increase; no rewrites"
    assert after[:2] == before, "prior revisions remain byte-identical"


def test_stored_revisions_are_isolated_from_later_caller_mutation(service, store):
    case, source, template, r1 = save_min_draft(service, store)
    original_digest = r1["digest"]
    r1["content"]["blocks"][0]["text"] = "TAMPERED AFTER APPEND"
    fetched = service.revision_by_id(case["id"], r1["revision_id"])
    assert fetched["content"]["blocks"][0]["text"] != "TAMPERED AFTER APPEND", "store serializes/copies on write"
    assert fetched["digest"] == original_digest


def test_http_stale_draft_put_returns_409_with_current_revision_and_by_id_read_round_trips(client, settings, store):
    svc, case, source, template = http_seed(settings, store)
    url = f"/api/cases/{case['id']}/deliverables/FULL_CREDIT/draft"
    body = draft_request(template, source).model_dump(mode="json")
    first = client.put(url, json=body, headers=ANALYST_H)
    assert first.status_code in (200, 201)
    saved = first.json()["current"]
    assert saved["version"] == 1 and first.json()["history"], "the PUT returns the full workspace envelope"
    stale = client.put(url, json=body, headers=ANALYST_H)  # expected_version still 0
    assert stale.status_code == 409
    detail = stale.json()["detail"]
    assert detail["code"], "409 shape carries a typed code"
    assert detail["current"]["version"] == 1, "conflict embeds the full current revision for rebase"
    assert detail["current"]["id"] == saved["id"]
    by_id = client.get(f"/api/cases/{case['id']}/deliverables/revisions/{saved['id']}", headers=ANALYST_H)
    assert by_id.status_code == 200
    assert by_id.json()["content"] == saved["content"], "any revision stays fetchable by stable id"


# --- strict schema and template policy --------------------------------------------


def test_freeze_round_trips_on_nothing_but_what_the_draft_wire_serves(client, settings, store):
    """The freeze contract is keyed on `draft_id`, and the draft wire served only
    `id` (the revision). Every other freeze test builds its request from the
    service's own revision dict, so nothing noticed that a client reading the wire
    had no way to name the draft — the workbench sent the revision id and every
    freeze came back DELIVERABLE_DRAFT_STALE. This test may read only the wire."""
    svc, case, source, template = http_seed(settings, store)
    url = f"/api/cases/{case['id']}/deliverables/FULL_CREDIT/draft"
    saved = client.put(url, json=draft_request(template, source).model_dump(mode="json"),
                       headers=ANALYST_H).json()["current"]

    frozen = client.post(
        f"/api/cases/{case['id']}/deliverables/FULL_CREDIT/freeze",
        json={"draft_id": saved["draft_id"], "draft_version": saved["version"], "draft_digest": saved["digest"]},
        headers=ANALYST_H,
    )
    assert frozen.status_code == 201, frozen.text
    assert frozen.json()["draft_version"] == saved["version"]


def test_strict_schema_rejects_client_supplied_generated_values(client, settings, store):
    svc, case, source, template = http_seed(settings, store)
    body = draft_request(template, source).model_dump(mode="json")
    forged = optional_block(template, "GENERATED_METRIC")
    forged["values"] = {"total_leverage": 1.0}  # client-computed values are server-authoritative
    body["blocks"].append(forged)
    resp = client.put(f"/api/cases/{case['id']}/deliverables/FULL_CREDIT/draft", json=body, headers=ANALYST_H)
    assert resp.status_code == 422, "extra=forbid rejects client-generated values before any node runs"
    assert svc.workspace(case["id"], "FULL_CREDIT")["draft"] is None


@pytest.mark.parametrize("violation", ["heading_in_optional_slot", "narrative_in_optional_slot", "invented_slot_id"])
def test_client_invented_optional_kind_or_slot_is_rejected(service, store, violation):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    if violation == "heading_in_optional_slot":
        rogue = {"block_id": "blk-rogue", "slot_id": optional_slot(template, "LIMITATIONS"), "kind": "HEADING", "text": "Sneaky heading"}
    elif violation == "narrative_in_optional_slot":
        rogue = {"block_id": "blk-rogue", "slot_id": optional_slot(template, "LIMITATIONS"), "kind": "NARRATIVE",
                 "text": "Sneaky narrative", "content_mode": "ANALYST_JUDGMENT", "citations": []}
    else:
        rogue = optional_block(template, "LIMITATIONS", slot_id="appendix-invented-1")
    with pytest.raises(Exception, match="DELIVERABLE_SLOT_INVALID"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source, extra_blocks=[rogue]), actor="analyst")
    assert service.workspace(case["id"], "FULL_CREDIT")["draft"] is None, "no state change on rejection"


def test_optional_blocks_out_of_template_declared_order_are_rejected(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case)
    declared = [p["kind"] for p in template["optional_blocks"]]
    assert declared.index("GENERATED_METRIC") < declared.index("GENERATED_TABLE"), "premise: metric declared before table"
    out_of_order = [optional_block(template, "GENERATED_TABLE"), optional_block(template, "GENERATED_METRIC")]
    with pytest.raises(Exception, match="DELIVERABLE_TEMPLATE_ORDER_INVALID"):
        service.save_draft(
            case["id"], "FULL_CREDIT",
            draft_request(template, source, model_selection=revision_selection(model), extra_blocks=out_of_order),
            actor="analyst",
        )


def test_declared_limitations_slot_saves_without_model_identity_on_optional_pathway(service, store):
    case, source, _ = seed_ready_case(service, store)
    pathway = "RELATIVE_VALUE"
    template = service.templates()[pathway]
    revision = service.save_draft(
        case["id"], pathway,
        draft_request(template, source, extra_blocks=[optional_block(template, "LIMITATIONS")]),
        actor="analyst",
    )
    assert revision["content"]["model_identity"] is None, "non-model slots need no model"
    assert revision["content"]["blocks"][-1]["kind"] == "LIMITATIONS"


@pytest.mark.parametrize("pathway", ["FULL_CREDIT", "DISTRESSED_RESTRUCTURING"])
def test_model_required_templates_reject_required_blocks_without_a_selected_model(
    service, store, pathway,
):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()[pathway]
    assert template["model_requirement"] == "REQUIRED"

    with pytest.raises(Exception, match="MODEL_REQUIRED"):
        service.save_draft(
            case["id"], pathway, draft_request(template, source), actor="analyst",
        )

    assert service.workspace(case["id"], pathway)["draft"] is None


# --- evidence citations -----------------------------------------------------------


def test_citations_resolve_same_case_existing_non_withdrawn_blocks(service, store):
    case, source, _ = seed_ready_case(service, store)
    other_case, other_source = seed_case_with_source(store, body=b"foreign case evidence")
    template = service.templates()["FULL_CREDIT"]

    def with_citation(source_id, block_id):
        blocks = required_blocks(template, source)
        register = next(b for b in blocks if b["kind"] == "EVIDENCE_REGISTER")
        register["citations"] = [{"source_id": source_id, "block_ids": [block_id], "claim": "cited claim"}]
        return draft_request(template, blocks=blocks)

    with pytest.raises(Exception, match="EVIDENCE_CASE_MISMATCH"):
        service.save_draft(case["id"], "FULL_CREDIT", with_citation(other_source["id"], "b00001"), actor="analyst")
    with pytest.raises(Exception, match="EVIDENCE_BLOCK_MISMATCH"):
        service.save_draft(case["id"], "FULL_CREDIT", with_citation(source["id"], "b99999"), actor="analyst")
    store.withdraw(case["id"], source["id"], "analyst")
    with pytest.raises(Exception, match="EVIDENCE_SOURCE_WITHDRAWN"):
        service.save_draft(case["id"], "FULL_CREDIT", with_citation(source["id"], "b00001"), actor="analyst")
    assert service.workspace(case["id"], "FULL_CREDIT")["draft"] is None, "every failed save aborts with nothing appended"


# --- model selection and generated blocks -----------------------------------------


def test_model_selection_pins_current_revision_and_fallback_requires_acknowledged_no_revision_state(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    first = seed_model(service, case)
    second = service.seed_signed_revision_for_tests(case["id"], outputs={"total_leverage": 3.9})  # head advances
    metric = [optional_block(template, "GENERATED_METRIC")]
    with pytest.raises(Exception, match="MODEL_REVISION_STALE"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source, model_selection=revision_selection(first), extra_blocks=metric), actor="analyst")
    fallback = {"kind": "APPLICATION_BUILD", "build_id": second["build_id"], "fallback_acknowledged": True}
    with pytest.raises(Exception, match="FALLBACK|REVISION"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source, model_selection=fallback, extra_blocks=metric), actor="analyst")

    # eligible fallback: a case whose build has no signed revision, explicitly acknowledged
    bare_case, bare_source, _ = seed_ready_case(service, store)
    build = service.seed_application_build_for_tests(bare_case["id"], outputs={"total_leverage": 4.2})
    eligible = {"kind": "APPLICATION_BUILD", "build_id": build["build_id"], "fallback_acknowledged": True}
    revision = service.save_draft(bare_case["id"], "FULL_CREDIT", draft_request(template, bare_source, model_selection=eligible, extra_blocks=metric), actor="analyst")
    identity = revision["content"]["model_identity"]
    assert identity["kind"] == "APPLICATION_BUILD"
    assert identity["build_id"] == build["build_id"]
    assert {"name", "version", "sha256"} <= set(identity["calculation_runtime"]), "fallback pins the full build identity incl. calculation runtime"


def test_selection_must_resolve_to_exact_stored_revision_and_build_records(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case)
    metric = [optional_block(template, "GENERATED_METRIC")]
    ghost_revision = {"kind": "ANALYST_REVISION", "build_id": model["build_id"], "revision_id": "rev-missing"}
    with pytest.raises(Exception, match="MODEL_REVISION_STALE"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source, model_selection=ghost_revision, extra_blocks=metric), actor="analyst")

    bare_case, bare_source, _ = seed_ready_case(service, store)
    ghost_build = {"kind": "APPLICATION_BUILD", "build_id": "build-missing", "fallback_acknowledged": True}
    with pytest.raises(Exception, match="MODEL_BUILD_STALE"):
        service.save_draft(bare_case["id"], "FULL_CREDIT", draft_request(template, bare_source, model_selection=ghost_build, extra_blocks=metric), actor="analyst")


@pytest.mark.parametrize("kind", ["GENERATED_METRIC", "GENERATED_TABLE", "GENERATED_CHART", "MODEL_APPENDIX"])
def test_model_dependent_block_kinds_require_selection_and_are_digest_pinned(service, store, kind):
    from caos.contracts import digest

    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    extra = [optional_block(template, kind)]
    with pytest.raises(Exception, match="MODEL_REQUIRED"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source, extra_blocks=extra), actor="analyst")
    model = seed_model(service, case)
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, source, model_selection=revision_selection(model), extra_blocks=extra),
        actor="analyst",
    )
    stored = revision["content"]["blocks"][-1]
    assert stored["kind"] == kind
    assert stored["model_digest"] == digest(revision["content"]["model_identity"]), "generated block pinned to the model identity"


def test_generated_metric_outputs_are_rebuilt_server_side_from_pinned_revision(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case, outputs={"total_leverage": 4.2})
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, source, model_selection=revision_selection(model), extra_blocks=[optional_block(template, "GENERATED_METRIC")]),
        actor="analyst",
    )
    stored = revision["content"]["blocks"][-1]
    assert stored["values"]["total_leverage"] == 4.2, "server recomputes from the exact pinned revision"


def test_forged_scenario_outputs_raise_calculation_mismatch(service, store):
    from caos.contracts import digest

    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case)
    selection = revision_selection(model)
    preview = service.preview_scenario_for_tests(case["id"], build_id=model["build_id"], base_revision_id=model["revision_id"], shocks=[SHOCK])

    # the honest preview saves and pins model identity
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, source, model_selection=selection, extra_blocks=[scenario_exhibit_block(template, preview)]),
        actor="analyst",
    )
    stored = revision["content"]["blocks"][-1]
    assert stored["model_digest"] == digest(revision["content"]["model_identity"])
    assert stored["scenario"]["build_id"] == model["build_id"]
    exhibit = revision["content"]["document_sections"][-2]
    assert exhibit["kind"] == "chart"
    assert exhibit["title"] == "Downside revenue shock"
    assert exhibit["origin"] == {
        "kind": "MODEL",
        "authority_id": stored["scenario_digest"],
        "block_ids": [],
    }

    # forged outputs with self-consistent digests must still fail the server recompute
    forged = copy.deepcopy(preview)
    forged["scenario"]["outputs"]["total_leverage"] = 99.9
    forged["scenario"]["outputs_digest"] = digest(forged["scenario"]["outputs"])
    forged["scenario_digest"] = digest(forged["scenario"])
    with pytest.raises(Exception, match="SCENARIO_EXHIBIT_CALCULATION_MISMATCH"):
        service.save_draft(
            case["id"], "FULL_CREDIT",
            draft_request(template, source, expected_version=1, model_selection=selection, extra_blocks=[scenario_exhibit_block(template, forged)]),
            actor="analyst",
        )


def test_scenario_exhibit_requires_selected_model_before_any_calculation(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case)
    preview = service.preview_scenario_for_tests(case["id"], build_id=model["build_id"], base_revision_id=model["revision_id"], shocks=[SHOCK])
    with pytest.raises(Exception, match="MODEL_REQUIRED"):
        service.save_draft(
            case["id"], "FULL_CREDIT",
            draft_request(template, source, extra_blocks=[scenario_exhibit_block(template, preview)]),  # no model_selection
            actor="analyst",
        )


def test_scenario_base_identity_mismatch_fails_before_calculation_with_zero_residue(service, store):
    from caos.contracts import digest

    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case)
    preview = service.preview_scenario_for_tests(case["id"], build_id=model["build_id"], base_revision_id=model["revision_id"], shocks=[SHOCK])
    mismatched = copy.deepcopy(preview)
    mismatched["scenario"]["base_revision_id"] = "rev-someone-elses"  # binds to a different model than the selection
    mismatched["scenario_digest"] = digest(mismatched["scenario"])
    service.forbid_scenario_calculation_for_tests()  # any calculator invocation now raises AssertionError
    with pytest.raises(Exception, match="IDENTITY"):
        service.save_draft(
            case["id"], "FULL_CREDIT",
            draft_request(template, source, model_selection=revision_selection(model), extra_blocks=[scenario_exhibit_block(template, mismatched)]),
            actor="analyst",
        )
    assert service.workspace(case["id"], "FULL_CREDIT")["draft"] is None, "no revision persisted"
    saved = [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.draft.saved"]
    assert saved == [], "no audit residue on the failed save"


def test_application_build_fallback_scenario_accepts_null_base_revision(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    build = service.seed_application_build_for_tests(case["id"], outputs={"total_leverage": 4.2})
    preview = service.preview_scenario_for_tests(case["id"], build_id=build["build_id"], base_revision_id=None, shocks=[SHOCK])
    assert preview["scenario"]["base_revision_id"] is None
    fallback = {"kind": "APPLICATION_BUILD", "build_id": build["build_id"], "fallback_acknowledged": True}
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, source, model_selection=fallback, extra_blocks=[scenario_exhibit_block(template, preview)]),
        actor="analyst",
    )
    assert revision["content"]["model_identity"]["kind"] == "APPLICATION_BUILD"
    assert revision["content"]["blocks"][-1]["scenario"]["base_revision_id"] is None


def test_ungoverned_metric_ids_and_chart_recipe_fields_are_rejected(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case)
    selection = revision_selection(model)
    rogue_metric = [optional_block(template, "GENERATED_METRIC", metric_ids=["made_up_metric"])]
    with pytest.raises(Exception, match="GENERATED_FIELD_INVALID"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source, model_selection=selection, extra_blocks=rogue_metric), actor="analyst")
    rogue_chart = [optional_block(template, "GENERATED_CHART", recipe={"chart_kind": "line", "fields": ["not_a_governed_field"]})]
    with pytest.raises(Exception, match="GENERATED_FIELD_INVALID|RECIPE"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source, model_selection=selection, extra_blocks=rogue_chart), actor="analyst")
    rogue_table = [optional_block(
        template, "GENERATED_TABLE", table_id="annual_model", field_ids=["not_a_governed_field"],
    )]
    with pytest.raises(Exception, match="GENERATED_FIELD_INVALID"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(
            template, source, model_selection=selection, extra_blocks=rogue_table,
        ), actor="analyst")


def test_scenario_digest_must_equal_server_computed_digest(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case)
    preview = service.preview_scenario_for_tests(case["id"], build_id=model["build_id"], base_revision_id=model["revision_id"], shocks=[SHOCK])
    with pytest.raises(Exception, match="SCENARIO_EXHIBIT_DIGEST_INVALID"):
        service.save_draft(
            case["id"], "FULL_CREDIT",
            draft_request(template, source, model_selection=revision_selection(model),
                          extra_blocks=[scenario_exhibit_block(template, preview, scenario_digest="0" * 64)]),
            actor="analyst",
        )


def test_live_model_builder_authority_feeds_eligibility_selection_scenario_and_freeze(settings, store, engine, tmp_path):
    """Model Builder is the system of record when nothing is seeded: eligibility,
    the ANALYST_REVISION selection, scenario revalidation through the real
    calculator, generated blocks, and the frozen model embedding all resolve live."""
    import asyncio

    from caos.contracts import ModelPreviewRequest, ModelScenarioRequest, ModelSignOffRequest
    from caos.deliverables.service import DeliverableService
    from caos.models.service import ModelService

    models = ModelService(store=store, vault_dir=settings.storage_dir, engine=engine)
    service = DeliverableService(store=store, vault_dir=tmp_path / "deliverable-vault", models=models)
    case, source = seed_case_with_source(store)

    async def accept_scripted():
        run = await engine.run_scripted_for_tests(case["id"])
        await engine.accept(run["id"], actor="analyst")

    asyncio.run(accept_scripted())
    build = models.run_build_for_tests(models.queue_build(case["id"], "analyst")["id"])
    registry = models.assumption_registry(case["id"], build["id"])
    preview = models.preview(case["id"], ModelPreviewRequest.model_validate({
        "build_id": build["id"], "parent_revision_id": None, "registry_version": registry["version"],
        "registry_digest": registry["digest"], "assumptions": registry["defaults"], "draft_generation": 1,
    }))
    signed = models.sign_off(case["id"], ModelSignOffRequest.model_validate({
        "build_id": build["id"], "parent_revision_id": None, "registry_version": registry["version"],
        "registry_digest": registry["digest"], "assumptions": registry["defaults"], "draft_generation": 1,
        "preview_digest": preview["preview_digest"], "expected_head_revision_id": None,
        "note": "Live authority spec sign-off.",
    }), actor="approver-user")

    eligibility = service.model_eligibility(case["id"])
    assert eligibility["active_revision"]["revision_id"] == signed["id"]
    assert eligibility["application_build"]["build_id"] == build["id"]
    assert eligibility["fallback_acknowledgement_required"] is False
    selection = eligibility["default_model_selection"]
    assert selection == {"kind": "ANALYST_REVISION", "build_id": build["id"], "revision_id": signed["id"]}

    shock_row = next(row for row in registry["defaults"] if row["status"] == "READY" and row["case"] == "BASE")
    shocks = [{"assumption_id": shock_row["assumption_id"], "case": shock_row["case"],
               "period_id": shock_row["period_id"], "value": 0.07}]
    calculated = models.scenario(case["id"], ModelScenarioRequest.model_validate({
        "build_id": build["id"], "base_revision_id": signed["id"], "registry_version": registry["version"],
        "registry_digest": registry["digest"], "shocks": shocks, "draft_generation": 1,
    }))
    template = service.templates()["FULL_CREDIT"]
    scenario_slot = optional_slot(template, "SCENARIO_EXHIBIT")
    revision = service.save_draft(case["id"], "FULL_CREDIT", draft_request(
        template, source, model_selection=selection,
        extra_blocks=[
            optional_block(template, "GENERATED_METRIC", metric_ids=["total_leverage"]),
            {"block_id": scenario_slot, "slot_id": scenario_slot, "kind": "SCENARIO_EXHIBIT",
             "title": "Downside revenue shock", "shocks": shocks,
             "scenario": calculated["scenario"], "scenario_digest": calculated["scenario_digest"]},
        ],
    ), actor="analyst")

    content = service.workspace(case["id"], "FULL_CREDIT")["draft"]["content"]
    assert content["model_selection"] == selection
    metric_slot = optional_slot(template, "GENERATED_METRIC")
    assert content["generated_blocks"][f"blk-{metric_slot}"]["status"] == "READY"
    assert content["generated_blocks"][scenario_slot]["outputs"] == calculated["scenario"]["outputs"]

    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    assert frozen["payload"]["model"]["outputs"] == signed["outputs"], "live revision outputs embedded verbatim"
    assert frozen["payload"]["authority"]["build_id"] == build["id"]
    assert frozen["payload"]["methodology"]["build_id"] == build[
        "methodology_build_id"
    ], "signed-revision publishing retains the originating methodology identity"
    assert frozen["payload"]["template"]["title"] == "Investment Committee Credit Memo"


def test_filing_rejects_a_frozen_fallback_after_an_analyst_revision_becomes_active(
    settings,
    store,
    engine,
    tmp_path,
):
    import asyncio

    from caos.contracts import ModelPreviewRequest, ModelSignOffRequest
    from caos.deliverables.service import DeliverableService
    from caos.models.service import ModelService

    models = ModelService(store=store, vault_dir=settings.storage_dir, engine=engine)
    service = DeliverableService(
        store=store,
        vault_dir=tmp_path / "deliverable-vault",
        models=models,
    )
    case, source = seed_case_with_source(store)

    async def accept_scripted():
        run = await engine.run_scripted_for_tests(case["id"])
        await engine.accept(run["id"], actor="analyst")

    asyncio.run(accept_scripted())
    build = models.run_build_for_tests(models.queue_build(case["id"], "analyst")["id"])
    fallback = {
        "kind": "APPLICATION_BUILD",
        "build_id": build["id"],
        "fallback_acknowledged": True,
    }
    distressed_template = service.templates()["DISTRESSED_RESTRUCTURING"]
    with pytest.raises(ValueError, match="DELIVERABLE_PATHWAY_AUTHORITY_MISMATCH"):
        service.save_draft(
            case["id"],
            "DISTRESSED_RESTRUCTURING",
            draft_request(distressed_template, source, model_selection=fallback),
            actor="analyst",
        )
    template = service.templates()["FULL_CREDIT"]
    draft = service.save_draft(
        case["id"],
        "FULL_CREDIT",
        draft_request(template, source, model_selection=fallback),
        actor="analyst",
    )
    frozen = service.freeze(case["id"], freeze_request(draft), actor="analyst")

    registry = models.assumption_registry(case["id"], build["id"])
    preview_request = ModelPreviewRequest.model_validate({
        "build_id": build["id"],
        "parent_revision_id": None,
        "registry_version": registry["version"],
        "registry_digest": registry["digest"],
        "assumptions": registry["defaults"],
        "draft_generation": 1,
    })
    preview = models.preview(case["id"], preview_request)
    models.sign_off(case["id"], ModelSignOffRequest.model_validate({
        **preview_request.model_dump(mode="json"),
        "preview_digest": preview["preview_digest"],
        "expected_head_revision_id": None,
        "note": "Revision made active while the filing gate was parked.",
    }))

    assert service.model_eligibility(case["id"])["default_model_selection"]["kind"] == "ANALYST_REVISION"
    with pytest.raises(ValueError, match="FROZEN_MODEL_AUTHORITY_STALE"):
        service.approve_filing(
            case["id"],
            frozen["deliverable_id"],
            file_request(frozen),
            actor="approver-user",
        )
    assert service.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN"


@pytest.mark.parametrize("pathway", ["EARNINGS_UPDATE", "COVENANT_REFINANCING"])
@pytest.mark.parametrize("selection_kind", ["APPLICATION_BUILD", "ANALYST_REVISION"])
async def test_live_incremental_pathway_publishes_against_a_validated_prior_full_credit_model(
    settings,
    store,
    engine,
    tmp_path,
    pathway,
    selection_kind,
):
    from caos.contracts import ModelPreviewRequest, ModelSignOffRequest, digest
    from caos.deliverables.service import DeliverableService
    from caos.models.service import ModelService

    models = ModelService(store=store, vault_dir=settings.storage_dir, engine=engine)
    service = DeliverableService(
        store=store,
        vault_dir=tmp_path / "incremental-deliverable-vault",
        engine=engine,
        models=models,
    )
    case, source = seed_case_with_source(store)
    store.add_member(
        case["id"],
        "analyst",
        "approver-user",
        "APPROVER",
        actor_role="ADMIN",
    )

    full_credit = await engine.run_scripted_for_tests(case["id"])
    base_snapshot = await engine.accept(full_credit["id"], actor="analyst")
    queued = next(
        build
        for build in models.list_builds(case["id"])
        if build["snapshot_id"] == base_snapshot["id"]
    )
    base_build = models.run_build_for_tests(queued["id"])
    assert base_build["status"] == "READY"

    signed = None
    if selection_kind == "ANALYST_REVISION":
        registry = models.assumption_registry(case["id"], base_build["id"])
        preview_request = ModelPreviewRequest.model_validate(
            {
                "build_id": base_build["id"],
                "parent_revision_id": None,
                "registry_version": registry["version"],
                "registry_digest": registry["digest"],
                "assumptions": registry["defaults"],
                "draft_generation": 1,
            }
        )
        preview = models.preview(case["id"], preview_request)
        signed = models.sign_off(
            case["id"],
            ModelSignOffRequest.model_validate(
                {
                    **preview_request.model_dump(mode="json"),
                    "preview_digest": preview["preview_digest"],
                    "expected_head_revision_id": None,
                    "note": "Prior Full Credit authority for incremental publication.",
                }
            ),
            actor="analyst",
        )
        selection = {
            "kind": "ANALYST_REVISION",
            "build_id": base_build["id"],
            "revision_id": signed["id"],
        }
    else:
        selection = {
            "kind": "APPLICATION_BUILD",
            "build_id": base_build["id"],
            "fallback_acknowledged": True,
        }

    incremental = await engine.run_scripted_for_tests(
        case["id"],
        pathway=pathway,
    )
    incremental_snapshot = await engine.accept(incremental["id"], actor="analyst")
    assert incremental_snapshot["previous_snapshot_id"] == base_snapshot["id"]

    eligibility = service.model_eligibility(case["id"])
    assert eligibility["application_build"]["build_id"] == base_build["id"]
    if signed is not None:
        assert eligibility["default_model_selection"] == selection
    else:
        assert eligibility["default_model_selection"] is None
        assert eligibility["fallback_acknowledgement_required"] is True

    template = service.templates()[pathway]
    if signed is not None:
        with pytest.raises(ValueError, match="MODEL_FALLBACK_INELIGIBLE"):
            service.save_draft(
                case["id"],
                pathway,
                draft_request(
                    template,
                    source,
                    model_selection={
                        "kind": "APPLICATION_BUILD",
                        "build_id": base_build["id"],
                        "fallback_acknowledged": True,
                    },
                ),
                actor="analyst",
            )
    revision = service.save_draft(
        case["id"],
        pathway,
        draft_request(template, source, model_selection=selection),
        actor="analyst",
    )
    expected_model_authority = {
        "relationship": "PRIOR_FULL_CREDIT_BASE",
        "snapshot_id": base_snapshot["id"],
        "snapshot_digest": base_snapshot["digest"],
        "run_id": base_snapshot["run_id"],
        "source_set_id": base_snapshot["source_set_id"],
        "source_set_version": base_snapshot["source_set_version"],
        "input_fingerprint": base_build["input_fingerprint"],
        "payload_digest": base_build["payload_digest"],
    }
    assert revision["content"]["model_identity"]["model_authority"] == (
        expected_model_authority
    )

    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    assert frozen["payload"]["authority"]["accepted_snapshot_id"] == (
        incremental_snapshot["id"]
    )
    assert frozen["payload"]["model"]["build_id"] == base_build["id"]
    assert frozen["payload"]["model"]["model_authority"] == (
        expected_model_authority
    )
    assert frozen["payload"]["methodology"]["build_id"] == base_build[
        "methodology_build_id"
    ]
    assert frozen["input_fingerprint"] == digest(
        {
            "snapshot_id": incremental_snapshot["id"],
            "source_set_id": incremental_snapshot["source_set_id"],
            "source_set_version": incremental_snapshot["source_set_version"],
            "build_id": base_build["id"],
            "methodology_build_id": base_build["methodology_build_id"],
        }
    )

    filed = service.approve_filing(
        case["id"],
        frozen["deliverable_id"],
        file_request(frozen),
        actor="approver-user",
    )
    assert filed["status"] == "FILED"
    assert filed["payload"] == frozen["payload"]


# --- reader authorization ---------------------------------------------------------


def test_case_reader_reads_workspace_but_cannot_write_draft(client, settings, store):
    svc, case, source, template = http_seed(settings, store)
    store.add_member(case["id"], "analyst", "reader-user", "READER", actor_role="ADMIN")
    reader = {"x-caos-role": "ANALYST", "x-forwarded-user": "reader-user"}
    url = f"/api/cases/{case['id']}/deliverables/FULL_CREDIT/draft"
    assert client.get(url, headers=reader).status_code == 200, "reader may read the workspace"
    write = client.put(url, json=draft_request(template, source).model_dump(mode="json"), headers=reader)
    assert write.status_code == 403, "reader never writes drafts"
    assert svc.workspace(case["id"], "FULL_CREDIT")["draft"] is None


def test_downgraded_global_reader_loses_filing_authority_despite_case_standing(client, settings, store):
    """Filing is two-dimensional. Stored case standing is not a second, staler
    identity: once the IdP downgrades the account to READER, the highest-integrity
    human gate must close even though the case row still says APPROVER."""
    svc, case, source, template = http_seed(settings, store)
    revision = svc.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst")
    frozen = svc.freeze(case["id"], freeze_request(revision), actor="analyst")
    store.add_member(case["id"], "analyst", "approver-user", "APPROVER", actor_role="ADMIN")
    downgraded = {"x-caos-role": "READER", "x-forwarded-user": "approver-user"}
    body = {"preview_digest": frozen["preview_digest"], "input_fingerprint": frozen["input_fingerprint"]}
    base = f"/api/cases/{case['id']}/deliverables/by-id/{frozen['deliverable_id']}"

    assert client.post(f"{base}/approve", json=body, headers=downgraded).status_code == 403
    assert client.post(f"{base}/request-changes", json={**body, "comment": "No."},
                       headers=downgraded).status_code == 403
    assert svc.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN", "nothing was filed"
    # The other direction is unchanged: a global APPROVER without case standing
    # still never files — case standing never escalates and never substitutes.
    assert client.post(f"{base}/approve", json=body,
                       headers={"x-caos-role": "APPROVER", "x-forwarded-user": "stranger"}).status_code == 404


# --- freeze and rendered exports --------------------------------------------------


def test_canonical_document_is_the_only_cross_format_export_source(service, store):
    case, source, _ = seed_ready_case(service, store)
    governed_tail = "END-OF-GOVERNED-NARRATIVE"
    unicode_marker = "“Downside”—£95m 债务重组 fi fl ffi ffl"  # ligature pairs must extract verbatim
    long_narrative = f"Committee-ready analysis {unicode_marker} {'x' * 140} {governed_tail}"
    model = seed_model(service, case, outputs={
        "BASE": {"FY2027": {"total_leverage": 3.8}},
        "DOWNSIDE": {"FY2027": {"total_leverage": 5.1}},
    })
    template = service.templates()["FULL_CREDIT"]
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(
            template,
            blocks=required_blocks(template, source, narrative_text=long_narrative),
            model_selection=revision_selection(model),
        ),
        actor="analyst",
    )
    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    payload = copy.deepcopy(frozen["payload"])
    expected_titles = [section["title"] for section in payload["content"]["document_sections"]]
    assert payload["renderer"]["version"] == "caos.deliverable-renderer.v2"

    # Canonical sections are sufficient: no renderer may fall back to the old
    # generated/block projections once the versioned document exists.
    payload["content"].pop("blocks")
    payload["content"].pop("generated_blocks")
    model_section = next(section for section in payload["content"]["document_sections"] if section["section_id"] == "base_downside_model")
    model_section["items"][1][0]["rows"].append(["BASE / FY2028 / total_leverage", "1e309"])
    from caos.publishing.renderers import render_frozen_export

    markdown = render_frozen_export(payload, "md").decode("utf-8")
    markdown_titles = [line[3:] for line in markdown.splitlines() if line.startswith("## ")]
    assert markdown_titles == expected_titles

    from pypdf import PdfReader

    pdf = render_frozen_export(payload, "pdf")
    assert pdf == render_frozen_export(payload, "pdf"), "Unicode PDF bytes are deterministic"
    pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(pdf)).pages)
    pdf_titles = [line[3:] for line in pdf_text.splitlines() if line.startswith("## ")]
    assert pdf_titles == expected_titles

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(render_frozen_export(payload, "xlsx")))
    reviewed = workbook["Reviewed Deliverable"]
    xlsx_titles = [row[2].value for row in reviewed.iter_rows() if row[0].value == "SECTION"]
    assert xlsx_titles == expected_titles
    assert "FY2027" in markdown and "FY2027" in pdf_text
    assert governed_tail in markdown and governed_tail in pdf_text
    assert unicode_marker in markdown and unicode_marker in pdf_text
    assert any(cell.value == "BASE / FY2027 / total_leverage" for row in reviewed.iter_rows() for cell in row)
    assert any(cell.value == 3.8 for row in reviewed.iter_rows() for cell in row)
    assert any(cell.value == "1e309" for row in reviewed.iter_rows() for cell in row)
    assert any(unicode_marker in str(cell.value) for row in reviewed.iter_rows() for cell in row)
    evidence = list(workbook["Evidence Register"].iter_rows(values_only=True))
    assert evidence[:2] == [
        ("Source", "Blocks", "Claim"),
        (source["id"], "b00001", "Pinned evidence line supports the opinion."),
    ]


@pytest.mark.parametrize("pathway", SIX_PATHWAYS)
def test_each_pathway_frozen_payload_renders_substantive_md_pdf_xlsx(service, store, pathway):
    case, source, revision, frozen = freeze_min(service, store, pathway)

    first_section = service.templates()[pathway]["blocks"][0]["title"]
    md, md_sha = service.export(frozen["deliverable_id"], "md")
    assert hashlib.sha256(md).hexdigest() == md_sha
    text = md.decode("utf-8")
    assert first_section in text and "Leverage is manageable." in text

    pdf, _ = service.export(frozen["deliverable_id"], "pdf")
    assert pdf[:5] == b"%PDF-", "structurally a PDF"
    assert first_section in service.export_text_for_tests(frozen["deliverable_id"], "pdf")

    xlsx, _ = service.export(frozen["deliverable_id"], "xlsx")
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(xlsx))
    assert {"Cover", "Reviewed Deliverable", "Evidence Register", "Revision Record"} <= set(book.sheetnames)


def test_freeze_embeds_selected_revision_outputs_assumptions_and_build_payload_verbatim(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case, outputs={"total_leverage": 4.2})
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, source, model_selection=revision_selection(model), extra_blocks=[optional_block(template, "GENERATED_METRIC")]),
        actor="analyst",
    )
    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    pinned = frozen["payload"]["model"]
    assert pinned["outputs"] == model["outputs"], "signed revision outputs embedded verbatim"
    assert pinned["effective_assumptions"] == model["assumptions"], "effective assumptions embedded verbatim"
    assert pinned["application_build"]["payload"] == model["build_payload"], "application-build payload embedded verbatim"
    assert pinned["application_build"]["qa"] == model["build_qa"], "build QA embedded verbatim"
    assert "4.2" in service.export_text_for_tests(frozen["deliverable_id"], "md"), "exports show the exact pinned figures"


def test_freeze_is_idempotent_under_race_with_one_record_one_thread_one_audit_event(service, store):
    case, source, template, revision = save_min_draft(service, store)
    with ThreadPoolExecutor(max_workers=2) as pool:
        raced = list(pool.map(lambda _: service.freeze(case["id"], freeze_request(revision), actor="analyst"), range(2)))
    retried = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    everyone = raced + [retried]
    assert len({r["deliverable_id"] for r in everyone}) == 1, "one content-addressed frozen record"
    assert len({r["thread_id"] for r in everyone}) == 1, "racing freezes converge on one thread (§10.7)"
    frozen_events = [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.frozen"]
    assert len(frozen_events) == 1, "single audit event under race and retry"


def test_divergent_render_for_same_freeze_identity_is_freeze_conflict(tmp_path, store):
    vault = tmp_path / "shared-vault"
    service = make_service(store, vault)
    case, source, template, revision = save_min_draft(service, store)
    service.freeze(case["id"], freeze_request(revision), actor="analyst")
    divergent = make_service(store, vault, renderer_for_tests=lambda payload, fmt: b"DIFFERENT RENDER BYTES")
    with pytest.raises(Exception, match="DELIVERABLE_FREEZE_CONFLICT"):
        divergent.freeze(case["id"], freeze_request(revision), actor="analyst")


def test_freeze_retry_recovers_after_files_publish_but_before_record_commit(
    service, store, monkeypatch
):
    case, _source, _template, revision = save_min_draft(service, store)
    original_insert = service.records.insert_frozen

    def fail_before_record(*args, **kwargs):
        raise RuntimeError("injected record commit failure")

    monkeypatch.setattr(service.records, "insert_frozen", fail_before_record)
    with pytest.raises(RuntimeError, match="injected record commit failure"):
        service.freeze(case["id"], freeze_request(revision), actor="analyst")
    assert service.workspace(case["id"], "FULL_CREDIT")["frozen"] == []

    monkeypatch.setattr(service.records, "insert_frozen", original_insert)
    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    assert set(frozen["exports"]) == {"md", "pdf", "xlsx"}
    assert all(service.export(frozen["deliverable_id"], format_name)[0]
               for format_name in frozen["exports"])


def test_filing_thread_id_is_a_deterministic_digest_including_build_id():
    from caos.deliverables.graph import filing_thread_id

    args = dict(case_id="case-1", pathway="FULL_CREDIT", draft_version=3, draft_digest="a" * 64, build_id="build-1")
    assert filing_thread_id(**args) == filing_thread_id(**args), "deterministic"
    baseline = filing_thread_id(**args)
    assert filing_thread_id(**{**args, "build_id": "build-2"}) != baseline, "thread identity includes build_id (§12.23)"
    assert filing_thread_id(**{**args, "draft_version": 4}) != baseline
    assert filing_thread_id(**{**args, "draft_digest": "b" * 64}) != baseline
    assert filing_thread_id(**{**args, "case_id": "case-2"}) != baseline


def test_draft_fails_closed_without_upstream_accepted_authority(service, store):
    case, source = seed_case_with_source(store)  # deliberately: no accepted authority seeded
    template = service.templates()["FULL_CREDIT"]
    bind_default_model_for_tests(service, case, template)
    with pytest.raises(Exception, match="AUTHORITY|SNAPSHOT|UPSTREAM"):
        service.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst")
    assert service.workspace(case["id"], "FULL_CREDIT")["frozen"] == [], "no frozen record without pinned upstream identity"


# --- the filing gate --------------------------------------------------------------


def test_concurrent_filers_yield_exactly_one_filed_with_payload_equal_to_frozen(service, store):
    case, revision, frozen = frozen_deliverable(service, store)

    def approve(_):
        try:
            return ("filed", service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user"))
        except Exception as exc:  # noqa: BLE001 — the losing racer's typed refusal
            return ("refused", exc)

    with ThreadPoolExecutor(max_workers=2) as pool:
        outcomes = list(pool.map(approve, range(2)))
    assert sorted(kind for kind, _ in outcomes) == ["filed", "refused"], "resume ticket: exactly one filer wins (§12.21)"
    filed = next(value for kind, value in outcomes if kind == "filed")
    assert filed["status"] == "FILED"
    assert filed["payload"] == frozen["payload"], "filed payload is the frozen payload, bit-identical"
    filed_events = [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.filed"]
    assert len(filed_events) == 1


def test_approval_binds_exact_preview_digest_and_fingerprint_mismatch_leaves_frozen_retryable(service, store):
    case, revision, frozen = frozen_deliverable(service, store)
    with pytest.raises(Exception, match="STALE_PREVIEW|PREVIEW"):
        service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen, preview_digest="0" * 64), actor="approver-user")
    with pytest.raises(Exception, match="STALE_PREVIEW|FINGERPRINT"):
        service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen, input_fingerprint="1" * 64), actor="approver-user")
    record = service.frozen_record(case["id"], frozen["deliverable_id"])
    assert record["status"] == "FROZEN", "mismatch keeps the gate parked with the record intact"
    filed = service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user")
    assert filed["status"] == "FILED", "the record stayed retryable: an exact approval still lands"


def test_approval_leaves_frozen_content_bit_identical_and_appends_one_audit_event(service, store):
    case, revision, frozen = frozen_deliverable(service, store)
    before = copy.deepcopy(service.frozen_record(case["id"], frozen["deliverable_id"]))
    assert before["status"] == "FROZEN"
    service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user")
    after = service.frozen_record(case["id"], frozen["deliverable_id"])
    assert after["payload"] == before["payload"], "approval never edits the reviewed content"
    assert after["preview_digest"] == before["preview_digest"]
    assert after["input_fingerprint"] == before["input_fingerprint"]
    assert after["build_id"] == before["build_id"]
    assert after["status"] == "FILED"
    assert after["filed_by"] == "approver-user" and after["filed_at"]
    changed = {key for key in after if after.get(key) != before.get(key)}
    assert changed <= {"status", "filed_by", "filed_at"}, "only phase and approval fields may change"
    filed_events = [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.filed"]
    assert len(filed_events) == 1


def test_gate_approval_hash_is_canonical_sha256_form(service, store):
    from caos.deliverables.graph import APPROVAL_HASH_PATTERN, canonical_approval_hash

    assert APPROVAL_HASH_PATTERN == r"^sha256:[0-9a-f]{64}$"
    case, revision, frozen = frozen_deliverable(service, store)
    canonical = canonical_approval_hash(frozen)
    assert re.fullmatch(APPROVAL_HASH_PATTERN, canonical)
    assert canonical == f"sha256:{frozen['preview_digest']}", "gate hash binds the exact frozen preview digest"
    for alias in (canonical.upper(), canonical[:-1], "md5:" + "a" * 32, frozen["preview_digest"]):
        with pytest.raises(Exception):
            service.resume_filing_with_hash_for_tests(frozen["thread_id"], alias)
    filed = service.resume_filing_with_hash_for_tests(frozen["thread_id"], canonical)
    assert filed["status"] == "FILED", "only the canonical form resumes the gate"


def test_http_approval_payload_rejects_caller_owned_authority_fields(client, settings, store):
    svc, case, source, template = http_seed(settings, store)
    revision = svc.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst")
    frozen = svc.freeze(case["id"], freeze_request(revision), actor="analyst")
    store.add_member(case["id"], "analyst", "approver-user", "APPROVER", actor_role="ADMIN")
    url = f"/api/cases/{case['id']}/deliverables/by-id/{frozen['deliverable_id']}/approve"
    base = {"preview_digest": frozen["preview_digest"], "input_fingerprint": frozen["input_fingerprint"]}
    forgeries = (
        ("approved_by", "attacker"),
        ("approved_at", "2026-01-01T00:00:00Z"),
        ("authority_hash", "sha256:" + "a" * 64),
        ("plan_hash", "sha256:" + "b" * 64),
        ("budgets", {"turns": 10_000}),
        ("model", "claude-someone-elses"),
        ("scope", ["*"]),
    )
    for field, value in forgeries:
        resp = client.post(url, json={**base, field: value}, headers=APPROVER_H)
        assert resp.status_code == 422, f"extra=forbid must reject caller-owned field {field!r}"
    assert svc.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN", "no forged request may file"


def test_gate_rejects_source_set_drift_between_freeze_and_filing(service, store):
    case, revision, frozen = frozen_deliverable(service, store)
    ingest_second_source(store, case["id"])  # evidence base moves while the gate is parked
    with pytest.raises(Exception, match="SOURCE_SET_CHANGED|FROZEN_AUTHORITY_STALE"):
        service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user")
    assert service.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN"
    assert not [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.filed"]


async def test_gate_revalidates_accepted_artifact_graph_after_freeze(settings, store, engine, tmp_path):
    service = make_service(store, tmp_path / "deliverable-vault", engine=engine)
    case, source = seed_case_with_source(store)
    run = await engine.run_scripted_for_tests(case["id"], pathway="RELATIVE_VALUE")
    await engine.accept(run["id"], actor="analyst")
    template = service.templates()["RELATIVE_VALUE"]
    revision = service.save_draft(
        case["id"], "RELATIVE_VALUE", draft_request(template, source), actor="analyst",
    )
    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    add_approver(store, case)
    engine.forge_node_artifact_for_tests(run["id"], module_id="CP-0", digest="0" * 64)

    with pytest.raises(Exception, match="DELIVERABLE_COMPOSITION_REQUIRED_VALUE_MISSING:artifact.CP-0"):
        service.approve_filing(
            case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user",
        )
    assert service.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN"


def test_gate_recomputes_stored_digest_and_refuses_tampered_frozen_content(service, store):
    case, revision, frozen = frozen_deliverable(service, store)
    service.tamper_frozen_payload_for_tests(frozen["deliverable_id"])
    with pytest.raises(Exception, match="PREVIEW|INTEGRITY"):
        service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user")
    assert not [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.filed"], "tamper-evident: nothing filed"


def test_freeze_refuses_stored_revision_content_tampered_without_its_indexed_digest(
    service, store,
):
    import sqlalchemy as sa

    from caos.storage.deliverables import deliverable_revisions

    case, _source, _template, revision = save_min_draft(service, store)
    tampered = copy.deepcopy(revision["content"])
    tampered["unreviewed_direct_row_edit"] = True
    with service.records.engine.begin() as connection:
        connection.execute(
            sa.update(deliverable_revisions)
            .where(deliverable_revisions.c.revision_id == revision["revision_id"])
            .values(content=tampered)
        )

    with pytest.raises(ValueError, match="DELIVERABLE_REVISION_INTEGRITY_FAILED"):
        service.freeze(case["id"], freeze_request(revision), actor="analyst")

    assert service.records.frozen_for_pathway(case["id"], "FULL_CREDIT") == []


def test_filing_rechecks_frozen_draft_digest_even_if_other_preview_digests_are_rebound(
    service, store,
):
    import sqlalchemy as sa

    from caos.contracts import digest
    from caos.deliverables.graph import frozen_approval_digest
    from caos.storage.deliverables import deliverable_frozen

    case, _revision, frozen = frozen_deliverable(service, store)
    payload = copy.deepcopy(frozen["payload"])
    payload["content"]["unreviewed_direct_row_edit"] = True
    payload["preview_digest"] = digest({
        key: value for key, value in payload.items() if key != "preview_digest"
    })
    rebound = {**frozen, "payload": payload}
    rebound["preview_digest"] = frozen_approval_digest(rebound)
    with service.records.engine.begin() as connection:
        connection.execute(
            sa.update(deliverable_frozen)
            .where(deliverable_frozen.c.deliverable_id == frozen["deliverable_id"])
            .values(payload=payload, preview_digest=rebound["preview_digest"])
        )

    with pytest.raises(ValueError, match="DELIVERABLE_PREVIEW_INTEGRITY_FAILED"):
        service.approve_filing(
            case["id"], frozen["deliverable_id"], file_request(rebound), actor="approver-user",
        )

    assert service.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN"
    assert service.thread_state_for_tests(frozen["thread_id"])["status"] == "PARKED"


def test_resume_outside_a_parked_filing_gate_is_refused(service, store):
    case, revision, frozen = frozen_deliverable(service, store)
    service.terminate_filing_thread_for_tests(frozen["thread_id"])  # thread no longer paused at the gate
    with pytest.raises(Exception, match="RESUME_NOT_APPLIED|NOT_PAUSED"):
        service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user")
    assert not [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.filed"]


def test_approval_revalidates_current_accepted_snapshot_and_stale_leaves_no_residue(service, store):
    case, revision, frozen = frozen_deliverable(service, store)
    service.supersede_accepted_authority_for_tests(case["id"])  # newer accepted snapshot after freeze
    with pytest.raises(Exception, match="FROZEN_AUTHORITY_STALE"):
        service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user")
    record = service.frozen_record(case["id"], frozen["deliverable_id"])
    assert record["status"] == "FROZEN", "stale rejection leaves the record intact and retryable (after re-freeze)"
    assert not [e for e in service.audit_events_for_tests(case["id"]) if e["action"] == "deliverable.filed"], "no audit residue"


def test_authority_move_during_approval_cannot_cross_the_filing_cas(
    service,
    store,
    monkeypatch,
):
    case, _revision, frozen = frozen_deliverable(service, store)
    resolve = service._resolve_stored_selection

    def move_authority(case_id, stored):
        resolved = resolve(case_id, stored)
        service.supersede_accepted_authority_for_tests(case_id)
        return resolved

    monkeypatch.setattr(service, "_resolve_stored_selection", move_authority)
    with pytest.raises(ValueError, match="FROZEN_AUTHORITY_STALE"):
        service.approve_filing(
            case["id"],
            frozen["deliverable_id"],
            file_request(frozen),
            actor="approver-user",
        )
    assert service.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN"


def test_stale_or_duplicate_resume_returns_resume_not_applied_never_success(client, settings, store):
    svc, case, source, template = http_seed(settings, store)
    revision = svc.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst")
    frozen = svc.freeze(case["id"], freeze_request(revision), actor="analyst")
    store.add_member(case["id"], "analyst", "approver-user", "APPROVER", actor_role="ADMIN")
    url = f"/api/cases/{case['id']}/deliverables/by-id/{frozen['deliverable_id']}/approve"
    body = {"preview_digest": frozen["preview_digest"], "input_fingerprint": frozen["input_fingerprint"]}
    first = client.post(url, json=body, headers=APPROVER_H)
    assert first.status_code == 200 and first.json()["status"] == "FILED"
    second = client.post(url, json=body, headers=APPROVER_H)
    assert second.status_code == 409, "a resume that advances nothing must never surface as success (§12.22)"
    detail = second.json()["detail"]
    assert detail["code"] == "RESUME_NOT_APPLIED"
    assert "current_interrupt_id" in detail
    assert svc.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FILED", "still exactly one filing"


def test_later_filing_supersedes_prior_with_pointer_and_terminalizes_its_thread(service, store):
    case, source, r1, frozen_a = freeze_min(service, store)  # frozen A parks its gate, never filed
    add_approver(store, case)
    template = service.templates()["FULL_CREDIT"]
    template[_TEST_MODEL_SELECTION] = r1["content"]["model_selection"]

    r2 = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, blocks=required_blocks(template, source, narrative_text="Updated opinion."), expected_version=r1["version"]),
        actor="analyst",
    )
    frozen_b = service.freeze(case["id"], freeze_request(r2), actor="analyst")
    service.approve_filing(case["id"], frozen_b["deliverable_id"], file_request(frozen_b), actor="approver-user")

    prior = service.frozen_record(case["id"], frozen_a["deliverable_id"])
    assert prior["status"] == "SUPERSEDED", "a parked frozen record is superseded by the later filing"
    assert prior["superseded_by_id"] == frozen_b["deliverable_id"], "supersession carries a pointer"
    thread = service.thread_state_for_tests(frozen_a["thread_id"])
    assert thread["outcome"] == "SUPERSEDED", "prior thread terminalized with the typed outcome — no zombie interrupts (§10.5)"

    r3 = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, blocks=required_blocks(template, source, narrative_text="Final opinion."), expected_version=r2["version"]),
        actor="analyst",
    )
    frozen_c = service.freeze(case["id"], freeze_request(r3), actor="analyst")
    service.approve_filing(case["id"], frozen_c["deliverable_id"], file_request(frozen_c), actor="approver-user")
    prior_filed = service.frozen_record(case["id"], frozen_b["deliverable_id"])
    assert prior_filed["status"] == "SUPERSEDED", "a FILED record is likewise superseded by the next filing"
    assert prior_filed["superseded_by_id"] == frozen_c["deliverable_id"]


def test_http_request_changes_requires_approver_and_nonblank_comment_and_appends_commented_draft(client, settings, store):
    svc, case, source, template = http_seed(settings, store)
    revision = svc.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst")
    frozen = svc.freeze(case["id"], freeze_request(revision), actor="analyst")
    store.add_member(case["id"], "analyst", "approver-user", "APPROVER", actor_role="ADMIN")
    url = f"/api/cases/{case['id']}/deliverables/by-id/{frozen['deliverable_id']}/request-changes"
    base = {"preview_digest": frozen["preview_digest"], "input_fingerprint": frozen["input_fingerprint"]}
    assert client.post(url, json={**base, "comment": "   "}, headers=APPROVER_H).status_code == 422, "blank comment refused"
    good = {**base, "comment": "Tighten the covenant discussion."}
    assert client.post(url, json=good, headers=ANALYST_H).status_code == 403, "request-changes is approver-only"
    resp = client.post(url, json=good, headers=APPROVER_H)
    assert resp.status_code == 200
    assert resp.json()["frozen"]["status"] == "CHANGES_REQUESTED"
    assert resp.json()["frozen"]["change_request"]["comment"] == good["comment"]
    assert resp.json()["draft"]["version"] == revision["version"] + 1
    draft = svc.workspace(case["id"], "FULL_CREDIT")["draft"]
    assert draft["version"] == revision["version"] + 1, "a traceable replacement draft is appended"
    assert "Tighten the covenant discussion." in str(draft["content"]), "the replacement draft carries the comment"


def test_change_request_revision_conflict_rolls_back_frozen_and_thread_transition(
    service, store, monkeypatch,
):
    import threading

    from caos.contracts import RequestDeliverableChangesRequest
    from caos.storage.deliverables import DeliverableVersionConflict

    case, source, revision, frozen = freeze_min(service, store)
    head_read = threading.Event()
    release_request = threading.Event()
    real_head_revision = service.records.head_revision

    def pause_change_request_after_head_read(case_id, pathway):
        head = real_head_revision(case_id, pathway)
        if threading.current_thread().name == "change-request":
            head_read.set()
            assert release_request.wait(5), "test did not release the paused change request"
        return head

    monkeypatch.setattr(service.records, "head_revision", pause_change_request_after_head_read)
    request = RequestDeliverableChangesRequest(
        preview_digest=frozen["preview_digest"],
        input_fingerprint=frozen["input_fingerprint"],
        comment="Address the newer analyst view.",
    )
    request_error = []

    def request_changes():
        try:
            service.request_changes(
                case["id"], frozen["deliverable_id"], request, actor="approver-user",
            )
        except Exception as exc:  # noqa: BLE001 — asserted after joining the thread
            request_error.append(exc)

    worker = threading.Thread(target=request_changes, name="change-request")
    worker.start()
    assert head_read.wait(5), "change request did not reach its revision CAS"
    template = service.templates()["FULL_CREDIT"]
    try:
        analyst_revision = service.save_draft(
            case["id"],
            "FULL_CREDIT",
            draft_request(
                template,
                blocks=required_blocks(template, source, narrative_text="Concurrent analyst revision."),
                expected_version=revision["version"],
                model_selection=revision["content"]["model_selection"],
            ),
            actor="analyst",
        )
    finally:
        release_request.set()
    worker.join(5)

    assert not worker.is_alive()
    assert len(request_error) == 1 and isinstance(request_error[0], DeliverableVersionConflict)
    assert request_error[0].current["revision_id"] == analyst_revision["revision_id"]
    assert service.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN"
    assert service.thread_state_for_tests(frozen["thread_id"])["status"] == "PARKED"
    assert [item["revision_id"] for item in service.revision_history(case["id"], "FULL_CREDIT")] == [
        revision["revision_id"], analyst_revision["revision_id"],
    ]
    assert not [
        event for event in service.audit_events_for_tests(case["id"])
        if event["action"] == "deliverable.changes_requested"
    ]


def test_http_filing_gate_denies_outsiders_404_readers_403_analysts_403_for_every_global_role(client, settings, store):
    svc, case, source, template = http_seed(settings, store)
    revision = svc.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst")
    frozen = svc.freeze(case["id"], freeze_request(revision), actor="analyst")
    store.add_member(case["id"], "analyst", "stored-reader", "READER", actor_role="ADMIN")
    store.add_member(case["id"], "analyst", "stored-analyst", "ANALYST", actor_role="ADMIN")
    url = f"/api/cases/{case['id']}/deliverables/by-id/{frozen['deliverable_id']}/approve"
    body = {"preview_digest": frozen["preview_digest"], "input_fingerprint": frozen["input_fingerprint"]}
    for global_role in ("READER", "ANALYST", "APPROVER", "ADMIN"):
        outsider = client.post(url, json=body, headers={"x-caos-role": global_role, "x-forwarded-user": "outsider"})
        assert outsider.status_code == 404, f"non-member with global {global_role} must not learn the case exists"
        reader = client.post(url, json=body, headers={"x-caos-role": global_role, "x-forwarded-user": "stored-reader"})
        assert reader.status_code == 403, f"case READER never approves — global {global_role} does not escalate"
    analyst_member = client.post(url, json=body, headers={"x-caos-role": "APPROVER", "x-forwarded-user": "stored-analyst"})
    assert analyst_member.status_code == 403, "filing is approver-gated: case ANALYST standing is insufficient (DECISIONS §2)"
    assert svc.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN"


def test_case_approver_standing_suffices_across_all_global_writer_roles(client, settings, store):
    """Positive half of the gate authz matrix, adapted to the approver-gated filing
    interrupt: every approver-capable stored case role (APPROVER, ADMIN) crossed with
    every global writer role can approve — case standing governs, global role is inert."""
    for case_role in ("APPROVER", "ADMIN"):
        for global_role in ("ANALYST", "APPROVER", "ADMIN"):
            svc, case, source, template = http_seed(settings, store)
            revision = svc.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst")
            frozen = svc.freeze(case["id"], freeze_request(revision), actor="analyst")
            subject = f"{case_role.lower()}-{global_role.lower()}"
            store.add_member(case["id"], "analyst", subject, case_role, actor_role="ADMIN")
            url = f"/api/cases/{case['id']}/deliverables/by-id/{frozen['deliverable_id']}/approve"
            body = {"preview_digest": frozen["preview_digest"], "input_fingerprint": frozen["input_fingerprint"]}
            resp = client.post(url, json=body, headers={"x-caos-role": global_role, "x-forwarded-user": subject})
            assert resp.status_code == 200, f"case {case_role} + global {global_role} must be able to approve"
            assert resp.json()["status"] == "FILED"


# --- filed exports ----------------------------------------------------------------


def test_filed_exports_are_byte_exact_sha_verified_and_never_rerendered(client, settings, store):
    svc, case, source, template = http_seed(settings, store)
    revision = svc.save_draft(case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst")
    frozen = svc.freeze(case["id"], freeze_request(revision), actor="analyst")
    store.add_member(case["id"], "analyst", "approver-user", "APPROVER", actor_role="ADMIN")
    export_url = f"/api/cases/{case['id']}/deliverables/by-id/{frozen['deliverable_id']}/export/md"
    assert client.get(export_url, headers=ANALYST_H).status_code == 409, "pre-filing downloads are refused"

    approve_url = f"/api/cases/{case['id']}/deliverables/by-id/{frozen['deliverable_id']}/approve"
    body = {"preview_digest": frozen["preview_digest"], "input_fingerprint": frozen["input_fingerprint"]}
    assert client.post(approve_url, json=body, headers=APPROVER_H).status_code == 200

    resp = client.get(export_url, headers=ANALYST_H)
    assert resp.status_code == 200
    recorded = frozen["exports"]["md"]
    assert hashlib.sha256(resp.content).hexdigest() == recorded["sha256"], "served bytes match the frozen record"
    assert len(resp.content) == recorded["size"]

    # never re-rendered: a fresh instance whose renderer explodes still serves identical bytes
    exploding = make_service(store, settings.storage_dir, renderer_for_tests=_never_render)
    data, sha = exploding.export(frozen["deliverable_id"], "md")
    assert data == resp.content and sha == recorded["sha256"]


def test_tampered_or_missing_stored_export_fails_closed_with_no_audit_residue(service, store):
    case, revision, frozen = frozen_deliverable(service, store)
    service.approve_filing(case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user")
    before = service.audit_events_for_tests(case["id"])
    service.tamper_export_for_tests(frozen["deliverable_id"], "md")
    with pytest.raises(Exception, match="EXPORT_INTEGRITY_FAILED"):
        service.export(frozen["deliverable_id"], "md")
    service.delete_export_for_tests(frozen["deliverable_id"], "pdf")
    with pytest.raises(Exception, match="EXPORT_UNAVAILABLE"):
        service.export(frozen["deliverable_id"], "pdf")
    assert service.audit_events_for_tests(case["id"]) == before, "failed downloads mutate nothing and leave no audit residue"


def test_filing_refuses_a_missing_frozen_export_and_keeps_the_gate_parked(service, store):
    case, _revision, frozen = frozen_deliverable(service, store)
    service.delete_export_for_tests(frozen["deliverable_id"], "pdf")

    with pytest.raises(ValueError, match="DELIVERABLE_EXPORT_UNAVAILABLE"):
        service.approve_filing(
            case["id"], frozen["deliverable_id"], file_request(frozen), actor="approver-user"
        )

    assert service.frozen_record(case["id"], frozen["deliverable_id"])["status"] == "FROZEN"
    assert service.thread_state_for_tests(frozen["thread_id"])["status"] == "PARKED"


def test_filing_refuses_export_metadata_substituted_from_another_reviewed_draft(service, store):
    import sqlalchemy as sa

    from caos.storage.deliverables import deliverable_frozen

    case, source, _authority = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    bind_default_model_for_tests(service, case, template)
    first_revision = service.save_draft(
        case["id"], "FULL_CREDIT", draft_request(template, source), actor="analyst"
    )
    first = service.freeze(case["id"], freeze_request(first_revision), actor="analyst")
    second_revision = service.save_draft(
        case["id"],
        "FULL_CREDIT",
        draft_request(
            template,
            blocks=required_blocks(template, source, narrative_text="Different reviewed conclusion."),
            expected_version=first_revision["version"],
        ),
        actor="analyst",
    )
    second = service.freeze(case["id"], freeze_request(second_revision), actor="analyst")
    assert service.export(first["deliverable_id"], "md")[0] != service.export(second["deliverable_id"], "md")[0]

    with service.records.engine.begin() as connection:
        connection.execute(
            sa.update(deliverable_frozen)
            .where(deliverable_frozen.c.deliverable_id == first["deliverable_id"])
            .values(exports=second["exports"])
        )
    audit_before = service.audit_events_for_tests(case["id"])

    with pytest.raises(ValueError, match="DELIVERABLE_PREVIEW_INTEGRITY_FAILED"):
        service.approve_filing(
            case["id"], first["deliverable_id"], file_request(first), actor="approver-user"
        )

    assert service.frozen_record(case["id"], first["deliverable_id"])["status"] == "FROZEN"
    assert service.thread_state_for_tests(first["thread_id"])["status"] == "PARKED"
    assert service.audit_events_for_tests(case["id"]) == audit_before

    with service.records.engine.begin() as connection:
        connection.execute(
            sa.update(deliverable_frozen)
            .where(deliverable_frozen.c.deliverable_id == first["deliverable_id"])
            .values(exports=first["exports"])
        )
    service.approve_filing(
        case["id"], first["deliverable_id"], file_request(first), actor="approver-user"
    )
    with service.records.engine.begin() as connection:
        connection.execute(
            sa.update(deliverable_frozen)
            .where(deliverable_frozen.c.deliverable_id == first["deliverable_id"])
            .values(exports=second["exports"])
        )
    with pytest.raises(ValueError, match="DELIVERABLE_PREVIEW_INTEGRITY_FAILED"):
        service.export(first["deliverable_id"], "pdf")


def test_xlsx_export_neutralizes_formula_text_and_preserves_typed_model_values(service, store):
    case, source, _ = seed_ready_case(service, store)
    template = service.templates()["FULL_CREDIT"]
    model = seed_model(service, case, outputs={"total_leverage": 4.2})
    blocks = required_blocks(template, source, narrative_text='=HYPERLINK("http://evil.example","click me") injected narrative')
    revision = service.save_draft(
        case["id"], "FULL_CREDIT",
        draft_request(template, blocks=blocks + [optional_block(template, "GENERATED_METRIC")], model_selection=revision_selection(model)),
        actor="analyst",
    )
    frozen = service.freeze(case["id"], freeze_request(revision), actor="analyst")
    data, _ = service.export(frozen["deliverable_id"], "xlsx")
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(data))
    injected = None
    numeric = None
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                assert cell.data_type != "f", f"live formula in {sheet.title}!{cell.coordinate}"
                if isinstance(cell.value, str) and "HYPERLINK" in cell.value:
                    injected = cell
                if cell.value == 4.2:
                    numeric = cell
    assert injected is not None and injected.data_type == "s", "=-prefixed analyst text is an escaped literal string"
    assert numeric is not None and isinstance(numeric.value, float), "generated model numbers stay typed numbers"


# ROW MAPPING
# <legacy test> -> <spec test(s) in this file>
#
# test_deliverables.py:
#   test_six_pathway_templates_have_stable_identity_and_evidence_register -> test_template_registry_serves_six_pathways_with_stable_identity_and_evidence_register
#   test_deliverable_revisions_are_append_only_and_conflict_atomically -> test_stale_expected_version_is_atomic_cas_conflict_carrying_current_head
#   test_strict_draft_contract_rejects_client_generated_values -> test_strict_schema_rejects_client_supplied_generated_values
#   test_domain_validates_template_order_and_saves_complete_shared_draft -> test_saved_revision_digest_is_digest_of_content_and_order_violation_appends_nothing
#   test_restore_creates_a_new_revision_without_mutating_history -> test_restore_appends_new_revision_and_history_stays_byte_identical
#   test_evidence_citations_reject_cross_case_withdrawn_and_unknown_blocks -> test_citations_resolve_same_case_existing_non_withdrawn_blocks
#   test_active_revision_and_acknowledged_fallback_model_eligibility -> test_model_selection_pins_current_revision_and_fallback_requires_acknowledged_no_revision_state
#   test_selected_model_must_resolve_to_exact_pinned_record -> test_selection_must_resolve_to_exact_stored_revision_and_build_records
#   test_all_model_dependent_optional_blocks_require_and_pin_selected_model -> test_model_dependent_block_kinds_require_selection_and_are_digest_pinned (parametrized x4)
#   test_template_owns_exact_optional_slot_kind_and_order_policy -> test_template_owns_optional_block_kind_stem_cap_order_and_model_dependence
#   test_template_rejects_client_invented_optional_kind_or_slot -> test_client_invented_optional_kind_or_slot_is_rejected (parametrized x3)
#   test_template_rejects_invalid_optional_order -> test_optional_blocks_out_of_template_declared_order_are_rejected
#   test_template_accepts_declared_non_model_limitations_slot -> test_declared_limitations_slot_saves_without_model_identity
#   test_generated_metric_and_scenario_are_rebuilt_from_exact_current_model -> test_generated_metric_outputs_are_rebuilt_server_side_from_pinned_revision + test_forged_scenario_outputs_raise_calculation_mismatch (split: metric rebuild vs scenario forgery)
#   test_scenario_exhibit_requires_selected_model_before_calculation -> test_scenario_exhibit_requires_selected_model_before_any_calculation
#   test_scenario_base_binds_exactly_to_selected_model_without_persistence -> test_scenario_base_identity_mismatch_fails_before_calculation_with_zero_residue
#   test_application_build_scenario_accepts_null_base_revision -> test_application_build_fallback_scenario_accepts_null_base_revision
#   test_generated_blocks_reject_ungoverned_metric_and_recipe_fields -> test_ungoverned_metric_ids_and_chart_recipe_fields_are_rejected
#   test_scenario_exhibit_requires_exact_server_calculation_digest -> test_scenario_digest_must_equal_server_computed_digest
#   test_http_deliverable_current_history_by_id_and_recoverable_conflict -> test_http_stale_draft_put_returns_409_with_current_revision_and_by_id_read_round_trips
#   test_http_case_reader_can_read_but_cannot_write_deliverable -> test_case_reader_reads_workspace_but_cannot_write_draft
#
# test_deliverable_exports.py:
#   test_all_six_pathways_render_substantive_semantic_exports -> test_each_pathway_frozen_payload_renders_substantive_md_pdf_xlsx (parametrized x6)
#   test_freeze_pins_and_renders_exact_active_revision_model_authority -> test_freeze_embeds_selected_revision_outputs_assumptions_and_build_payload_verbatim
#   test_frozen_filed_history_and_request_changes_are_atomic -> test_concurrent_filers_yield_exactly_one_filed_with_payload_equal_to_frozen + test_later_filing_supersedes_prior_with_pointer_and_terminalizes_its_thread + test_http_request_changes_requires_approver_and_nonblank_comment_and_appends_commented_draft (split: race, supersession, change-request)
#   test_exact_draft_freeze_is_one_idempotent_record_under_race -> test_freeze_is_idempotent_under_race_with_one_record_one_thread_one_audit_event + test_divergent_render_for_same_freeze_identity_is_freeze_conflict (split: idempotency vs conflict)
#   test_http_freeze_file_and_download_exact_stored_bytes -> test_filed_exports_are_byte_exact_sha_verified_and_never_rerendered + test_tampered_or_missing_stored_export_fails_closed_with_no_audit_residue (approver gating of the filing itself is asserted in test_http_filing_gate_denies_outsiders_404_readers_403_analysts_403_for_every_global_role)
#   test_approval_revalidates_current_accepted_authority_without_residue -> test_approval_revalidates_current_accepted_snapshot_and_stale_leaves_no_residue
#   test_http_request_changes_requires_approver_comment_and_appends_draft -> test_http_request_changes_requires_approver_and_nonblank_comment_and_appends_commented_draft
#   test_xlsx_neutralizes_formula_text_and_preserves_typed_model_values -> test_xlsx_export_neutralizes_formula_text_and_preserves_typed_model_values
#
# test_ledger_contracts.py (deliverable rows):
#   test_deliverable_revision_lookup_preserves_immutable_history -> test_stored_revisions_are_isolated_from_later_caller_mutation
#   test_report_freeze_and_approval_require_exact_preview -> test_approval_binds_exact_preview_digest_and_fingerprint_mismatch_leaves_frozen_retryable (re-hosted from the cut report era onto the deliverable filing gate per DECISIONS §1/§10.10; bit-identical + audit halves also in test_approval_leaves_frozen_content_bit_identical_and_appends_one_audit_event)
#
# test_cp_dr_planning.py (eight approval-gate rows re-hosted onto the filing interrupt per DECISIONS §11.9):
#   test_research_brief_rejects_caller_owned_authority_fields -> test_http_approval_payload_rejects_caller_owned_authority_fields
#   test_plan_approval_request_requires_exact_canonical_hash -> test_gate_approval_hash_is_canonical_sha256_form
#   test_exact_plan_approval_preserves_identity_and_phase4_fails_closed -> test_approval_leaves_frozen_content_bit_identical_and_appends_one_audit_event
#   test_plan_approval_rejects_wrong_double_wrong_phase_and_changed_source_set -> test_approval_binds_exact_preview_digest_and_fingerprint_mismatch_leaves_frozen_retryable (wrong hash) + test_gate_rejects_source_set_drift_between_freeze_and_filing (changed source set) + test_resume_outside_a_parked_filing_gate_is_refused (wrong phase)
#   test_plan_approval_rejects_a_post_hash_plan_change_and_missing_pause -> test_gate_recomputes_stored_digest_and_refuses_tampered_frozen_content (post-hash change) + test_resume_outside_a_parked_filing_gate_is_refused (missing pause)
#   test_research_planning_requires_upstream_artifact_identity -> test_freeze_fails_closed_without_upstream_accepted_authority
#   test_research_plan_route_denies_cross_case_outsider_and_stored_reader -> test_http_filing_gate_denies_outsiders_404_readers_403_analysts_403_for_every_global_role
#   test_all_case_writer_role_combinations_can_approve_research_plan -> test_case_approver_standing_suffices_across_all_global_writer_roles
#       ADAPTATION: the filing gate is approver-gated (DECISIONS §2), so the approver-capable case-writer set is {APPROVER, ADMIN} x all global writer roles; the legacy matrix's case-ANALYST-can-approve third is deliberately inverted and asserted as 403 in the denial test above. Not inexpressible — re-scoped.
#
# DECISIONS-only guarantees carried by this file (no single legacy row):
#   §12.23 deliverable thread identity includes build_id; filing never re-renders -> test_filing_thread_id_is_a_deterministic_digest_including_build_id + test_filed_exports_are_byte_exact_sha_verified_and_never_rerendered
#   §12.21 resume ticket single-effect -> test_concurrent_filers_yield_exactly_one_filed_with_payload_equal_to_frozen
#   §12.22 RESUME_NOT_APPLIED on stale/duplicate resume -> test_stale_or_duplicate_resume_returns_resume_not_applied_never_success
#   §10.5 supersession terminalizes the parked thread (typed SUPERSEDED) -> test_later_filing_supersedes_prior_with_pointer_and_terminalizes_its_thread
#
# Inexpressible rows: none.


def test_frozen_pdf_is_structurally_complete_under_optimized_python():
    """The /Pages object used to be created by the call inside an `assert`.
    `python -O` drops the whole expression, so every page dangles against a
    parent that was never written — a corrupt PDF whose sha256 still matches
    the frozen record. Render in a real -O interpreter and read it back."""
    import subprocess
    import sys
    from pathlib import Path

    server = str(Path(__file__).resolve().parents[2] / "server")
    script = (
        "import sys, io, json;"
        f"sys.path.insert(0, {server!r});"
        "from caos.publishing.renderers import render_frozen_pdf;"
        "payload={'pathway':'FULL_CREDIT','content':{'blocks':[{'block_id':'b','slot_id':'s',"
        "'kind':'NARRATIVE','text':'Pinned narrative body.'}]},'template':{'block_titles':{}},"
        "'draft':{'version':1,'digest':'d'},'preview_digest':'p','input_fingerprint':'f',"
        "'methodology':{'build_id':'b'}};"
        "content=render_frozen_pdf(payload);"
        "from pypdf import PdfReader;"
        "reader=PdfReader(io.BytesIO(content));"
        "print(json.dumps({'pages':len(reader.pages),"
        "'text':''.join(page.extract_text() or '' for page in reader.pages)}))"
    )
    completed = subprocess.run([sys.executable, "-O", "-c", script], capture_output=True, text=True)
    assert completed.returncode == 0, completed.stderr
    result = json.loads(completed.stdout)
    assert result["pages"] == 1 and "Pinned narrative body." in result["text"]
