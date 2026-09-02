"""Source-complete modelling for every pathway (Task 9; ETR-B12; CALC-001–020).

One answer-keyed provider double at the ordinary provider port drives the
document-first journey end to end: `POST /api/intake` admits a pack, the host
classifies it and starts the route, every supplied document is read through
`read_evidence`, the canonical modules emit the golden CP-MODEL fixtures
re-identified from the documents that are actually present, acceptance queues
the pathway's declared model effect, and the build carries the source lineage
that proves which document reached which model input, calculation or cited
analysis. The double's identity is deliberately not `host_control` and no test
here calls `run_scripted_for_tests` or any placeholder capability.

The answer key is the pack itself: an annual report is the canonical data
source, a quarterly report reaches the CP-1B snapshot, management guidance is
the forecast-driver authority, a restated annual supersedes the original and
changes the forecast margin, a press clipping is evidence the host records but
nothing consumes. The metamorphic cases below remove or add one class at a
time and assert exactly what the key says changes — input, artifact, model
fingerprint, output, limitation or refusal, and audit lineage — and that the
irrelevant file perturbs no result while never being silently discarded.
"""

from __future__ import annotations

import copy
import io
import json
import sys
from datetime import date
from pathlib import Path
from typing import Any

import pytest
import sqlalchemy as sa
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

SERVER = Path(__file__).resolve().parents[2] / "server"
if str(SERVER) not in sys.path:
    sys.path.insert(0, str(SERVER))
TESTS = Path(__file__).resolve().parents[1]
if str(TESTS) not in sys.path:
    sys.path.insert(0, str(TESTS))

from calculator_fixtures import VALID_CALCULATION_INPUTS, calculation_ref  # noqa: E402
from test_intake_spec import (  # noqa: E402
    ANALYST,
    ISSUER,
    TEXT,
    XLSX,
    amendment,
    annual_report,
    credit_agreement,
    earnings_release,
    guidance,
    quarterly_report,
    research_brief,
    submit,
)
from test_loan_universe_spec import _row, _sheet  # noqa: E402

FIXTURES = TESTS / "fixtures" / "cp_model"
MODEL_FIXTURES = {
    "CP-1": "cp1.md", "CP-1A": "cp1a.md", "CP-1B": "cp1b.md",
    "CP-2": "cp2.md", "CP-2A": "cp2a.md", "CP-2G": "cp2g.md",
}
HEADINGS = ("Audit Summary", "Analysis", "Evidence Trace", "Source Registry", "Gaps & Conflicts", "QA Validation")
# Filename prefixes are the double's answer key; the bytes carry the signals
# the host classifier reads. The two must agree for a pack to be meaningful.
# Filenames are fixed per kind; the issuer varies so a comparison pack lands in
# a fresh case instead of resolving into the baseline case by issuer.
OTHER_ISSUER = "Southstar Holdings"
FILENAMES = {
    "annual": "annual-fy2024.txt",
    "restated": "annual-restated-fy2024.txt",
    "quarterly": "quarterly-q1-2025.txt",
    "quarterly_q2": "quarterly-q2-2025.txt",
    "guidance": "guidance-fy2025.txt",
    "agreement": "agreement-credit.txt",
    "amendment": "amendment-2.txt",
    "earnings": "earnings-q2-2025.txt",
    "other": "other-press-clipping.txt",
    "brief": "brief-research.json",
    "marks": "marks-REF_CP-3_Sector_RV.xlsx",
}


def doc(kind: str, issuer: str = ISSUER) -> tuple[str, bytes, str]:
    content = {
        "annual": lambda: annual_report(issuer),
        "restated": lambda: annual_report(issuer, restated=True),
        "quarterly": lambda: quarterly_report(issuer, quarter=1),
        "quarterly_q2": lambda: quarterly_report(issuer, quarter=2),
        "guidance": lambda: guidance(issuer),
        "agreement": lambda: credit_agreement(issuer),
        "amendment": lambda: amendment(issuer),
        "earnings": lambda: earnings_release(issuer, quarter=2),
        "other": lambda: f"{issuer}\nSponsorship announcement for a regatta. No financial content.\n".encode(),
        "brief": lambda: research_brief(),
    }[kind]()
    media = "application/json" if kind == "brief" else TEXT
    return FILENAMES[kind], content, media


ANNUAL, QUARTERLY, GUIDANCE, AGREEMENT = doc("annual"), doc("quarterly"), doc("guidance"), doc("agreement")
RESTATED, AMENDMENT, EARNINGS, QUARTERLY_Q2, OTHER, BRIEF = (
    doc("restated"), doc("amendment"), doc("earnings"), doc("quarterly_q2"), doc("other"), doc("brief"),
)
BASE_PACK = [ANNUAL, QUARTERLY, GUIDANCE, AGREEMENT]
QUARTERLY_SNAPSHOT_ROW = (
    "| historical_performance | Revenue and EBITDA increased through FY2024 "
    "| READY | SRC-1 | Annual report 2024 p. 42 | 2024-12-31 |"
)


def _kind(filename: str) -> str:
    return "restated" if filename.startswith("annual-restated") else filename.split("-", 1)[0]


def marks_workbook(workbook_date: date = date(2026, 8, 24)) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "IT Services", [_row()], workbook_date=workbook_date)
    _sheet(
        workbook, "Healthcare IT",
        [_row(borrower="FinThrive Inc", bloomberg="BLS1989347", figi="BBG01THRIVE1")],
        workbook_date=workbook_date,
    )
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class AnswerKeyProvider:
    """Answer-keyed provider at the ordinary port: reads every pinned source,
    runs every assigned calculator with the fixture inputs, and writes the
    canonical modules from the golden CP-MODEL fixtures re-identified to the
    documents actually in the pack."""

    def __init__(self) -> None:
        from caos.contracts import digest
        from caos.engine.provider import ProviderIdentity

        self.identity = ProviderIdentity(
            provider_name="answer_key_fixture",
            model="source-complete-answer-key-v1",
            provider_version="1",
            adapter_version="caos.answer-key-fixture.v1",
            parameter_context_digest=digest({"provider": "answer-key-fixture-v1"}),
            qualification_record_id=None,
            qualification_record_digest=None,
            qualification_status="unqualified",
            qualification_expires_at=None,
        )
        self.calls = 0
        self.modules_by_run: dict[str, set[str]] = {}
        # Kinds the double never reads (the host requires every delivered block
        # to be cited, so "unused" means "never requested"): the clipping by
        # answer key, plus whatever a test adds to prove the lineage oracle bites.
        self.unread_kinds: set[str] = {"other"}

    def count_tokens(self, _request: Any) -> int:
        return 100

    @staticmethod
    def _prompt(request: Any) -> dict[str, Any]:
        content = request.messages[0]["content"]
        return json.loads(content.split("\n", 1)[1])

    @staticmethod
    def _tool_results(request: Any) -> list[Any]:
        return [
            json.loads(block["content"])
            for message in request.messages
            if isinstance(message.get("content"), list)
            for block in message["content"]
            if isinstance(block, dict) and block.get("type") == "tool_result"
        ]

    def create_message(self, request: Any):
        from caos.engine.provider import ProviderBlock

        self.calls += 1
        prompt = self._prompt(request)
        module_id, run_id = prompt["host_identity"]["module_id"], prompt["host_identity"]["run_id"]
        self.modules_by_run.setdefault(run_id, set()).add(module_id)
        manifest = [source for source in prompt["source_metadata_manifest"] if source.get("blocks")]
        results = self._tool_results(request)
        evidence_results = [result for result in results if isinstance(result, list)]
        delivered = {row["source_id"] for result in evidence_results for row in result}
        remaining = [
            source for source in manifest
            if source["source_id"] not in delivered and _kind(source["filename"]) not in self.unread_kinds
        ]
        if remaining:
            source = remaining[0]
            return self._tool_call("read_evidence", {
                "source_id": source["source_id"], "block_ids": [source["blocks"][0]["block_id"]],
            })
        records = [result for result in results if isinstance(result, dict) and "output_digest" in result]
        calculation_tool = next(
            (tool for tool in request.effective_tools() if tool["name"] == "run_methodology_calculation"),
            None,
        )
        if calculation_tool is not None:
            assigned = calculation_tool["input_schema"]["properties"]["calculator_id"]["enum"]
            if len(records) < len(assigned):
                calculator_id = assigned[len(records)]
                return self._tool_call("run_methodology_calculation", {
                    "calculator_id": calculator_id,
                    "input_json": json.dumps(VALID_CALCULATION_INPUTS[calculator_id], sort_keys=True),
                })
        rows = [row for result in evidence_results for row in result]
        docs = {
            source["source_id"]: {
                "kind": _kind(source["filename"]),
                "source_id": source["source_id"],
                "block_id": source["blocks"][0]["block_id"],
                "sha256": source["sha256"],
            }
            for source in manifest
        }
        cited = rows  # every delivered block is cited: the host refuses anything else
        annual = next((doc for doc in docs.values() if doc["kind"] == "annual"), None)
        markdown, gate = self._markdown(module_id, run_id, docs, cited, prompt["host_identity"])
        body = {
            "markdown": markdown,
            "evidence_refs": [{"source_id": row["source_id"], "block_id": row["block_id"]} for row in cited],
            "calculation_refs": [calculation_ref(record) for record in records],
            "lineage_counts": {"directly_sourced": len(cited)},
            "fields_present": len(cited),
            "fields_total": max(1, len(cited)),
            "source_gate": gate,
            "findings": {},
        }
        del annual
        return self._message([ProviderBlock(type="text", text=json.dumps(body))], stop_reason="end_turn")

    def _markdown(self, module_id, run_id, docs, cited, host_identity) -> tuple[str, str]:
        kinds = {doc["kind"]: doc for doc in docs.values()}
        primary = kinds.get("restated") or kinds.get("annual")
        if module_id in MODEL_FIXTURES:
            if primary is None:
                # No annual report: the canonical data foundation has no
                # authority to build from, and the provider says so through the
                # declared source gate (§14.10), never through invented rows.
                return self._generic(module_id, cited), "fail"
            text = (FIXTURES / MODEL_FIXTURES[module_id]).read_text(encoding="utf-8")
            text = text.replace('"run-cp-model-fixture"', json.dumps(run_id))
            if module_id == "CP-1B" and "quarterly" in kinds:
                quarterly = kinds["quarterly"]
                # The snapshot row keeps the annual as its first source and
                # gains the quarterly as a second: both ids ride the model table.
                text = text.replace(QUARTERLY_SNAPSHOT_ROW, (
                    "| historical_performance | Revenue and EBITDA increased through FY2024; "
                    f"FY2025 Q1 trading in line | READY | SRC-1; {quarterly['source_id']} "
                    "| Annual report 2024 p. 42; Quarterly report p. 3 | 2025-03-31 |"
                ))
            if module_id == "CP-2G":
                text = "\n".join(self._forecast_row(line, kinds) for line in text.splitlines()) + "\n"
            text = (
                text.replace("SRC-1", primary["source_id"])
                .replace("block-1", primary["block_id"])
                .replace("b" * 64, primary["sha256"])
                .replace("Acme Credit Ltd", host_identity["issuer_name"])
                .replace("Acme-Credit", host_identity["issuer_id"])
            )
            return text, "pass"
        return self._generic(module_id, cited), "pass"

    @staticmethod
    def _forecast_row(line: str, kinds: dict[str, dict[str, Any]]) -> str:
        if "| operating.adjusted_ebitda_margin |" not in line:
            return line
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        # driver_id, slot_id, case, period_id, fiscal_year, value, unit,
        # assumption_id, status, source_id, source_locator, as_of, gap_code
        guidance = kinds.get("guidance")
        if guidance is None:
            cells[5], cells[8], cells[9], cells[10], cells[11] = "", "UNAVAILABLE", "", "", ""
            cells[12] = "MANAGEMENT_GUIDANCE_UNAVAILABLE"
        else:
            cells[9], cells[10] = guidance["source_id"], "Guidance p. 1"
            if "restated" in kinds and cells[2] == "BASE" and cells[3] == "FY2025":
                cells[5] = "0.22"
        return "| " + " | ".join(cells) + " |"

    @staticmethod
    def _generic(module_id: str, cited: list[dict[str, Any]]) -> str:
        trace = "\n".join((
            "| source_id | block_id |", "|---|---|",
            *(f"| {row['source_id']} | {row['block_id']} |" for row in cited),
        ))
        sections = {
            "Audit Summary": f"{module_id} completed over the supplied documents.",
            "Analysis": "Answer-keyed interpretation under the verified module authority.",
            "Evidence Trace": trace,
            "Source Registry": trace,
            "Gaps & Conflicts": "None keyed for this pack." if cited else "No usable document was supplied.",
            "QA Validation": "Every relevant supplied document was read through the pinned evidence tool.",
        }
        return "\n\n".join(f"## {heading}\n\n{sections[heading]}" for heading in HEADINGS)

    def _tool_call(self, name: str, arguments: dict[str, Any]):
        from caos.engine.provider import ProviderBlock

        return self._message(
            [ProviderBlock(type="tool_use", id=f"tool-{self.calls}", name=name, input=arguments)],
            stop_reason="tool_use",
        )

    def _message(self, content, *, stop_reason: str):
        from caos.engine.provider import ProviderMessage, ProviderUsage

        return ProviderMessage(
            content=content, stop_reason=stop_reason,
            usage=ProviderUsage(input_tokens=100, output_tokens=200),
            request_id=f"answer-key-{self.calls}",
            observed_model=self.identity.model,
            observed_provider_version=self.identity.provider_version,
        )


# --- harness ------------------------------------------------------------------------------------


class Harness:
    def __init__(self, tmp_path: Path, settings, store) -> None:
        from caos.api import create_app
        from caos.engine.runtime import Engine
        from caos.models.service import ModelService

        self.store = store
        self.provider = AnswerKeyProvider()
        self.engine = Engine.create(
            settings=settings, store=store,
            checkpoint_path=tmp_path / "answer-key-checkpoints.db", provider=self.provider,
        )
        self.models = ModelService(store=store, vault_dir=settings.storage_dir, engine=self.engine)
        self.client = TestClient(create_app(settings=settings, store=store, engine=self.engine))
        self.client.__enter__()

    async def aclose(self) -> None:
        self.client.__exit__(None, None, None)
        await self.engine.aclose()

    def intake(self, pack, *, case_id: str | None = None):
        response = submit(self.client, pack, case_id=case_id, headers=ANALYST)
        return response

    async def run_pack(self, pack, *, case_id: str | None = None) -> tuple[dict[str, Any], dict[str, Any]]:
        body = self.intake(pack, case_id=case_id).json()
        assert body["status"] == "started", body.get("refusal")
        run = await self.engine.wait(body["run"]["id"])
        return body, run

    async def accept(self, run_id: str) -> dict[str, Any]:
        return await self.engine.accept(run_id, actor="analyst")

    def build_for(self, case_id: str, snapshot_id: str) -> dict[str, Any]:
        readiness = self.models.readiness(case_id)
        assert readiness["status"] in {"READY_TO_BUILD", "READY"}, readiness
        queued = next(build for build in self.models.list_builds(case_id) if build["snapshot_id"] == snapshot_id)
        return self.models.run_build_for_tests(queued["id"])

    async def ready_full_credit(self, pack=BASE_PACK) -> tuple[str, dict[str, Any], dict[str, Any], dict[str, Any]]:
        body, run = await self.run_pack(pack)
        assert run["status"] == "succeeded", run.get("error")
        snapshot = await self.accept(run["id"])
        build = self.build_for(body["case_id"], snapshot["id"])
        assert build["status"] == "READY", build.get("error")
        return body["case_id"], body, snapshot, build

    def lineage(self, build: dict[str, Any], body: dict[str, Any]) -> dict[str, dict[str, Any]]:
        """Lineage rows keyed by the filename the intake manifest admitted them under."""
        by_source = {row["source_id"]: row for row in build["payload"]["source_lineage"]}
        return {
            document["filename"]: by_source[document["source_id"]]
            for document in body["documents"]
            if document["source_id"] in by_source
        }

    def audit(self, action: str) -> list[dict[str, Any]]:
        return [row for row in self.store.audit_trail() if row["action"] == action]


@pytest.fixture()
async def harness(tmp_path, settings, store):
    value = Harness(tmp_path, settings, store)
    try:
        yield value
    finally:
        await value.aclose()


def _digest(value: Any) -> str:
    from caos.contracts import digest

    return digest(value)


def _assumption_values(harness: Harness, case_id: str, build: dict[str, Any]) -> list[tuple[Any, ...]]:
    registry = harness.models.assumption_registry(case_id, build["id"])
    return [
        (row["assumption_id"], row["case"], row["period_id"], row["value"], row["status"], row["gap_code"])
        for row in registry["defaults"]
    ]


def _model_tabs(build: dict[str, Any]) -> list[dict[str, Any]]:
    """The calculated worksheets; `_AUDIT` carries the input files' run-bound
    digests and moves with every run by construction."""
    return [tab for tab in build["payload"]["tabs"] if tab["title"] != "_AUDIT"]


# --- the complete Full Credit model from the complete relevant manifest (CALC-001) --------------


async def test_full_credit_build_binds_every_relevant_document_to_the_model_or_the_analysis(harness):
    case_id, body, snapshot, build = await harness.ready_full_credit()
    assert harness.provider.modules_by_run[snapshot["run_id"]] >= {"CP-1", "CP-1B", "CP-2G"}
    lineage = harness.lineage(build, body)
    assert {name: row["binding"] for name, row in lineage.items()} == {
        ANNUAL[0]: "MODEL_INPUT",
        QUARTERLY[0]: "MODEL_INPUT",
        GUIDANCE[0]: "MODEL_INPUT",
        AGREEMENT[0]: "CITED_ANALYSIS",
    }
    assert {name: row["document_type"] for name, row in lineage.items()} == {
        ANNUAL[0]: "annual_report", QUARTERLY[0]: "quarterly_report",
        GUIDANCE[0]: "management_guidance", AGREEMENT[0]: "credit_agreement",
    }
    assert all(row["disposition"] == "used" and row["reason"] for row in lineage.values())
    assert lineage[AGREEMENT[0]]["cited_by"], "the legal document reaches the cited analysis"
    assert "CP-4" in lineage[AGREEMENT[0]]["cited_by"]
    assert lineage[GUIDANCE[0]]["consumers"] == ["CP-PARSE", "CP-0", "CP-1B", "CP-MODEL"]
    # The lineage is part of the model's identity and of its audit trail.
    assert build["payload_digest"] == _digest(build["payload"])
    ready = harness.audit("model.build_ready")
    assert [(row["build_id"], row["snapshot_id"], row["run_id"], row["sha256"]) for row in ready] == [
        (build["id"], snapshot["id"], snapshot["run_id"], build["payload_digest"]),
    ]
    assert harness.audit("snapshot.accepted")[0]["snapshot_id"] == snapshot["id"]
    assert harness.audit("intake.admitted")[0]["case_id"] == case_id
    # The worksheet an analyst reviews and the workbook they export carry the same lineage sheet.
    registry = harness.models.assumption_registry(case_id, build["id"])
    from caos.contracts import ModelPreviewRequest

    preview = harness.models.preview(case_id, ModelPreviewRequest.model_validate({
        "build_id": build["id"], "parent_revision_id": None,
        "registry_version": registry["version"], "registry_digest": registry["digest"],
        "assumptions": registry["defaults"], "draft_generation": 1,
    }))
    assert "Source Lineage" in [tab["title"] for tab in preview["worksheet"]["tabs"]]
    harness.models.queue_export(build["id"], "analyst")
    harness.models.run_export_for_tests(build["id"])
    content, _sha = harness.models.download(case_id, build["id"])
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    try:
        assert "Source Lineage" in workbook.sheetnames
        rows = dict(workbook["Source Lineage"].iter_rows(min_row=2, values_only=True))
        bindings = {value for path, value in rows.items() if path.endswith(".binding")}
        assert bindings == {'"MODEL_INPUT"', '"CITED_ANALYSIS"'}
    finally:
        workbook.close()


async def test_a_used_document_that_nothing_consumed_is_a_typed_incomplete_model_never_ready(harness):
    """The lineage oracle bites: a relevant document the analysis neither cited
    nor tabled leaves the model NOT_READY with the typed code, and the build
    queue refuses under the same code."""
    harness.provider.unread_kinds.add("agreement")
    body, run = await harness.run_pack(BASE_PACK)
    assert run["status"] == "succeeded", run.get("error")
    snapshot = await harness.accept(run["id"])
    readiness = harness.models.readiness(body["case_id"])
    assert readiness["status"] == "NOT_READY"
    blocker, = readiness["blockers"]
    assert blocker["code"] == "MODEL_SOURCE_LINEAGE_INCOMPLETE"
    agreement_id = next(document["source_id"] for document in body["documents"] if document["filename"] == AGREEMENT[0])
    assert agreement_id in blocker["detail"] and "1 admitted document" in blocker["detail"]
    assert harness.models.list_builds(body["case_id"]) == []
    with pytest.raises(ValueError, match="MODEL_NOT_READY: MODEL_SOURCE_LINEAGE_INCOMPLETE"):
        harness.models.queue_build(body["case_id"], "analyst")
    assert harness.store.get_case(body["case_id"])["accepted_snapshot_id"] == snapshot["id"]


# --- metamorphic cases (ENTERPRISE_TESTING_READINESS.md, deterministic models) ------------------


async def test_removing_the_annual_report_refuses_at_the_source_gate(harness):
    """Without an annual report the canonical data foundation has no authority:
    the run ends with the provider-declared source gate, nothing is accepted,
    and Model Builder reports the absence of authority — never a model."""
    body = harness.intake([QUARTERLY, GUIDANCE, AGREEMENT]).json()
    assert body["status"] == "started" and body["route"]["pathway"] == "FULL_CREDIT"
    run = await harness.engine.wait(body["run"]["id"])
    assert run["status"] == "failed"
    assert run["error"]["code"] == "SOURCE_EVIDENCE_INSUFFICIENT"
    assert run["accepted_snapshot_id"] is None
    assert harness.models.readiness(body["case_id"])["blockers"] == [{
        "code": "ACCEPTED_FULL_CREDIT_REQUIRED",
        "detail": "Accept a completed Full Credit run before building a model.",
    }]
    assert harness.models.list_builds(body["case_id"]) == []
    assert {row["document_type"] for row in body["documents"]} == {"quarterly_report", "management_guidance", "credit_agreement"}


async def test_removing_the_quarterly_report_changes_the_input_and_the_lineage_but_no_output(harness):
    _base_case, base_body, _base_snapshot, base_build = await harness.ready_full_credit()
    case_id, body, snapshot, build = await harness.ready_full_credit(
        [doc("annual", OTHER_ISSUER), doc("guidance", OTHER_ISSUER), doc("agreement", OTHER_ISSUER)],
    )

    assert build["outputs_digest"] == base_build["outputs_digest"]
    # Assumption values, statuses and gaps are identical; only the provenance
    # ids differ, because the comparison pack lives in its own case.
    assert _assumption_values(harness, case_id, build) == _assumption_values(harness, _base_case, base_build)
    assert build["input_fingerprint"] != base_build["input_fingerprint"]
    assert QUARTERLY[0] not in harness.lineage(build, body)
    assert set(harness.lineage(build, body)) == {ANNUAL[0], GUIDANCE[0], AGREEMENT[0]}
    cp1b = next(a for a in harness.engine.artifacts_for_run(snapshot["run_id"]) if a["module_id"] == "CP-1B")
    base_cp1b = next(a for a in harness.engine.artifacts_for_run(base_build["accepted_run_id"]) if a["module_id"] == "CP-1B")
    assert "FY2025 Q1 trading in line" in base_cp1b["markdown"]
    assert "FY2025 Q1 trading in line" not in cp1b["markdown"]


async def test_removing_the_forecast_source_is_a_typed_model_refusal_never_a_default(harness):
    """Management guidance is the forecast-driver authority. Without it the
    provider marks the margin driver UNAVAILABLE with a named gap, and the
    pinned CP-MODEL validator refuses the build: a null forecast input is a
    refusal, never zero and never an interpolated default (CALC-003, CALC-004)."""
    body, run = await harness.run_pack([ANNUAL, QUARTERLY, AGREEMENT])
    assert run["status"] == "succeeded", run.get("error")
    snapshot = await harness.accept(run["id"])
    readiness = harness.models.readiness(body["case_id"])
    assert readiness["status"] == "CANONICAL_MODEL_INPUTS_INVALID", readiness
    assert readiness["blockers"][0]["code"] == "CANONICAL_MODEL_INPUTS_INVALID"
    assert harness.models.list_builds(body["case_id"]) == []
    cp2g = next(a for a in harness.engine.artifacts_for_run(snapshot["run_id"]) if a["module_id"] == "CP-2G")
    assert "MANAGEMENT_GUIDANCE_UNAVAILABLE" in cp2g["markdown"]
    assert "| 0.2 | PERCENT_DECIMAL | operating.adjusted_ebitda_margin |" not in cp2g["markdown"]


async def test_an_irrelevant_document_changes_no_result_and_is_never_silently_discarded(harness):
    case_id, base_body, _base_snapshot, base_build = await harness.ready_full_credit()

    # Dropped into the same case: the run pins the widened set, so the model's
    # identity moves while every result stays exactly where it was.
    body = harness.intake([OTHER], case_id=case_id).json()
    assert body["status"] == "started", body.get("refusal")
    clipping, = body["documents"]
    assert clipping["document_type"] == "other" and clipping["disposition"] == "used"
    assert clipping["reason"] and clipping["consumers"] == ["CP-PARSE", "CP-0"]
    run = await harness.engine.wait(body["run"]["id"])
    assert run["status"] == "succeeded", run.get("error")
    snapshot = await harness.accept(run["id"])
    build = harness.build_for(case_id, snapshot["id"])
    assert build["status"] == "READY"

    assert build["input_fingerprint"] != base_build["input_fingerprint"]
    assert _model_tabs(build) == _model_tabs(base_build)
    assert build["outputs_digest"] == base_build["outputs_digest"]
    assert build["assumptions_digest"] == base_build["assumptions_digest"]
    # QA is identical apart from the input files' run-bound digests.
    assert {k: v for k, v in build["qa"].items() if k != "source_manifest"} == {
        k: v for k, v in base_build["qa"].items() if k != "source_manifest"
    }
    rows = {row["source_id"]: row for row in build["payload"]["source_lineage"]}
    other = rows[clipping["source_id"]]
    assert other["binding"] == "NOT_REQUIRED" and other["cited_by"] == [] and other["model_input"] is False
    assert other["document_type"] == "other" and "no model consumer for document type other" in other["reason"]
    base_rows = {row["source_id"]: row for row in base_build["payload"]["source_lineage"]}
    assert {sid: row["binding"] for sid, row in rows.items() if sid != clipping["source_id"]} == {
        sid: row["binding"] for sid, row in base_rows.items()
    }


async def test_a_restated_annual_report_supersedes_the_original_and_moves_the_forecast(harness):
    _base_case, _base_body, _base_snapshot, base_build = await harness.ready_full_credit()
    case_id, body, snapshot, build = await harness.ready_full_credit(
        [doc(kind, OTHER_ISSUER) for kind in ("annual", "quarterly", "guidance", "agreement", "restated")],
    )

    dispositions = {document["filename"]: document["disposition"] for document in body["documents"]}
    assert dispositions[ANNUAL[0]] == "superseded" and dispositions[RESTATED[0]] == "used"
    lineage = harness.lineage(build, body)
    assert lineage[ANNUAL[0]]["binding"] == "SUPERSEDED"
    assert lineage[ANNUAL[0]]["reason"].startswith("superseded by the restated document")
    assert lineage[RESTATED[0]]["binding"] == "MODEL_INPUT" and lineage[RESTATED[0]]["version_status"] == "restated"
    assert build["outputs_digest"] != base_build["outputs_digest"], "the restated margin moves the forecast"
    assert build["assumptions_digest"] != base_build["assumptions_digest"]
    registry = harness.models.assumption_registry(case_id, build["id"])
    margin = next(
        row for row in registry["defaults"]
        if row["assumption_id"] == "operating.adjusted_ebitda_margin" and row["case"] == "BASE" and row["period_id"] == "FY2025"
    )
    assert margin["value"] == 0.22
    assert margin["source_context"]["provenance"][0]["source_id"] == next(
        document["source_id"] for document in body["documents"] if document["filename"] == GUIDANCE[0]
    )


async def test_a_conflicting_document_refuses_the_whole_pack_and_admits_nothing(harness):
    conflicting = (ANNUAL[0], annual_report(fiscal_year=2023), TEXT)
    response = harness.intake([*BASE_PACK, conflicting])
    assert response.status_code == 422, response.text
    assert response.json()["detail"]["code"] == "INTAKE_SOURCE_CONFLICT"
    assert harness.store.list_cases("analyst") == []
    assert harness.audit("intake.refused")[0]["code"] == "INTAKE_SOURCE_CONFLICT"


async def test_withdrawing_a_bound_source_revokes_the_model_and_leaves_the_audit_trail(harness):
    case_id, body, snapshot, build = await harness.ready_full_credit()
    annual_id = next(document["source_id"] for document in body["documents"] if document["filename"] == ANNUAL[0])

    harness.store.withdraw(case_id, annual_id, "analyst")

    readiness = harness.models.readiness(case_id)
    assert readiness["status"] == "CANONICAL_MODEL_INPUTS_INVALID"
    with pytest.raises(ValueError, match="MODEL_BUILD_NOT_READY"):
        harness.models.assumption_registry(case_id, build["id"])
    withdrawn = harness.audit("source.withdrawn")
    assert [(row["case_id"], row["source_id"]) for row in withdrawn] == [(case_id, annual_id)]
    assert harness.store.get_case(case_id)["accepted_snapshot_id"] == snapshot["id"], "acceptance is history, not revoked"
    # A re-run over the reduced set is the source-gate refusal, not a thinner model.
    started = await harness.engine.start_run(case_id=case_id, pathway="FULL_CREDIT", depth="full", actor="analyst")
    rerun = await harness.engine.wait(started["id"])
    assert rerun["status"] == "failed" and rerun["error"]["code"] == "SOURCE_EVIDENCE_INSUFFICIENT"


async def test_corrupting_a_bound_source_revokes_the_model_before_any_use(harness):
    from caos.storage.store import sources

    case_id, body, _snapshot, build = await harness.ready_full_credit()
    annual_id = next(document["source_id"] for document in body["documents"] if document["filename"] == ANNUAL[0])
    row = harness.store.get_source_private(annual_id)
    blocks = copy.deepcopy(row["blocks"])
    blocks[0]["text"] = blocks[0]["text"] + " [tampered]"
    with harness.store.engine.begin() as conn:
        conn.execute(sa.update(sources).where(sources.c.id == annual_id).values(blocks=blocks))

    assert harness.models.readiness(case_id)["status"] == "CANONICAL_MODEL_INPUTS_INVALID"
    with pytest.raises(ValueError, match="MODEL_BUILD_NOT_READY"):
        harness.models.assumption_registry(case_id, build["id"])


async def test_redropping_a_source_route_document_cannot_wave_it_out_of_the_lineage(harness):
    """A document admitted through the source route has no intake row; dropping
    the same bytes again yields only a `duplicate` row. That row lends its
    classification and nothing else: the annual stays `used`, so it must still
    reach the model — and it does, as a model input."""
    created = harness.client.post(
        "/api/cases", json={"name": "Manual first", "issuer": ISSUER, "sector": "Services"}, headers=ANALYST,
    )
    assert created.status_code == 201, created.text
    case_id = created.json()["id"]
    uploaded = harness.client.post(
        f"/api/cases/{case_id}/sources", files={"file": (ANNUAL[0], ANNUAL[1], TEXT)}, headers=ANALYST,
    )
    assert uploaded.status_code == 201, uploaded.text
    body, run = await harness.run_pack(BASE_PACK, case_id=case_id)
    annual_row = next(document for document in body["documents"] if document["filename"] == ANNUAL[0])
    assert annual_row["disposition"] == "duplicate" and annual_row["source_id"] == uploaded.json()["id"]
    assert run["status"] == "succeeded", run.get("error")
    snapshot = await harness.accept(run["id"])
    build = harness.build_for(case_id, snapshot["id"])
    rows = {row["source_id"]: row for row in build["payload"]["source_lineage"]}
    annual = rows[uploaded.json()["id"]]
    assert annual["disposition"] == "used" and annual["document_type"] == "annual_report"
    assert annual["binding"] == "MODEL_INPUT"
    assert annual["reason"].startswith("admitted through the source route")


# --- every other pathway's declared model effect -------------------------------------------------


async def test_every_pathway_overlays_its_declared_effect_on_the_full_credit_model(harness):
    """One case, the golden journey for each pathway in turn: every overlay
    reuses the base tabs byte for byte, carries its effect under its own
    fingerprint, and binds every supplied document."""
    case_id, base_body, base_snapshot, base_build = await harness.ready_full_credit()
    effects: dict[str, dict[str, Any]] = {}

    # Earnings Update: periods and forecast variance.
    body, run = await harness.run_pack([QUARTERLY_Q2, EARNINGS], case_id=case_id)
    assert body["route"]["pathway"] == "EARNINGS_UPDATE" and run["status"] == "succeeded", run.get("error")
    snapshot = await harness.accept(run["id"])
    build = harness.build_for(case_id, snapshot["id"])
    assert build["status"] == "READY", build.get("error")
    effect, = build["payload"]["pathway_effects"]
    effects["EARNINGS_UPDATE"] = effect
    assert effect["effect_id"] == "EARNINGS_PERIOD_FORECAST_VARIANCE"
    assert effect["accepted_authority"]["snapshot_id"] == snapshot["id"]
    assert {record["calculator_id"] for record in effect["calculations"]} == {"credit_metrics"}
    assert {record["module_id"] for record in effect["calculations"]} == {"CP-1", "CP-1B"}
    assert [(update["period_id"], update["module_id"], update["authority"]) for update in effect["period_updates"]] == [
        ("FY2025", "CP-1", "REPORTED_ACTUAL"), ("FY2025", "CP-1B", "REPORTED_ACTUAL"),
    ]
    assert effect["period_updates"][0]["reported"]["revenue"] == "1000"
    revenue = next(
        row for row in effect["forecast_variance"]
        if row["period_id"] == "FY2025" and row["case"] == "BASE" and row["metric_id"] == "revenue"
    )
    assert revenue == {
        "period_id": "FY2025", "case": "BASE", "metric_id": "revenue", "reported_metric": "revenue",
        "reported": "1000", "forecast": "483.00", "variance": "517.00",
        "variance_ratio": revenue["variance_ratio"], "gap_code": None,
        "reported_authority": "REPORTED_ACTUAL", "forecast_authority": "ANALYST_FORECAST",
    }
    assert revenue["variance_ratio"].startswith("1.0703933747412008")
    assert {row["gap_code"] for row in effect["forecast_variance"]} == {None}
    lineage = harness.lineage(build, body)
    assert lineage[QUARTERLY_Q2[0]]["binding"] == "CITED_ANALYSIS" and lineage[EARNINGS[0]]["binding"] == "CITED_ANALYSIS"

    # Covenant & Refinancing: covenant tests, refinancing wall, proposed assumptions.
    body, run = await harness.run_pack([AMENDMENT], case_id=case_id)
    assert body["route"]["pathway"] == "COVENANT_REFINANCING" and run["status"] == "succeeded", run.get("error")
    snapshot = await harness.accept(run["id"])
    build = harness.build_for(case_id, snapshot["id"])
    assert build["status"] == "READY", build.get("error")
    effect, = build["payload"]["pathway_effects"]
    effects["COVENANT_REFINANCING"] = effect
    assert effect["effect_id"] == "COVENANT_REFINANCING_ASSUMPTIONS"
    assert {(record["module_id"], record["calculator_id"]) for record in effect["calculations"]} == {
        ("CP-4", "covenant_headroom"), ("CP-4C", "funding_gap"), ("CP-4C", "recovery_waterfall"),
    }
    test, = effect["covenant_updates"]
    assert (test["test"], test["test_type"], test["threshold"], test["current_ratio"], test["headroom"], test["status"]) == (
        "Leverage", "max-ratio", "5", "4", "1", "Compliant",
    )
    assert test["authority"] == "DOCUMENTARY_COVENANT_TERMS"
    wall, = effect["refinancing_updates"]
    assert wall["view"] == "as_of_balance_sheet_date" and wall["gap"]["funding_gap"] == 150
    updates = effect["assumption_updates"]
    assert {(row["case"], row["period_id"]) for row in updates} == {
        (case, period) for case in ("BASE", "DOWNSIDE") for period in ("FY2025", "FY2026", "FY2027")
    }
    assert all(
        row["assumption_id"] == "covenant.max_total_leverage" and row["base_status"] == "UNAVAILABLE"
        and row["base_value"] is None and row["proposed_value"] == "5"
        and row["base_gap_code"] == "COVENANT_DEFINITION_UNAVAILABLE"
        and row["treatment"] == "PROPOSED_REQUIRES_FULL_CREDIT_HANDOFF"
        for row in updates
    )
    # Nothing is applied silently: the base registry still carries the slot as
    # the accepted handoff left it, and the model's own covenant headroom stays
    # unavailable until a Full Credit re-run sources the definition.
    from caos.contracts import ModelPreviewRequest

    registry = harness.models.assumption_registry(case_id, build["id"])
    covenant_rows = [row for row in registry["defaults"] if row["assumption_id"] == "covenant.max_total_leverage"]
    assert {(row["status"], row["value"]) for row in covenant_rows} == {("UNAVAILABLE", None)}
    preview = harness.models.preview(case_id, ModelPreviewRequest.model_validate({
        "build_id": build["id"], "parent_revision_id": None, "registry_version": registry["version"],
        "registry_digest": registry["digest"], "assumptions": registry["defaults"], "draft_generation": 1,
    }))
    assert preview["outputs"]["BASE"]["BASE::FY2025"]["covenant_headroom"] is None
    assert "Pathway Effects" in [tab["title"] for tab in preview["worksheet"]["tabs"]]
    assert build["payload"]["tabs"] == base_build["payload"]["tabs"], "the signed model is untouched"

    # Relative Value: time-aligned market marks supplied by upload.
    marks = (FILENAMES["marks"], marks_workbook(), XLSX)
    body, run = await harness.run_pack([marks], case_id=case_id)
    assert body["route"]["pathway"] == "RELATIVE_VALUE" and run["status"] == "succeeded", run.get("error")
    marks_id = next(document["source_id"] for document in body["documents"] if document["filename"] == marks[0])
    assert run["plan"]["loan_universe"]["source_id"] == marks_id
    snapshot = await harness.accept(run["id"])
    build = harness.build_for(case_id, snapshot["id"])
    assert build["status"] == "READY", build.get("error")
    effect, = build["payload"]["pathway_effects"]
    effects["RELATIVE_VALUE"] = effect
    assert effect["effect_id"] == "RELATIVE_VALUE_MARKET_MARKS"
    active = harness.store.active_loan_universe(case_id)
    assert effect["market_marks"]["universe_digest"] == active["universe_digest"]
    assert effect["market_marks"]["source_id"] == marks_id and effect["market_marks"]["row_count"] == 2
    assert {row["borrower_name"] for row in effect["market_marks"]["rows"]} == {"Access CIG LLC", "FinThrive Inc"}
    assert effect["market_marks"]["rows"][0]["bid_points"] == 88 and effect["market_marks"]["authority"] == "SUPPLIED_MARKET_MARKS"
    assert effect["time_alignment"] == {
        "workbook_date": "2026-08-24", "latest_reported_period_end": "2024-12-31",
        "analysis_date": run["created_at"][:10], "status": "ALIGNED",
    }
    assert effect["limitations"] == []
    assert {(record["module_id"], record["calculator_id"]) for record in effect["calculations"]} == {
        ("CP-3", "recovery_waterfall"), ("CP-1C", "peer_statistics"),
    }
    assert harness.lineage(build, body)[marks[0]]["binding"] == "MARKET_MARKS"

    # Deep Research: revalidates the base and declares no numeric effect.
    body = harness.intake([BRIEF], case_id=case_id).json()
    assert body["route"]["pathway"] == "DEEP_RESEARCH", body.get("refusal")
    paused = await harness.engine.wait(body["run"]["id"])
    assert paused["status"] == "paused" and paused["error"]["code"] == "PLAN_APPROVAL_REQUIRED"
    await harness.engine.approve_research_plan(
        paused["id"], plan_hash=paused["research"]["proposed_plan_hash"], actor="analyst",
    )
    run = await harness.engine.wait(paused["id"])
    assert run["status"] == "succeeded", run.get("error")
    snapshot = await harness.accept(run["id"])
    build = harness.build_for(case_id, snapshot["id"])
    assert build["status"] == "READY", build.get("error")
    effect, = build["payload"]["pathway_effects"]
    effects["DEEP_RESEARCH"] = effect
    assert effect["effect_id"] == "DEEP_RESEARCH_REVALIDATION"
    assert effect["numeric_effect"] == "NONE" and effect["base_model_revalidated"] is True
    assert effect["calculations"] == []
    assert effect["research"]["approved_plan_hash"] == run["research"]["approved_plan_hash"]
    assert effect["research"]["brief_digest"] == run["research"]["brief_digest"]
    assert harness.lineage(build, body)[BRIEF[0]]["binding"] == "RESEARCH_BRIEF"

    # Every overlay: base tabs byte-identical, distinct fingerprints, base identity bound.
    fingerprints = set()
    for build in harness.models.list_builds(case_id):
        assert build["status"] == "READY"
        fingerprints.add(build["input_fingerprint"])
        if build["id"] == base_build["id"]:
            continue
        effect, = build["payload"]["pathway_effects"]
        assert build["payload"]["tabs"] == base_build["payload"]["tabs"]
        assert effect["base_model"]["build_id"] == base_build["id"]
        assert effect["base_model"]["payload_digest"] == base_build["payload_digest"]
        assert effect["schema_version"] == "caos.model-pathway-effect.v1"
        assert build["assumptions_digest"] == base_build["assumptions_digest"]
        assert build["outputs_digest"] == base_build["outputs_digest"]
    assert len(fingerprints) == 5
    assert set(effects) == {"EARNINGS_UPDATE", "COVENANT_REFINANCING", "RELATIVE_VALUE", "DEEP_RESEARCH"}
    assert len(harness.audit("model.build_ready")) == 5
    assert harness.models.readiness(case_id)["status"] == "READY"


async def test_market_marks_dated_before_the_latest_reported_period_are_a_named_limitation(harness):
    case_id, _body, _snapshot, _base = await harness.ready_full_credit()
    marks = (FILENAMES["marks"], marks_workbook(date(2024, 6, 30)), XLSX)
    body, run = await harness.run_pack([marks], case_id=case_id)
    assert run["status"] == "succeeded", run.get("error")
    snapshot = await harness.accept(run["id"])
    build = harness.build_for(case_id, snapshot["id"])
    effect, = build["payload"]["pathway_effects"]
    assert effect["time_alignment"]["status"] == "PRECEDES_LATEST_REPORTED_PERIOD"
    assert effect["time_alignment"]["workbook_date"] == "2024-06-30"
    assert effect["limitations"] == [{"module_id": "CP-3", "limitation": {"code": "MARKET_MARKS_PRECEDES_LATEST_REPORTED_PERIOD"}}]


async def test_relative_value_without_a_pinned_workbook_names_the_missing_marks(harness):
    case_id, _body, _snapshot, _base = await harness.ready_full_credit()
    started = await harness.engine.start_run(case_id=case_id, pathway="RELATIVE_VALUE", depth="full", actor="analyst")
    run = await harness.engine.wait(started["id"])
    assert run["status"] == "succeeded", run.get("error")
    assert "loan_universe" not in run["plan"]
    await harness.accept(run["id"])
    readiness = harness.models.readiness(case_id)
    assert readiness["status"] == "NOT_READY"
    assert [blocker["code"] for blocker in readiness["blockers"]] == ["RELATIVE_VALUE_MARKET_MARKS_REQUIRED"]
    with pytest.raises(ValueError, match="MODEL_NOT_READY: RELATIVE_VALUE_MARKET_MARKS_REQUIRED"):
        harness.models.queue_build(case_id, "analyst")


async def test_deep_research_without_a_full_credit_model_declares_no_numeric_effect(harness):
    body = harness.intake([ANNUAL, BRIEF]).json()
    assert body["route"]["pathway"] == "DEEP_RESEARCH", body.get("refusal")
    paused = await harness.engine.wait(body["run"]["id"])
    await harness.engine.approve_research_plan(
        paused["id"], plan_hash=paused["research"]["proposed_plan_hash"], actor="analyst",
    )
    run = await harness.engine.wait(paused["id"])
    assert run["status"] == "succeeded", run.get("error")
    await harness.accept(run["id"])
    readiness = harness.models.readiness(body["case_id"])
    assert readiness["status"] == "NOT_READY"
    assert readiness["blockers"] == [{
        "code": "DEEP_RESEARCH_NO_NUMERIC_EFFECT",
        "detail": "Deep Research declares no numeric model effect, and this case has no Full Credit model to revalidate.",
    }]
    assert harness.models.list_builds(body["case_id"]) == []


@pytest.mark.parametrize("pathway", ["FULL_CREDIT", "EARNINGS_UPDATE", "COVENANT_REFINANCING", "RELATIVE_VALUE"])
async def test_screen_depth_reads_as_a_depth_precondition_not_as_corruption(harness, pathway):
    case_id, _body, _snapshot, _base = await harness.ready_full_credit()
    started = await harness.engine.start_run(case_id=case_id, pathway=pathway, depth="screen", actor="analyst")
    run = await harness.engine.wait(started["id"])
    assert run["status"] == "succeeded", run.get("error")
    await harness.accept(run["id"])
    readiness = harness.models.readiness(case_id)
    assert readiness["status"] == "NOT_READY", readiness
    assert [blocker["code"] for blocker in readiness["blockers"]] == ["FULL_DEPTH_REQUIRED"]
    with pytest.raises(ValueError, match="MODEL_NOT_READY: FULL_DEPTH_REQUIRED"):
        harness.models.queue_build(case_id, "analyst")


async def test_an_overlay_effect_is_rebuilt_from_the_records_it_names_and_refuses_a_forged_one(harness):
    """CALC-020 for overlays: the effect's calculation records are re-executed
    against the pinned calculator before use; a record whose output digest was
    edited after acceptance cannot resolve, so the overlay can never be READY."""
    from caos.contracts import digest

    case_id, _body, _snapshot, _base = await harness.ready_full_credit()
    body, run = await harness.run_pack([AMENDMENT], case_id=case_id)
    assert run["status"] == "succeeded", run.get("error")
    cp4 = next(a for a in harness.engine.artifacts_for_run(run["id"]) if a["module_id"] == "CP-4")
    payload = copy.deepcopy(cp4["payload"])
    payload["calculations"][0]["canonical_output"]["headroom"][0]["headroom"] = 9.5
    payload["calculations"][0]["output_digest"] = digest(payload["calculations"][0]["canonical_output"])
    harness.engine.runs.update_artifact_for_tests(run["id"], "CP-4", payload=payload, digest=digest(payload))
    with pytest.raises(Exception, match="RUN_NOT_READY"):
        await harness.accept(run["id"])
    assert harness.engine.runs.snapshot_for_run(run["id"]) is None


# --- the pure calculation contract (CALC-005, CALC-006, CALC-012) ---------------------------------


def _bundle():
    from caos.config import Settings
    from caos.models.engine import CpModelBundle

    return CpModelBundle(Settings().deploy_v_root)


def _fixture_paths() -> dict[str, Path]:
    return {module: FIXTURES / name for module, name in {
        "CP-1": "cp1.md", "CP-1A": "cp1a.md", "CP-1B": "cp1b.md", "CP-2": "cp2.md", "CP-2B": "cp2b.md", "CP-2G": "cp2g.md",
    }.items()}


def test_derived_periods_follow_the_governed_formulas_and_never_overwrite_a_reported_one():
    """CALC-005 and CALC-006: quarter, year-to-date, last-twelve-month, pro
    forma, base and downside columns are derived from the four reported
    quarters; the reported quarters keep their reported values."""
    model, calculations = _bundle().calculate(_fixture_paths())
    groups = {column.group for column in calculations.columns}
    assert groups == {"QUARTER", "YTD", "LTM", "PF", "BASE", "DOWNSIDE"}
    assert list(model.reported_periods) == ["FY2024_Q1", "FY2024_Q2", "FY2024_Q3", "FY2024_Q4"]
    assert model.annuals == (), "no annual period is reported, so none may be invented"
    reported = {period_id: model.accounts[("revenue", period_id)].value for period_id in model.reported_periods}
    assert {k: str(v) for k, v in reported.items()} == {"FY2024_Q1": "100", "FY2024_Q2": "110", "FY2024_Q3": "120", "FY2024_Q4": "130"}
    for period_id, value in reported.items():
        assert calculations.for_column(period_id).values["revenue"] == value
    assert str(calculations.for_column("LTM_FY2024_Q4").values["revenue"]) == "460"
    assert str(calculations.for_column("YTD_2024_Q4").values["revenue"]) == "460"
    assert str(calculations.for_column("BASE::FY2025").values["revenue"]) == "483.00"


@pytest.mark.parametrize(("value", "accepted"), [
    ("1.9999", True), ("2", True), ("2.0001", False),
    ("-0.7499", True), ("-0.75", True), ("-0.7501", False),
])
def test_hard_bounds_apply_one_value_below_at_and_above_each_boundary(value, accepted):
    """CALC-012 on the registry's division revenue-growth hard bounds
    (`hard_min` -0.75, `hard_max` 2): one value inside, at, and outside each."""
    from caos.models.engine import ModelInputError

    bundle = _bundle()
    cp2g = (FIXTURES / "cp2g.md").read_text(encoding="utf-8").replace(
        "| 0.05 | PERCENT_DECIMAL | operating.revenue_growth.division_1 |",
        f"| {value} | PERCENT_DECIMAL | operating.revenue_growth.division_1 |",
        1,
    )
    docs = [
        (FIXTURES / name).read_text(encoding="utf-8") if module != "CP-2G" else cp2g
        for module, name in (("CP-1", "cp1.md"), ("CP-1A", "cp1a.md"), ("CP-1B", "cp1b.md"), ("CP-2", "cp2.md"), ("CP-2B", "cp2b.md"), ("CP-2G", "cp2g.md"))
    ]
    errors = list(bundle.validate(*docs).errors)
    if accepted:
        assert not any("outside registry bounds" in error for error in errors), errors
    else:
        assert any("outside registry bounds" in error for error in errors), errors
        del ModelInputError

