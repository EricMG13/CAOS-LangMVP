from __future__ import annotations

import hashlib
import io
import json
from pathlib import Path
import re
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader

from caos.api import create_app
from caos.contracts import (
    DeliverableDraftRequest,
    FileDeliverableRequest,
    FreezeDeliverableRequest,
    digest,
)
from caos.deliverables.service import DeliverableService
from caos.engine.graphs import compiled_route
from caos.engine.provider import (
    ProviderBlock,
    ProviderIdentity,
    ProviderMessage,
    ProviderUsage,
)
from caos.engine.runtime import Engine
from caos.models.service import ModelService


FIXTURES = Path(__file__).resolve().parents[1] / "fixtures" / "cp_model"
MODEL_FIXTURES = {
    "CP-1": "cp1.md",
    "CP-1A": "cp1a.md",
    "CP-1B": "cp1b.md",
    "CP-2": "cp2.md",
    "CP-2A": "cp2a.md",
    "CP-2G": "cp2g.md",
}
HEADINGS = (
    "Audit Summary",
    "Analysis",
    "Evidence Trace",
    "Source Registry",
    "Gaps & Conflicts",
    "QA Validation",
)


class OrdinaryFixtureProvider:
    """Answer-key provider that uses the ordinary provider/evidence/tool path.

    Its identity is deliberately not ``host_control`` and the test never calls
    a scripted or placeholder engine capability.  It stands in for a model at
    the provider port while keeping the regression deterministic.
    """

    def __init__(self) -> None:
        self.identity = ProviderIdentity(
            provider_name="ordinary_fixture",
            model="credit-answer-key-v1",
            provider_version="1",
            adapter_version="caos.ordinary-fixture.v1",
            parameter_context_digest=digest({"provider": "ordinary-fixture-v1"}),
            qualification_record_id=None,
            qualification_record_digest=None,
            qualification_status="unqualified",
            qualification_expires_at=None,
        )
        self.evidence: list[dict[str, str]] = []
        self.modules_by_run: dict[str, set[str]] = {}
        self.calls = 0

    def bind(self, sources: list[dict[str, Any]]) -> None:
        self.evidence = [
            {
                "source_id": source["id"],
                "block_id": source["blocks"][0]["block_id"],
                "source_digest": source["sha256"],
            }
            for source in sources
        ]

    def count_tokens(self, _request: Any) -> int:
        return 100

    def create_message(self, request: Any) -> ProviderMessage:
        self.calls += 1
        module_id, run_id = self._host_identity(request)
        self.modules_by_run.setdefault(run_id, set()).add(module_id)
        tool_results = [
            json.loads(block["content"])
            for message in request.messages
            if isinstance(message.get("content"), list)
            for block in message["content"]
            if isinstance(block, dict) and block.get("type") == "tool_result"
        ]
        evidence_results = [result for result in tool_results if isinstance(result, list)]
        if len(evidence_results) < len(self.evidence):
            target = self.evidence[len(evidence_results)]
            return self._tool_call(
                "read_evidence",
                {
                    "source_id": target["source_id"],
                    "block_ids": [target["block_id"]],
                },
            )

        calculation_records = [
            result for result in tool_results if isinstance(result, dict)
        ]
        calculation_tool = next(
            (
                tool
                for tool in request.effective_tools()
                if tool["name"] == "run_methodology_calculation"
            ),
            None,
        )
        if calculation_tool is not None:
            calculator_ids = calculation_tool["input_schema"]["properties"][
                "calculator_id"
            ]["enum"]
            if len(calculation_records) < len(calculator_ids):
                calculator_id = calculator_ids[len(calculation_records)]
                return self._tool_call(
                    "run_methodology_calculation",
                    {
                        "calculator_id": calculator_id,
                        "input_json": json.dumps(
                            self._calculation_input(calculator_id),
                            sort_keys=True,
                        ),
                    },
                )

        evidence_rows = [row for result in evidence_results for row in result]
        body = {
            "markdown": self._markdown(module_id, run_id, evidence_rows),
            "evidence_refs": [
                {"source_id": row["source_id"], "block_id": row["block_id"]}
                for row in evidence_rows
            ],
            "calculation_refs": [
                {
                    field: record[field]
                    for field in (
                        "calculator_id",
                        "script_digest",
                        "calculator_digest",
                        "input_digest",
                        "output_digest",
                    )
                }
                for record in calculation_records
            ],
            "lineage_counts": {"directly_sourced": len(evidence_rows)},
            "fields_present": len(evidence_rows),
            "fields_total": len(evidence_rows),
            "source_gate": "pass",
            "findings": {},
        }
        return self._message(
            [ProviderBlock(type="text", text=json.dumps(body))],
            stop_reason="end_turn",
        )

    @staticmethod
    def _host_identity(request: Any) -> tuple[str, str]:
        content = request.messages[0]["content"]
        payload = json.loads(content.split("\n", 1)[1])
        identity = payload["host_identity"]
        return identity["module_id"], identity["run_id"]

    def _markdown(
        self,
        module_id: str,
        run_id: str,
        evidence_rows: list[dict[str, Any]],
    ) -> str:
        fixture = MODEL_FIXTURES.get(module_id)
        if fixture is not None:
            primary = evidence_rows[0]
            return (
                (FIXTURES / fixture)
                .read_text(encoding="utf-8")
                .replace('"run-cp-model-fixture"', json.dumps(run_id))
                .replace("SRC-1", primary["source_id"])
                .replace("block-1", primary["block_id"])
                .replace("b" * 64, primary["source_digest"])
            )

        trace = "\n".join(
            (
                "| source_id | block_id | source_digest |",
                "|---|---|---|",
                *(
                    f"| {row['source_id']} | {row['block_id']} | {row['source_digest']} |"
                    for row in evidence_rows
                ),
            )
        )
        registry = "\n".join(
            (
                "| source_id | source_digest |",
                "|---|---|",
                *(
                    f"| {row['source_id']} | {row['source_digest']} |"
                    for row in evidence_rows
                ),
            )
        )
        sections = {
            "Audit Summary": f"{module_id} completed through ordinary semantic execution.",
            "Analysis": "The supplied documents were interpreted under the verified module authority.",
            "Evidence Trace": trace,
            "Source Registry": registry,
            "Gaps & Conflicts": "No answer-keyed blocking gap.",
            "QA Validation": "Every supplied source was read through the pinned evidence tool.",
        }
        return "\n\n".join(
            f"## {heading}\n\n{sections[heading]}" for heading in HEADINGS
        )

    @staticmethod
    def _calculation_input(calculator_id: str) -> dict[str, Any]:
        if calculator_id == "credit_metrics":
            return {"periods": {"FY2025": {
                "revenue": 1_000,
                "adjusted_ebitda": 200,
                "total_debt": 600,
                "cash_and_equivalents": 100,
            }}}
        if calculator_id == "peer_statistics":
            return {
                "metric": "EV/EBITDA",
                "peers": [
                    {"name": "Peer A", "value": 5, "comparability": "Comparable"},
                    {"name": "Peer B", "value": 6, "comparability": "Comparable"},
                ],
            }
        if calculator_id == "rate_fx_sensitivity":
            return {
                "gross_floating_rate_debt": 500,
                "hedged_floating_rate_debt": 300,
                "total_debt": 1_000,
            }
        if calculator_id == "liquidity_bridge":
            return {
                "beginning_accessible_liquidity": 100,
                "operating_cash_flow": 20,
                "working_capital_movement": 0,
                "cash_interest": 5,
                "cash_taxes": 2,
                "mandatory_capex": 3,
                "debt_amortisation_and_maturities": 4,
                "other_cash_uses": 1,
                "committed_inflows": 0,
                "period_months": 12,
            }
        if calculator_id == "bond_analytics":
            return {"price": 98.5, "coupon": 6, "years_to_maturity": 5}
        if calculator_id == "covenant_headroom":
            return {"tests": [{
                "test": "Leverage",
                "test_type": "max-ratio",
                "threshold": 5,
                "current_ratio": 4,
            }]}
        if calculator_id == "funding_gap":
            return {
                "horizon_years": 2,
                "cash": 100,
                "currency": "USD",
                "forecast_fcf": 50,
                "instruments": [
                    {
                        "instrument": "Near-term notes",
                        "amount": 300,
                        "years_to_maturity": 1,
                        "currency": "USD",
                    }
                ],
            }
        if calculator_id == "recovery_waterfall":
            return {
                "enterprise_value": 130,
                "claims": [
                    *[
                        {
                            "claim_id": f"FILLER-{index:03d}",
                            "class": "Senior Secured",
                            "amount": 1,
                        }
                        for index in range(80)
                    ],
                    {"claim_id": "SUN", "class": "Senior Unsecured", "amount": 200},
                ],
            }
        return {}

    def _tool_call(self, name: str, arguments: dict[str, Any]) -> ProviderMessage:
        return self._message(
            [
                ProviderBlock(
                    type="tool_use",
                    id=f"tool-{self.calls}",
                    name=name,
                    input=arguments,
                )
            ],
            stop_reason="tool_use",
        )

    def _message(
        self,
        content: list[ProviderBlock],
        *,
        stop_reason: str,
    ) -> ProviderMessage:
        return ProviderMessage(
            content=content,
            stop_reason=stop_reason,
            usage=ProviderUsage(input_tokens=100, output_tokens=200),
            request_id=f"ordinary-fixture-{self.calls}",
            observed_model=self.identity.model,
            observed_provider_version=self.identity.provider_version,
        )


def _upload(
    client: TestClient,
    case_id: str,
    filename: str,
    content: bytes,
) -> dict[str, Any]:
    response = client.post(
        f"/api/cases/{case_id}/sources",
        files={"file": (filename, content, "text/plain")},
    )
    assert response.status_code == 201, response.text
    source = response.json()
    assert source["blocks"]
    return source


async def _complete(
    engine: Engine,
    provider: OrdinaryFixtureProvider,
    case_id: str,
    pathway: str,
    depth: str,
) -> dict[str, Any]:
    started = await engine.start_run(
        case_id=case_id,
        pathway=pathway,
        depth=depth,
        actor="analyst",
    )
    completed = await engine.wait(started["id"])
    assert completed["status"] == "succeeded", completed.get("error")
    assert provider.modules_by_run[completed["id"]] == set(
        compiled_route(pathway, depth).nodes
    )
    assert completed["provider_identity"] == provider.identity.as_dict()
    return completed


def _assert_uploaded_lineage(
    engine: Engine,
    run_id: str,
    expected_source_ids: set[str],
) -> None:
    run_date = engine.get_run(run_id)["created_at"][:10]
    artifacts = engine.artifacts_for_run(run_id)
    assert artifacts
    for artifact in artifacts:
        payload = artifact["payload"]
        assert payload["schema_version"] == "caos.canonical.artifact.v1"
        handoff = payload["handoff_metadata"]
        assert handoff["module_id"] == artifact["module_id"]
        assert handoff["run_id"] == run_id
        assert handoff["reporting_period"] == run_date
        assert handoff["analysis_date"] == run_date
        if artifact["module_id"] not in {"CP-PARSE", "CP-0"}:
            assert "CP-0" in {
                upstream["module_id"]
                for upstream in handoff["upstream_artifacts_used"]
            }
        assert {
            ref["source_id"] for ref in payload["evidence_refs"]
        } == expected_source_ids
        for calculation in payload["calculations"]:
            assert calculation["input_digest"] == digest(
                calculation["canonical_input"]
            )
            assert calculation["output_digest"] == digest(
                calculation["canonical_output"]
            )


def _required_deliverable_blocks(
    template: dict[str, Any],
    source: dict[str, Any],
) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    for definition in template["blocks"]:
        identity = {
            "block_id": definition["block_id"],
            "slot_id": definition["slot_id"],
        }
        if definition["kind"] == "NARRATIVE":
            blocks.append(
                {
                    **identity,
                    "kind": "NARRATIVE",
                    "text": (
                        f"Analyst interpretation for {definition['title']} based on "
                        "the accepted distressed evidence and calculated overlay."
                    ),
                    "content_mode": "ANALYST_JUDGMENT",
                    "citations": [],
                }
            )
            continue
        assert definition["kind"] == "EVIDENCE_REGISTER"
        blocks.append(
            {
                **identity,
                "kind": "EVIDENCE_REGISTER",
                "citations": [
                    {
                        "source_id": source["id"],
                        "block_ids": [source["blocks"][0]["block_id"]],
                        "claim": "The current distress update supports this published opinion.",
                    }
                ],
            }
        )
    return blocks


def _publish_overlay(
    *,
    service: DeliverableService,
    vault_dir: Path,
    engine: Engine,
    models: ModelService,
    case_id: str,
    source: dict[str, Any],
    snapshot: dict[str, Any],
    overlay: dict[str, Any],
) -> dict[str, Any]:
    """Publish through the ordinary required-model draft/freeze/file APIs."""

    pathway = "DISTRESSED_RESTRUCTURING"
    workspace = service.workspace(case_id, pathway)
    template = workspace["template"]
    assert template["model_requirement"] == "REQUIRED"
    previous = workspace["draft"]
    selection = {
        "kind": "APPLICATION_BUILD",
        "build_id": overlay["id"],
        "fallback_acknowledged": True,
    }
    revision = service.save_draft(
        case_id,
        pathway,
        DeliverableDraftRequest.model_validate(
            {
                "expected_version": previous["version"] if previous else 0,
                "template_id": template["template_id"],
                "template_version": template["template_version"],
                "model_selection": selection,
                "blocks": _required_deliverable_blocks(template, source),
            }
        ),
        actor="analyst",
    )
    assert revision["content"]["model_selection"] == selection
    assert revision["content"]["model_identity"]["build_id"] == overlay["id"]

    frozen = service.freeze(
        case_id,
        FreezeDeliverableRequest(
            draft_id=revision["draft_id"],
            draft_version=revision["version"],
            draft_digest=revision["digest"],
        ),
        actor="analyst",
    )
    source_set = service.store.current_source_set(case_id)
    expected_authority = {
        "accepted_snapshot_id": snapshot["id"],
        "source_set_id": source_set["id"],
        "source_set_version": source_set["version"],
        "build_id": overlay["id"],
    }
    assert frozen["payload"]["authority"] == expected_authority
    assert frozen["input_fingerprint"] == digest(
        {
            "snapshot_id": snapshot["id"],
                "source_set_id": source_set["id"],
                "source_set_version": source_set["version"],
                "build_id": overlay["id"],
                "methodology_build_id": engine.bundle.build_id,
            }
        )
    payload_without_preview = {
        key: value
        for key, value in frozen["payload"].items()
        if key != "preview_digest"
    }
    assert frozen["payload"]["preview_digest"] == digest(payload_without_preview)
    assert frozen["preview_digest"] == digest({
        "schema_version": "caos.frozen-approval.v1",
        "deliverable_id": frozen["deliverable_id"],
        "thread_id": frozen["thread_id"],
        "case_id": case_id,
        "pathway": "DISTRESSED_RESTRUCTURING",
        "build_id": overlay["id"],
        "draft_version": frozen["draft_version"],
        "draft_digest": frozen["draft_digest"],
        "content_digest": frozen["payload"]["preview_digest"],
        "input_fingerprint": frozen["input_fingerprint"],
        "exports": frozen["exports"],
        "authority": frozen["authority"],
    })
    pinned_model = frozen["payload"]["model"]
    assert pinned_model["kind"] == "APPLICATION_BUILD"
    assert pinned_model["build_id"] == overlay["id"]
    assert pinned_model["application_build"]["payload"] == {
        "payload_digest": overlay["payload_digest"]
    }
    assert pinned_model["application_build"]["qa"] == overlay["qa"]

    filed = service.approve_filing(
        case_id,
        frozen["deliverable_id"],
        FileDeliverableRequest(
            preview_digest=frozen["preview_digest"],
            input_fingerprint=frozen["input_fingerprint"],
        ),
        actor="approver-user",
    )
    assert filed["status"] == "FILED"
    assert filed["payload"] == frozen["payload"]

    # A fresh service reconstructs the exact governed record and the original
    # export bytes; filing never re-renders or mutates reviewed content.
    reopened = DeliverableService(
        store=service.store,
        vault_dir=vault_dir,
        engine=engine,
        models=models,
    )
    reconstructed = reopened.frozen_record(case_id, frozen["deliverable_id"])
    assert reconstructed["status"] == "FILED"
    assert reconstructed["payload"] == frozen["payload"]
    assert reconstructed["preview_digest"] == frozen["preview_digest"]
    assert reconstructed["input_fingerprint"] == frozen["input_fingerprint"]
    exported: dict[str, bytes] = {}
    for format_name, metadata in frozen["exports"].items():
        content, recorded_digest = reopened.export(
            frozen["deliverable_id"], format_name
        )
        exported[format_name] = content
        assert recorded_digest == metadata["sha256"]
        assert hashlib.sha256(content).hexdigest() == metadata["sha256"]
        assert len(content) == metadata["size"]

    events = service.audit_events_for_tests(case_id)
    draft_events = [
        event
        for event in events
        if event["action"] == "deliverable.draft.saved"
        and event.get("revision_id") == revision["revision_id"]
    ]
    assert len(draft_events) == 1
    assert draft_events[0]["actor"] == "analyst"
    assert draft_events[0]["version"] == revision["version"]
    gate_events: dict[str, dict[str, Any]] = {}
    for action, actor in (
        ("deliverable.frozen", "analyst"),
        ("deliverable.filed", "approver-user"),
    ):
        matched = [
            event
            for event in events
            if event["action"] == action
            and event.get("deliverable_id") == frozen["deliverable_id"]
        ]
        assert len(matched) == 1
        assert matched[0]["actor"] == actor
        gate_events[action] = matched[0]
    assert gate_events["deliverable.frozen"]["preview_digest"] == frozen[
        "preview_digest"
    ]
    assert events.index(gate_events["deliverable.filed"]) < events.index(
        gate_events["deliverable.frozen"]
    ) < events.index(draft_events[0])
    return {
        "filed": filed,
        "methodology": frozen["payload"]["methodology"],
        "exports": exported,
    }


def _searchable_export_text(format_name: str, content: bytes) -> str:
    if format_name == "md":
        return content.decode("utf-8")
    if format_name == "pdf":
        return "\n".join(
            page.extract_text() or ""
            for page in PdfReader(io.BytesIO(content)).pages
        )
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )


def _assert_visible_distressed_calculations(
    depth: str,
    exports: dict[str, bytes],
) -> None:
    for format_name, content in exports.items():
        text = _searchable_export_text(format_name, content)
        folded = text.casefold()
        assert "funding_gap" in folded, f"{depth} {format_name} omits funding_gap"
        assert re.search(
            r"(?<![\w.])150(?:\.0+)?(?![\w.])", text
        ), f"{depth} {format_name} omits funding_gap=150"
        assert re.search(r"\bSUN\b", text), (
            f"{depth} {format_name} omits the SUN recovery class"
        )
        assert "recovery_pct" in folded, (
            f"{depth} {format_name} omits recovery_pct"
        )
        assert re.search(r"(?<![\w.])0\.25(?:0+)?(?![\w.])", text), (
            f"{depth} {format_name} omits SUN recovery_pct=0.25"
        )


@pytest.mark.parametrize(
    "publishing_requirement",
    ("methodology_identity", "visible_distressed_calculations"),
)
async def test_user_uploads_flow_through_ordinary_full_and_screen_distressed_models(
    tmp_path,
    settings,
    store,
    publishing_requirement,
):
    """Full and screen Distressed share one accepted-model overlay contract.

    Full Credit first creates the immutable READY operating model.  Both
    Distressed depths then use ordinary semantic runs and accepted snapshots;
    ModelService deliberately resolves each through the same prior Full Credit
    base rather than creating a weaker screen-only model or deliverable path.
    """

    provider = OrdinaryFixtureProvider()
    engine = Engine.create(
        settings=settings,
        store=store,
        checkpoint_path=tmp_path / "ordinary-distressed-checkpoints.db",
        provider=provider,
    )
    models = ModelService(
        store=store,
        vault_dir=settings.storage_dir,
        engine=engine,
    )
    deliverable_vault = tmp_path / "ordinary-distressed-deliverable-vault"
    deliverables = DeliverableService(
        store=store,
        vault_dir=deliverable_vault,
        engine=engine,
        models=models,
    )
    try:
        with TestClient(create_app(settings=settings, store=store, engine=engine)) as client:
            created = client.post(
                "/api/cases",
                json={
                    "name": "Ordinary provider distressed case",
                    "issuer": "Acme Credit Ltd",
                    "sector": "Business services",
                },
            )
            assert created.status_code == 201, created.text
            case_id = created.json()["id"]
            assert store.add_member(
                case_id,
                "analyst",
                "approver-user",
                "APPROVER",
                actor_role="ADMIN",
            )
            sources = [
                _upload(
                    client,
                    case_id,
                    "annual-report.txt",
                    b"FY2024 annual report: revenue, EBITDA, debt, cash and facilities.",
                ),
                _upload(
                    client,
                    case_id,
                    "quarterly-update.txt",
                    b"FY2025 Q1 update: trading, cash flow and leverage movement.",
                ),
                _upload(
                    client,
                    case_id,
                    "management-forecast.txt",
                    b"FY2025-FY2027 management forecast with base and downside cases.",
                ),
            ]
            provider.bind(sources)
            base_source_ids = {source["id"] for source in sources}

            full_credit = await _complete(
                engine,
                provider,
                case_id,
                "FULL_CREDIT",
                "full",
            )
            _assert_uploaded_lineage(engine, full_credit["id"], base_source_ids)
            base_snapshot = await engine.accept(full_credit["id"], actor="analyst")
            assert models.readiness(case_id)["status"] == "READY_TO_BUILD", (
                models.readiness(case_id)
            )
            base_queued = next(
                build
                for build in models.list_builds(case_id)
                if build["snapshot_id"] == base_snapshot["id"]
            )
            base_build = models.run_build_for_tests(base_queued["id"])
            assert base_build["status"] == "READY"
            assert base_build["qa"]["status"] == "PASS"
            assert base_build["payload_digest"] == digest(base_build["payload"])

            distress_source = _upload(
                client,
                case_id,
                "distressed-update.txt",
                b"Near-term maturity, liquidity need, claim stack and recovery assumptions.",
            )
            sources.append(distress_source)
            provider.bind(sources)
            distressed_source_ids = {source["id"] for source in sources}

            overlays: dict[str, tuple[dict[str, Any], dict[str, Any]]] = {}
            published: dict[str, dict[str, Any]] = {}
            for depth in ("full", "screen"):
                run = await _complete(
                    engine,
                    provider,
                    case_id,
                    "DISTRESSED_RESTRUCTURING",
                    depth,
                )
                _assert_uploaded_lineage(engine, run["id"], distressed_source_ids)
                snapshot = await engine.accept(run["id"], actor="analyst")
                queued = next(
                    build
                    for build in models.list_builds(case_id)
                    if build["snapshot_id"] == snapshot["id"]
                )
                overlay = models.run_build_for_tests(queued["id"])
                assert overlay["status"] == "READY"
                assert overlay["payload_digest"] == digest(overlay["payload"])
                effect, = overlay["payload"]["pathway_effects"]
                assert effect["effect_id"] == "DISTRESSED_SCENARIO_RECOVERY"
                assert effect["base_model"] == {
                    "build_id": base_build["id"],
                    "run_id": base_snapshot["run_id"],
                    "snapshot_id": base_snapshot["id"],
                    "snapshot_digest": base_snapshot["digest"],
                    "source_set_id": base_snapshot["source_set_id"],
                    "source_set_version": base_snapshot["source_set_version"],
                    "input_fingerprint": base_build["input_fingerprint"],
                    "payload_digest": base_build["payload_digest"],
                    "assumptions_digest": base_build["assumptions_digest"],
                    "outputs_digest": base_build["outputs_digest"],
                    "qa_digest": digest(base_build["qa"]),
                    "methodology_build_id": base_build["methodology_build_id"],
                    "calculation_runtime": base_build["calculation_runtime"],
                }
                assert {
                    record["calculator_id"] for record in effect["calculations"]
                } == {"funding_gap", "recovery_waterfall"}
                assert effect["calculation_records_digest"] == digest(
                    effect["calculations"]
                )
                cp4c_reference = next(
                    reference
                    for reference in snapshot["artifacts"]
                    if reference["module_id"] == "CP-4C"
                )
                assert effect["distressed_authority"] == {
                    "run_id": snapshot["run_id"],
                    "snapshot_id": snapshot["id"],
                    "snapshot_digest": snapshot["digest"],
                    "source_set_id": snapshot["source_set_id"],
                    "source_set_version": snapshot["source_set_version"],
                    "provider_identity_digest": snapshot["provider_identity"][
                        "identity_digest"
                    ],
                    "artifacts": [
                        {
                            "module_id": reference["module_id"],
                            "artifact_id": reference["id"],
                            "digest": reference["digest"],
                        }
                        for reference in snapshot["artifacts"]
                    ],
                    "cp4c_artifact_id": cp4c_reference["id"],
                    "cp4c_artifact_digest": cp4c_reference["digest"],
                }
                assert effect["methodology_build_id"] == engine.bundle.build_id
                assert effect["cp_model_runtime"] == models.bundle.calculation_runtime
                assert overlay["payload"]["tabs"] == base_build["payload"]["tabs"]
                publication = _publish_overlay(
                    service=deliverables,
                    vault_dir=deliverable_vault,
                    engine=engine,
                    models=models,
                    case_id=case_id,
                    source=distress_source,
                    snapshot=snapshot,
                    overlay=overlay,
                )
                filed = publication["filed"]
                assert filed["payload"]["authority"]["accepted_snapshot_id"] == (
                    snapshot["id"]
                )
                overlays[depth] = (snapshot, overlay)
                published[depth] = publication

            full_snapshot, full_overlay = overlays["full"]
            screen_snapshot, screen_overlay = overlays["screen"]
            assert full_snapshot["previous_snapshot_id"] == base_snapshot["id"]
            assert screen_snapshot["previous_snapshot_id"] == full_snapshot["id"]
            assert screen_overlay["input_fingerprint"] != full_overlay["input_fingerprint"]
            assert models.readiness(case_id)["status"] == "READY"
            assert models.readiness(case_id)["snapshot_id"] == screen_snapshot["id"]

            # The model reads the separately indexed Markdown column, so that
            # column must remain byte-bound to canonical_output in the digested
            # payload.  A schema-preserving comment mutation still fails closed.
            base_cp1 = next(
                artifact
                for artifact in engine.artifacts_for_run(full_credit["id"])
                if artifact["module_id"] == "CP-1"
            )
            engine.runs.update_artifact_for_tests(
                full_credit["id"],
                "CP-1",
                markdown=base_cp1["markdown"]
                + "\n<!-- MARKDOWN-COLUMN-TAMPER -->\n",
            )
            tampered_readiness = models.readiness(case_id)
            assert tampered_readiness["status"] == "NOT_READY", tampered_readiness
            assert [
                blocker["code"] for blocker in tampered_readiness["blockers"]
            ] == ["DISTRESSED_BASE_MODEL_REQUIRED"]

            if publishing_requirement == "methodology_identity":
                for depth, (_snapshot, overlay) in overlays.items():
                    assert published[depth]["methodology"]["build_id"] == overlay[
                        "methodology_build_id"
                    ]
            else:
                for depth, publication in published.items():
                    _assert_visible_distressed_calculations(
                        depth, publication["exports"]
                    )
    finally:
        await engine.aclose()
