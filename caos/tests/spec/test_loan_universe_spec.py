"""CP-3 loan-universe specification: deterministic normalization with per-row
provenance, fail-closed workbook validation, source pinning, and the pinned
identity triple bound into the CP-3 artifact.

Sources: TEST_INVENTORY.md CONTRACTUAL rows for test_loan_universe.py (17), plus
the ledger rows test_loan_universe_versions_supersede_reject_and_withdraw_portably
and test_postgres_loan_import_and_withdrawal_serialize_without_deadlock (expressed
here without Postgres: both sequential orders plus an injected interleaving).

Spec-first: every import of an unbuilt module (caos.artifacts.loan_universe,
caos.api via the conftest client fixture, caos.engine.runtime via engine) happens
inside a helper called from a test body or inside a fixture, so every test FAILS
today with ModuleNotFoundError. A passing test is a defect in this suite.
"""

from __future__ import annotations

import io
import zipfile
from datetime import date

import pytest
from openpyxl import Workbook

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --- minimal workbook builders (rewritten, not copied) ----------------------------


def _row(
    *,
    borrower: str = "Access CIG LLC",
    bloomberg: str = "BLS202439",
    figi: str = "BBG01WMCP303",
    margin: object = 400,
    maturity: object = date(2030, 8, 19),
    change_1d: object = 0.5,
) -> list[object]:
    """One 25-column CP-3 sector row in template order (source units preserved)."""
    return [
        "Access CIG", borrower, "Records management services.", "Business Services",
        "Records Management", "Private", bloomberg, figi, "B1", "1L Gtd. Sr. Secd",
        "B3 / B", 1475, margin, maturity, 88, 90, change_1d, 0.5, 1, -2, -4.13, 1,
        -7.5, 11.2, 851,
    ]


def _sheet(workbook: Workbook, title: str, rows: list[list[object]], *, workbook_date: object = date(2026, 8, 24)) -> None:
    from caos.artifacts.loan_universe import HEADERS  # the template header row is module authority

    sheet = workbook.create_sheet(title)
    sheet["B1"] = "Date"
    sheet["B2"] = workbook_date
    for column, header in enumerate(HEADERS, start=1):
        sheet.cell(row=5, column=column, value=header)
    for row_number, values in enumerate(rows, start=6):
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_number, column=column, value=value)
    sheet.cell(row=6 + len(rows) + 1, column=1, value="Index Statistics")


def _workbook_bytes(
    *,
    first_rows: list[list[object]] | None = None,
    second_rows: list[list[object]] | None = None,
    second_date: object = date(2026, 8, 24),
) -> bytes:
    """Two visible sector sheets plus one hidden sheet that must never surface."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "IT Services", first_rows or [_row()])
    _sheet(
        workbook,
        "Healthcare IT",
        second_rows or [_row(borrower="FinThrive Inc", bloomberg="BLS1989347", figi="BBG01THRIVE1")],
        workbook_date=second_date,
    )
    _sheet(workbook, "Hidden Support", [_row(borrower="Hidden", bloomberg="HIDDEN", figi="BBG00HIDDEN1")])
    workbook["Hidden Support"].sheet_state = "hidden"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _invalid_template_bytes() -> bytes:
    """A partial-template workbook: one header cell drifted to 'Coupon'."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "IT Services", [_row()])
    workbook["IT Services"].cell(row=5, column=13, value="Coupon")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _parse(content: bytes, sha: str = "a" * 64) -> dict:
    from caos.artifacts.loan_universe import parse_loan_workbook

    return parse_loan_workbook(content, source_id="src_1", source_sha256=sha)


def _codes(error) -> set[str]:
    return {finding["code"] for finding in error.findings}


def _create_case(client, name: str = "Loan RV") -> str:
    return client.post(
        "/api/cases", json={"name": name, "issuer": "Issuer", "sector": "Services"}
    ).json()["id"]


def _upload_workbook(client, case_id: str, content: bytes, name: str = "REF_CP-3_Sector_RV.xlsx") -> dict:
    response = client.post(
        f"/api/cases/{case_id}/sources", files={"file": (name, content, XLSX_MEDIA_TYPE)}
    )
    assert response.status_code == 201
    return response.json()


def _import(client, case_id: str, source_id: str):
    return client.post(f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": source_id})


def _active(client, case_id: str):
    return client.get(f"/api/cases/{case_id}/rv/loan-universes/active")


# --- pure parser: deterministic normalization with provenance ---------------------


def test_identical_workbook_bytes_yield_identical_universe_digest():
    content = _workbook_bytes()
    assert _parse(content)["universe_digest"] == _parse(content)["universe_digest"]
    changed = _workbook_bytes(first_rows=[_row(margin=425)])
    assert _parse(changed)["universe_digest"] != _parse(content)["universe_digest"]


def test_visible_rows_map_with_source_units_and_locators_while_hidden_sheets_never_appear():
    parsed = _parse(
        _workbook_bytes(
            first_rows=[
                _row(),
                _row(borrower="Apex Group", bloomberg="BLS5005287", figi="BBG01S807689", margin="#N/A", change_1d="N/A"),
            ]
        )
    )
    assert parsed["workbook_date"] == "2026-08-24"
    assert parsed["row_count"] == 3
    access = next(row for row in parsed["rows"] if row["borrower_name"] == "Access CIG LLC")
    assert access["instrument_key"] == "FIGI:BBG01WMCP303"
    assert access["size_mn"] == 1475.0 and access["margin_bps"] == 400.0, "source units preserved"
    assert access["maturity_date"] == "2030-08-19"
    assert access["mid_3y_dm_bps"] == 851.0
    assert access["source_locators"] == [{"sheet": "IT Services", "row": 6}]
    apex = next(row for row in parsed["rows"] if row["borrower_name"] == "Apex Group")
    assert apex["margin_bps"] is None and apex["change_1d_points"] is None, "#N/A stays null"
    assert all(row["borrower_name"] != "Hidden" for row in parsed["rows"])


def test_duplicate_rows_collapse_and_preserve_every_locator():
    duplicate = _row()
    parsed = _parse(_workbook_bytes(first_rows=[duplicate, duplicate]), sha="b" * 64)
    access = next(row for row in parsed["rows"] if row["borrower_name"] == "Access CIG LLC")
    assert access["source_locators"] == [
        {"sheet": "IT Services", "row": 6},
        {"sheet": "IT Services", "row": 7},
    ], "collapse never discards provenance"
    assert parsed["row_count"] == 2, "row_count counts pre-collapse originals"


def test_blank_optional_cells_remain_null():
    parsed = _parse(_workbook_bytes(first_rows=[_row(margin=None)]), sha="0" * 64)
    access = next(row for row in parsed["rows"] if row["borrower_name"] == "Access CIG LLC")
    assert access["margin_bps"] is None, "blank is null — never coerced to zero"


def test_bloomberg_alias_reconciles_a_missing_figi_before_duplicate_collapse():
    parsed = _parse(
        _workbook_bytes(first_rows=[_row(bloomberg="LOAN1", figi="#N/A"), _row(bloomberg="LOAN1", figi="FIGI1")]),
        sha="9" * 64,
    )
    access = next(row for row in parsed["rows"] if row["bloomberg_loan_id"] == "LOAN1")
    assert access["instrument_key"] == "FIGI:FIGI1" and access["figi"] == "FIGI1"
    assert access["source_locators"] == [
        {"sheet": "IT Services", "row": 6},
        {"sheet": "IT Services", "row": 7},
    ], "one instrument identity, both locators"
    assert parsed["row_count"] == 2


# --- pure parser: fail-closed validation ------------------------------------------


def test_conflicting_duplicates_and_identifier_mappings_reject_the_candidate():
    from caos.artifacts.loan_universe import LoanWorkbookValidationError

    with pytest.raises(LoanWorkbookValidationError) as raised:
        _parse(
            _workbook_bytes(
                first_rows=[_row(), _row(margin=425), _row(borrower="Other", bloomberg="OTHER", figi="BBG01WMCP303")]
            ),
            sha="c" * 64,
        )
    assert {"RV_DUPLICATE_CONFLICT", "RV_ID_CONFLICT"} <= _codes(raised.value)


def test_partial_headers_and_conflicting_dates_reject_the_complete_workbook():
    from caos.artifacts.loan_universe import LoanWorkbookValidationError

    with pytest.raises(LoanWorkbookValidationError) as date_error:
        _parse(_workbook_bytes(second_date="25/08/2026"), sha="d" * 64)
    with pytest.raises(LoanWorkbookValidationError) as header_error:
        _parse(_invalid_template_bytes(), sha="e" * 64)
    assert "RV_WORKBOOK_DATE_CONFLICT" in _codes(date_error.value)
    assert {"RV_TEMPLATE_PARTIAL", "RV_TEMPLATE_MISSING"} <= _codes(header_error.value)


@pytest.mark.parametrize(
    ("row", "expected_code"),
    [
        (_row(borrower="", bloomberg="", figi=""), "RV_BORROWER_MISSING"),
        (_row(margin="not-a-number"), "RV_NUMBER_INVALID"),
        (_row(margin="1e999"), "RV_NUMBER_NON_FINITE"),
        (_row(maturity="31/31/2030"), "RV_MATURITY_INVALID"),
        (_row(maturity=0), "RV_MATURITY_INVALID"),
    ],
)
def test_invalid_rows_fail_closed(row, expected_code):
    from caos.artifacts.loan_universe import LoanWorkbookValidationError

    with pytest.raises(LoanWorkbookValidationError) as raised:
        _parse(_workbook_bytes(first_rows=[row]), sha="f" * 64)
    assert expected_code in _codes(raised.value)


def test_formula_without_cached_value_becomes_null_without_execution():
    parsed = _parse(_workbook_bytes(first_rows=[_row(margin="=200+200")]), sha="1" * 64)
    access = next(row for row in parsed["rows"] if row["borrower_name"] == "Access CIG LLC")
    assert access["margin_bps"] is None, "formulas are never evaluated"


# --- package screening runs before parsing ----------------------------------------


def test_unsafe_package_parts_reject_before_workbook_parsing():
    from caos.artifacts.loan_universe import LoanWorkbookValidationError, screen_package

    content = io.BytesIO(_workbook_bytes())
    with zipfile.ZipFile(content, "a") as archive:
        archive.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")
    tampered = content.getvalue()
    with pytest.raises(LoanWorkbookValidationError) as screened:
        screen_package(tampered)
    assert "RV_PACKAGE_ACTIVE_CONTENT" in _codes(screened.value)
    with pytest.raises(LoanWorkbookValidationError) as parsed:
        _parse(tampered, sha="2" * 64)
    assert "RV_PACKAGE_ACTIVE_CONTENT" in _codes(parsed.value), "screening precedes parsing"


def test_external_relationship_with_xml_whitespace_is_rejected():
    from caos.artifacts.loan_universe import LoanWorkbookValidationError, screen_package

    content = io.BytesIO(_workbook_bytes())
    with zipfile.ZipFile(content, "a") as archive:
        archive.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            """<?xml version="1.0" encoding="UTF-8"?>
            <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
              <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" Target="https://example.invalid" TargetMode = "External"/>
            </Relationships>""",
        )
    with pytest.raises(LoanWorkbookValidationError) as raised:
        screen_package(content.getvalue())
    assert "RV_PACKAGE_EXTERNAL_LINK" in _codes(raised.value), "whitespace around TargetMode is still caught"


def test_workbook_sheet_limit_rejects_over_cap_counts():
    from caos.artifacts.loan_universe import LoanWorkbookValidationError

    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "IT Services", [_row()])
    for index in range(64):
        workbook.create_sheet(f"Hidden {index:02d}").sheet_state = "hidden"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    with pytest.raises(LoanWorkbookValidationError) as raised:
        _parse(output.getvalue(), sha="3" * 64)
    assert "RV_WORKSHEET_LIMIT" in _codes(raised.value)


# --- import lifecycle over the API ------------------------------------------------


def test_import_is_idempotent_on_content_identity_and_serves_the_active_rows(client):
    case_id = _create_case(client)
    source = _upload_workbook(client, case_id, _workbook_bytes())
    imported = _import(client, case_id, source["id"])
    repeated = _import(client, case_id, source["id"])
    active = _active(client, case_id)
    assert imported.status_code == 201 and repeated.status_code == 200, "repeat is 200, not 201"
    assert repeated.json()["id"] == imported.json()["id"]
    assert active.status_code == 200 and active.json()["status"] == "ACTIVE"
    assert active.json()["universe"]["id"] == imported.json()["id"]
    assert active.json()["universe"]["source_id"] == source["id"]
    assert len(active.json()["rows"]) == 2
    assert active.json()["rows"][0]["mid_3y_dm_bps"] == 851.0, "pinned rows served with units intact"


def test_invalid_import_returns_structured_findings_and_preserves_prior_active_universe(client):
    case_id = _create_case(client)
    valid = _upload_workbook(client, case_id, _workbook_bytes())
    active_id = _import(client, case_id, valid["id"]).json()["id"]
    invalid = _upload_workbook(client, case_id, _invalid_template_bytes(), "changed-template.xlsx")
    rejected = _import(client, case_id, invalid["id"])
    repeated = _import(client, case_id, invalid["id"])
    assert rejected.status_code == repeated.status_code == 422
    assert rejected.json()["detail"]["code"] == "RV_WORKBOOK_INVALID"
    assert {finding["code"] for finding in rejected.json()["detail"]["findings"]} >= {
        "RV_TEMPLATE_PARTIAL",
        "RV_TEMPLATE_MISSING",
    }
    assert repeated.json()["detail"]["universe_id"] == rejected.json()["detail"]["universe_id"], "rejection is idempotent with a stable id"
    assert _active(client, case_id).json()["universe"]["id"] == active_id


def test_import_rejects_vault_bytes_that_do_not_match_the_source_digest(client, store):
    case_id = _create_case(client)
    source = _upload_workbook(client, case_id, _workbook_bytes())
    store.replace_vault_bytes_for_tests(source["id"], _workbook_bytes(first_rows=[_row(margin=425)]))
    rejected = _import(client, case_id, source["id"])
    assert rejected.status_code == 409
    assert rejected.json()["detail"]["code"] == "RV_SOURCE_INTEGRITY_MISMATCH"
    assert _active(client, case_id).json()["status"] == "NO_ACTIVE_UNIVERSE", "nothing created from tampered bytes"


def test_loan_universe_api_is_case_scoped_and_reader_safe(client, store):
    case_id = _create_case(client)
    other_case_id = _create_case(client, "Other")
    source = _upload_workbook(client, case_id, _workbook_bytes())
    assert _import(client, other_case_id, source["id"]).status_code == 404, "sources are unusable across cases"
    store.add_member(case_id, "analyst", "reader-user", "READER", actor_role="ADMIN")
    # case READER outranks the claimed global role
    reader = {"x-forwarded-user": "reader-user", "x-caos-role": "ANALYST"}
    assert client.get(f"/api/cases/{case_id}/rv/loan-universes/active", headers=reader).status_code == 200
    assert _import(client, case_id, source["id"]).status_code == 201  # writer path still works
    assert client.post(
        f"/api/cases/{case_id}/rv/loan-universes", json={"source_id": source["id"]}, headers=reader
    ).status_code == 403
    outsider = {"x-forwarded-user": "outsider"}
    assert client.get(f"/api/cases/{case_id}/rv/loan-universes/active", headers=outsider).status_code == 404


def test_source_withdrawal_deactivates_loan_universe(client):
    case_id = _create_case(client)
    source = _upload_workbook(client, case_id, _workbook_bytes())
    assert _import(client, case_id, source["id"]).status_code == 201
    assert client.post(f"/api/cases/{case_id}/sources/{source['id']}/withdraw").status_code == 200
    assert _active(client, case_id).json() == {
        "status": "NO_ACTIVE_UNIVERSE",
        "universe": None,
        "rows": [],
    }, "derived artifacts cannot outlive their evidence"


# --- versioning portability (ledger row) ------------------------------------------


def test_new_import_supersedes_the_prior_active_and_keeps_exactly_one_active(client, store):
    case_id = _create_case(client)
    first = _upload_workbook(client, case_id, _workbook_bytes())
    imported = _import(client, case_id, first["id"])
    assert imported.status_code == 201
    replay = _import(client, case_id, first["id"])
    assert replay.status_code == 200 and replay.json()["id"] == imported.json()["id"]
    second = _upload_workbook(client, case_id, _workbook_bytes(first_rows=[_row(margin=425)]), "updated.xlsx")
    superseding = _import(client, case_id, second["id"])
    assert superseding.status_code == 201
    assert _active(client, case_id).json()["universe"]["id"] == superseding.json()["id"]
    records = {record["id"]: record for record in store.list_loan_universes(case_id)}
    assert records[imported.json()["id"]]["status"] == "SUPERSEDED"
    assert [record["status"] for record in records.values()].count("ACTIVE") == 1


def test_rejected_candidates_never_activate_and_withdrawn_sources_refuse_new_imports(client, store):
    case_id = _create_case(client)
    invalid = _upload_workbook(client, case_id, _invalid_template_bytes(), "invalid.xlsx")
    assert _import(client, case_id, invalid["id"]).status_code == 422
    assert all(record["status"] != "ACTIVE" for record in store.list_loan_universes(case_id))
    valid = _upload_workbook(client, case_id, _workbook_bytes())
    assert client.post(f"/api/cases/{case_id}/sources/{valid['id']}/withdraw").status_code == 200
    refused = _import(client, case_id, valid["id"])
    assert refused.status_code in {409, 422}
    assert refused.json()["detail"]["code"] == "RV_SOURCE_NOT_ACTIVE"
    assert _active(client, case_id).json()["status"] == "NO_ACTIVE_UNIVERSE"


# --- import racing withdrawal (ledger row, expressed without Postgres) ------------


def test_withdraw_then_import_refuses_with_a_typed_source_not_active(client):
    case_id = _create_case(client)
    source = _upload_workbook(client, case_id, _workbook_bytes())
    assert client.post(f"/api/cases/{case_id}/sources/{source['id']}/withdraw").status_code == 200
    refused = _import(client, case_id, source["id"])
    assert refused.status_code in {409, 422}
    assert refused.json()["detail"]["code"] == "RV_SOURCE_NOT_ACTIVE"
    assert _active(client, case_id).json()["status"] == "NO_ACTIVE_UNIVERSE"


def test_import_then_withdraw_leaves_the_universe_withdrawn_and_inactive(client, store):
    case_id = _create_case(client)
    source = _upload_workbook(client, case_id, _workbook_bytes())
    assert _import(client, case_id, source["id"]).status_code == 201
    assert client.post(f"/api/cases/{case_id}/sources/{source['id']}/withdraw").status_code == 200
    assert _active(client, case_id).json()["status"] == "NO_ACTIVE_UNIVERSE"
    records = store.list_loan_universes(case_id)
    assert any(record["status"] == "WITHDRAWN" for record in records)
    assert all(record["status"] != "ACTIVE" for record in records)


def test_withdrawal_injected_inside_the_import_window_never_yields_an_active_universe(client, store):
    """Portable restaging of the Postgres serialization race: the withdrawal lands in
    the window between the import's source-status check and its universe write. The
    only legal end states are {refused, withdrawn} — never an active universe on a
    withdrawn source."""
    case_id = _create_case(client)
    source = _upload_workbook(client, case_id, _workbook_bytes())
    store.interpose_before_universe_write_for_tests(
        lambda: store.withdraw(case_id, source["id"], "analyst")
    )
    raced = _import(client, case_id, source["id"])
    if raced.status_code in {409, 422}:
        assert raced.json()["detail"]["code"] in {"RV_SOURCE_NOT_ACTIVE", "RV_SOURCE_INTEGRITY_MISMATCH"}
    else:
        assert raced.json()["status"] != "ACTIVE", "the loser must not commit an active universe"
    assert _active(client, case_id).json() == {
        "status": "NO_ACTIVE_UNIVERSE",
        "universe": None,
        "rows": [],
    }
    assert all(record["status"] != "ACTIVE" for record in store.list_loan_universes(case_id))
    assert client.get(f"/api/cases/{case_id}/sources/{source['id']}").json()["withdrawn"] is True


# --- CP-3 artifact binds the pinned universe identity (engine-level) --------------


async def test_cp3_artifact_binds_the_pinned_normalized_loan_universe(client, store, engine):
    # Amended 2026-08-27 with user sign-off: CP-3 rides only full-depth routes
    # (the verified catalog LITE_RELATIVE_VALUE selection is CP-0/CP-L10/CP-1C,
    # and adding a node is a methodology change), so this binds on the
    # RELATIVE_VALUE full route via the scripted-canonical run seam.
    case_id = _create_case(client)
    source = _upload_workbook(client, case_id, _workbook_bytes())
    universe = _import(client, case_id, source["id"]).json()
    active_rows = _active(client, case_id).json()["rows"]
    run = await engine.run_scripted_for_tests(case_id, pathway="RELATIVE_VALUE")
    assert run["status"] == "succeeded"
    cp3 = next(artifact for artifact in engine.artifacts_for_run(run["id"]) if artifact["module_id"] == "CP-3")
    identity = cp3["payload"]["lineage"]["loan_universe"]
    assert identity == {
        "id": universe["id"],
        "universe_digest": universe["universe_digest"],
        "source_id": source["id"],
    }, "the identity triple is pinned"
    assert cp3["payload"]["provenance"]["loan_universe"] == identity
    assert cp3["payload"]["inputs"]["loan_universe"] == {"identity": identity, "rows": active_rows}, (
        "the artifact consumed exactly the pinned normalized rows"
    )
    assert {row["source_locators"][0]["sheet"] for row in active_rows} == {"IT Services", "Healthcare IT"}


# ROW MAPPING (TEST_INVENTORY.md -> this file)
# test_cp3_workbook_maps_all_visible_sector_rows_with_source_units -> test_identical_workbook_bytes_yield_identical_universe_digest + test_visible_rows_map_with_source_units_and_locators_while_hidden_sheets_never_appear
# test_duplicate_rows_collapse_and_preserve_every_locator -> test_duplicate_rows_collapse_and_preserve_every_locator
# test_blank_optional_cells_remain_null_and_keep_their_column_locator -> test_blank_optional_cells_remain_null
# test_bloomberg_alias_reconciles_a_missing_figi_before_duplicate_collapse -> test_bloomberg_alias_reconciles_a_missing_figi_before_duplicate_collapse
# test_conflicting_duplicates_and_identifier_mappings_reject_the_candidate -> test_conflicting_duplicates_and_identifier_mappings_reject_the_candidate
# test_partial_headers_and_conflicting_dates_reject_the_complete_workbook -> test_partial_headers_and_conflicting_dates_reject_the_complete_workbook
# test_invalid_rows_fail_closed -> test_invalid_rows_fail_closed (x5 parametrized, non-finite 1e999 guard included)
# test_formula_without_cached_value_becomes_null_without_execution -> test_formula_without_cached_value_becomes_null_without_execution
# test_unsafe_package_parts_reject_before_workbook_parsing -> test_unsafe_package_parts_reject_before_workbook_parsing
# test_external_relationship_with_xml_whitespace_is_rejected -> test_external_relationship_with_xml_whitespace_is_rejected
# test_workbook_sheet_limit_rejects_without_scanning_beyond_the_cap -> test_workbook_sheet_limit_rejects_over_cap_counts (the not-scanning-beyond claim was never directly asserted in legacy either)
# test_case_api_imports_reads_and_idempotently_returns_the_active_universe -> test_import_is_idempotent_on_content_identity_and_serves_the_active_rows
# test_invalid_import_returns_structured_findings_and_preserves_prior_active_universe -> test_invalid_import_returns_structured_findings_and_preserves_prior_active_universe
# test_import_rejects_vault_bytes_that_do_not_match_the_source_digest -> test_import_rejects_vault_bytes_that_do_not_match_the_source_digest
# test_loan_universe_api_is_case_scoped_and_reader_safe -> test_loan_universe_api_is_case_scoped_and_reader_safe
# test_source_withdrawal_deactivates_loan_universe -> test_source_withdrawal_deactivates_loan_universe
# test_cp3_artifact_binds_the_pinned_normalized_loan_universe -> test_cp3_artifact_binds_the_pinned_normalized_loan_universe
# test_loan_universe_versions_supersede_reject_and_withdraw_portably (test_ledger_contracts.py) -> test_new_import_supersedes_the_prior_active_and_keeps_exactly_one_active + test_rejected_candidates_never_activate_and_withdrawn_sources_refuse_new_imports
# test_postgres_loan_import_and_withdrawal_serialize_without_deadlock (test_ledger_contracts.py) -> test_withdraw_then_import_refuses_with_a_typed_source_not_active + test_import_then_withdraw_leaves_the_universe_withdrawn_and_inactive + test_withdrawal_injected_inside_the_import_window_never_yields_an_active_universe (the two-connection Postgres deadlock staging is inexpressible without Postgres; the serialization guarantee is asserted via both sequential orders plus an injected interleaving)
# test_loan_universe_migration_has_atomic_identity_and_active_constraints -> not ported: MECHANISM row (asserts legacy migration SQL text; the constraint intent — identity uniqueness, one ACTIVE — is asserted behaviorally above)
# test_loan_universe_source_foreign_key_migration_backfills_before_validation -> not ported: MECHANISM row (legacy Postgres migration-file text with no behavioral counterpart)


def test_relationships_part_declaring_a_dtd_is_refused_without_expanding_it():
    """The screen parses attacker-supplied XML. `ElementTree.fromstring` expands
    internal entities, so a few hundred bytes of .rels became tens of megabytes
    of allocation inside the very function that exists to refuse hostile
    packages. An OOXML part carries no DTD: refuse the declaration."""
    import time
    import tracemalloc

    from caos.artifacts.loan_universe import LoanWorkbookValidationError, screen_package

    entities = ['<!ENTITY lol "lol">'] + [
        f'<!ENTITY lol{level} "{"&%s;" % ("lol" if level == 1 else f"lol{level - 1}") * 10}">'
        for level in range(1, 10)
    ]
    bomb = ('<?xml version="1.0"?>\n<!DOCTYPE Relationships [\n' + "\n".join(entities)
            + "\n]>\n<Relationships>&lol9;</Relationships>").encode()
    content = io.BytesIO(_workbook_bytes())
    with zipfile.ZipFile(content, "a") as archive:
        archive.writestr("xl/worksheets/_rels/sheet1.xml.rels", bomb)

    tracemalloc.start()
    started = time.monotonic()
    try:
        with pytest.raises(LoanWorkbookValidationError) as raised:
            screen_package(content.getvalue())
        peak = tracemalloc.get_traced_memory()[1]
    finally:
        tracemalloc.stop()
    assert "RV_PACKAGE_INVALID" in _codes(raised.value)
    assert peak < 8_000_000, f"the declaration must be refused, not expanded (peak {peak} bytes)"
    assert time.monotonic() - started < 1.0
