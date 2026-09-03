"""Frozen-deliverable export renderers (md / pdf / xlsx) — Task 10.

Every export walks one structure, `publication_view(payload)`: the masthead,
the pages of canonical sections (opinion first, the pathway's sections, then
the Evidence & QA Control Sheet) and the revision record. Nothing below reads
draft blocks, model records or the store, so the four formats — these three
and the browser, which draws `payload["publication"]` directly — carry the
same facts, numbers, units, citations, origin labels, limitations, model
identity and opinion by construction.

PDF: pango-view shapes every page from Pango markup. Fonts are the vendored
DejaVu bundle under `fonts/` — verified on the bytes at use, served through a
hermetic fontconfig with hinting off and pango pinned to the fontconfig
backend — so Latin, Greek and Cyrillic layout is byte-stable across a
developer Mac, CI and the image. Scripts DejaVu lacks fall back to the host's
fonts (`fonts-noto-cjk` ships in the image) on absolute per-span line heights,
so a fallback face never moves a page break. Pages are paginated by measuring
candidate pages, tables repeat their header
row across a page split, and a rotated transparent `PENDING APPROVAL`
watermark is merged under each page with pypdf. XLSX: openpyxl, typed numeric
cells for model-owned values, formula prefixes neutralised, no formulas, frozen
header rows and filters on every table sheet, deterministic zip re-pack.
Renders are content-addressed, so nothing here reads the clock.
"""

from __future__ import annotations

import hashlib
import html
import io
import math
import os
import re
import shutil
import subprocess
import tempfile
import textwrap
from decimal import Decimal
from pathlib import Path
from typing import Any

from .markdown import (  # noqa: F401 — re-exported for callers and tests
    FORMULA_PREFIXES,
    MASTHEAD_FIELDS,
    PENDING_APPROVAL,
    _PLAIN_NUMBER,
    _is_numeric,
    _masthead_line,
    _origin_label,
    _section_rows,
    _short_id,
    _walk_sections,
    publication_view,
    render_frozen_markdown,
)


# --- PDF ---------------------------------------------------------------------------

PAGE_WIDTH, PAGE_HEIGHT, MARGIN = 612, 792, 54
BODY_WIDTH = PAGE_WIDTH - 2 * MARGIN
FOOTER_HEIGHT = 22  # the footer is its own layer, pinned above the bottom margin
BODY_HEIGHT = PAGE_HEIGHT - 2 * MARGIN - FOOTER_HEIGHT
MONO_COLUMNS = 100  # 8pt monospace on a 504pt line
_FEATURES = 'font_features="liga=0, clig=0, dlig=0"'
_INK, _META, _RULE = "#16161e", "#5c5c66", "#9c998e"
# The font pin (DejaVu 2.37, Bitstream Vera licence in fonts/LICENSE). The
# bytes are the pin: a host copy of the same family name never substitutes.
FONT_DIR = Path(__file__).resolve().parent / "fonts"
FONT_BUNDLE = {
    "DejaVuSans.ttf": "7da195a74c55bef988d0d48f9508bd5d849425c1770dba5d7bfc6ce9ed848954",
    "DejaVuSans-Bold.ttf": "e6476c1b80502924294eed40894c5b18e06c181444ca953e5334262df9c27724",
    "DejaVuSansMono.ttf": "b4a6c3e4faab8773f4ff761d56451646409f29abedd68f05d38c2df667d3c582",
    "DejaVuSansMono-Bold.ttf": "bce60f1b4421acd9ea51ba6623d7024ecbe6817a953e3654df62a5e6bdf8f769",
}
SANS, MONO = "DejaVu Sans", "DejaVu Sans Mono"
LINE_HEIGHT = 1.2  # absolute line height per span, as a factor of its size
_HOST_FONT_CONFIGS = ("/etc/fonts/fonts.conf", "/opt/homebrew/etc/fonts/fonts.conf", "/usr/local/etc/fonts/fonts.conf")


def _font_environment(workspace: Path) -> dict[str, str]:
    """Process environment for pango-view: a hermetic fontconfig that lists the
    verified bundle first and the host's own configuration after it, for the
    scripts the bundle lacks. Debian's DejaVu package is rejected by path so the
    vendored bytes answer even where the host ships the same family;
    PANGOCAIRO_BACKEND pins pango to fontconfig where the host default is
    CoreText (macOS), so one file answers every font request everywhere."""
    config = workspace / "fonts.conf"
    if not config.exists():
        for filename, expected in FONT_BUNDLE.items():
            path = FONT_DIR / filename
            if not path.is_file() or hashlib.sha256(path.read_bytes()).hexdigest() != expected:
                raise ValueError("PDF_FONT_BUNDLE_INVALID")
        includes = "".join(f'<include ignore_missing="yes">{_esc(host)}</include>' for host in _HOST_FONT_CONFIGS)
        config.write_text(
            '<?xml version="1.0"?><!DOCTYPE fontconfig SYSTEM "fonts.dtd"><fontconfig>'
            f"<dir>{_esc(FONT_DIR)}</dir><cachedir>{_esc(workspace / 'fontconfig-cache')}</cachedir>{includes}"
            "<selectfont><rejectfont><glob>/usr/share/fonts/truetype/dejavu/*</glob></rejectfont></selectfont>"
            "</fontconfig>",
            encoding="utf-8",
        )
    return {**os.environ, "SOURCE_DATE_EPOCH": "0", "FONTCONFIG_FILE": str(config), "PANGOCAIRO_BACKEND": "fontconfig"}


def _pango(executable: str, workspace: Path, name: str, text: str, *options: str) -> Path:
    """One pango-view call under the pinned font environment: hinting off,
    metrics unhinted and glyphs positioned at subpixel precision, so advances
    are the design values on every host and a kerned pair ("y.") is not
    rounded into a gap that text extraction reads as a space."""
    source, rendered = workspace / f"{name}.txt", workspace / f"{name}.pdf"
    source.write_text(text, encoding="utf-8")
    command = [executable, "--no-display", "--pixels", "--hinting=none", "--hint-metrics=off", "--subpixel-positions",
               *options, f"--output={rendered}", str(source)]
    try:
        subprocess.run(command, check=True, capture_output=True, timeout=60, env=_font_environment(workspace))
    except (OSError, subprocess.SubprocessError) as exc:
        raise ValueError("PDF_UNICODE_RENDERER_FAILED") from exc
    return rendered


def _esc(text: str) -> str:
    return html.escape(str(text), quote=False)


def _span(text: str, *, size: float, weight: str = "normal", family: str = SANS, color: str = _INK) -> str:
    # No letter-spacing anywhere: Pango tracking is emitted as per-glyph
    # positioning and text extraction then reads "A N A L Y S T", which breaks
    # search, copy and the cross-format parity check.
    # `line_height` as an integer above 1024 is absolute Pango units: the line
    # is exactly LINE_HEIGHT × size tall whichever face shaped it, so a fallback
    # font for a script DejaVu lacks never changes where a page breaks.
    attributes = [f'font_family="{family}"', f'size="{int(size * 1024)}"', f'weight="{weight}"',
                  f'foreground="{color}"', f'line_height="{int(size * LINE_HEIGHT * 1024)}"', _FEATURES]
    return f"<span {' '.join(attributes)}>{_esc(text)}</span>"


def _rule(strong: bool = False) -> str:
    return _span("─" * MONO_COLUMNS, size=7, family=MONO, color=_INK if strong else _RULE)


class _Block:
    """One indivisible unit of a page: a list of markup lines, whether it must
    stay with the next block, and the table header it should repeat after a
    page split (for table rows)."""

    __slots__ = ("lines", "keep_with_next", "repeat_header")

    def __init__(self, lines: list[str], *, keep_with_next: bool = False, repeat_header: list[str] | None = None) -> None:
        self.lines = lines
        self.keep_with_next = keep_with_next
        self.repeat_header = repeat_header


def _wrap_cell(text: str, width: int) -> list[str]:
    """Wrap on spaces; break inside a token only when the token itself is
    wider than the column (digests, ids), so no ordinary word is ever split."""
    lines: list[str] = []
    for paragraph in str(text).split("\n") or [""]:
        words = paragraph.split(" ")
        current = ""
        for word in words:
            while len(word) > width:
                if current:
                    lines.append(current)
                    current = ""
                lines.append(word[:width])
                word = word[width:]
            candidate = f"{current} {word}".strip() if current else word
            if len(candidate) <= width:
                current = candidate
            else:
                lines.append(current)
                current = word
        lines.append(current)
    return lines or [""]


def _record_rows(rows: list[list[str]]) -> list[_Block]:
    """Wide tables (wider than the monospace line) become one record per row:
    `Column  value` lines that never interleave across columns, so a digest or
    a long disposition reads whole. The header block lists the columns."""
    header = rows[0]
    label_width = min(max(len(str(column)) for column in header) + 1, 24)
    value_width = MONO_COLUMNS - label_width - 2
    header_lines = [
        _span("Columns: " + " · ".join(str(column) for column in header), size=8, family=MONO, weight="bold"),
        _rule(),
    ]
    blocks = [_Block(header_lines, keep_with_next=True)]
    for row in rows[1:]:
        lines: list[str] = []
        for index, column in enumerate(header):
            value = str(row[index]) if index < len(row) else ""
            wrapped = _wrap_cell(value, value_width)
            lines.append(_span(str(column).ljust(label_width), size=8, family=MONO, color=_META) + "  " + _span(wrapped[0], size=8, family=MONO))
            lines.extend(_span(" " * label_width, size=8, family=MONO) + "  " + _span(part, size=8, family=MONO) for part in wrapped[1:])
        lines.append(_span("·", size=6, color=_RULE))
        blocks.append(_Block(lines, repeat_header=header_lines))
    return blocks


def _mono_table(rows: list[list[str]], *, model_owned: bool) -> list[_Block]:
    """Fixed-width monospace grid when the natural widths fit the line:
    numerics right-aligned, cells wrapped on spaces inside their column, header
    repeated across page splits. Wider tables use the record layout."""
    if not rows:
        return []
    columns = len(rows[0])
    natural = [max(len(str(row[index])) if index < len(row) else 0 for row in rows) for index in range(columns)]
    if sum(natural) + 2 * (columns - 1) > MONO_COLUMNS:
        return _record_rows(rows)
    numeric = [
        all(_is_numeric(str(row[index])) for row in rows[1:] if index < len(row)) and len(rows) > 1
        for index in range(columns)
    ]
    widths = list(natural)

    def cell_lines(row: list[str]) -> list[list[str]]:
        wrapped = [_wrap_cell(str(row[index]) if index < len(row) else "", widths[index]) for index in range(columns)]
        height = max(len(part) for part in wrapped)
        return [
            [
                (wrapped[index][line] if line < len(wrapped[index]) else "").rjust(widths[index])
                if numeric[index] else
                (wrapped[index][line] if line < len(wrapped[index]) else "").ljust(widths[index])
                for index in range(columns)
            ]
            for line in range(height)
        ]

    def markup(parts: list[str], *, bold: bool) -> str:
        return _span("  ".join(parts), size=8, family=MONO, weight="bold" if bold else "normal")

    header_lines = [markup(parts, bold=True) for parts in cell_lines(rows[0])] + [_rule()]
    blocks = [_Block(header_lines, keep_with_next=True)]
    for row in rows[1:]:
        blocks.append(_Block([markup(parts, bold=False) for parts in cell_lines(row)], repeat_header=header_lines))
    return blocks


PROSE_COLUMNS = 94  # DejaVu Sans 9.5pt prose on a 504pt line; pango wraps the rare wider line itself


def _prose_lines(text: str, *, indent: str = "") -> list[str]:
    """A paragraph as one span per visual line. A paragraph handed to pango as
    a single line cannot be divided at a page break — the paginator halves
    blocks by their lines — so a long narrative overprinted the footer and ran
    off the page. Pre-wrapped lines paginate exactly."""
    parts = textwrap.wrap(text, width=PROSE_COLUMNS, break_on_hyphens=False, subsequent_indent=indent) if text.strip() else []
    return [_span(part, size=9.5) for part in parts] or [_span(" ", size=9.5)]


def _section_blocks(section: dict[str, Any], depth: int) -> list[_Block]:
    heading = [
        _span(section["title"], size=10.5 - depth, weight="bold")
        + "  " + _span(_origin_label(section), size=7, family=MONO, color=_META),
        _rule(strong=depth == 0),
    ]
    blocks = [_Block(heading, keep_with_next=True)]
    kind = section["kind"]
    if kind == "columns":
        for column in section["items"]:
            for item in column:
                blocks.extend(_section_blocks(item, depth + 1))
        return blocks
    rows = _section_rows(section)
    model_owned = (section.get("origin") or {}).get("kind") == "MODEL"
    if kind in {"table", "chart"}:
        if kind == "chart":
            blocks.append(_Block([_span(
                f"Chart exhibit · {section.get('recipe', {}).get('chart_kind', 'chart')} · authoritative data table",
                size=8, color=_META)]))
        if rows[1:]:
            blocks.extend(_mono_table(rows, model_owned=model_owned))
        else:
            blocks.extend(_mono_table(rows, model_owned=model_owned))
            blocks.append(_Block([_span("No rows.", size=8, color=_META)]))
    elif kind == "profile":
        width = min(max(len(label) for label, _ in rows) + 1, 28)
        for label, value in rows:
            wrapped = textwrap.wrap(str(value), width=MONO_COLUMNS - width - 2, break_long_words=True, break_on_hyphens=False) or [""]
            lines = [
                _span(str(label).ljust(width), size=8, family=MONO, color=_META) + "  " + _span(wrapped[0], size=8, family=MONO)
            ] + [
                _span(" " * width, size=8, family=MONO) + "  " + _span(part, size=8, family=MONO)
                for part in wrapped[1:]
            ]
            blocks.append(_Block(lines))
    elif kind == "list":
        for row in rows:
            blocks.append(_Block(_prose_lines(f"• {row[0]}", indent="  ")))
    else:
        for paragraph in str(section["body"]).split("\n"):
            blocks.append(_Block(_prose_lines(paragraph)))
    if section.get("note"):
        blocks.append(_Block([_span(section["note"], size=7.5, color=_META)]))
    blocks.append(_Block([_span(" ", size=6)]))
    return blocks


def _pango_view() -> str:
    executable = shutil.which("pango-view")
    if executable is None:
        for candidate in (Path("/opt/homebrew/bin/pango-view"), Path("/usr/local/bin/pango-view")):
            if candidate.is_file():
                executable = str(candidate)
                break
    if executable is None:
        raise ValueError("PDF_UNICODE_RENDERER_UNAVAILABLE")
    return executable


def _shape(executable: str, workspace: Path, name: str, markup: str, *, height: int | None,
           extra: tuple[str, ...] = ()) -> Path:
    return _pango(
        executable, workspace, name, markup, "--markup", f"--font={SANS} 9.5", f"--margin={MARGIN}",
        f"--width={BODY_WIDTH}", "--wrap=word-char", "--background=transparent", f"--foreground={_INK}",
        *extra, *([f"--height={height}"] if height is not None else []),
    )


def _meta_lines(text: str, size: float) -> list[str]:
    """Masthead metadata wrapped here, at spaces only: pango's word-char
    wrapping may break a build id at its hyphen, and an identifier split
    across two lines is neither searchable nor recognisable. DejaVu Sans Mono
    advances 0.602 em; 0.61 leaves the margin that keeps the line inside."""
    columns = int(BODY_WIDTH / (0.61 * size))
    return [_span(part, size=size, family=MONO, color=_META)
            for part in textwrap.wrap(text, width=columns, break_on_hyphens=False) or [" "]]


def _page_markup(masthead: dict[str, Any], page_name: str, blocks: list[_Block], page_number: int, page_count: int,
                 *, first: bool) -> str:
    lines = [
        _span(f"{masthead.get('issuer', '')} — {masthead.get('report_type', '')}", size=18 if first else 11, weight="bold"),
        *_meta_lines(_masthead_line(masthead, page_name), size=7),
        _rule(strong=True),
    ]
    if first:
        lines.extend(_meta_lines(
            f"Opinion owner {masthead.get('opinion_owner', '')} / signed {masthead.get('opinion_signed_at', '')} / "
            f"model {masthead.get('model_identity', '')} / methodology {masthead.get('methodology_build_id', '')}",
            size=7.5))
        lines.extend(_meta_lines(f"Machine assistance: {masthead.get('machine_assistance', '')}", size=7.5))
    lines.append(_span(" ", size=6))
    for block in blocks:
        lines.extend(block.lines)
    return "\n".join(lines)


def _footer_markup(masthead: dict[str, Any], page_number: int, page_count: int) -> str:
    footer_left = f"CAOS / {_short_id(masthead.get('deliverable_id', ''))} / {masthead.get('approval_state', PENDING_APPROVAL)} / content digest in Revision Record"
    footer_right = f"PAGE {page_number} OF {page_count}"
    gap = max(1, MONO_COLUMNS - len(footer_left) - len(footer_right))
    return _rule(strong=True) + "\n" + _span(footer_left + " " * gap + footer_right, size=7, family=MONO, color=_META)


def _measure(executable: str, workspace: Path, markup: str) -> float:
    from pypdf import PdfReader

    rendered = _shape(executable, workspace, "measure", markup, height=None)
    page = PdfReader(rendered).pages[0]
    return float(page.mediabox.height) - 2 * MARGIN


def _footer_page(executable: str, workspace: Path, markup: str):
    from pypdf import PdfReader

    rendered = _pango(executable, workspace, "footer", markup, "--markup", "--margin=0", f"--width={BODY_WIDTH}",
                      "--background=transparent", f"--foreground={_INK}")
    return PdfReader(rendered).pages[0]


_SIZE_ATTRIBUTE = re.compile(r'size="(\d+)"')
_TAG = re.compile(r"<[^>]+>")


def _estimate(lines: list[str]) -> float:
    """Cheap height bound for a list of markup lines: font size × 1.4 per
    wrapped visual line, wrapping prose at ~95 characters of 9.5pt sans. It
    only decides when to pay for a real measurement, never the layout."""
    total = 0.0
    for line in lines:
        sizes = [int(match) / 1024 for match in _SIZE_ATTRIBUTE.findall(line)]
        size = max(sizes) if sizes else 9.5
        characters = len(_TAG.sub("", line))
        per_line = 100 if MONO in line else max(40, int(95 * 9.5 / size))
        total += max(1, -(-characters // per_line)) * size * 1.45
    return total


def _paginate(executable: str, workspace: Path, masthead: dict[str, Any], pages: list[dict[str, Any]]) -> list[tuple[str, list[_Block]]]:
    """Flow every logical page's blocks into physical pages by measurement.

    Logical pages (Decision, Financials, Control, …) open with a band block
    and flow on; a physical page breaks only when the measured content no
    longer fits, so short logical pages never leave most of a sheet blank.
    Blocks are packed by a cheap estimate, then measured, the tail moved
    forward until it fits, then topped up while more still measures in. A
    table row that moves forward drags a copy of its header; a heading never
    ends a page.
    """
    stream: list[tuple[str, _Block]] = []
    for page in pages:
        band = _Block([
            _span(page["name"].upper(), size=8, weight="bold", family=MONO, color=_META),
            _span(" ", size=4),
        ], keep_with_next=True)
        stream.append((page["name"], band))
        for section in page["sections"]:
            for block in _section_blocks(section, 0):
                stream.append((page["name"], block))
    laid_out: list[tuple[str, list[_Block]]] = []
    chrome = 6 * 9.5 * 1.45 + 2 * 18 * 1.45  # masthead and rules
    pending = list(stream)
    while pending:
        first = not laid_out
        page_name = pending[0][0]
        current: list[tuple[str, _Block]] = []
        estimate = chrome
        while pending and estimate + _estimate(pending[0][1].lines) <= BODY_HEIGHT:
            item = pending.pop(0)
            current.append(item)
            estimate += _estimate(item[1].lines)
        if not current:
            current.append(pending.pop(0))

        def blocks_of(items: list[tuple[str, _Block]]) -> list[_Block]:
            return [block for _name, block in items]

        def fit_prefix(placed: list[tuple[str, _Block]], name: str, block: _Block, *, floor: int) -> int:
            """Largest number of `block`'s leading lines that still measure in
            after `placed` (bisection; `floor` lines are accepted unmeasured)."""
            low, high = floor, max(floor, len(block.lines) - 1)
            while low < high:
                middle = (low + high + 1) // 2
                candidate = blocks_of([*placed, (name, _Block(block.lines[:middle]))])
                if _measure(executable, workspace, _page_markup(masthead, page_name, candidate, 99, 99, first=first)) <= BODY_HEIGHT:
                    low = middle
                else:
                    high = middle - 1
            return low

        while True:
            markup = _page_markup(masthead, page_name, blocks_of(current), 99, 99, first=first)
            if _measure(executable, workspace, markup) <= BODY_HEIGHT or (len(current) == 1 and len(current[0][1].lines) == 1):
                break
            carried = [current.pop()]
            while current and current[-1][1].keep_with_next:
                carried.insert(0, current.pop())
            if not current:
                # Everything carried is a keep-with-next chain ending in the
                # block that overflowed. The chain stays; the overflowing block
                # is split at the largest prefix that measures in (bisection),
                # so a long table or paragraph fills the page instead of
                # leaving it heading-only, blank, or overprinted.
                name, oversized = carried.pop()
                current.extend(carried)
                carried = []
                low = fit_prefix(current, name, oversized, floor=1)
                current.append((name, _Block(oversized.lines[:low])))
                if not oversized.lines[low:]:
                    break  # indivisible: one line taller than the body; accept it rather than loop
                carried.append((name, _Block(oversized.lines[low:], keep_with_next=oversized.keep_with_next,
                                             repeat_header=oversized.repeat_header)))
            if carried and carried[0][1].repeat_header:
                carried.insert(0, (carried[0][0], _Block(list(carried[0][1].repeat_header), keep_with_next=True)))
            pending = carried + pending
        # Top up in estimated batches (one measurement per batch, halving on
        # overflow) instead of one measurement per block: a 400-row table costs
        # a handful of pango calls rather than hundreds.
        measured = _measure(executable, workspace, _page_markup(masthead, page_name, blocks_of(current), 99, 99, first=first))
        while pending:
            room = BODY_HEIGHT - measured
            batch: list[tuple[str, _Block]] = []
            spent = 0.0
            for item in pending:
                cost = _estimate(item[1].lines)
                if batch and spent + cost > room:
                    break
                batch.append(item)
                spent += cost
            while batch:
                markup = _page_markup(masthead, page_name, blocks_of([*current, *batch]), 99, 99, first=first)
                height = _measure(executable, workspace, markup)
                if height <= BODY_HEIGHT:
                    current.extend(batch)
                    del pending[:len(batch)]
                    measured = height
                    break
                batch = batch[: len(batch) // 2]
            if not batch:
                # Nothing more fits whole. Split the next block into the room
                # that is left — at least two lines, so a heading is never left
                # alone at a page foot and a paragraph never opens with a
                # single orphaned line on the page after it.
                name, block = pending[0]
                if len(block.lines) >= 4:
                    low = fit_prefix(current, name, block, floor=0)
                    if 2 <= low <= len(block.lines) - 2:
                        current.append((name, _Block(block.lines[:low])))
                        pending[0] = (name, _Block(block.lines[low:], keep_with_next=block.keep_with_next,
                                                   repeat_header=block.repeat_header))
                break
        carried = []
        while len(current) > 1 and current[-1][1].keep_with_next and pending:
            carried.insert(0, current.pop())
        pending = carried + pending
        # A table continuing on the next page repeats its header row there.
        if pending and pending[0][1].repeat_header and not (current and current[-1][1].keep_with_next):
            pending.insert(0, (pending[0][0], _Block(list(pending[0][1].repeat_header), keep_with_next=True)))
        laid_out.append((page_name, blocks_of(current)))
    if not laid_out:
        laid_out.append((pages[0]["name"] if pages else "Document", []))
    return laid_out


def _watermark_page(executable: str, workspace: Path, text: str):
    from pypdf import PdfReader

    markup = f'<span font_family="{SANS}" size="{46 * 1024}" weight="bold" foreground="#be5410" alpha="14%" {_FEATURES}>{_esc(text)}</span>'
    rendered = _pango(executable, workspace, "watermark", markup, "--markup", "--margin=0", "--rotate=-32",
                      "--background=transparent")
    return PdfReader(rendered).pages[0]


def _white_page(executable: str, workspace: Path):
    """An opaque white letter page: viewers that composite transparency onto
    black (thumbnailers, some print pipelines) must still show paper."""
    from pypdf import PdfReader

    rendered = _pango(executable, workspace, "paper", " ", "--margin=0", f"--width={PAGE_WIDTH}",
                      f"--height={PAGE_HEIGHT}", "--background=#ffffff")
    return PdfReader(rendered).pages[0]


def render_frozen_pdf(payload: dict[str, Any]) -> bytes:
    from pypdf import PageObject, PdfReader, PdfWriter, Transformation

    view = publication_view(payload)
    masthead = view["masthead"]
    executable = _pango_view()
    revision_page = {
        "name": "Revision Record",
        "sections": [{
            "kind": "profile", "section_id": "revision_record", "title": "Revision Record", "page": "Revision Record",
            "editable": False, "origin": {"kind": "SYSTEM", "authority_id": str(masthead.get("deliverable_id", "")), "block_ids": []},
            "rows": [{"label": label, "value": value or "Unavailable"} for label, value in view["revision"]],
        }],
    }
    writer = PdfWriter()
    with tempfile.TemporaryDirectory(prefix="caos-pdf-") as directory:
        workspace = Path(directory)
        laid_out = _paginate(executable, workspace, masthead, [*view["pages"], revision_page])
        watermark = _watermark_page(executable, workspace, str(masthead.get("watermark") or PENDING_APPROVAL))
        paper = _white_page(executable, workspace)
        offset_x = (PAGE_WIDTH - float(watermark.mediabox.width)) / 2
        offset_y = (PAGE_HEIGHT - float(watermark.mediabox.height)) / 2
        for index, (page_name, blocks) in enumerate(laid_out, start=1):
            markup = _page_markup(masthead, page_name, blocks, index, len(laid_out), first=index == 1)
            rendered = _shape(executable, workspace, f"page-{index:04d}", markup, height=BODY_HEIGHT)
            content = PdfReader(rendered).pages[0]
            footer = _footer_page(executable, workspace, _footer_markup(masthead, index, len(laid_out)))
            base = PageObject.create_blank_page(width=PAGE_WIDTH, height=PAGE_HEIGHT)
            base.merge_page(paper)
            base.merge_transformed_page(watermark, Transformation().translate(offset_x, offset_y))
            # pango-view sizes its page to body + margins; anchor it to the top
            # of the letter page so the top margin is exact and the footer band
            # below the body is never overprinted.
            base.merge_transformed_page(content, Transformation().translate(0, PAGE_HEIGHT - float(content.mediabox.height)))
            base.merge_transformed_page(footer, Transformation().translate(MARGIN, MARGIN - 4))
            writer.add_page(base)
        output = io.BytesIO()
        writer.write(output)
    return output.getvalue()


# --- XLSX ----------------------------------------------------------------------------


def _safe_cell(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, Decimal)):
        return value
    text = str(value)
    if text.startswith(FORMULA_PREFIXES):
        return "'" + text
    return text


def _model_cell(value: str, *, model_value: bool) -> Any:
    if model_value and _PLAIN_NUMBER.fullmatch(value):
        number = float(value)
        if math.isfinite(number):
            return number if any(character in value for character in ".eE") else int(value)
    return _safe_cell(value)


MODEL_NUMBER_FORMAT = "#,##0.00;[Red](#,##0.00);0.00"


def _sheet_title(used: set[str], title: str) -> str:
    """Excel caps sheet names at 31 characters; cut at a word boundary only
    when a title is longer, and suffix a counter only on collision."""
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", title).strip() or "Sheet"
    if len(cleaned) > 31:
        cut = cleaned[:31]
        cleaned = cut[: cut.rfind(" ")] if " " in cut[8:] else cut
    candidate = cleaned
    counter = 2
    while candidate in used:
        suffix = f" {counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(candidate)
    return candidate


def render_frozen_xlsx(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    view = publication_view(payload)
    masthead = view["masthead"]
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="E9E7DF")
    wrap = Alignment(wrap_text=True, vertical="top")
    workbook = Workbook()
    used_titles: set[str] = set()

    def finish(sheet: Any, *, header_rows: int = 1, widths: dict[int, int] | None = None, filters: bool = False) -> None:
        sheet.freeze_panes = f"A{header_rows + 1}"
        for row in sheet.iter_rows(min_row=1, max_row=header_rows):
            for cell in row:
                cell.font = bold
                cell.fill = head_fill
        lengths: dict[int, int] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                lengths[cell.column] = max(lengths.get(cell.column, 0), min(len(str(cell.value)), 60))
                if isinstance(cell.value, str) and len(cell.value) > 40:
                    cell.alignment = wrap
        for column, length in lengths.items():
            sheet.column_dimensions[get_column_letter(column)].width = (widths or {}).get(column, max(10, length + 2))
        if filters and sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = f"A{header_rows}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        sheet.sheet_view.showGridLines = False

    cover = workbook.active
    cover.title = _sheet_title(used_titles, "Cover & Control")
    cover.append(["Field", "Value"])
    cover.append(["Document", f"{masthead.get('issuer', '')} — {masthead.get('report_type', '')}"])
    for label, key in MASTHEAD_FIELDS:
        cover.append([label, _safe_cell(str(masthead.get(key, "")))])
    cover.append(["Watermark", _safe_cell(str(masthead.get("watermark", PENDING_APPROVAL)))])
    cover.append(["Approver identity", "Recorded in the detached filing receipt and the audit chain, never in this workbook"])
    for label, value in view["revision"]:
        cover.append([label, _safe_cell(value)])
    disclosures = view["disclosures"]
    for key in ("content_origin", "sources", "machine_assistance", "analyst_opinion_owner", "approval_state", "as_of_date", "version"):
        if key in disclosures:
            cover.append([f"Disclosure · {key.replace('_', ' ')}", _safe_cell(str(disclosures[key]))])
    for item in disclosures.get("limitations") or []:
        cover.append(["Disclosure · limitation", _safe_cell(str(item))])
    cover.append(["Sheets", "Report (narrative and profiles), one sheet per table, Revision Record"])
    finish(cover, widths={1: 28, 2: 100})

    report = workbook.create_sheet(_sheet_title(used_titles, "Report"))
    report.append(["Page", "Section", "Origin", "Authority", "Field", "Content"])
    tables: list[tuple[str, dict[str, Any], str]] = []
    for page in view["pages"]:
        for section, _depth in _walk_sections(page["sections"]):
            origin = section.get("origin") or {}
            base = [page["name"], section["title"], _origin_label(section).split(" · ")[0], origin.get("authority_id", "")]
            kind = section["kind"]
            if kind in {"table", "chart"}:
                tables.append((page["name"], section, _sheet_title(used_titles, section["title"])))
                report.append([*base, "Table", f"See sheet '{tables[-1][2]}' ({len(_section_rows(section)) - 1} rows)"])
            elif kind == "profile":
                for label, value in _section_rows(section):
                    report.append([*base, _safe_cell(label), _safe_cell(value)])
            elif kind == "list":
                for row in _section_rows(section):
                    report.append([*base, "Item", _safe_cell(row[0])])
            elif kind == "text":
                report.append([*base, "Text", _safe_cell(section["body"])])
            elif kind == "columns":
                report.append([*base, "Group", f"{sum(len(column) for column in section['items'])} sections follow"])
            if section.get("note"):
                report.append([*base, "Note", _safe_cell(section["note"])])
    finish(report, widths={1: 14, 2: 34, 3: 18, 4: 34, 5: 22, 6: 110}, filters=True)

    for page_name, section, title in tables:
        sheet = workbook.create_sheet(title)
        rows = _section_rows(section)
        model_owned = (section.get("origin") or {}).get("kind") == "MODEL"
        sheet.append([_safe_cell(value) for value in rows[0]])
        for row in rows[1:]:
            values = [
                _model_cell(str(value), model_value=model_owned and index > 0)
                for index, value in enumerate(row)
            ]
            sheet.append(values)
            for index, value in enumerate(values, start=1):
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    sheet.cell(row=sheet.max_row, column=index).number_format = MODEL_NUMBER_FORMAT
        note = [f"{page_name} · {section['title']} · {_origin_label(section)}"]
        if section.get("note"):
            note.append(section["note"])
        sheet.append([])
        sheet.append([_safe_cell(" · ".join(note))])
        finish(sheet, filters=True)

    record = workbook.create_sheet(_sheet_title(used_titles, "Revision Record"))
    record.append(["Field", "Value"])
    for label, value in view["revision"]:
        record.append([label, _safe_cell(value)])
    finish(record, widths={1: 22, 2: 80})

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise ValueError("DELIVERABLE_EXPORT_UNSAFE: a formula reached the workbook")

    from datetime import datetime

    workbook.properties.created = datetime(2026, 1, 1)
    workbook.properties.modified = datetime(2026, 1, 1)
    workbook.properties.title = f"{masthead.get('issuer', '')} — {masthead.get('report_type', '')}"
    workbook.properties.subject = str(payload.get("preview_digest", ""))
    workbook.properties.creator = "CAOS"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return _deterministic_zip(output.getvalue())


def _deterministic_zip(content: bytes) -> bytes:
    """Re-pack an xlsx with sorted entries, fixed zip metadata, and pinned
    document timestamps — openpyxl stamps entry mtimes AND rewrites
    docProps/core.xml created/modified from the wall clock at save time."""
    import zipfile

    source = zipfile.ZipFile(io.BytesIO(content))
    output = io.BytesIO()
    try:
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
            for name in sorted(source.namelist()):
                data = source.read(name)
                if name == "docProps/core.xml":
                    data = re.sub(
                        rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                        rb"\g<1>2026-01-01T00:00:00Z\g<2>",
                        data,
                    )
                info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                target.writestr(info, data)
    finally:
        source.close()
    return output.getvalue()


def render_frozen_export(payload: dict[str, Any], format_name: str) -> bytes:
    if format_name == "md":
        return render_frozen_markdown(payload)
    if format_name == "pdf":
        return render_frozen_pdf(payload)
    if format_name == "xlsx":
        return render_frozen_xlsx(payload)
    raise ValueError(f"DELIVERABLE_EXPORT_UNAVAILABLE: unknown format {format_name!r}")
