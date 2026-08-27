"""CP-3 loan-universe workbook parsing and import (DECISIONS §5: the parser
ports from LEGACY artifacts/loan_universe.py; the import lifecycle is rewritten
against the framework store). Deterministic normalization with per-row
provenance; every validation fails closed with structured findings."""

from __future__ import annotations

import io
import hashlib
import math
import zipfile
import copy  # noqa: F401 — kept from the ported module surface
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

from ..contracts import digest


TEMPLATE_VERSION = "cp3-sector-rv-v1"
IMPORTER_VERSION = "loan-rv-importer-v1"
MAX_WORKSHEETS = 64
# ponytail: the MVP returns and binds the full universe; paginate the API and
# CP-3 handoff before raising this ceiling.
MAX_ROWS = 2_000
MAX_COLUMNS = 64
MAX_CELL_TEXT = 32 * 1024
MAX_FINDINGS = 200
HEADER_SCAN_ROWS = 200

HEADERS = (
    "Company",
    "Borrower Name",
    "Core Business Description",
    "Sub-Sector",
    "Sub-Group",
    "Public/Private",
    "Bloomberg",
    "FIGI",
    "Loan Type",
    "Ranking",
    "Ratings",
    "Size ($Mn)",
    "Margin",
    "Maturity",
    "Bid",
    "Ask",
    "Δ 1D",
    "Δ 1W",
    "Δ 1M",
    "Δ 3M",
    "Δ 6M",
    "Δ 1YR",
    "Δ YTD",
    "Mid YTM",
    "Mid 3Y DM",
)

FIELDS = (
    "company",
    "borrower_name",
    "business_description",
    "sub_sector",
    "sub_group",
    "public_private",
    "bloomberg_loan_id",
    "figi",
    "loan_type",
    "ranking",
    "ratings",
    "size_mn",
    "margin_bps",
    "maturity_date",
    "bid_points",
    "ask_points",
    "change_1d_points",
    "change_1w_points",
    "change_1m_points",
    "change_3m_points",
    "change_6m_points",
    "change_1yr_points",
    "change_ytd_points",
    "mid_ytm_pct",
    "mid_3y_dm_bps",
)

NUMERIC_FIELDS = {
    "size_mn",
    "margin_bps",
    "bid_points",
    "ask_points",
    "change_1d_points",
    "change_1w_points",
    "change_1m_points",
    "change_3m_points",
    "change_6m_points",
    "change_1yr_points",
    "change_ytd_points",
    "mid_ytm_pct",
    "mid_3y_dm_bps",
}

MISSING_TEXT = {"", "#N/A", "#N/A N/A", "N/A", "NA", "-", "—"}
FORBIDDEN_PACKAGE_PATHS = (
    "xl/vbaproject.bin",
    "xl/externallinks/",
    "xl/embeddings/",
    "xl/activex/",
    "xl/oleobjects/",
)


class LoanWorkbookValidationError(ValueError):
    def __init__(self, findings: list[dict[str, Any]]) -> None:
        super().__init__("RV_WORKBOOK_INVALID")
        self.findings = findings


class LoanUniverseImportRejected(ValueError):
    def __init__(self, record: dict[str, Any]) -> None:
        super().__init__("RV_WORKBOOK_INVALID")
        self.record = record


class LoanUniverseSourceError(ValueError):
    def __init__(self, code: str, detail: str) -> None:
        super().__init__(code)
        self.code = code
        self.detail = detail


class _Findings:
    def __init__(self) -> None:
        self.items: list[dict[str, Any]] = []

    def add(
        self,
        code: str,
        detail: str,
        *,
        sheet: str | None = None,
        row: int | None = None,
        column: str | None = None,
    ) -> None:
        if len(self.items) >= MAX_FINDINGS:
            return
        self.items.append(
            {
                "code": code[:80],
                "detail": detail[:300],
                "sheet": sheet,
                "row": row,
                "column": column,
            }
        )


def _validate_package(content: bytes, findings: _Findings) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            names = [name.lower() for name in archive.namelist()]
            if "[content_types].xml" not in names or "xl/workbook.xml" not in names:
                findings.add(
                    "RV_PACKAGE_INVALID",
                    "XLSX package is missing required workbook parts.",
                )
                return
            for name in names:
                if any(
                    name == prefix or name.startswith(prefix)
                    for prefix in FORBIDDEN_PACKAGE_PATHS
                ):
                    findings.add(
                        "RV_PACKAGE_ACTIVE_CONTENT",
                        f"Unsupported active package content: {name}",
                    )
            for info in archive.infolist():
                if not info.filename.lower().endswith(".rels"):
                    continue
                relationships = ElementTree.fromstring(archive.read(info))
                if any(
                    relationship.attrib.get("TargetMode", "").casefold() == "external"
                    for relationship in relationships.iter()
                ):
                    findings.add(
                        "RV_PACKAGE_EXTERNAL_LINK",
                        "External package relationships are not allowed.",
                    )
                    break
            content_types = archive.read("[Content_Types].xml")
            if b"macroEnabled" in content_types or b"vbaProject" in content_types:
                findings.add(
                    "RV_PACKAGE_MACRO", "Macro-enabled workbooks are not allowed."
                )
    except (
        ElementTree.ParseError,
        KeyError,
        OSError,
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
    ):
        findings.add("RV_PACKAGE_INVALID", "XLSX package is malformed.")


def screen_package(content: bytes) -> None:
    """Package screening runs before any workbook parsing: active content,
    external relationships, and macro parts reject the bytes outright."""
    findings = _Findings()
    _validate_package(content, findings)
    if findings.items:
        raise LoanWorkbookValidationError(findings.items)


def _header_text(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""


def _text(
    value: Any, findings: _Findings, *, sheet: str, row: int, column: str
) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        normalized = value.strip()
    elif isinstance(value, bool):
        normalized = "TRUE" if value else "FALSE"
    elif isinstance(value, (int, float)):
        normalized = (
            str(int(value))
            if isinstance(value, int) or value.is_integer()
            else str(value)
        )
    else:
        normalized = str(value).strip()
    if len(normalized) > MAX_CELL_TEXT:
        findings.add(
            "RV_CELL_TEXT_LIMIT",
            "Cell text exceeds the 32 KB limit.",
            sheet=sheet,
            row=row,
            column=column,
        )
        return None
    return None if normalized.upper() in MISSING_TEXT else normalized


def _number(
    value: Any, findings: _Findings, *, sheet: str, row: int, column: str
) -> float | None:
    if value is None or (
        isinstance(value, str) and value.strip().upper() in MISSING_TEXT
    ):
        return None
    if isinstance(value, bool):
        findings.add(
            "RV_NUMBER_INVALID",
            "Boolean value is not a market number.",
            sheet=sheet,
            row=row,
            column=column,
        )
        return None
    try:
        normalized = (
            float(value.replace(",", "").strip())
            if isinstance(value, str)
            else float(value)
        )
    except (TypeError, ValueError, OverflowError):
        findings.add(
            "RV_NUMBER_INVALID",
            "Market value is not numeric.",
            sheet=sheet,
            row=row,
            column=column,
        )
        return None
    if not math.isfinite(normalized):
        findings.add(
            "RV_NUMBER_NON_FINITE",
            "Market value must be finite.",
            sheet=sheet,
            row=row,
            column=column,
        )
        return None
    return normalized


def _date_value(
    value: Any,
    findings: _Findings,
    *,
    sheet: str,
    row: int,
    column: str,
    epoch: datetime,
    code: str,
) -> str | None:
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            excel_date = from_excel(value, epoch)
            if isinstance(excel_date, datetime):
                parsed = excel_date.date()
            elif isinstance(excel_date, date):
                parsed = excel_date
        except (TypeError, ValueError, OverflowError):
            parsed = None
    elif isinstance(value, str) and value.strip().upper() not in MISSING_TEXT:
        candidate = value.strip()
        for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%b-%y", "%d-%b-%Y"):
            try:
                parsed = datetime.strptime(candidate, pattern).date()
                break
            except ValueError:
                continue
    if parsed is None:
        if value is not None and not (
            isinstance(value, str) and value.strip().upper() in MISSING_TEXT
        ):
            findings.add(
                code, "Date value is invalid.", sheet=sheet, row=row, column=column
            )
        return None
    return parsed.isoformat()


def _find_headers(sheet: Any) -> tuple[list[tuple[int, int]], bool]:
    matches: list[tuple[int, int]] = []
    partial = False
    max_row = min(sheet.max_row or 0, HEADER_SCAN_ROWS)
    max_column = min(sheet.max_column or 0, MAX_COLUMNS)
    for cells in sheet.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_column
    ):
        values = [_header_text(cell.value) for cell in cells]
        canonical = set(values).intersection(HEADERS)
        partial = partial or ("Borrower Name" in values and len(canonical) >= 5)
        for start in range(0, max(0, len(values) - len(HEADERS) + 1)):
            if tuple(values[start : start + len(HEADERS)]) == HEADERS:
                matches.append((cells[0].row, start + 1))
    return matches, partial


def _workbook_date(
    sheet: Any, header_row: int, findings: _Findings, epoch: datetime
) -> str | None:
    for cells in sheet.iter_rows(
        min_row=1,
        max_row=min(header_row, 20),
        min_col=1,
        max_col=min(sheet.max_column or 1, 12),
    ):
        for cell in cells:
            if _header_text(cell.value) != "Date":
                continue
            candidates = ((cell.row + 1, cell.column), (cell.row, cell.column + 1))
            for row, column in candidates:
                if row > sheet.max_row or column > sheet.max_column:
                    continue
                value = sheet.cell(row=row, column=column).value
                if value is None:
                    continue
                parsed = _date_value(
                    value,
                    findings,
                    sheet=sheet.title,
                    row=row,
                    column=sheet.cell(row=row, column=column).column_letter,
                    epoch=epoch,
                    code="RV_WORKBOOK_DATE_INVALID",
                )
                if parsed:
                    return parsed
    findings.add(
        "RV_WORKBOOK_DATE_MISSING",
        "Worksheet is missing the fixed Date field.",
        sheet=sheet.title,
    )
    return None


def _normalize_row(
    sheet: Any,
    row_number: int,
    cells: tuple[Any, ...],
    findings: _Findings,
    epoch: datetime,
    start_column: int,
) -> dict[str, Any]:
    row: dict[str, Any] = {
        "sector": sheet.title,
        "source_locators": [{"sheet": sheet.title, "row": row_number}],
    }
    for offset, field in enumerate(FIELDS):
        cell = cells[offset]
        column = get_column_letter(start_column + offset)
        location = {"sheet": sheet.title, "row": row_number, "column": column}
        if field in NUMERIC_FIELDS:
            row[field] = _number(cell.value, findings, **location)
        elif field == "maturity_date":
            row[field] = _date_value(
                cell.value,
                findings,
                epoch=epoch,
                code="RV_MATURITY_INVALID",
                **location,
            )
        else:
            row[field] = _text(cell.value, findings, **location)
    for identifier in ("figi", "bloomberg_loan_id"):
        if row[identifier]:
            row[identifier] = row[identifier].upper()
    if not row["borrower_name"]:
        findings.add(
            "RV_BORROWER_MISSING",
            "Borrower Name is required.",
            sheet=sheet.title,
            row=row_number,
            column=get_column_letter(start_column + 1),
        )
    if not row["figi"] and not row["bloomberg_loan_id"]:
        findings.add(
            "RV_INSTRUMENT_ID_MISSING",
            "FIGI or Bloomberg loan ID is required.",
            sheet=sheet.title,
            row=row_number,
        )
        row["instrument_key"] = f"INVALID:{sheet.title}:{row_number}"
    else:
        row["instrument_key"] = (
            f"FIGI:{row['figi']}" if row["figi"] else f"BBG:{row['bloomberg_loan_id']}"
        )
    return row


def parse_loan_workbook(
    content: bytes, *, source_id: str, source_sha256: str
) -> dict[str, Any]:
    findings = _Findings()
    _validate_package(content, findings)
    if findings.items:
        raise LoanWorkbookValidationError(findings.items)
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise LoanWorkbookValidationError(
            [
                {
                    "code": "RV_PACKAGE_INVALID",
                    "detail": "XLSX workbook cannot be opened.",
                    "sheet": None,
                    "row": None,
                    "column": None,
                }
            ]
        ) from exc
    try:
        sheets = workbook.worksheets
        if len(sheets) > MAX_WORKSHEETS:
            findings.add(
                "RV_WORKSHEET_LIMIT",
                f"Workbook exceeds the {MAX_WORKSHEETS}-worksheet limit.",
            )
        rows: list[dict[str, Any]] = []
        raw_row_count = 0
        workbook_dates: list[tuple[str, str]] = []
        recognized_sheets = 0
        for sheet in sheets[:MAX_WORKSHEETS]:
            if sheet.sheet_state != "visible":
                continue
            if (sheet.max_column or 0) > MAX_COLUMNS:
                findings.add(
                    "RV_COLUMN_LIMIT",
                    f"Worksheet exceeds the {MAX_COLUMNS}-column limit.",
                    sheet=sheet.title,
                )
                continue
            matches, partial = _find_headers(sheet)
            if len(matches) > 1:
                findings.add(
                    "RV_MULTIPLE_TABLES",
                    "Worksheet contains more than one issuer table.",
                    sheet=sheet.title,
                )
                continue
            if not matches:
                if partial:
                    findings.add(
                        "RV_TEMPLATE_PARTIAL",
                        "Worksheet contains a partial or changed issuer-table header.",
                        sheet=sheet.title,
                    )
                continue
            recognized_sheets += 1
            header_row, start_column = matches[0]
            sheet_date = _workbook_date(sheet, header_row, findings, workbook.epoch)
            if sheet_date:
                workbook_dates.append((sheet.title, sheet_date))
            seen_data = False
            instrument_rows = sheet.iter_rows(
                min_row=header_row + 1,
                max_row=sheet.max_row or header_row,
                min_col=start_column,
                max_col=start_column + len(HEADERS) - 1,
            )
            for row_number, cells in enumerate(instrument_rows, start=header_row + 1):
                if all(
                    cell.value is None
                    or (isinstance(cell.value, str) and not cell.value.strip())
                    for cell in cells
                ):
                    if seen_data:
                        break
                    continue
                seen_data = True
                raw_row_count += 1
                if raw_row_count > MAX_ROWS:
                    findings.add(
                        "RV_ROW_LIMIT",
                        f"Workbook exceeds the {MAX_ROWS}-instrument-row limit.",
                        sheet=sheet.title,
                        row=row_number,
                    )
                    break
                rows.append(
                    _normalize_row(
                        sheet, row_number, cells, findings, workbook.epoch, start_column
                    )
                )
        if recognized_sheets == 0:
            findings.add(
                "RV_TEMPLATE_MISSING",
                "Workbook contains no recognized visible sector worksheet.",
            )
        elif raw_row_count == 0:
            findings.add(
                "RV_ROWS_MISSING",
                "Recognized sector worksheets contain no instrument rows.",
            )
        if workbook_dates:
            expected_date = workbook_dates[0][1]
            for sheet_name, sheet_date in workbook_dates[1:]:
                if sheet_date != expected_date:
                    findings.add(
                        "RV_WORKBOOK_DATE_CONFLICT",
                        f"Worksheet date {sheet_date} does not match {expected_date}.",
                        sheet=sheet_name,
                    )
        else:
            expected_date = None

        normalized: dict[str, dict[str, Any]] = {}
        figi_to_bloomberg: dict[str, str] = {}
        bloomberg_to_figi: dict[str, str] = {}
        for row in rows:
            figi = row.get("figi")
            bloomberg = row.get("bloomberg_loan_id")
            if figi and bloomberg:
                if figi in figi_to_bloomberg and figi_to_bloomberg[figi] != bloomberg:
                    findings.add(
                        "RV_ID_CONFLICT",
                        "FIGI maps to multiple Bloomberg loan IDs.",
                        **row["source_locators"][0],
                    )
                if (
                    bloomberg in bloomberg_to_figi
                    and bloomberg_to_figi[bloomberg] != figi
                ):
                    findings.add(
                        "RV_ID_CONFLICT",
                        "Bloomberg loan ID maps to multiple FIGIs.",
                        **row["source_locators"][0],
                    )
                figi_to_bloomberg[figi] = bloomberg
                bloomberg_to_figi[bloomberg] = figi

        identities: dict[str, tuple[str | None, str | None]] = {}
        for row in rows:
            figi = row.get("figi")
            bloomberg = row.get("bloomberg_loan_id")
            if figi and not bloomberg:
                row["bloomberg_loan_id"] = bloomberg = figi_to_bloomberg.get(figi)
            elif bloomberg and not figi:
                row["figi"] = figi = bloomberg_to_figi.get(bloomberg)
            if figi or bloomberg:
                row["instrument_key"] = f"FIGI:{figi}" if figi else f"BBG:{bloomberg}"
            key = row["instrument_key"]
            identity = (figi, bloomberg)
            if key in identities and identities[key] != identity:
                findings.add(
                    "RV_ID_CONFLICT",
                    "Instrument key has conflicting identifiers.",
                    **row["source_locators"][0],
                )
            identities[key] = identity
            existing = normalized.get(key)
            if existing is None:
                normalized[key] = row
                continue
            comparable = {
                name: value for name, value in row.items() if name != "source_locators"
            }
            existing_comparable = {
                name: value
                for name, value in existing.items()
                if name != "source_locators"
            }
            if comparable != existing_comparable:
                findings.add(
                    "RV_DUPLICATE_CONFLICT",
                    "Duplicate instrument rows contain different values.",
                    **row["source_locators"][0],
                )
            else:
                existing["source_locators"].extend(row["source_locators"])

        if findings.items:
            raise LoanWorkbookValidationError(findings.items)
        ordered_rows = sorted(
            normalized.values(),
            key=lambda row: (
                row["sector"],
                row["borrower_name"] or "",
                row["instrument_key"],
            ),
        )
        canonical = {
            "source_id": source_id,
            "source_sha256": source_sha256,
            "template_version": TEMPLATE_VERSION,
            "importer_version": IMPORTER_VERSION,
            "workbook_date": expected_date,
            "rows": ordered_rows,
        }
        return {
            **canonical,
            "row_count": len(ordered_rows),
            "universe_digest": digest(canonical),
        }
    finally:
        workbook.close()


def import_loan_source(store: Any, case_id: str, source_id: str, actor: str) -> tuple[dict[str, Any], bool]:
    """Import lifecycle over the framework store: content-identity idempotent,
    fail-closed on withdrawn sources and vault/digest drift, exactly one ACTIVE
    universe per case. Returns (public record, created)."""
    source = store.get_source_private(source_id)
    if source is None or source.get("case_id") != case_id:
        raise LoanUniverseSourceError("RV_SOURCE_NOT_FOUND", "source is not part of this case")
    if source.get("withdrawn"):
        raise LoanUniverseSourceError("RV_SOURCE_NOT_ACTIVE", "withdrawn sources cannot seed a loan universe")
    vault_path = source.get("vault_path")
    if not vault_path or not Path(vault_path).is_file():
        raise LoanUniverseSourceError("RV_SOURCE_INTEGRITY_MISMATCH", "source bytes are unavailable")
    content = Path(vault_path).read_bytes()
    if hashlib.sha256(content).hexdigest() != source["sha256"]:
        raise LoanUniverseSourceError("RV_SOURCE_INTEGRITY_MISMATCH", "vault bytes do not match the source digest")

    existing = store.find_loan_universe(case_id, source["sha256"])
    if existing is not None:
        if existing["status"] == "REJECTED":
            raise LoanUniverseImportRejected(existing)
        return existing, False

    base = {
        "case_id": case_id,
        "source_id": source_id,
        "source_filename": source["filename"],
        "source_sha256": source["sha256"],
        "template_version": TEMPLATE_VERSION,
        "importer_version": IMPORTER_VERSION,
        "created_by": actor,
    }
    try:
        parsed = parse_loan_workbook(content, source_id=source_id, source_sha256=source["sha256"])
    except LoanWorkbookValidationError as exc:
        record = store.save_loan_universe({
            **base,
            "status": "REJECTED",
            "workbook_date": None,
            "universe_digest": None,
            "row_count": 0,
            "findings": exc.findings,
            "rows": [],
        }, actor)
        raise LoanUniverseImportRejected(record) from exc
    record = store.save_loan_universe({
        **base,
        "status": "ACTIVE",
        "workbook_date": parsed["workbook_date"],
        "universe_digest": parsed["universe_digest"],
        "row_count": parsed["row_count"],
        "findings": [],
        "rows": parsed["rows"],
    }, actor)
    return record, True


