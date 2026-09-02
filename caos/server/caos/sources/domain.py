from __future__ import annotations

import hashlib
import io
import json
import math
import os
import secrets
import socket
import struct
import zipfile
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile

from ..atomic_files import (
    VaultFileIntegrityError,
    VaultFileUnavailable,
    read_verified_vault_bytes,
)
from ..config import Settings
from ..storage.store import DomainStore
from ..storage.store import now_iso


ALLOWED_SUFFIXES = {".pdf", ".xlsx", ".json", ".txt", ".md", ".csv"}
MAX_ZIP_ENTRIES = 2000
MAX_ZIP_RATIO = 100
MAX_ZIP_UNCOMPRESSED = 100_000_000
# A real annual report, credit agreement or lender presentation is 0.5-3 MB of
# extracted text; 12 MB leaves headroom without admitting a container that
# bundles many unrelated exhibits rather than one document. The 25 MB route
# ceiling still bounds what is read off the wire.
MAX_SOURCE_TEXT = 12_000_000
# A block is the quotable unit and one row of every module's source manifest,
# so both its size and its count are bounded. A line longer than MAX_BLOCK_CHARS
# is split, never refused: an inline-XBRL filing is a whole document on one
# line, and refusing it loses a real filing over its whitespace.
MAX_BLOCK_CHARS = 20_000
# The run manifest carries one row per block into every module prompt, so block
# count must not track document size. Past this many, lines are grouped into
# blocks wide enough to hold the count down. It is a switch point, not a hard
# cap: once the width needed would exceed MAX_BLOCK_CHARS the count grows again,
# to at most MAX_SOURCE_TEXT / MAX_BLOCK_CHARS = 600 for any shape of document.
MAX_BLOCKS_PER_SOURCE = 320
# Floor on a grouped block, so a merely-large document is not chopped into
# hundreds of fragments when a few hundred would index it just as well.
TARGET_BLOCK_CHARS = 4_000
MAX_XLSX_EXTRACT_WORKSHEETS = 64
MAX_XLSX_EXTRACT_ROWS = 25_000
MAX_XLSX_EXTRACT_COLUMNS = 64
EXTRACTION_LIMIT_DETAIL = "source exceeds safe extraction limits"


class Vault:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.root = settings.storage_dir.resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _key(sha256: str) -> str:
        if (
            not isinstance(sha256, str)
            or len(sha256) != 64
            or any(character not in "0123456789abcdef" for character in sha256)
        ):
            raise ValueError("invalid source digest")
        return f"sources/{sha256[:2]}/{sha256}"

    def put(self, content: bytes, sha256: str) -> str:
        actual = hashlib.sha256(content).hexdigest()
        if actual != sha256:
            raise ValueError("content digest mismatch")
        key = self._key(sha256)
        directory = self.root / "sources" / sha256[:2]
        directory.mkdir(parents=True, exist_ok=True)
        target = directory / sha256
        if target.exists():
            try:
                read_verified_vault_bytes(
                    self.root,
                    key,
                    expected_sha256=sha256,
                    expected_size=len(content),
                    max_bytes=self.settings.max_upload_bytes,
                )
            except (VaultFileIntegrityError, VaultFileUnavailable) as exc:
                raise ValueError("existing source identity mismatch") from exc
            return str(target)
        temp = directory / f".{sha256}.{secrets.token_hex(8)}.tmp"
        try:
            with temp.open("xb") as handle:
                handle.write(content)
                handle.flush()
                os.fsync(handle.fileno())
            os.replace(temp, target)
        finally:
            temp.unlink(missing_ok=True)
        return str(target)

    def verify(self, source: dict[str, Any]) -> bytes:
        sha256 = source.get("sha256")
        key = self._key(sha256)
        expected_path = self.root.joinpath(*key.split("/"))
        if source.get("vault_path") != str(expected_path):
            raise ValueError("source vault identity mismatch")
        try:
            return read_verified_vault_bytes(
                self.root,
                key,
                expected_sha256=sha256,
                expected_size=source.get("bytes"),
                max_bytes=self.settings.max_upload_bytes,
            )
        except (VaultFileIntegrityError, VaultFileUnavailable) as exc:
            raise ValueError("source vault integrity mismatch") from exc


def scan_content(content: bytes, settings: Settings) -> None:
    if b"EICAR-STANDARD-ANTIVIRUS-TEST-FILE" in content:
        raise HTTPException(status_code=422, detail="malware detected")
    if settings.environment != "production":
        return
    if not settings.clamav_host:
        raise HTTPException(status_code=503, detail="malware scanner is not configured")
    try:
        with socket.create_connection(
            (settings.clamav_host, settings.clamav_port), timeout=15
        ) as connection:
            connection.sendall(b"zINSTREAM\0")
            for start in range(0, len(content), 1024 * 1024):
                chunk = content[start : start + 1024 * 1024]
                connection.sendall(struct.pack(">I", len(chunk)) + chunk)
            connection.sendall(b"\0\0\0\0")
            response = connection.recv(4096).lower()
    except OSError as exc:
        raise HTTPException(
            status_code=503, detail="malware scanner unavailable"
        ) from exc
    if b"found" in response:
        raise HTTPException(status_code=422, detail="malware detected")
    if b"ok" not in response:
        raise HTTPException(
            status_code=503, detail="malware scanner returned an unusable result"
        )


def validate_archive(content: bytes) -> None:
    if not content[:2] == b"PK":
        return
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ZIP_ENTRIES:
                raise HTTPException(
                    status_code=422, detail="archive has too many entries"
                )
            total_size = sum(info.file_size for info in infos)
            if (
                total_size > MAX_ZIP_UNCOMPRESSED
                or total_size > max(len(content), 1) * MAX_ZIP_RATIO
            ):
                raise HTTPException(
                    status_code=422, detail="archive expanded size exceeds limit"
                )
            for info in infos:
                if (
                    info.filename.startswith(("/", "\\"))
                    or ".." in Path(info.filename).parts
                ):
                    raise HTTPException(
                        status_code=422, detail="archive path traversal"
                    )
                if (
                    info.compress_size
                    and info.file_size / info.compress_size > MAX_ZIP_RATIO
                ):
                    raise HTTPException(
                        status_code=422,
                        detail="archive compression ratio exceeds limit",
                    )
    except zipfile.BadZipFile as exc:
        raise HTTPException(status_code=422, detail="invalid archive") from exc


def extract_blocks(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".json":

        def reject_non_finite(_constant: str) -> None:
            raise ValueError("non-finite JSON number")

        def reject_overflowing_number(token: str) -> float:
            # `parse_constant` only fires on the NaN/Infinity literals. A finite-looking
            # token that overflows the double range (1e999) is admitted by the parser as
            # `inf`, which then re-serialises to a bare `Infinity` this same reader would
            # reject -- a figure that no longer round-trips through its own evidence.
            value = float(token)
            if not math.isfinite(value):
                raise ValueError("non-finite JSON number")
            return value

        def reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
            data = dict(pairs)
            if len(data) != len(pairs):
                raise ValueError("duplicate JSON key")
            return data

        try:
            data = json.loads(
                content.decode("utf-8"),
                parse_constant=reject_non_finite,
                parse_float=reject_overflowing_number,
                object_pairs_hook=reject_duplicate_keys,
            )
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
            raise HTTPException(status_code=422, detail="invalid JSON source") from exc
        if (
            data is None
            or (isinstance(data, str) and not data.strip())
            or (isinstance(data, (dict, list)) and not data)
        ):
            raise HTTPException(
                status_code=422, detail="source contains no extractable evidence"
            )
        text = json.dumps(data, sort_keys=True, indent=2)
    elif suffix == ".pdf":
        try:
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(content))
            page_count = len(reader.pages)
            if not page_count:
                raise ValueError("PDF contains no pages")
        except Exception as exc:
            raise HTTPException(status_code=422, detail="invalid PDF source") from exc
        try:
            text = "\n".join(
                reader.pages[index].extract_text() or "" for index in range(page_count)
            )
        except Exception:
            text = ""
        if not text.strip():
            text = "PDF source stored; no text layer was available for safe extraction."
    elif suffix == ".xlsx":
        workbook = None
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(
                io.BytesIO(content), read_only=True, data_only=False
            )
            if len(workbook.worksheets) > MAX_XLSX_EXTRACT_WORKSHEETS:
                raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
            lines: list[str] = []
            text_length = 0

            def append_xlsx_line(line: str) -> None:
                nonlocal text_length
                next_length = text_length + len(line) + (1 if lines else 0)
                if next_length > MAX_SOURCE_TEXT:
                    raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
                lines.append(line)
                text_length = next_length

            has_cell_value = False
            extracted_rows = 0
            for sheet in workbook.worksheets:
                sheet.reset_dimensions()
                append_xlsx_line(f"[sheet:{sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    extracted_rows += 1
                    if extracted_rows > MAX_XLSX_EXTRACT_ROWS:
                        raise HTTPException(
                            status_code=422, detail=EXTRACTION_LIMIT_DETAIL
                        )
                    if len(row) > MAX_XLSX_EXTRACT_COLUMNS:
                        raise HTTPException(
                            status_code=422, detail=EXTRACTION_LIMIT_DETAIL
                        )
                    values = ["" if value is None else str(value) for value in row]
                    values.extend([""] * (MAX_XLSX_EXTRACT_COLUMNS - len(values)))
                    has_cell_value = has_cell_value or any(
                        value.strip() for value in values
                    )
                    line = "\t".join(values)
                    append_xlsx_line(line)
            if not has_cell_value:
                raise HTTPException(
                    status_code=422, detail="source contains no extractable evidence"
                )
            text = "\n".join(lines)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=422, detail="invalid XLSX source") from exc
        finally:
            if workbook is not None:
                workbook.close()
    else:
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError as exc:
            raise HTTPException(
                status_code=422, detail="text source must be UTF-8"
            ) from exc
    if suffix in {".txt", ".md"} and not text.strip():
        raise HTTPException(
            status_code=422, detail="source contains no extractable evidence"
        )
    if suffix == ".csv" and not text.strip(' \t\r\n,"'):
        raise HTTPException(
            status_code=422, detail="source contains no extractable evidence"
        )
    if len(text) > MAX_SOURCE_TEXT:
        raise HTTPException(status_code=422, detail=EXTRACTION_LIMIT_DETAIL)
    return pack_blocks(text) or [
        {
            "block_id": "b00001",
            "locator": {"line": 1},
            "text": "",
            "extractor_version": "builtin-v1",
            "confidence": "LOW",
            "untrusted_data": True,
        }
    ]


def _fragments(text: str) -> list[tuple[int, int, str]]:
    """Non-blank lines as (line, part, text), splitting any line wider than a
    block. `part` is 0 for a line that was not split, so an ordinary document
    keeps exactly the block ids and locators it had before blocks were bounded."""
    fragments: list[tuple[int, int, str]] = []
    for index, line in enumerate(text.splitlines() or [text], start=1):
        if not line.strip():
            continue
        if len(line) <= MAX_BLOCK_CHARS:
            fragments.append((index, 0, line))
            continue
        for part, start in enumerate(range(0, len(line), MAX_BLOCK_CHARS), start=1):
            fragments.append((index, part, line[start:start + MAX_BLOCK_CHARS]))
    return fragments


def _block(block_id: str, locator: dict[str, Any], text: str, version: str) -> dict[str, Any]:
    return {
        "block_id": block_id,
        "locator": locator,
        "text": text,
        "extractor_version": version,
        "confidence": "MEDIUM",
        "untrusted_data": True,
    }


def _line_block(index: int, part: int, fragment: str) -> dict[str, Any]:
    """One line, or one slice of an over-wide line. An unsplit line keeps
    exactly the id and locator the extractor produced before blocks were
    bounded, which is what leaves ordinary documents untouched."""
    if not part:
        return _block(f"b{index:05d}", {"line": index}, fragment, "builtin-v1")
    return _block(f"b{index:05d}p{part:02d}", {"line": index, "part": part}, fragment, "builtin-v2")


def pack_blocks(text: str) -> list[dict[str, Any]]:
    """Evidence blocks: one per line while a document is small, bounded groups
    of lines once it is not.

    Block count has to stay bounded independently of document size. The run's
    source manifest carries one row per block into *every* module prompt, so a
    300-page credit agreement indexed line by line costs more context than a
    module is given to think with -- which is why such a run used to die in
    `bound_manifest` before its first provider call. Grouping keeps that index
    affordable; MAX_BLOCK_CHARS keeps any single block quotable and keeps one
    evidence read inside its byte budget.
    """
    fragments = _fragments(text)
    if not fragments:
        return []
    if len(fragments) <= MAX_BLOCKS_PER_SOURCE:
        return [_line_block(*fragment) for fragment in fragments]

    width = min(MAX_BLOCK_CHARS, max(TARGET_BLOCK_CHARS, math.ceil(len(text) / MAX_BLOCKS_PER_SOURCE)))
    blocks: list[dict[str, Any]] = []
    buffer: list[str] = []
    first = last = fragments[0][0]
    size = 0

    def flush() -> None:
        nonlocal buffer, size
        if buffer:
            blocks.append(_block(f"b{len(blocks) + 1:05d}", {"lines": [first, last]},
                                 "\n".join(buffer), "builtin-v2"))
            buffer, size = [], 0

    for index, _part, fragment in fragments:
        if buffer and size + len(fragment) + 1 > width:
            flush()
        if not buffer:
            first = index
        buffer.append(fragment)
        size += len(fragment) + 1
        last = index
    flush()
    return blocks


async def ingest_upload(
    catalog: DomainStore,
    vault: Vault,
    case_id: str,
    actor: str,
    upload: UploadFile,
    max_bytes: int,
) -> dict[str, Any]:
    filename = Path(upload.filename or "source.bin").name
    suffix = Path(filename).suffix.lower()
    if suffix not in ALLOWED_SUFFIXES:
        raise HTTPException(status_code=415, detail="unsupported source type")
    content = await upload.read(max_bytes + 1)
    if not content:
        raise HTTPException(status_code=422, detail="source is empty")
    if len(content) > max_bytes:
        raise HTTPException(status_code=413, detail="source exceeds upload limit")
    scan_content(content, vault.settings)
    validate_archive(content)
    sha256 = hashlib.sha256(content).hexdigest()
    blocks = extract_blocks(filename, content)
    vault_path = vault.put(content, sha256)
    source = {
        "case_id": case_id,
        "filename": filename,
        "media_type": upload.content_type or "application/octet-stream",
        "bytes": len(content),
        "sha256": sha256,
        "vault_path": vault_path,
        "blocks": blocks,
        "created_by": actor,
        "created_at": now_iso(),
        "withdrawn": False,
    }
    try:
        return catalog.ingest(source, actor)
    except ValueError as exc:
        if str(exc) == "source content already active":
            raise HTTPException(status_code=409, detail=str(exc)) from exc
        raise


def list_sources(catalog: DomainStore, case_id: str) -> list[dict[str, Any]]:
    return catalog.list_sources(case_id)


def current_source_set(catalog: DomainStore, case_id: str) -> dict[str, Any] | None:
    return catalog.current_source_set(case_id)


def pathway_fit(catalog: DomainStore, case_id: str) -> dict[str, Any]:
    # Filenames only: this runs once per case on the case-list route, and the
    # full source row carries the blocks column with it.
    filenames = catalog.list_source_filenames(case_id)
    types = sorted({Path(filename).suffix.lower().lstrip(".") for filename in filenames})
    return {
        "source_count": len(filenames),
        "file_types": types,
        "fit": "READY" if filenames else "NEEDS_SOURCE",
        "language": "Pathway fit is a source-coverage signal, not CP-0 readiness.",
        "message": "Source coverage supports pathway selection."
        if filenames
        else "Upload governed source material before selecting a route.",
    }
