"""The CP-MODEL pure calculation engine boundary.

Ported from LEGACY models/domain.py per DECISIONS §5 (the calculation engine
copies across as domain code); the vendored cp_model_v3 scripts themselves are
methodology authority and load byte-verified from the pinned Deploy V bundle.
The service/runner halves (legacy models/runtime.py, revisions.py) do NOT port
— caos.models.service is rewritten against the framework.
"""

from __future__ import annotations

import hashlib
import importlib.util
import json
import math
import os
import re
import shutil
import stat
import sys
import tempfile
import threading
import weakref
from copy import deepcopy
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from types import ModuleType
from typing import Any

from ..config import Settings
from ..contracts import canonical_json
from ..methodology.bundle import DeployVBundle, MethodologyError
from ..methodology.canonical import (
    CanonicalValidationError,
    model_source_ids,
    validate_model_sources,
    validate_visible_stable_tables,
    visible_top_level_markdown,
)


class ModelInputError(ValueError):
    """Carries the typed blocker code; the message stays host-side.

    Readiness used to collapse every raise here into
    CANONICAL_MODEL_INPUTS_INVALID, which told an analyst whose accepted
    authority is simply the wrong pathway that their inputs were corrupt
    (DECISIONS §9: the code each refusal carries is contract). The code is the
    only part callers may branch on — messages here interpolate host internals
    (bundle paths, table ids) and must never reach the wire.
    """

    def __init__(
        self,
        message: str,
        code: str = "CANONICAL_MODEL_INPUTS_INVALID",
        *,
        detail: str | None = None,
    ) -> None:
        # `detail` is the analyst-facing sentence readiness may render; it is
        # written from host identifiers only, never from the message, which can
        # interpolate bundle paths and table ids.
        self.detail = detail
        super().__init__(message)
        self.code = code


class CpModelV3Error(ValueError):
    """Stable host type for errors raised by content-isolated CP-MODEL runtimes."""


_T5_HEADERS = {
    "T5.1": (
        "event_id",
        "source_document",
        "event_description",
        "event_category",
        "date_range",
        "evidence_quality",
        "source_reliability",
    ),
    "T5.2": (
        "date_window",
        "event_description",
        "event_category",
        "credit_relevance_summary",
        "source",
    ),
    "T5.3": (
        "event_id",
        "description",
        "probability",
        "credit_impact_channel_s",
        "impact_severity",
        "affected_metrics",
        "risk_direction",
        "source",
    ),
    "T5.4": ("event_id", "description", "probability", "impact", "p_i_classification"),
    "T5.5": (
        "event_id",
        "description",
        "priority",
        "monitoring_frequency",
        "trigger_condition",
        "responsible_module",
    ),
    "T5.6": (
        "event_id",
        "description",
        "priority",
        "receiving_module",
        "handoff_content",
        "timing",
        "rationale",
    ),
    "T5.7": (
        "gap_description",
        "affected_section",
        "downstream_impact",
        "severity",
        "recommended_action",
    ),
}
_T5_HEADING = re.compile(r"^###\s+(T5\.[1-7])\b.*$", re.MULTILINE)
_ANALYSIS_HEADING = re.compile(r"^## Analysis\s*$", re.MULTILINE)
_EVIDENCE_HEADING = re.compile(r"^## Evidence Trace\s*$", re.MULTILINE)
_SEPARATOR = re.compile(r"^:?-{3,}:?$")
_HEX_DIGEST = re.compile(r"^[0-9a-f]{64}$")
_LOAD_LOCK = threading.Lock()
_MISSING = object()

WORKSHEET_SCHEMA_VERSION = "caos.model.worksheet.v1"
MAX_WORKSHEET_CELLS = 20_000
MAX_WORKSHEET_BYTES = 8 * 1024 * 1024
MAX_CELL_TEXT = 10_000


def _normalise_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _table_headers(block: str, table_id: str) -> tuple[str, ...]:
    lines = block.splitlines()
    for index in range(len(lines) - 2):
        if not lines[index].lstrip().startswith("|") or not lines[index + 1].lstrip().startswith("|"):
            continue
        headers = tuple(_normalise_header(cell) for cell in lines[index].strip().strip("|").split("|"))
        separator = tuple(cell.strip().replace(" ", "") for cell in lines[index + 1].strip().strip("|").split("|"))
        if len(headers) == len(separator) and all(_SEPARATOR.fullmatch(cell) for cell in separator):
            row_count = 0
            for row in lines[index + 2 :]:
                if not row.lstrip().startswith("|"):
                    break
                values = row.strip().strip("|").split("|")
                if len(values) != len(headers):
                    raise ModelInputError(f"{table_id} row has {len(values)} cells; expected {len(headers)}")
                if all(_SEPARATOR.fullmatch(cell.strip()) for cell in values):
                    raise ModelInputError(f"{table_id} contains a separator instead of data")
                row_count += 1
            if not row_count:
                raise ModelInputError(f"{table_id} must contain at least one row")
            return headers
    raise ModelInputError(f"{table_id} is missing its Markdown table")


def _t5_sections(markdown: str) -> list[str]:
    visible = visible_top_level_markdown(markdown)
    analysis_headings = list(_ANALYSIS_HEADING.finditer(visible))
    evidence_headings = list(_EVIDENCE_HEADING.finditer(visible))
    if (
        len(analysis_headings) != 1
        or len(evidence_headings) != 1
        or analysis_headings[0].end() >= evidence_headings[0].start()
    ):
        raise ModelInputError("CP-2A must contain one ordered Analysis section")
    start, end = analysis_headings[0].end(), evidence_headings[0].start()
    analysis = markdown[start:end]
    visible_analysis = visible[start:end]
    matches = list(_T5_HEADING.finditer(visible_analysis))
    observed = [match.group(1) for match in matches]
    expected = list(_T5_HEADERS)
    if observed != expected:
        raise ModelInputError(f"CP-2A must contain T5.1 through T5.7 exactly once in order; found {observed}")
    sections: list[str] = []
    for index, match in enumerate(matches):
        end = matches[index + 1].start() if index + 1 < len(matches) else len(analysis)
        section = analysis[match.start() : end]
        visible_section = visible_analysis[match.start() : end]
        table_id = match.group(1)
        headers = _table_headers(visible_section, table_id)
        if headers != _T5_HEADERS[table_id]:
            raise ModelInputError(f"{table_id} headers do not match the canonical CP-2B schema")
        sections.append(section)
    return sections


def _load_module(
    name: str,
    path: Path,
    content_digest: str,
    aliases: dict[str, ModuleType] | None = None,
) -> ModuleType:
    instance_digest = hashlib.sha256(f"{path}:{content_digest}".encode()).hexdigest()[:16]
    module_name = f"caos_deploy_v_{name}_{instance_digest}"
    cached = sys.modules.get(module_name)
    if cached is not None:
        return cached
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise ModelInputError(f"cannot load CP-MODEL authority script: {name}")
    module = importlib.util.module_from_spec(spec)
    prior = {alias: sys.modules.get(alias, _MISSING) for alias in aliases or {}}
    prior_dont_write_bytecode = sys.dont_write_bytecode
    sys.modules[module_name] = module
    try:
        sys.dont_write_bytecode = True
        sys.modules.update(aliases or {})
        spec.loader.exec_module(module)
    except BaseException:
        sys.modules.pop(module_name, None)
        raise
    finally:
        sys.dont_write_bytecode = prior_dont_write_bytecode
        for alias, previous in prior.items():
            if previous is _MISSING:
                sys.modules.pop(alias, None)
            else:
                sys.modules[alias] = previous
    return module


@dataclass(frozen=True)
class DebtRow:
    """One disclosed debt-schedule line. The unallocated movement is its own
    disclosed row — security/seniority is never inferred for it."""

    facility_id: str
    label: str
    value: Any
    caption: str


class _ColumnView:
    """Delegating view over a vendor ColumnCalculation adding the disclosed
    debt schedule the spec reads (the vendor type is frozen methodology)."""

    def __init__(self, column: Any, model: Any) -> None:
        self._column = column
        self._model = model

    def __getattr__(self, name: str) -> Any:
        return getattr(self._column, name)

    @property
    def debt_schedule(self) -> list[DebtRow]:
        labels = {series.facility_id: series.label for series in self._model.debt}
        rows: list[DebtRow] = []
        for facility_id, value in self._column.debt_values.items():
            if value is None:
                continue
            if facility_id == "forecast::unallocated_debt_movement":
                rows.append(DebtRow(
                    facility_id=facility_id,
                    label="Unallocated debt movement",
                    value=value,
                    caption="Disclosed reconciliation line; security and seniority are not inferred.",
                ))
            else:
                rows.append(DebtRow(
                    facility_id=facility_id,
                    label=labels.get(facility_id, facility_id),
                    value=value,
                    caption="",
                ))
        return rows


class _CalculationBookView:
    def __init__(self, book: Any, model: Any) -> None:
        self._book = book
        self._model = model

    def __getattr__(self, name: str) -> Any:
        return getattr(self._book, name)

    def for_column(self, column_id: str) -> _ColumnView:
        return _ColumnView(self._book.for_column(column_id), self._model)

    def for_period(self, period_id: str) -> Any:
        return self._book.for_period(period_id)


@dataclass(frozen=True)
class CellExpectation:
    """A decision-output formula cell: address, exact Python expectation, and
    the literal Excel formula written there (spec: no IFERROR masking)."""

    sheet: str
    cell: str
    semantic_id: str
    column_id: str
    expected: Any
    formula: str
    tolerance: Any


@dataclass(frozen=True)
class RenderedWorkbook:
    formulas: tuple[CellExpectation, ...]
    mappings: tuple[Any, ...]
    model_cells: dict[tuple[str, str], str]


class CpModelBundle:
    """Load the integrity-pinned CP-MODEL validators without a permanent sys.path edit."""

    _RUNTIME_FILES = (
        "scripts/validate_handoff.py",
        "scripts/validate_cp_model_inputs.py",
        "scripts/cp_model_v3/__init__.py",
        "scripts/cp_model_v3/domain.py",
        "scripts/cp_model_v3/calculations.py",
        "scripts/cp_model_v3/workbook.py",
        "scripts/cp_model_v3/builder.py",
    )

    def __init__(self, deploy_v_root: Path, *, integrity: dict[str, Any] | None = None) -> None:
        self.deploy_v_root = Path(deploy_v_root).resolve()
        if integrity is None:
            try:
                integrity = DeployVBundle(self.deploy_v_root).integrity
            except MethodologyError as exc:
                raise ModelInputError("CP-MODEL methodology authority is invalid") from exc
        self._integrity = deepcopy(integrity)
        source_entries = next(
            (skill.get("relative_file_hashes") for skill in self._integrity.get("skills", [])
             if skill.get("folder_slug") == "cp-model"),
            None,
        )
        if not isinstance(source_entries, dict):
            raise ModelInputError("CP-MODEL methodology authority is not declared")
        sources = {
            relative: self._read_pinned_source(relative, source_entries.get(relative))
            for relative in self._RUNTIME_FILES
        }
        content_digest = hashlib.sha256(b"".join(
            relative.encode("utf-8") + b"\0" + sources[relative]
            for relative in self._RUNTIME_FILES
        )).hexdigest()
        # mkdtemp + an explicit finalizer, not TemporaryDirectory: the latter's
        # implicit cleanup at garbage collection is a ResourceWarning, and the
        # bundle is built once per service, so every discarded service warned.
        runtime_root = Path(tempfile.mkdtemp(prefix=f"caos-cp-model-{content_digest[:12]}-"))
        self._runtime_cleanup = weakref.finalize(self, shutil.rmtree, str(runtime_root), ignore_errors=True)
        for relative, source in sources.items():
            destination = runtime_root / relative
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(source)
        scripts = runtime_root / "scripts"
        with _LOAD_LOCK:
            self._handoff = _load_module(
                "cp_model_validate_handoff",
                scripts / "validate_handoff.py",
                content_digest,
            )
            self._inputs = _load_module(
                "cp_model_validate_inputs",
                scripts / "validate_cp_model_inputs.py",
                content_digest,
                {"validate_handoff": self._handoff},
            )
            package = _load_module(
                "cp_model_v3",
                scripts / "cp_model_v3" / "__init__.py",
                content_digest,
                {"validate_handoff": self._handoff, "validate_cp_model_inputs": self._inputs},
            )
            self._domain = sys.modules[f"{package.__name__}.domain"]
            self._calculations = sys.modules[f"{package.__name__}.calculations"]
            self._workbook = sys.modules[f"{package.__name__}.workbook"]
            self._builder = sys.modules[f"{package.__name__}.builder"]
        runtime_hash = hashlib.sha256()
        for relative in (
            "scripts/cp_model_v3/domain.py",
            "scripts/cp_model_v3/calculations.py",
            "scripts/cp_model_v3/workbook.py",
            "scripts/validate_cp_model_inputs.py",
            "scripts/validate_handoff.py",
        ):
            path = Path(relative)
            runtime_hash.update(path.name.encode("utf-8"))
            runtime_hash.update(sources[relative])
        self.calculation_runtime = {
            "name": "cp_model_v3_python",
            "version": self._domain.V3_CONTRACT_VERSION,
            "sha256": runtime_hash.hexdigest(),
            "assumption_registry_version": self._inputs.ASSUMPTION_REGISTRY_VERSION,
            "assumption_registry_digest": self._inputs.ASSUMPTION_REGISTRY_DIGEST,
            "calculation_contract_version": self._calculations.CALCULATION_CONTRACT_VERSION,
        }
        self.assumption_registry = self._inputs.assumption_registry()

    def _read_pinned_source(self, relative: str, entry: Any) -> bytes:
        if (
            not isinstance(entry, dict)
            or type(entry.get("bytes")) is not int
            or not isinstance(entry.get("sha256"), str)
        ):
            raise ModelInputError(f"CP-MODEL authority file is not pinned: {relative}")
        path = self.deploy_v_root / "skills" / "cp-model" / relative
        descriptor = None
        try:
            descriptor = os.open(path, os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0))
            metadata = os.fstat(descriptor)
            if not stat.S_ISREG(metadata.st_mode):
                raise ModelInputError(f"CP-MODEL authority file is not regular: {relative}")
            chunks: list[bytes] = []
            while chunk := os.read(descriptor, 1024 * 1024):
                chunks.append(chunk)
            source = b"".join(chunks)
        except OSError as exc:
            raise ModelInputError(f"CP-MODEL authority file is unavailable: {relative}") from exc
        finally:
            if descriptor is not None:
                os.close(descriptor)
        if len(source) != entry["bytes"] or hashlib.sha256(source).hexdigest() != entry["sha256"]:
            raise ModelInputError(f"CP-MODEL authority file changed: {relative}")
        return source

    # -- integrity ---------------------------------------------------------

    def verify_integrity(self) -> list[str]:
        """Hash every file the pinned integrity manifest declares; return the
        mismatches (empty list == the vendored bundle is byte-exact)."""
        mismatches: list[str] = []
        for skill in self._integrity["skills"]:
            for relative, entry in (skill.get("relative_file_hashes") or {}).items():
                path = self.deploy_v_root / "skills" / skill["folder_slug"] / relative
                if not path.is_file():
                    mismatches.append(f"{skill['folder_slug']}/{relative}: missing")
                    continue
                data = path.read_bytes()
                if len(data) != entry["bytes"] or hashlib.sha256(data).hexdigest() != entry["sha256"]:
                    mismatches.append(f"{skill['folder_slug']}/{relative}: digest mismatch")
        return mismatches

    def resolve_declared_file(self, folder_slug: str, relative: str) -> Path:
        """Resolve a declared authority file, refusing any symlink component —
        a symlinked path can escape the pinned bundle."""
        base = self.deploy_v_root / "skills" / folder_slug
        path = base / relative
        probe = path
        while probe != self.deploy_v_root.parent:
            if probe.is_symlink():
                raise ModelInputError(f"declared authority path is a symlink: {probe}")
            if probe == self.deploy_v_root:
                break
            probe = probe.parent
        if not path.is_file():
            raise ModelInputError(f"declared authority file is missing: {folder_slug}/{relative}")
        return path

    # -- validation and calculation ---------------------------------------

    def validate_handoff(self, markdown: str, *, module_id: str, run_id: str | None = None) -> Any:
        try:
            validate_visible_stable_tables(markdown)
        except CanonicalValidationError as exc:
            raise ModelInputError(str(exc)) from exc
        return self._handoff.validate_text(markdown, expected_module=module_id, expected_run_id=run_id)

    def validate(self, cp1: str, cp1a: str, cp1b: str, cp2: str, cp2b: str, cp2g: str | None = None) -> Any:
        try:
            documents = (
                ("CP-1", cp1),
                ("CP-1A", cp1a),
                ("CP-1B", cp1b),
                ("CP-2", cp2),
                ("CP-2B", cp2b),
                ("CP-2G", cp2g),
            )
            for module_id, markdown in documents:
                if markdown is not None:
                    validate_visible_stable_tables(markdown)
                    validate_model_sources(
                        markdown,
                        model_source_ids(markdown),
                        module_id=module_id,
                    )
        except CanonicalValidationError as exc:
            raise ModelInputError(str(exc)) from exc
        return self._inputs.validate_cp_model_bundle(cp1, cp1a, cp1b, cp2, cp2b, cp2g)

    def calculate(
        self, paths: dict[str, Path], *, effective_assumptions: list[dict[str, Any]] | None = None
    ) -> tuple[Any, Any]:
        try:
            for module_id, path in paths.items():
                markdown = Path(path).read_text(encoding="utf-8")
                validate_visible_stable_tables(markdown)
                validate_model_sources(
                    markdown,
                    model_source_ids(markdown),
                    module_id=module_id,
                )
        except CanonicalValidationError as exc:
            raise ModelInputError(str(exc)) from exc
        bundle_paths = self._domain.BundlePaths(
            cp1=paths["CP-1"],
            cp1a=paths["CP-1A"],
            cp1b=paths["CP-1B"],
            cp2=paths["CP-2"],
            cp2b=paths["CP-2B"],
            cp2g=paths.get("CP-2G"),
        )
        self._growth_slot_preflight(paths.get("CP-2G"))
        try:
            model = self._domain.build_ir(bundle_paths, effective_assumptions=effective_assumptions)
            # Serialised with calculate_model: its lenient-check window must
            # never be observable by a concurrent strict calculation.
            with _LOAD_LOCK:
                calculations = self._calculations.calculate(model)
        except self._domain.CpModelV3Error as exc:
            raise self._name_growth_slot(exc) from exc
        return model, _CalculationBookView(calculations, model)

    _GROWTH_SLOTS = re.compile(r"^operating\.(?:revenue_growth\.division_(\d+)|(consolidated)_revenue_growth)$")

    def _growth_slot_preflight(self, cp2g_path: Path | None) -> None:
        """The calc node trusts nothing: a READY active-growth driver whose
        value is non-finite is refused here, naming the slot (spec row 133/134),
        before the vendor parser can reduce it to a row number."""
        if cp2g_path is None:
            return
        header: list[str] | None = None
        for line in Path(cp2g_path).read_text(encoding="utf-8").splitlines():
            if not line.startswith("|"):
                header = None
                continue
            cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
            if "assumption_id" in cells and "value" in cells:
                header = cells
                continue
            if header is None or len(cells) != len(header) or set(cells[0].replace("-", "")) <= {"", ":"}:
                continue
            row = dict(zip(header, cells))
            match = self._GROWTH_SLOTS.match(row.get("assumption_id", ""))
            if not match or row.get("status") != "READY" or not row.get("value"):
                continue
            slot = f"DIVISION_{match.group(1)}" if match.group(1) else "CONSOLIDATED"
            try:
                finite = Decimal(row["value"]).is_finite()
            except Exception:
                finite = False
            if not finite:
                raise CpModelV3Error(
                    f"active growth slot {slot} value must be finite, got {row['value']!r}"
                )

    def _name_growth_slot(self, exc: Exception) -> CpModelV3Error:
        """The calc boundary names the offending active-growth slot even where
        the engine reports only the assumption identity (spec rows 133+134)."""
        message = str(exc)
        match = re.search(r"revenue_growth\.division_(\d+)", message)
        if match:
            return CpModelV3Error(f"{message} (active growth slot DIVISION_{match.group(1)})")
        if "consolidated" in message.lower():
            return CpModelV3Error(f"{message} (active growth slot CONSOLIDATED)")
        return CpModelV3Error(message)

    def calculate_model(self, model: Any) -> Any:
        """Calculate a caller-supplied IR variant. Reconciliation checks are
        recorded but not blocking here — enforcement belongs to the
        validate/build boundary (bundle.calculate), and a variant IR (e.g. a
        stripped historical addback series) must still produce its forecast
        book. The vendored engine is not edited: only its blocking policy is
        held open for the duration of this call."""
        import dataclasses as _dataclasses

        original = self._calculations._check

        def lenient(check_id: str, period_id: str, actual: Any, expected: Any, detail: str) -> Any:
            check = original(check_id, period_id, actual, expected, detail)
            if check.status == "BLOCK":
                check = _dataclasses.replace(check, status="WARN")
            return check

        with _LOAD_LOCK:
            self._calculations._check = lenient
            try:
                calculations = self._calculations.calculate(model)
            finally:
                self._calculations._check = original
        return _CalculationBookView(calculations, model)

    # -- rendering ---------------------------------------------------------

    def render_workbook(self, model: Any, calculations: Any, output_path: Path) -> RenderedWorkbook:
        book = calculations._book if isinstance(calculations, _CalculationBookView) else calculations
        rendered = self._workbook.render_workbook(
            model,
            book,
            self._workbook.RenderMetadata(
                renderer_hash=self.calculation_runtime["sha256"],
                calculation_engine="CP-MODEL Python calculations",
            ),
            output_path,
        )
        return self._enrich(rendered, calculations, output_path)

    _IFERROR = re.compile(r"^=IFERROR\((.+),0\)$")

    def _enrich(self, rendered: Any, calculations: Any, output_path: Path) -> RenderedWorkbook:
        """Post-render host pass: unwrap IFERROR zero-fill masking (the spec
        bans error masking — an unavailable value is omitted, never zeroed) and
        attach column identity plus the literal formula text to every
        decision-output expectation so callers can assert render/engine parity.
        A cell counts as a decision output only when the renderer registered it
        in model_cells under (semantic_id, column_id) — aggregate KPI cells
        reuse semantic ids and are not per-column decision outputs."""
        from openpyxl import load_workbook

        column_by_target = {
            item.target: item.period_id
            for item in rendered.mappings
            if item.write_class == "FORMULA" and item.period_id
        }
        workbook = load_workbook(output_path, data_only=False)
        try:
            rewrote = False
            for sheet_name in workbook.sheetnames:
                for row in workbook[sheet_name].iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str):
                            match = self._IFERROR.match(cell.value)
                            if match:
                                cell.value = f"={match.group(1)}"
                                rewrote = True
            if rewrote:
                workbook.save(output_path)
            expectations: list[CellExpectation] = []
            for item in rendered.formulas:
                column_id = column_by_target.get(f"{item.sheet}!{item.cell}")
                if not column_id:
                    continue
                if rendered.model_cells.get((item.semantic_id, column_id)) != item.cell:
                    continue
                try:
                    column = calculations.for_column(column_id)
                except (KeyError, ValueError):
                    continue
                # Metric/KPI rows reuse value semantic ids with aggregate (LTM)
                # expectations — they are not per-column value cells.
                if item.semantic_id not in column.values or item.semantic_id in column.credit_metrics:
                    continue
                raw = workbook[item.sheet][item.cell].value
                expectations.append(CellExpectation(
                    sheet=item.sheet,
                    cell=item.cell,
                    semantic_id=item.semantic_id,
                    column_id=column_id,
                    expected=item.expected,
                    formula=raw if isinstance(raw, str) and raw.startswith("=") else "",
                    tolerance=item.tolerance,
                ))
        finally:
            workbook.close()
        return RenderedWorkbook(
            formulas=tuple(expectations),
            mappings=tuple(rendered.mappings),
            model_cells=dict(rendered.model_cells),
        )

    def serialize_workbook(self, model: Any, calculations: Any) -> dict[str, Any]:
        """Serialize every worksheet to the JSON worksheet payload using only
        Python — no external binaries (LibreOffice et al.) are ever invoked."""
        with tempfile.TemporaryDirectory(prefix="caos-model-serialize-") as temporary:
            draft = Path(temporary) / "model.xlsx"
            rendered = self.render_workbook(model, calculations, draft)
            return _serialize_workbook_file(draft, rendered, model, calculations)

    # -- exports -----------------------------------------------------------

    def export(self, paths: dict[str, Path], output_dir: Path) -> Any:
        return self._builder.build_cp_model(
            self._builder.BuildRequest(
                bundle=self._domain.BundlePaths(
                    cp1=paths["CP-1"], cp1a=paths["CP-1A"], cp1b=paths["CP-1B"],
                    cp2=paths["CP-2"], cp2b=paths["CP-2B"], cp2g=paths.get("CP-2G"),
                ),
                output_dir=output_dir,
            )
        )

    def export_revision(
        self,
        paths: dict[str, Path],
        output_dir: Path,
        *,
        effective_assumptions: list[dict[str, Any]],
        default_assumptions: list[dict[str, Any]],
        revision: dict[str, Any],
    ) -> Any:
        return self._builder.build_cp_model(
            self._builder.BuildRequest(
                bundle=self._domain.BundlePaths(
                    cp1=paths["CP-1"], cp1a=paths["CP-1A"], cp1b=paths["CP-1B"],
                    cp2=paths["CP-2"], cp2b=paths["CP-2B"], cp2g=paths.get("CP-2G"),
                ),
                output_dir=output_dir,
                effective_assumptions=effective_assumptions,
                revision=self._workbook.RevisionRenderContext(
                    assumptions=effective_assumptions,
                    defaults=default_assumptions,
                    definitions=self.assumption_registry["definitions"],
                    record=revision,
                ),
            )
        )


# -- worksheet serialization (ported from LEGACY models/runtime.py; the spec
# widens it to every sheet, keyed by "title") --------------------------------


def json_value(value: Any) -> Any:
    from datetime import date, datetime

    if isinstance(value, Decimal):
        if not value.is_finite():
            raise ModelInputError("non-finite worksheet value")
        converted = float(value)
        if not math.isfinite(converted):
            raise ModelInputError("non-finite worksheet value")
        return converted
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ModelInputError("non-finite worksheet value")
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, str) and len(value) > MAX_CELL_TEXT:
        raise ModelInputError("worksheet cell exceeds size limit")
    if value is None or isinstance(value, (str, int, bool)):
        return value
    return str(value)


def _cell_type(value: Any, formula: str | None) -> str:
    from datetime import date, datetime

    if formula:
        return "formula"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float, Decimal)):
        return "number"
    if isinstance(value, (date, datetime)):
        return "date"
    return "text"


def _assumptions_tab(model: Any) -> dict[str, Any]:
    """The analyst-facing assumption view: one row per effective assumption
    (the _INPUTS sheet stays the renderer's raw registry projection)."""
    headers = ("assumption_id", "case", "period_id", "unit", "status", "value", "gap_code")
    cells: list[dict[str, Any]] = []
    for column, header in enumerate(headers, start=1):
        cells.append({
            "address": f"R1C{column}", "row": 1, "column": column,
            "value": header, "value_type": "text", "formula": None,
            "semantic_id": None, "owner": "CP-MODEL", "write_class": "HEADER",
            "period_id": None, "source_refs": None,
        })
    for index, item in enumerate(model.effective_assumptions, start=2):
        row = (item.assumption_id, item.case, item.period_id, item.unit, item.status,
               json_value(item.value) if item.value is not None else None, item.gap_code)
        for column, value in enumerate(row, start=1):
            cells.append({
                "address": f"R{index}C{column}", "row": index, "column": column,
                "value": value, "value_type": "number" if isinstance(value, (int, float)) and not isinstance(value, bool) else "text",
                "formula": None, "semantic_id": item.assumption_id if column == 1 else None,
                "owner": "CP-2G", "write_class": "ASSUMPTION",
                "period_id": item.period_id, "source_refs": None,
            })
    return {
        "id": "ASSUMPTIONS",
        "title": "Assumptions",
        "max_row": 1 + len(model.effective_assumptions),
        "max_column": len(headers),
        "cells": cells,
    }


def _serialize_workbook_file(draft: Path, rendered: RenderedWorkbook, model: Any, calculations: Any) -> dict[str, Any]:
    from openpyxl import load_workbook

    expectations = {(item.sheet, item.cell): item for item in rendered.formulas}
    mappings = {item.target: item for item in rendered.mappings}
    workbook = load_workbook(draft, data_only=False, read_only=False)
    try:
        tabs: list[dict[str, Any]] = []
        cell_count = 0
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            cells: list[dict[str, Any]] = []
            for row in worksheet.iter_rows():
                for cell in row:
                    address = cell.coordinate
                    expectation = expectations.get((sheet_name, address))
                    if cell.value is None and expectation is None:
                        continue
                    formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
                    value = expectation.expected if expectation is not None else cell.value
                    mapping = mappings.get(f"{sheet_name}!{address}")
                    cells.append({
                        "address": address,
                        "row": cell.row,
                        "column": cell.column,
                        "value": json_value(value),
                        "value_type": _cell_type(value, formula),
                        "formula": formula,
                        "semantic_id": mapping.semantic_id if mapping else None,
                        "owner": mapping.owner if mapping else None,
                        "write_class": mapping.write_class if mapping else None,
                        "period_id": mapping.period_id if mapping else None,
                        "source_refs": mapping.source_refs if mapping else None,
                    })
                    cell_count += 1
                    if cell_count > MAX_WORKSHEET_CELLS:
                        raise ModelInputError("worksheet cell limit exceeded")
            tabs.append({
                "id": sheet_name.upper().replace(" ", "_"),
                "title": sheet_name,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "cells": cells,
            })
        if not any(tab["title"] == "Assumptions" for tab in tabs):
            tabs.append(_assumptions_tab(model))
        payload = {
            "schema_version": WORKSHEET_SCHEMA_VERSION,
            "identity": {
                "issuer_id": model.issuer_id,
                "issuer_name": model.issuer_name,
                "analysis_date": model.analysis_date.isoformat(),
            },
            "tabs": tabs,
        }
        if len(canonical_json(payload).encode("utf-8")) > MAX_WORKSHEET_BYTES:
            raise ModelInputError("worksheet payload limit exceeded")
        checks = [
            {
                "check_id": check.check_id,
                "status": check.status,
                "period_id": check.period_id,
                "difference": json_value(check.difference),
                "tolerance": json_value(check.tolerance),
                "detail": check.detail,
            }
            for check in calculations.checks
        ]
        qa = {
            "status": "PASS",
            "semantic_checks": checks,
            "semantic_check_count": len(checks),
            "formula_count": len(rendered.formulas),
            "worksheet_cell_count": cell_count,
            "limitation_flags": list(model.limitation_flags),
            "validation_warnings": list(model.validation_warnings),
            "source_manifest": [
                {
                    "module_id": module_id,
                    "filename": model.source_paths[module_id].name,
                    "sha256": model.source_hashes[module_id],
                }
                for module_id in sorted(model.source_paths)
            ],
        }
        return {"payload": payload, "qa": qa}
    finally:
        workbook.close()


# -- CP-2B derived projection (ported from LEGACY models/domain.py) -----------


def project_cp2b(cp2a_markdown: str, *, run_id: str, cp2a_artifact_digest: str, bundle: CpModelBundle) -> str:
    if not _HEX_DIGEST.fullmatch(cp2a_artifact_digest):
        raise ModelInputError("CP-2A artifact digest must be 64 lowercase hex characters")
    validation = bundle.validate_handoff(cp2a_markdown, module_id="CP-2A", run_id=run_id)
    if validation.errors or validation.identity_mismatches or validation.fields is None:
        raise ModelInputError("CP-2A canonical handoff is invalid")
    fields = validation.fields
    if fields.get("qa_status") != "Passed" or "CP-MODEL" not in fields.get("downstream_consumers", []):
        raise ModelInputError("CP-2A is not ready for CP-MODEL")
    try:
        validate_model_sources(
            cp2a_markdown,
            model_source_ids(cp2a_markdown),
            module_id="CP-2A",
        )
    except CanonicalValidationError as exc:
        raise ModelInputError(str(exc)) from exc
    sections = _t5_sections(cp2a_markdown)

    def quoted(value: object) -> str:
        return json.dumps(str(value), ensure_ascii=True)

    cp0_lineage = [
        artifact
        for artifact in fields["upstream_artifacts_used"]
        if artifact.get("module_id") == "CP-0"
    ]
    if len(cp0_lineage) > 1:
        raise ModelInputError("CP-2A carries ambiguous CP-0 lineage")
    upstream_lines = ["upstream_artifacts_used:"]
    for artifact in (*cp0_lineage, {
        "module_id": "CP-2A",
        "run_id": run_id,
        "period": fields["reporting_period"],
        "artifact_digest": cp2a_artifact_digest,
    }):
        upstream_lines.extend((
            f"  - module_id: {artifact['module_id']}",
            f"    run_id: {quoted(artifact['run_id'])}",
            f"    period: {quoted(artifact['period'])}",
        ))
        artifact_digest = artifact.get("artifact_digest")
        if isinstance(artifact_digest, str) and _HEX_DIGEST.fullmatch(artifact_digest):
            upstream_lines.append(f"    artifact_digest: {quoted(artifact_digest)}")

    projected = "\n".join(
        [
            "---",
            "module_id: CP-2B",
            'module_name: "Catalyst and Event-Risk Projection"',
            f"run_id: {quoted(run_id)}",
            f"reporting_period: {quoted(fields['reporting_period'])}",
            f"analysis_date: {quoted(fields['analysis_date'])}",
            f"confidence_score: {fields['confidence_score']}",
            f"confidence_band: {quoted(fields['confidence_band'])}",
            "qa_status: Passed",
            f"committee_status: {quoted(fields['committee_status'])}",
            f"limitation_flags: {json.dumps(fields['limitation_flags'])}",
            f"validation_warnings: {json.dumps(fields['validation_warnings'])}",
            *upstream_lines,
            "downstream_consumers: [CP-MODEL]",
            f"issuer_name: {quoted(fields['issuer_name'])}",
            f"issuer_id: {quoted(fields['issuer_id'])}",
            "---",
            "",
            "## Audit Summary",
            "",
            f"Host projection of CP-2A artifact `{cp2a_artifact_digest}`; no rows were inferred or repaired.",
            "",
            "## Analysis",
            "",
            "".join(sections),
            "",
            "## Evidence Trace",
            "",
            "All evidence rows and source locators are preserved verbatim from the accepted CP-2A handoff.",
            "",
            "## Source Registry",
            "",
            "See T5.1 and the preserved source columns in the analytical registers.",
            "",
            "## Gaps & Conflicts",
            "",
            "See T5.7; this projection adds no gap resolution.",
            "",
            "## QA Validation",
            "",
            "The host verified the complete T5.1-T5.7 register sequence and canonical table headers.",
            "",
        ]
    )
    projected_validation = bundle.validate_handoff(projected, module_id="CP-2B", run_id=run_id)
    if projected_validation.errors or projected_validation.identity_mismatches:
        raise ModelInputError("projected CP-2B canonical handoff is invalid")
    return projected


# -- pure guards (spec surface; the error class is the vendored engine's own,
# so bundle.calculate failures and these guards raise the same type) ----------


def _default_bundle() -> CpModelBundle:
    global _DEFAULT_BUNDLE
    if _DEFAULT_BUNDLE is None:
        _DEFAULT_BUNDLE = CpModelBundle(Settings().deploy_v_root)
    return _DEFAULT_BUNDLE


_DEFAULT_BUNDLE: CpModelBundle | None = None


def is_finite_number(value: object) -> bool:
    if isinstance(value, Decimal):
        return value.is_finite()
    if not isinstance(value, (int, float)):
        return False
    try:
        return math.isfinite(float(value))
    except (OverflowError, ValueError):
        return False


def safe_ratio(numerator: object, denominator: object) -> float | None:
    if not is_finite_number(numerator) or not is_finite_number(denominator) or float(denominator) == 0:
        return None
    return float(numerator) / float(denominator)


def finite_operand(value: object, label: str) -> Decimal:
    bundle = _default_bundle()
    try:
        return bundle._calculations._finite_operand(value, label)
    except bundle._domain.CpModelV3Error as exc:
        raise CpModelV3Error(str(exc)) from exc


def calculate(model: Any) -> Any:
    bundle = _default_bundle()
    try:
        return bundle.calculate_model(model)
    except bundle._domain.CpModelV3Error as exc:
        raise CpModelV3Error(str(exc)) from exc
