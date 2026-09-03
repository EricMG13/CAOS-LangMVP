"""Model Builder service: readiness, build queue, previews, Sign-Off, scenarios,
sensitivities, rebase, and hash-verified exports (DECISIONS §§1, 10.6, 12).

Rewritten against the framework store — nothing from LEGACY models/runtime.py
or revisions.py machinery (leases, heartbeats, fences) ports; the calculation
authority is caos.models.engine over the pinned Deploy V bundle. Sign-Off stays
a store CAS transaction, never an interrupt (§2). Calculation runs in-process
under the legacy 30s aggregate request deadline — a recorded deviation from
§10.8's process isolation (DECISIONS entry 2026-08-27): golden-input
calculations complete in well under a second and nothing here executes
untrusted code.
"""

from __future__ import annotations

import copy
import json
import re
from contextlib import contextmanager
from contextvars import ContextVar
import tempfile
import threading
import time
from datetime import datetime
from decimal import Decimal, InvalidOperation
from functools import wraps
from pathlib import Path
from typing import Any, Callable

from ..contracts import (
    validate_boundary_text,
    ModelPreviewRequest,
    ModelRebasePreviewRequest,
    ModelScenarioRequest,
    ModelSignOffRequest,
    ModelTornadoRequest,
    OneWaySensitivityRequest,
    clean_json,
    digest,
)
from ..engine.provider import AgentError
from ..engine.state import source_set_digest
from ..atomic_files import (
    MAX_EXPORT_BYTES,
    VaultFileIntegrityError,
    VaultFileUnavailable,
    publish_hash_addressed_bytes,
    read_verified_vault_bytes,
)
from ..artifacts.loan_universe import universe_digest
from ..modules.registry import CP_MODEL_INPUT_MODULES
from ..methodology.canonical import model_facing_source_ids
from ..methodology.execution import MAX_CALCULATION_NODES, calculation_output_complete
from ..storage.models import ModelStore
from ..storage.store import DomainStore, now_iso
from .engine import CpModelBundle, ModelInputError, json_value, project_cp2b

# Per-request memo of validated snapshot artifacts (F23): set by the public
# entry points and by _resolve_snapshot when no scope is active.
_RESOLUTION_MEMO: ContextVar[dict[tuple, Any] | None] = ContextVar("caos_model_resolution_memo", default=None)


@contextmanager
def _resolution_scope():
    if _RESOLUTION_MEMO.get() is not None:
        yield
        return
    token = _RESOLUTION_MEMO.set({})
    try:
        yield
    finally:
        _RESOLUTION_MEMO.reset(token)

CANONICAL_MODULES = CP_MODEL_INPUT_MODULES
DISTRESSED_PATHWAY = "DISTRESSED_RESTRUCTURING"
DISTRESSED_CALCULATORS = frozenset({"funding_gap", "recovery_waterfall"})
PRIOR_FULL_CREDIT_PUBLICATION_PATHWAYS = frozenset({
    "EARNINGS_UPDATE",
    "COVENANT_REFINANCING",
})
# Task 9 (DECISIONS §14.18): every pathway declares one model effect. Full
# Credit builds the complete model; every other pathway overlays the nearest
# validated Full Credit model with one effect and its own input fingerprint.
PATHWAY_EFFECT_SCHEMA = "caos.model-pathway-effect.v1"
OVERLAY_PATHWAYS = frozenset({
    "EARNINGS_UPDATE", "COVENANT_REFINANCING", "RELATIVE_VALUE", DISTRESSED_PATHWAY, "DEEP_RESEARCH",
})
MODEL_EFFECT_PATHWAYS = frozenset({"FULL_CREDIT", *OVERLAY_PATHWAYS})
EFFECT_IDS = {
    "EARNINGS_UPDATE": "EARNINGS_PERIOD_FORECAST_VARIANCE",
    "COVENANT_REFINANCING": "COVENANT_REFINANCING_ASSUMPTIONS",
    "RELATIVE_VALUE": "RELATIVE_VALUE_MARKET_MARKS",
    DISTRESSED_PATHWAY: "DISTRESSED_SCENARIO_RECOVERY",
    "DEEP_RESEARCH": "DEEP_RESEARCH_REVALIDATION",
}
# The modules an effect consumes. An accepted route without them (screen
# depth) is a depth precondition, reported as such — never as input corruption.
EFFECT_MODULES = {
    "FULL_CREDIT": CANONICAL_MODULES,
    "EARNINGS_UPDATE": ("CP-1", "CP-1B"),
    "COVENANT_REFINANCING": ("CP-4", "CP-4C"),
    "RELATIVE_VALUE": ("CP-3",),
    DISTRESSED_PATHWAY: ("CP-4C",),
    "DEEP_RESEARCH": ("CP-DR",),
}
BASE_MODEL_CODES = {
    DISTRESSED_PATHWAY: "DISTRESSED_BASE_MODEL_REQUIRED",
    "DEEP_RESEARCH": "DEEP_RESEARCH_NO_NUMERIC_EFFECT",
}
DEFAULT_BASE_MODEL_CODE = "PRIOR_FULL_CREDIT_MODEL_REQUIRED"
# Document classes whose `used` disposition must reach a model input, a pinned
# market-marks workbook, the bound brief or the cited analysis before a build
# may read READY (ETR-B12). Anything else that is used but uncited is recorded
# as NOT_REQUIRED with its reason — explicit, never silently discarded.
RELEVANT_DOCUMENT_TYPES = frozenset({
    "annual_report", "quarterly_report", "earnings_release", "management_guidance",
    "credit_agreement", "amendment", "restructuring", "market_marks", "research_brief",
})
# Reported `credit_metrics` inputs and the base-model output each is compared with.
VARIANCE_METRICS = (
    ("revenue", "revenue"),
    ("adjusted_ebitda", "adjusted_ebitda_calc"),
    ("total_debt", "total_debt_reported"),
    ("cash_and_equivalents", "cash_and_equivalents"),
)
MARKET_MARK_FIELDS = (
    "instrument_key", "company", "borrower_name", "sector", "sub_sector", "ranking",
    "size_mn", "margin_bps", "maturity_date", "bid_points", "ask_points",
    "mid_ytm_pct", "mid_3y_dm_bps",
)
MAX_LINEAGE_DETAIL_IDS = 20
_ANNUAL_PERIOD = re.compile(r"^FY(\d{4})$")
# ponytail: bounds a corrupted prior-snapshot chain; raise only if a legitimate
# case ever approaches one thousand accepted analyses.
MAX_SNAPSHOT_ANCESTORS = 1_000
MAX_EFFECTIVE_ASSUMPTIONS = 256
MAX_SENSITIVITY_POINTS = 41
MAX_CALCULATION_SECONDS = 30.0
MAX_ERROR_DETAIL_CHARS = 2_000
MAX_PATHWAY_EFFECT_EXPORT_ROWS = (
    MAX_CALCULATION_NODES * len(DISTRESSED_CALCULATORS) * 2 + 1_024
)
MAX_EXCEL_AUDIT_CELL_CHARS = 32_000
LEGACY_TORNADO_DRIVERS = (
    (("operating.consolidated_revenue_growth", "operating.revenue_growth.division_1"), "Revenue growth", "0.025"),
    (("operating.adjusted_ebitda_margin",), "EBITDA margin", "0.015"),
    (("rates.base_rate",), "Interest rate", "0.01"),
    (("cash_flow.capex_pct_revenue",), "Capex % revenue", "0.01"),
)
# ponytail: the per-build caches hold a whole resolved snapshot (six canonical
# documents plus the assumption rows), and this service lives as long as the
# process. Work is newest-build-first, so a handful of entries covers every
# real access pattern; without a bound the API and the worker retain every
# build they ever resolved. Raise it if a workload proves it needs to.
MAX_CACHED_BUILDS = 8
BUILD_ERROR_CODES = {
    "MODEL_AUTHORITY_CHANGED",
    "MODEL_EXPORT_AUTHORITY_CHANGED",
    "MODEL_INPUT_INVALID",
    "MODEL_CALCULATION_FAILED",
    "MODEL_EXPORT_FAILED",
    "MODEL_REVISION_EXPORT_FAILED",
    "MODEL_REVISION_EXPORT_RUNTIME_UNAVAILABLE",
}


class ModelBuildStale(ValueError):
    """MODEL_BUILD_STALE — the refusal names the current build authority."""

    def __init__(self, current_build: dict[str, Any] | None) -> None:
        self.current_build = current_build
        super().__init__("MODEL_BUILD_STALE")


class BuildAuthorityChanged(ValueError):
    """A pinned source was withdrawn between validation and publication."""


class BuildIdentityChanged(ValueError):
    """The build row no longer carries the identity this result was computed
    from — a re-point requeued it, or another executor already published."""


class ModelCalculationTimeout(ValueError):
    def __init__(self) -> None:
        super().__init__("MODEL_CALCULATION_TIMEOUT")


# Model Builder renders each readiness blocker as humanizeCode(code) over
# `detail`; a null detail draws a heading above a blank line. These are written
# here rather than taken from the exception because ModelInputError messages
# interpolate host internals (bundle paths, table ids).
NO_ACCEPTED_AUTHORITY_DETAIL = "Accept a completed Full Credit run before building a model."
WRONG_PATHWAY_AUTHORITY_DETAIL = (
    "This case's accepted authority is not a Full Credit run. "
    "Accept a completed Full Credit run before building a model."
)
INPUTS_INVALID_DETAIL = (
    "The accepted Full Credit artifacts are missing, superseded, or fail "
    "CP-MODEL validation. Re-run Full Credit and accept it again."
)
DISTRESSED_BASE_MODEL_DETAIL = (
    "Accept and complete a Full Credit model before building a Distressed "
    "Restructuring overlay."
)
LINEAGE_INCOMPLETE_DETAIL = (
    "One or more admitted documents reached no model input, calculation or "
    "cited analysis in the accepted run, so the model is not source-complete."
)
# Readiness renders these as NOT_READY with the code as the blocker: each is a
# precondition the analyst can act on, never input corruption.
PRECONDITION_DETAILS = {
    "ACCEPTED_FULL_CREDIT_REQUIRED": WRONG_PATHWAY_AUTHORITY_DETAIL,
    "DISTRESSED_BASE_MODEL_REQUIRED": DISTRESSED_BASE_MODEL_DETAIL,
    "PRIOR_FULL_CREDIT_MODEL_REQUIRED": (
        "This pathway's model effect overlays a Full Credit model. Accept and "
        "complete a Full Credit model in this case first."
    ),
    "DEEP_RESEARCH_NO_NUMERIC_EFFECT": (
        "Deep Research declares no numeric model effect, and this case has no "
        "Full Credit model to revalidate."
    ),
    "FULL_DEPTH_REQUIRED": (
        "The accepted run is a screen-depth route without the modules the "
        "model effect depends on. Run the pathway at full depth and accept it."
    ),
    "RELATIVE_VALUE_MARKET_MARKS_REQUIRED": (
        "The accepted Relative Value run pinned no market-marks workbook. "
        "Supply the CP-3 loan-universe workbook with the documents and run "
        "Relative Value again."
    ),
    "MODEL_SOURCE_LINEAGE_INCOMPLETE": LINEAGE_INCOMPLETE_DETAIL,
}


def _is_sha256(value: Any) -> bool:
    return (
        isinstance(value, str)
        and len(value) == 64
        and all(character in "0123456789abcdef" for character in value)
    )


def _is_aware_iso_timestamp(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).tzinfo is not None
    except ValueError:
        return False


def _export_identity_digest(
    case_id: str,
    target_id: str,
    export: dict[str, Any],
) -> str:
    return digest({
        "schema_version": "caos.model-export.v1",
        "case_id": case_id,
        "target_id": target_id,
        "vault_key": export.get("vault_key"),
        "filename": export.get("filename"),
        "sha256": export.get("sha256"),
        "size": export.get("size"),
    })


def _audit_rows(root: str, value: Any):
    """Flatten validated JSON into deterministic, formula-safe audit rows."""
    stack: list[tuple[str, Any]] = [(root, value)]
    row_count = 0
    while stack:
        path, value = stack.pop()
        if isinstance(value, dict) and value:
            stack.extend(
                (f"{path}.{key}", value[key])
                for key in reversed(sorted(value))
            )
            continue
        if isinstance(value, list) and value:
            stack.extend(
                (f"{path}.{index}", item)
                for index, item in reversed(list(enumerate(value)))
            )
            continue
        rendered = (
            value
            if isinstance(value, (int, float)) and not isinstance(value, bool)
            else json.dumps(value, ensure_ascii=False)
        )
        chunks = (
            [rendered]
            if not isinstance(rendered, str) or len(rendered) <= MAX_EXCEL_AUDIT_CELL_CHARS
            else [
                rendered[index:index + MAX_EXCEL_AUDIT_CELL_CHARS]
                for index in range(0, len(rendered), MAX_EXCEL_AUDIT_CELL_CHARS)
            ]
        )
        for index, chunk in enumerate(chunks, 1):
            row_count += 1
            if row_count > MAX_PATHWAY_EFFECT_EXPORT_ROWS:
                raise ModelInputError(f"{root} audit record is too large to export")
            chunk_path = path if len(chunks) == 1 else f"{path}.chunk.{index:04d}"
            yield chunk_path, chunk


def _audit_tab(tab_id: str, title: str, root: str, value: Any) -> dict[str, Any]:
    rows = [("Field", "Value"), *_audit_rows(root, value)]
    cells = []
    for row, values in enumerate(rows, 1):
        for column, cell_value in enumerate(values, 1):
            cells.append({
                "address": f"R{row}C{column}",
                "row": row,
                "column": column,
                "value": cell_value,
                "value_type": (
                    "number"
                    if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool)
                    else "text"
                ),
                "formula": None,
                "semantic_id": None,
                "owner": "CAOS",
                "write_class": "AUDIT",
                "period_id": None,
                "source_refs": None,
            })
    return {
        "id": tab_id,
        "title": title,
        "max_row": len(rows),
        "max_column": 2,
        "cells": cells,
    }


def _audit_tabs(payload: dict[str, Any] | None) -> list[dict[str, Any]]:
    """The audit worksheets a build's payload earns: the pathway effect when the
    build is an overlay, and the source lineage on every build."""
    tabs: list[dict[str, Any]] = []
    effects = (payload or {}).get("pathway_effects") or []
    if effects:
        tabs.append(_audit_tab("PATHWAY_EFFECTS", "Pathway Effects", "pathway_effects", effects))
    lineage = (payload or {}).get("source_lineage") or []
    if lineage:
        tabs.append(_audit_tab("SOURCE_LINEAGE", "Source Lineage", "source_lineage", lineage))
    return tabs


def _with_audit_tabs(worksheet: dict[str, Any], payload: dict[str, Any] | None) -> dict[str, Any]:
    tabs = _audit_tabs(payload)
    if not tabs:
        return worksheet
    result = copy.deepcopy(worksheet)
    titles = {tab.get("title") for tab in result.get("tabs", [])}
    for tab in tabs:
        if tab["title"] in titles:
            raise ModelInputError(f"duplicate {tab['title']} worksheet")
        result["tabs"].append(tab)
    return result


def _table_rows(markdown: str, table_id: str) -> list[dict[str, str]]:
    """Rows of one vendor-validated stable table, keyed by its header names.
    Bounded by the document the vendor already validated; a missing table is
    an empty list, never a guess."""
    marker = f"<!-- table-id: {table_id} -->"
    start = markdown.find(marker)
    if start < 0:
        return []
    lines = markdown[start + len(marker):].splitlines()
    header: list[str] | None = None
    rows: list[dict[str, str]] = []
    for line in lines:
        if not line.lstrip().startswith("|"):
            if header is not None:
                break
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if header is None:
            header = cells
            continue
        if all(set(cell.replace("-", "").replace(":", "")) == set() for cell in cells):
            continue
        if len(cells) == len(header):
            rows.append(dict(zip(header, cells)))
    return rows


def _decimal_text(value: Any) -> str | None:
    """A finite JSON number as its exact decimal text; None stays None and a
    non-finite operand is refused before use (invariant 7)."""
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        raise ModelInputError("non-numeric model effect operand")
    number = value if isinstance(value, Decimal) else Decimal(str(value))
    if not number.is_finite():
        raise ModelInputError("non-finite model effect operand")
    # Calculator outputs arrive as JSON floats (1000.0); the exact value, not
    # the float's spelling, is the record.
    return format(number.normalize(), "f")


def _materialize_workbook_tabs(path: Path, tabs: list[dict[str, Any]]) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    workbook = load_workbook(path, data_only=False)
    try:
        changed = False
        for tab in tabs:
            title = tab["title"]
            if title in workbook.sheetnames:
                continue
            sheet = workbook.create_sheet(title)
            for item in tab["cells"]:
                cell = sheet.cell(item["row"], item["column"])
                formula = item.get("formula")
                value = formula if isinstance(formula, str) else item.get("value")
                cell.value = value
                if isinstance(value, str) and formula is None:
                    cell.data_type = "s"
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            changed = True
        if changed:
            workbook.save(path)
    finally:
        workbook.close()


def _assert_workbook_semantics(path: Path, expected: dict[str, Any]) -> None:
    from openpyxl import load_workbook

    def normalized(value: Any) -> Any:
        return None if value in (None, "") else json_value(value)

    workbook = load_workbook(path, data_only=False)
    try:
        expected_titles = [tab["title"] for tab in expected["tabs"]]
        if workbook.sheetnames != expected_titles:
            raise ModelInputError("signed worksheet sheet set changed during export")
        for tab in expected["tabs"]:
            expected_cells = {
                (item["row"], item["column"]): normalized(
                    item["formula"] if isinstance(item.get("formula"), str) else item.get("value")
                )
                for item in tab["cells"]
                if item.get("formula") not in (None, "") or item.get("value") not in (None, "")
            }
            actual_cells = {
                (cell.row, cell.column): normalized(cell.value)
                for row in workbook[tab["title"]].iter_rows()
                for cell in row
                if cell.value not in (None, "")
            }
            if actual_cells != expected_cells:
                raise ModelInputError(f"signed worksheet changed during export: {tab['title']}")
    finally:
        workbook.close()


def _authority_guarded(method):
    @wraps(method)
    def guarded(self, *args, **kwargs):
        # The process-wide authority lock covers mutations only (queue, sign-off):
        # it linearizes them against ingest, withdrawal and run finalization
        # until Phase 5 replaces it with database locks. Transient reads run
        # unguarded and publication is bound by a compare-and-swap instead,
        # because the worker process never shares this lock (DECISIONS §14 D9).
        with self.store.authority_guard():
            return method(self, *args, **kwargs)

    return guarded


class ModelService:
    def __init__(self, *, store: DomainStore, vault_dir: Path, engine: Any) -> None:
        self.store = store
        self.vault_dir = Path(vault_dir)
        self.engine = engine
        self.bundle = CpModelBundle(
            engine.settings.deploy_v_root,
            integrity=engine.bundle.integrity,
        )
        self.builds = ModelStore(store.engine)
        self._sign_off_lock = threading.Lock()
        # test seams
        self._dispatch_log: list[str] = []
        self._on_queue: list[Callable[[str], None]] = []
        self._fail_next_queue = False
        self._fail_next_dispatch = False
        self._fail_next_export = False
        self._forced_readiness: dict[str, str] = {}
        self._validate_bundle_calls: list[dict[str, Any]] = []
        self._runtime_sha_override: str | None = None
        self._deadline_exceeded = False
        self._export_input_reads = 0
        self._evolutions: dict[str, dict[str, tuple[str, ...]]] = {}
        self._resolved_cache: dict[str, dict[str, Any]] = {}
        self._defaults_cache: dict[str, tuple[list[dict[str, Any]], dict[str, Any]]] = {}
        engine.register_model_service(self)

    # -- engine hooks -------------------------------------------------------

    def on_accepted(self, run: dict[str, Any], actor: str) -> None:
        """Accept-time auto-queue. Acceptance is already durable; any failure
        here (readiness, queue, dispatch) leaves it standing."""
        if run.get("pathway") not in MODEL_EFFECT_PATHWAYS:
            return
        self.queue_build(run["case_id"], actor)

    def active_job_count(self) -> int:
        """Model jobs share the run engine's MAX_ACTIVE_JOBS admission budget."""
        return self.builds.active_build_count()

    def _current_runtime(self) -> dict[str, Any]:
        runtime = dict(self.bundle.calculation_runtime)
        if self._runtime_sha_override is not None:
            runtime["sha256"] = self._runtime_sha_override
        return runtime

    # -- readiness and resolve ---------------------------------------------

    def _accepted_snapshot(self, case_id: str) -> dict[str, Any] | None:
        case = self.store.get_case(case_id)
        if case is None or not case.get("accepted_snapshot_id"):
            return None
        return self.engine.runs.get_snapshot(case["accepted_snapshot_id"])

    def _resolve_snapshot(
        self, snapshot: dict[str, Any], *, deadline: float | None = None,
    ) -> dict[str, Any]:
        """Resolve the accepted pathway's declared model effect: the complete
        Full Credit model, or one overlay on the nearest validated Full Credit
        model (DECISIONS §14.18)."""
        with _resolution_scope():
            run = self._validated_snapshot_run(snapshot)
            if run["pathway"] not in MODEL_EFFECT_PATHWAYS:
                raise ModelInputError(
                    "accepted pathway declares no model effect",
                    code="ACCEPTED_FULL_CREDIT_REQUIRED",
                )
            self._require_effect_modules(run)
            if run["pathway"] == "FULL_CREDIT":
                return self._resolve_full_credit_snapshot(snapshot, run)
            return self._resolve_overlay_snapshot(snapshot, run, deadline=deadline)

    @staticmethod
    def _require_effect_modules(run: dict[str, Any]) -> None:
        planned = {
            node.get("module_id")
            for node in (run.get("plan") or {}).get("nodes") or []
            if isinstance(node, dict)
        }
        if not set(EFFECT_MODULES[run["pathway"]]) <= planned:
            raise ModelInputError(
                "accepted route lacks the modules the model effect depends on",
                code="FULL_DEPTH_REQUIRED",
            )

    def _resolve_full_credit_snapshot(
        self,
        snapshot: dict[str, Any],
        run: dict[str, Any],
    ) -> dict[str, Any]:
        """§10.6 Full Credit input resolution and vendor validation."""
        if run["pathway"] != "FULL_CREDIT":
            raise ModelInputError("accepted FULL_CREDIT run required",
                                  code="ACCEPTED_FULL_CREDIT_REQUIRED")
        artifacts, _inventory = self._validated_snapshot_artifacts(snapshot, run)
        by_module = {
            artifact["module_id"]: artifact
            for artifact in artifacts
            if artifact["module_id"] in CANONICAL_MODULES
        }
        if set(by_module) != set(CANONICAL_MODULES):
            raise ModelInputError("canonical artifacts are incomplete")
        markdown = {module_id: by_module[module_id]["markdown"] for module_id in ("CP-1", "CP-1A", "CP-1B", "CP-2", "CP-2G")}
        if any(text is None for text in markdown.values()) or by_module["CP-2A"]["markdown"] is None:
            raise ModelInputError("canonical markdown is unavailable")
        # CP-2B is CP-2A's registry-declared derived projection, recomputed
        # byte-identically from the pinned artifact (§12.7, MODULE_GRANULARITY).
        markdown["CP-2B"] = project_cp2b(
            by_module["CP-2A"]["markdown"],
            run_id=snapshot["run_id"],
            cp2a_artifact_digest=by_module["CP-2A"]["digest"],
            bundle=self.bundle,
        )
        validation = self.bundle.validate(
            markdown["CP-1"], markdown["CP-1A"], markdown["CP-1B"],
            markdown["CP-2"], markdown["CP-2B"], markdown["CP-2G"],
        )
        self._validate_bundle_calls.append({"case_id": snapshot["case_id"], "snapshot_id": snapshot["id"]})
        if validation.errors:
            raise ModelInputError("CP-MODEL bundle validation failed")
        source_set = self.store.source_set(snapshot["source_set_id"])
        if source_set is None or source_set.get("case_id") != snapshot["case_id"]:
            raise ModelInputError("accepted source set is unavailable")
        inventory = [
            {"module_id": module_id, "artifact_id": by_module[module_id]["id"], "digest": by_module[module_id]["digest"], "status": "READY"}
            for module_id in CANONICAL_MODULES
        ]
        inventory.append({
            "module_id": "CP-2B",
            "artifact_id": by_module["CP-2A"]["id"],
            "digest": digest(markdown["CP-2B"]),
            "derived_from": by_module["CP-2A"]["digest"],
            "status": "READY",
        })
        model_markdown = {**markdown, "CP-2A": by_module["CP-2A"]["markdown"]}
        lineage = self._source_lineage(snapshot, run, artifacts, model_markdown=model_markdown)
        fingerprint = digest({
            "case_id": snapshot["case_id"],
            "accepted_run_id": snapshot["run_id"],
            "accepted_snapshot_id": snapshot["id"],
            "accepted_snapshot_digest": snapshot.get("digest"),
            "source_set": {"id": source_set["id"], "version": source_set["version"]},
            "artifacts": [{"module_id": item["module_id"], "digest": item["digest"]} for item in inventory],
            "source_lineage_digest": digest(lineage),
            "methodology_build_id": self.engine.bundle.build_id,
            "calculation_runtime": self.bundle.calculation_runtime,
        })
        return {
            "snapshot": snapshot,
            "source_set": {"id": source_set["id"], "version": source_set["version"]},
            "artifact_inventory": inventory,
            "markdown": markdown,
            "model_markdown": model_markdown,
            "input_fingerprint": fingerprint,
            "source_lineage": lineage,
        }

    def _validated_snapshot_run(
        self,
        snapshot: dict[str, Any],
        *,
        error_code: str = "CANONICAL_MODEL_INPUTS_INVALID",
        require_live_sources: bool = True,
    ) -> dict[str, Any]:
        """Validate one identity-bound chain node before following its prior pointer.

        Liveness of the pinned sources is required for the snapshot whose
        artifacts are consumed; an intermediate ancestor that merely links the
        chain is validated by identity only, so withdrawing a document that
        only it pinned does not revoke an intact base."""
        if not isinstance(snapshot, dict) or not isinstance(snapshot.get("id"), str):
            raise ModelInputError(
                "Accepted snapshot is unavailable",
                code=error_code,
            )
        preimage = {
            key: value
            for key, value in snapshot.items()
            if key not in {"digest", "id"}
        }
        run = self.engine.runs.get_run(snapshot.get("run_id"))
        plan = run.get("plan") if run is not None else None
        if (
            run is None
            or run.get("case_id") != snapshot.get("case_id")
            or run.get("status") != "succeeded"
            or run.get("accepted_snapshot_id") != snapshot["id"]
            or not _is_aware_iso_timestamp(snapshot.get("accepted_at"))
            or run.get("provider_identity") != snapshot.get("provider_identity")
            or not isinstance(plan, dict)
            or not _is_sha256(run.get("plan_digest"))
            or digest(plan) != run["plan_digest"]
            or plan.get("source_set_id") != snapshot.get("source_set_id")
            or plan.get("source_set_version") != snapshot.get("source_set_version")
            or not _is_sha256(snapshot.get("digest"))
            or digest(preimage) != snapshot["digest"]
        ):
            raise ModelInputError(
                "Accepted snapshot identity is invalid",
                code=error_code,
            )
        source_set = self.store.source_set(snapshot["source_set_id"])
        if (
            source_set is None
            or source_set.get("case_id") != snapshot["case_id"]
            or source_set.get("version") != snapshot["source_set_version"]
            or source_set_digest(source_set) != plan.get("source_set_digest")
        ):
            raise ModelInputError(
                "Accepted snapshot source authority is unavailable",
                code=error_code,
            )
        if require_live_sources:
            live = self.store.sources_for_live_set(
                snapshot["case_id"],
                snapshot["source_set_id"],
                snapshot["source_set_version"],
            )
            if live is None:
                raise ModelInputError(
                    "Accepted snapshot source authority is unavailable",
                    code=error_code,
                )
            memo = _RESOLUTION_MEMO.get()
            if memo is not None:
                memo[("live", snapshot["id"])] = live
        return run

    def _live_snapshot_sources(self, snapshot: dict[str, Any]) -> list[dict[str, Any]]:
        memo = _RESOLUTION_MEMO.get()
        cached = memo.get(("live", snapshot["id"])) if memo is not None else None
        if cached is not None:
            return cached
        live = self.store.sources_for_live_set(
            snapshot["case_id"], snapshot["source_set_id"], snapshot["source_set_version"],
        )
        if live is None:
            raise ModelInputError("Accepted snapshot source authority is unavailable")
        return live

    def _source_lineage(
        self,
        snapshot: dict[str, Any],
        run: dict[str, Any],
        artifacts: list[dict[str, Any]],
        *,
        model_markdown: dict[str, str] | None,
        market_marks_source_id: str | None = None,
    ) -> list[dict[str, Any]]:
        """One row per pinned source: the intake disposition and reason, the
        expected consumers, the artifacts that cite it, whether a model-facing
        table names it, and the binding that follows. A `used` relevant document
        bound to nothing makes the model incomplete (typed, never READY)."""
        manifest: dict[str, dict[str, Any]] = {}
        # Every intake the case admitted, oldest first: a later intake's row for
        # a source supersedes an earlier one. A duplicate row (the same bytes
        # dropped again) never speaks for the source: when it is the only row —
        # the bytes were admitted through the source route — it lends its
        # classification, and the source stays `used`, so a relevant document
        # cannot be waved through as "not required" by re-dropping it.
        for intake in self.store.intakes_for_case(snapshot["case_id"]):
            for row in ((intake.get("record") or {}).get("documents") or []):
                source_id = row.get("source_id") if isinstance(row, dict) else None
                if not isinstance(source_id, str):
                    continue
                if row.get("disposition") == "duplicate":
                    if source_id not in manifest:
                        manifest[source_id] = {
                            **row, "disposition": "used",
                            "reason": "admitted through the source route; classified when the same bytes were dropped again",
                        }
                    continue
                manifest[source_id] = row
        cited: dict[str, set[str]] = {}
        for artifact in artifacts:
            for ref in (artifact.get("payload") or {}).get("evidence_refs") or []:
                source_id = ref.get("source_id") if isinstance(ref, dict) else ref
                if isinstance(source_id, str):
                    cited.setdefault(source_id, set()).add(artifact["module_id"])
        model_inputs: set[str] = set()
        for module_id, text in (model_markdown or {}).items():
            if isinstance(text, str):
                model_inputs |= model_facing_source_ids(text, module_id=module_id)
        pinned_universe = ((run.get("plan") or {}).get("loan_universe") or {}).get("source_id")
        brief_bound = run.get("research") is not None
        rows: list[dict[str, Any]] = []
        for source in self._live_snapshot_sources(snapshot):
            source_id = source["id"]
            row = manifest.get(source_id)
            document_type = str(row.get("document_type") or "unclassified") if row else "unclassified"
            disposition = str(row.get("disposition") or "used") if row else "used"
            reason = str(row.get("reason") or "") if row else "supplied through the source route; no intake manifest"
            try:
                reason = validate_boundary_text(reason)
            except ValueError:
                reason = "manifest reason failed boundary validation"
            cited_by = sorted(cited.get(source_id, ()))
            in_model = source_id in model_inputs
            if disposition == "superseded":
                binding = "SUPERSEDED"
            elif disposition != "used":
                binding = "NOT_REQUIRED"
            elif in_model:
                binding = "MODEL_INPUT"
            elif source_id in {pinned_universe, market_marks_source_id}:
                binding = "MARKET_MARKS"
            elif document_type == "research_brief" and brief_bound:
                binding = "RESEARCH_BRIEF"
            elif cited_by:
                binding = "CITED_ANALYSIS"
            elif document_type not in RELEVANT_DOCUMENT_TYPES and document_type != "unclassified":
                binding = "NOT_REQUIRED"
                reason = f"{reason}; no model consumer for document type {document_type}"
            else:
                binding = "UNBOUND"
            period = (row or {}).get("period") if row else None
            rows.append({
                "source_id": source_id,
                "sha256": source["sha256"],
                "document_type": document_type,
                "period": period.get("label") if isinstance(period, dict) else None,
                "version_status": (row or {}).get("version_status"),
                "disposition": disposition,
                "reason": reason,
                "consumers": [str(item) for item in ((row or {}).get("consumers") or [])],
                "cited_by": cited_by,
                "model_input": in_model,
                "binding": binding,
            })
        rows.sort(key=lambda item: item["source_id"])
        unbound = [item["source_id"] for item in rows if item["binding"] == "UNBOUND"]
        if unbound:
            listed = ", ".join(unbound[:MAX_LINEAGE_DETAIL_IDS])
            raise ModelInputError(
                "admitted evidence reached no model input, calculation or cited analysis",
                code="MODEL_SOURCE_LINEAGE_INCOMPLETE",
                detail=(
                    f"{len(unbound)} admitted document(s) ({listed}) reached no model input, "
                    "calculation or cited analysis in the accepted run, so the model is not "
                    "source-complete. Re-run the analysis so every relevant document is used, "
                    "or withdraw the unused documents and run again."
                ),
            )
        return rows

    def _prior_full_credit_snapshot(
        self,
        snapshot: dict[str, Any],
        *,
        code: str = DEFAULT_BASE_MODEL_CODE,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        current = snapshot
        seen = {snapshot["id"]}
        for _ in range(MAX_SNAPSHOT_ANCESTORS):
            previous_id = current.get("previous_snapshot_id")
            if not isinstance(previous_id, str) or not previous_id or previous_id in seen:
                break
            seen.add(previous_id)
            previous = self.engine.runs.get_snapshot(previous_id)
            if previous is None or previous.get("case_id") != snapshot["case_id"]:
                break
            run = self._validated_snapshot_run(
                previous,
                error_code=code,
                require_live_sources=False,
            )
            if run["pathway"] == "FULL_CREDIT":
                self._validated_snapshot_run(previous, error_code=code)
                return previous, run
            current = previous
        raise ModelInputError(
            "A prior Full Credit snapshot is required for this pathway's model effect",
            code=code,
        )

    def _validated_base_build(
        self,
        snapshot: dict[str, Any],
        resolved: dict[str, Any],
        *,
        deadline: float | None = None,
        code: str = DEFAULT_BASE_MODEL_CODE,
    ) -> dict[str, Any]:
        build = self.builds.build_for_fingerprint(
            snapshot["case_id"], resolved["input_fingerprint"]
        )
        registry = self.bundle.assumption_registry
        if (
            build is None
            or build.get("status") != "READY"
            or build.get("accepted_run_id") != snapshot["run_id"]
            or build.get("snapshot_id") != snapshot["id"]
            or build.get("source_set_id") != snapshot["source_set_id"]
            or build.get("input_fingerprint") != resolved["input_fingerprint"]
            or build.get("methodology_build_id") != self.engine.bundle.build_id
            or build.get("calculation_runtime") != self.bundle.calculation_runtime
            or build.get("registry_version") != registry["version"]
            or build.get("registry_digest") != registry["digest"]
            or not _is_sha256(build.get("assumptions_digest"))
            or not _is_sha256(build.get("outputs_digest"))
            or not _is_sha256(build.get("payload_digest"))
            or not isinstance(build.get("payload"), dict)
            or not isinstance(build.get("qa"), dict)
            or build["qa"].get("status") != "PASS"
            or "pathway_effects" in build["payload"]
        ):
            raise ModelInputError(
                "A validated READY Full Credit model is required",
                code=code,
            )
        self._resolved_cache.pop(build["id"], None)
        self._defaults_cache.pop(build["id"], None)
        try:
            expected_result, expected_identity = self._compute_build_result(build, deadline=deadline)
        except (ModelInputError, ValueError):
            raise ModelInputError(
                "The prior Full Credit model calculation is invalid",
                code=code,
            ) from None
        if (
            build["payload"] != expected_result["payload"]
            or build["payload_digest"] != expected_result["payload_digest"]
            or build["qa"] != expected_result["qa"]
            or build["assumptions_digest"] != expected_identity["assumptions_digest"]
            or build["outputs_digest"] != expected_identity["outputs_digest"]
        ):
            raise ModelInputError(
                "The prior Full Credit model identity is invalid",
                code=code,
            ) from None
        return build

    @staticmethod
    def _publication_model_authority(
        snapshot: dict[str, Any],
        build: dict[str, Any],
        *,
        relationship: str,
    ) -> dict[str, Any]:
        return {
            "relationship": relationship,
            "snapshot_id": snapshot["id"],
            "snapshot_digest": snapshot["digest"],
            "run_id": snapshot["run_id"],
            "source_set_id": snapshot["source_set_id"],
            "source_set_version": snapshot["source_set_version"],
            "input_fingerprint": build["input_fingerprint"],
            "payload_digest": build["payload_digest"],
        }

    def _validated_prior_full_credit_publication_build(
        self,
        build_id: str | None,
        accepted_snapshot: dict[str, Any],
        accepted_run: dict[str, Any],
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        if accepted_run.get("pathway") not in PRIOR_FULL_CREDIT_PUBLICATION_PATHWAYS:
            raise ModelInputError("accepted pathway cannot reuse a prior Full Credit model")
        base_snapshot, base_run = self._prior_full_credit_snapshot(accepted_snapshot)
        resolved = self._resolve_full_credit_snapshot(base_snapshot, base_run)
        build = self._validated_base_build(base_snapshot, resolved)
        if build_id is not None and build["id"] != build_id:
            raise ModelBuildStale(build)
        return build, self._publication_model_authority(
            base_snapshot,
            build,
            relationship="PRIOR_FULL_CREDIT_BASE",
        )

    def _validated_snapshot_artifacts(
        self,
        snapshot: dict[str, Any],
        run: dict[str, Any],
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        # One resolution re-verifies each snapshot once: the Distressed chain
        # otherwise re-hashes the vault and replays every calculation for the
        # base four to twelve times per request.
        memo = _RESOLUTION_MEMO.get()
        key = (snapshot.get("id"), snapshot.get("digest"), run.get("id"))
        if memo is not None and key in memo:
            return memo[key]
        try:
            artifacts = self.engine.validated_snapshot_artifacts(snapshot, run)
        except AgentError as exc:
            raise ModelInputError("Accepted snapshot artifacts are invalid") from exc
        inventory = [
            {
                "module_id": artifact["module_id"],
                "artifact_id": artifact["id"],
                "digest": artifact["digest"],
                "status": "READY",
            }
            for artifact in artifacts
        ]
        if memo is not None:
            memo[key] = (artifacts, inventory)
        return artifacts, inventory

    def _validated_calculations(
        self,
        artifact: dict[str, Any],
        module_id: str,
        *,
        required: frozenset[str] = frozenset(),
    ) -> list[dict[str, Any]]:
        """Every calculation record an accepted artifact carries, re-verified
        against the pinned binding and re-executed before the model may consume
        it; `required` names the calculators the effect cannot do without."""
        records = artifact["payload"].get("calculations")
        if not isinstance(records, list) or (required and not records):
            raise ModelInputError(f"{module_id} calculation records are unavailable")
        try:
            bindings = {
                item["calculator_id"]: item
                for item in self.engine._calculation_runtime.binding_manifest()
                if item["module_id"] == module_id
            }
        except (AttributeError, KeyError, TypeError, ValueError):
            raise ModelInputError(f"{module_id} calculation authority is unavailable") from None
        if not required <= set(bindings):
            raise ModelInputError(f"{module_id} calculation authority is unavailable")
        calculator_ids: set[str] = set()
        validated: list[dict[str, Any]] = []
        for record in records:
            if not isinstance(record, dict):
                raise ModelInputError(f"{module_id} calculation record is invalid")
            calculator_id = record.get("calculator_id")
            if calculator_id in calculator_ids:
                raise ModelInputError(f"{module_id} calculation record is duplicated")
            dependencies = record.get("dependency_digests")
            binding = bindings.get(calculator_id)
            if (
                set(record) != {
                    "schema_version",
                    "methodology_build_id",
                    "module_id",
                    "calculator_id",
                    "script_digest",
                    "script_bytes",
                    "dependency_digests",
                    "calculator_digest",
                    "canonical_input",
                    "input_digest",
                    "output_digest",
                    "canonical_output",
                }
                or record.get("schema_version") != "caos.methodology-calculation.v1"
                or record.get("methodology_build_id") != self.engine.bundle.build_id
                or record.get("module_id") != module_id
                or binding is None
                or record.get("script_digest") != binding["script_digest"]
                or record.get("script_bytes") != binding["script_bytes"]
                or record.get("calculator_digest") != binding["calculator_digest"]
                or dependencies != binding["dependency_digests"]
                or not _is_sha256(record.get("input_digest"))
                or not _is_sha256(record.get("output_digest"))
            ):
                raise ModelInputError(f"{module_id} calculation record identity is invalid")
            try:
                canonical_input = clean_json(record["canonical_input"])
                canonical_output = clean_json(record["canonical_output"])
            except (KeyError, TypeError, ValueError):
                raise ModelInputError(f"{module_id} calculation output is invalid") from None
            if (
                type(canonical_input) is not dict
                or digest(canonical_input) != record["input_digest"]
                or digest(canonical_output) != record["output_digest"]
            ):
                raise ModelInputError(f"{module_id} calculation input or output digest is invalid")
            try:
                recalculated = self.engine._calculation_runtime.execute(
                    module_id,
                    calculator_id,
                    canonical_input,
                )
            except (AttributeError, KeyError, TypeError, ValueError):
                raise ModelInputError(f"{module_id} calculation cannot be reconstructed") from None
            if recalculated != record:
                raise ModelInputError(f"{module_id} calculation reconstruction differs")
            if not calculation_output_complete(
                module_id,
                calculator_id,
                recalculated["canonical_output"],
            ):
                raise ModelInputError(f"{module_id} calculation is incomplete")
            calculator_ids.add(calculator_id)
            validated.append(copy.deepcopy(record))
        if not required <= calculator_ids:
            raise ModelInputError(f"{module_id} required calculations are missing")
        return validated

    @staticmethod
    def _calculation_limitations(
        by_module: dict[str, dict[str, Any]],
        module_ids: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        """Host-declared calculation limitations of the consumed modules, copied
        verbatim: a calculator that stayed incomplete is a named gap on the
        effect, never a zero."""
        return [
            {"module_id": module_id, "limitation": copy.deepcopy(limitation)}
            for module_id in module_ids
            for limitation in ((by_module.get(module_id) or {}).get("payload") or {}).get("calculation_limitations") or []
        ]

    def _resolve_overlay_snapshot(
        self,
        snapshot: dict[str, Any],
        run: dict[str, Any],
        *,
        deadline: float | None = None,
    ) -> dict[str, Any]:
        """One overlay on the nearest validated Full Credit model: the base
        build is re-verified by recomputation, the accepted run's calculation
        records are re-executed, and the pathway's effect rides the payload
        under its own input fingerprint. The base tabs are copied unchanged, so
        no reported period, assumption or output is ever overwritten."""
        pathway = run["pathway"]
        code = BASE_MODEL_CODES.get(pathway, DEFAULT_BASE_MODEL_CODE)
        artifacts, inventory = self._validated_snapshot_artifacts(snapshot, run)
        by_module = {artifact["module_id"]: artifact for artifact in artifacts}
        base_snapshot, base_run = self._prior_full_credit_snapshot(snapshot, code=code)
        try:
            self._validated_snapshot_artifacts(base_snapshot, base_run)
            base_resolved = self._resolve_full_credit_snapshot(base_snapshot, base_run)
        except (ModelInputError, ValueError):
            raise ModelInputError("The prior Full Credit snapshot is invalid", code=code) from None
        base_build = self._validated_base_build(base_snapshot, base_resolved, deadline=deadline, code=code)
        source_set = self.store.source_set(snapshot["source_set_id"])
        if (
            source_set is None
            or source_set.get("case_id") != snapshot["case_id"]
            or source_set.get("version") != snapshot["source_set_version"]
        ):
            raise ModelInputError("accepted source set is unavailable")
        deadline = deadline if deadline is not None else self._new_deadline()
        builder = {
            "EARNINGS_UPDATE": self._earnings_effect,
            "COVENANT_REFINANCING": self._covenant_effect,
            "RELATIVE_VALUE": self._relative_value_effect,
            DISTRESSED_PATHWAY: self._distressed_effect,
            "DEEP_RESEARCH": self._deep_research_effect,
        }[pathway]
        content = builder(snapshot, run, by_module, base_build, base_resolved, source_set, deadline)
        authority = {
            "run_id": snapshot["run_id"],
            "snapshot_id": snapshot["id"],
            "snapshot_digest": snapshot["digest"],
            "source_set_id": snapshot["source_set_id"],
            "source_set_version": snapshot["source_set_version"],
            "provider_identity_digest": (snapshot.get("provider_identity") or {}).get("identity_digest"),
            "artifacts": [
                {"module_id": item["module_id"], "artifact_id": item["artifact_id"], "digest": item["digest"]}
                for item in inventory
            ],
            **content.get("authority_extra", {}),
        }
        calculations = content["calculations"]
        effect = {
            "schema_version": PATHWAY_EFFECT_SCHEMA,
            "effect_id": EFFECT_IDS[pathway],
            "pathway": pathway,
            "base_model": {
                "build_id": base_build["id"],
                "run_id": base_snapshot["run_id"],
                "snapshot_id": base_snapshot["id"],
                "snapshot_digest": base_snapshot["digest"],
                "source_set_id": base_snapshot["source_set_id"],
                "source_set_version": base_snapshot["source_set_version"],
                "input_fingerprint": base_build["input_fingerprint"],
                "payload_digest": base_build["payload_digest"],
                "assumptions_digest": base_build["assumptions_digest"],
                "outputs_digest": base_build["outputs_digest"],
                "qa_digest": digest(base_build["qa"]),
                "methodology_build_id": base_build["methodology_build_id"],
                "calculation_runtime": base_build["calculation_runtime"],
            },
            ("distressed_authority" if pathway == DISTRESSED_PATHWAY else "accepted_authority"): authority,
            "calculations": calculations,
            "calculation_records_digest": digest(calculations),
            "limitations": content["limitations"],
            "methodology_build_id": self.engine.bundle.build_id,
            "cp_model_runtime": self.bundle.calculation_runtime,
            **content["sections"],
        }
        lineage = self._source_lineage(
            snapshot, run, artifacts,
            model_markdown=base_resolved["model_markdown"],
            market_marks_source_id=content.get("market_marks_source_id"),
        )
        model_payload = copy.deepcopy(base_build["payload"])
        model_payload["pathway_effects"] = [effect]
        model_payload["source_lineage"] = lineage
        fingerprint = digest({
            "case_id": snapshot["case_id"],
            "accepted_run_id": snapshot["run_id"],
            "accepted_snapshot_id": snapshot["id"],
            "accepted_snapshot_digest": snapshot["digest"],
            "previous_snapshot_id": snapshot.get("previous_snapshot_id"),
            "source_set": {"id": source_set["id"], "version": source_set["version"]},
            "artifacts": [
                {"module_id": item["module_id"], "artifact_id": item["artifact_id"], "digest": item["digest"]}
                for item in inventory
            ],
            "base_model": effect["base_model"],
            "effect_digest": digest(effect),
            "source_lineage_digest": digest(lineage),
            "methodology_build_id": self.engine.bundle.build_id,
            "calculation_runtime": self.bundle.calculation_runtime,
        })
        return {
            "snapshot": snapshot,
            "source_set": {"id": source_set["id"], "version": source_set["version"]},
            "artifact_inventory": inventory,
            "markdown": base_resolved["markdown"],
            "model_markdown": base_resolved["model_markdown"],
            "input_fingerprint": fingerprint,
            "base_build": base_build,
            "pathway_effect": effect,
            "model_payload": model_payload,
            "source_lineage": lineage,
        }

    @staticmethod
    def _require_artifact(by_module: dict[str, dict[str, Any]], module_id: str) -> dict[str, Any]:
        artifact = by_module.get(module_id)
        if artifact is None:
            raise ModelInputError(f"{module_id} artifact is unavailable")
        return artifact

    def _distressed_effect(self, snapshot, run, by_module, base_build, base_resolved, source_set, deadline):
        cp4c = self._require_artifact(by_module, "CP-4C")
        calculations = self._validated_calculations(cp4c, "CP-4C", required=DISTRESSED_CALCULATORS)
        return {
            "calculations": calculations,
            "limitations": self._calculation_limitations(by_module, ("CP-4C",)),
            "authority_extra": {"cp4c_artifact_id": cp4c["id"], "cp4c_artifact_digest": cp4c["digest"]},
            "sections": {
                "treatment": "BASE_MODEL_REUSED_WITH_DISTRESSED_SCENARIO_OVERLAY",
                "scope": (
                    "The prior Full Credit operating model is reused unchanged; "
                    "only the CP-4C funding-gap and recovery analysis is overlaid."
                ),
            },
        }

    def _earnings_effect(self, snapshot, run, by_module, base_build, base_resolved, source_set, deadline):
        """Reported periods from the run's verified credit_metrics records, and
        the variance of each against the base model's forecast for the same
        fiscal year. Reported actuals and analyst forecasts keep distinct
        authority; a missing side is a named gap, never zero."""
        calculations: list[dict[str, Any]] = []
        period_updates: list[dict[str, Any]] = []
        for module_id in ("CP-1", "CP-1B"):
            artifact = self._require_artifact(by_module, module_id)
            records = self._validated_calculations(
                artifact, module_id,
                required=frozenset({"credit_metrics"}) if module_id == "CP-1" else frozenset(),
            )
            calculations.extend(records)
            for record in records:
                if record["calculator_id"] != "credit_metrics":
                    continue
                periods = record["canonical_output"].get("periods") or {}
                for period_id in sorted(periods):
                    period = periods[period_id] or {}
                    period_updates.append({
                        "period_id": period_id,
                        "module_id": module_id,
                        "authority": "REPORTED_ACTUAL",
                        "reported": {
                            key: _decimal_text(value)
                            for key, value in sorted((period.get("inputs") or {}).items())
                        },
                        "kpis": {
                            key: _decimal_text(value)
                            for key, value in sorted((period.get("kpis") or {}).items())
                        },
                        "not_calculable": [str(item) for item in period.get("not_calculable") or []],
                        "record_output_digest": record["output_digest"],
                    })
        _defaults, base_outputs = self._defaults(base_build, deadline)
        return {
            "calculations": calculations,
            "limitations": self._calculation_limitations(by_module, ("CP-1", "CP-1B")),
            "sections": {
                "treatment": "BASE_MODEL_REUSED_WITH_REPORTED_PERIODS_AND_FORECAST_VARIANCE",
                "period_updates": period_updates,
                "forecast_variance": _forecast_variance(period_updates, base_outputs),
                "scope": (
                    "The prior Full Credit model is reused unchanged; the accepted "
                    "earnings run's reported periods are attached and compared "
                    "with the model's forecast for the same fiscal years."
                ),
            },
        }

    def _covenant_effect(self, snapshot, run, by_module, base_build, base_resolved, source_set, deadline):
        """Covenant tests from CP-4 and the refinancing wall from CP-4C, both
        documentary, plus the base registry slots a test maps to as proposals
        for sign-off. Nothing is applied to the signed model here."""
        calculations: list[dict[str, Any]] = []
        covenant_updates: list[dict[str, Any]] = []
        refinancing_updates: list[dict[str, Any]] = []
        for module_id in ("CP-4", "CP-4C"):
            artifact = self._require_artifact(by_module, module_id)
            records = self._validated_calculations(artifact, module_id)
            calculations.extend(records)
            for record in records:
                output = record["canonical_output"]
                if record["calculator_id"] == "covenant_headroom":
                    for row in output.get("headroom") or []:
                        covenant_updates.append({
                            "module_id": module_id,
                            "test": row.get("test"),
                            "test_type": row.get("test_type"),
                            "threshold": _decimal_text(row.get("threshold")),
                            "current_ratio": _decimal_text(row.get("current_ratio")),
                            "headroom": _decimal_text(row.get("headroom")),
                            "status": row.get("status"),
                            "missing_inputs": [str(item) for item in row.get("missing_inputs") or []],
                            "authority": "DOCUMENTARY_COVENANT_TERMS",
                            "record_output_digest": record["output_digest"],
                        })
                elif record["calculator_id"] == "funding_gap":
                    for view in ("as_of_balance_sheet_date", "pro_forma_for_subsequent_events"):
                        if isinstance(output.get(view), dict):
                            refinancing_updates.append({
                                "module_id": module_id,
                                "view": view,
                                "maturity_wall": copy.deepcopy(output[view].get("maturity_wall")),
                                "liquidity": copy.deepcopy(output[view].get("liquidity")),
                                "gap": copy.deepcopy(output[view].get("gap")),
                                "authority": "DOCUMENTARY_INSTRUMENT_TERMS",
                                "record_output_digest": record["output_digest"],
                            })
        defaults, _outputs = self._defaults(base_build, deadline)
        return {
            "calculations": calculations,
            "limitations": self._calculation_limitations(by_module, ("CP-4", "CP-4C")),
            "sections": {
                "treatment": "BASE_MODEL_REUSED_WITH_COVENANT_AND_REFINANCING_UPDATES",
                "covenant_updates": covenant_updates,
                "refinancing_updates": refinancing_updates,
                "assumption_updates": _covenant_assumption_updates(covenant_updates, defaults),
                "scope": (
                    "The prior Full Credit model is reused unchanged; the accepted "
                    "run's covenant tests and refinancing wall are attached and the "
                    "registry slots they map to are proposed for sign-off."
                ),
            },
        }

    def _relative_value_effect(self, snapshot, run, by_module, base_build, base_resolved, source_set, deadline):
        """The market marks the run pinned at its gate, re-read from the store
        and digest-verified, attached with their time alignment against the
        base model's latest reported period and the run's analysis date."""
        pinned = (run.get("plan") or {}).get("loan_universe")
        if not isinstance(pinned, dict):
            raise ModelInputError(
                "no market-marks workbook is pinned to the accepted run",
                code="RELATIVE_VALUE_MARKET_MARKS_REQUIRED",
            )
        record = self.store.loan_universe(pinned.get("id"))
        if (
            record is None
            or record.get("case_id") != snapshot["case_id"]
            or record.get("source_id") != pinned.get("source_id")
            or record.get("source_id") not in (source_set.get("source_ids") or [])
            or universe_digest(record) != pinned.get("universe_digest")
            or not isinstance(record.get("rows"), list)
        ):
            raise ModelInputError("pinned market marks do not match the store record")
        calculations: list[dict[str, Any]] = []
        consumed: list[str] = []
        for module_id in ("CP-3", "CP-1C"):
            artifact = by_module.get(module_id)
            if artifact is None:
                continue
            consumed.append(module_id)
            calculations.extend(self._validated_calculations(artifact, module_id))
        rows = [
            {field: row.get(field) for field in MARKET_MARK_FIELDS}
            for row in record["rows"]
            if isinstance(row, dict)
        ]
        workbook_date = record.get("workbook_date")
        latest_end = _latest_reported_period_end(base_resolved["markdown"].get("CP-1") or "")
        analysis_date = str(run.get("created_at") or "")[:10]
        if not isinstance(workbook_date, str) or not workbook_date:
            status = "WORKBOOK_DATE_UNAVAILABLE"
        elif analysis_date and workbook_date > analysis_date:
            status = "POSTDATES_ANALYSIS"
        elif latest_end and workbook_date < latest_end:
            status = "PRECEDES_LATEST_REPORTED_PERIOD"
        else:
            status = "ALIGNED"
        limitations = self._calculation_limitations(by_module, tuple(consumed))
        if status != "ALIGNED":
            limitations.append({"module_id": "CP-3", "limitation": {"code": f"MARKET_MARKS_{status}"}})
        return {
            "calculations": calculations,
            "limitations": limitations,
            "market_marks_source_id": record["source_id"],
            "sections": {
                "treatment": "BASE_MODEL_REUSED_WITH_MARKET_MARKS_ATTACHED",
                "market_marks": {
                    "universe_id": record["id"],
                    "universe_digest": record["universe_digest"],
                    "source_id": record["source_id"],
                    "source_sha256": record["source_sha256"],
                    "template_version": record["template_version"],
                    "importer_version": record["importer_version"],
                    "workbook_date": workbook_date,
                    "row_count": len(rows),
                    "authority": "SUPPLIED_MARKET_MARKS",
                    "rows": rows,
                },
                "time_alignment": {
                    "workbook_date": workbook_date,
                    "latest_reported_period_end": latest_end,
                    "analysis_date": analysis_date or None,
                    "status": status,
                },
                "scope": (
                    "The prior Full Credit model is reused unchanged; the supplied "
                    "market marks are attached with their alignment to the model's "
                    "latest reported period and the run's analysis date."
                ),
            },
        }

    def _deep_research_effect(self, snapshot, run, by_module, base_build, base_resolved, source_set, deadline):
        """Deep Research declares no numeric effect; with a Full Credit model in
        the case it revalidates that model (recomputed and compared above) and
        binds the approved research scope to the record."""
        cp_dr = self._require_artifact(by_module, "CP-DR")
        research = run.get("research") or {}
        return {
            "calculations": [],
            "limitations": self._calculation_limitations(by_module, ("CP-DR",)),
            "sections": {
                "treatment": "BASE_MODEL_REVALIDATED_NO_NUMERIC_EFFECT",
                "numeric_effect": "NONE",
                "base_model_revalidated": True,
                "research": {
                    "brief_digest": research.get("brief_digest"),
                    "approved_plan_hash": research.get("approved_plan_hash"),
                    "cp_dr_artifact_id": cp_dr["id"],
                    "cp_dr_artifact_digest": cp_dr["digest"],
                },
                "scope": (
                    "Deep Research changes no model period, assumption or output; "
                    "the prior Full Credit model is revalidated and left unchanged."
                ),
            },
        }

    def readiness(self, case_id: str) -> dict[str, Any]:
        forced = self._forced_readiness.get(case_id)
        if forced:
            return {"status": forced, "module_id": "CP-MODEL", "snapshot_id": None, "blockers": [{"code": forced}]}
        snapshot = self._accepted_snapshot(case_id)
        if snapshot is None:
            return {"status": "NOT_READY", "module_id": "CP-MODEL", "snapshot_id": None,
                    "blockers": [{"code": "ACCEPTED_FULL_CREDIT_REQUIRED",
                                  "detail": NO_ACCEPTED_AUTHORITY_DETAIL}]}
        try:
            resolved = self._resolve_snapshot(snapshot)
        except (ModelInputError, ValueError) as exc:
            # A wrong-pathway precondition is not input corruption, and the two
            # need different statuses: only NOT_READY offers Model Builder's
            # "Open Run Console" way out, so collapsing them dead-ended the
            # commonest journey (accept an Earnings Update, open Model Builder)
            # behind an alarming, detail-less callout. Both details are written
            # here, never taken from the exception: the messages raised below
            # interpolate bundle paths and table ids.
            code = getattr(exc, "code", None)
            if code in PRECONDITION_DETAILS:
                detail = getattr(exc, "detail", None) or PRECONDITION_DETAILS[code]
                return {"status": "NOT_READY", "module_id": "CP-MODEL", "snapshot_id": snapshot["id"],
                        "blockers": [{"code": code, "detail": detail}]}
            return {"status": "CANONICAL_MODEL_INPUTS_INVALID", "module_id": "CP-MODEL",
                    "snapshot_id": snapshot["id"],
                    "blockers": [{"code": "CANONICAL_MODEL_INPUTS_INVALID",
                                  "detail": INPUTS_INVALID_DETAIL}]}
        build = self.builds.build_for_fingerprint(case_id, resolved["input_fingerprint"])
        return {
            # A queued/building/failed job does not change what the accepted
            # authority is ready for — only a READY build advances the status.
            "status": "READY" if build and build["status"] == "READY" else "READY_TO_BUILD",
            "module_id": "CP-MODEL",
            "snapshot_id": snapshot["id"],
            # The wire's own name for the same fact. The route read this key and
            # the service never produced it, so Model Builder's authority panel
            # read "Not accepted" beside a build derived from that very snapshot.
            "accepted_snapshot": {"id": snapshot["id"], "run_id": snapshot["run_id"],
                                  "digest": snapshot.get("digest")},
            "input_fingerprint": resolved["input_fingerprint"],
            "source_set": resolved["source_set"],
            "requirements": resolved["artifact_inventory"],
            "calculation_runtime": self.bundle.calculation_runtime,
            "blockers": [],
            "build": build,
        }

    # -- queue --------------------------------------------------------------

    @_authority_guarded
    def queue_build(self, case_id: str, actor: str) -> dict[str, Any]:
        for callback in list(self._on_queue):
            callback(case_id)
        if self._fail_next_queue:
            self._fail_next_queue = False
            raise ValueError("MODEL_QUEUE_FAILED: injected queue failure")
        if self._forced_readiness.get(case_id):
            raise ValueError(f"MODEL_NOT_READY: {self._forced_readiness[case_id]}")
        snapshot = self._accepted_snapshot(case_id)
        if snapshot is None:
            raise ValueError("MODEL_NOT_READY: accept a completed Full Credit run first")
        try:
            resolved = self._resolve_snapshot(snapshot)
        except (ModelInputError, ValueError) as exc:
            code = getattr(exc, "code", None)
            if code == "ACCEPTED_FULL_CREDIT_REQUIRED":
                raise ValueError("MODEL_NOT_READY: accept a completed Full Credit run first") from exc
            if code == "DISTRESSED_BASE_MODEL_REQUIRED":
                raise ValueError("DISTRESSED_BASE_MODEL_REQUIRED") from exc
            if code in PRECONDITION_DETAILS:
                raise ValueError(f"MODEL_NOT_READY: {code}") from exc
            raise ValueError("MODEL_BUILD_INVALID: canonical model inputs are invalid") from exc
        registry = self.bundle.assumption_registry
        identity = {
            "accepted_run_id": snapshot["run_id"],
            "snapshot_id": snapshot["id"],
            "source_set_id": snapshot["source_set_id"],
            "input_fingerprint": resolved["input_fingerprint"],
            "methodology_build_id": self.engine.bundle.build_id,
            "calculation_runtime": self.bundle.calculation_runtime,
            "registry_version": registry["version"],
            "registry_digest": registry["digest"],
        }
        # At most one active (queued/building) job per case: a newer accepted
        # snapshot re-points the standing job rather than stacking a second.
        active = next((b for b in self.builds.list_builds(case_id) if b["status"] in ("QUEUED", "BUILDING")), None)
        if active is not None:
            if active["input_fingerprint"] != resolved["input_fingerprint"]:
                # Re-pointing a BUILDING row abandons that computation: it goes
                # back to QUEUED so a worker recomputes under the new identity,
                # and the abandoned executor's fingerprint-bound completion
                # no-ops instead of publishing a stale calculation.
                self.builds.update_build(active["id"], expected_status=("QUEUED", "BUILDING"),
                                         status="QUEUED", **identity)
                self._resolved_cache.pop(active["id"], None)
                self._defaults_cache.pop(active["id"], None)
            record, created = self.builds.get_build(active["id"]), False
        else:
            # Model jobs and runs share one admission budget (Appendix A:
            # MAX_ACTIVE_JOBS checked at run AND build start). Re-pointing an
            # existing job never consumes a new slot.
            from ..engine.budget import MAX_ACTIVE_JOBS

            occupied = (
                self.engine.runs.active_admission_count()
                + getattr(self.engine, "_admission_offset", 0)
                + self.builds.active_build_count()
            )
            if occupied >= MAX_ACTIVE_JOBS:
                raise ValueError("ADMISSION_BUSY: active job ceiling reached")
            record, created = self.builds.queue_build({"case_id": case_id, **identity}, actor)
        self._dispatch(record["id"])
        return {**record, "created": created}

    def queue_build_pinned_for_tests(self, case_id: str, *, snapshot_id: str) -> dict[str, Any]:
        snapshot = self.engine.runs.get_snapshot(snapshot_id)
        case = self.store.get_case(case_id)
        if (
            snapshot is None
            or case is None
            or snapshot["case_id"] != case_id
            or case.get("accepted_snapshot_id") != snapshot_id
        ):
            raise ValueError("MODEL_BUILD_INVALID: snapshot is not this case's accepted authority")
        return self.queue_build(case_id, "analyst")

    def _dispatch(self, build_id: str) -> None:
        if self._fail_next_dispatch:
            self._fail_next_dispatch = False
            raise ValueError("MODEL_DISPATCH_FAILED: injected dispatch failure")
        self._dispatch_log.append(build_id)

    # -- reads --------------------------------------------------------------

    def build(self, build_id: str) -> dict[str, Any] | None:
        return self.builds.get_build(build_id)

    def list_builds(self, case_id: str) -> list[dict[str, Any]]:
        return self.builds.list_builds(case_id)

    def current_build(self, case_id: str) -> dict[str, Any] | None:
        return self.builds.current_build(case_id)

    def worksheet(self, build_id: str) -> dict[str, Any] | None:
        build = self.builds.get_build(build_id)
        return build.get("payload") if build else None

    # -- build execution -----------------------------------------------------

    def _resolved_inputs(
        self, build: dict[str, Any], *, deadline: float | None = None,
    ) -> dict[str, Any]:
        cached = self._resolved_cache.get(build["id"])
        if cached is not None:
            return cached
        snapshot = self.engine.runs.get_snapshot(build["snapshot_id"])
        if snapshot is None or snapshot["case_id"] != build["case_id"]:
            raise ModelInputError("model snapshot is unavailable")
        resolved = self._resolve_snapshot(snapshot, deadline=deadline)
        if resolved["input_fingerprint"] != build["input_fingerprint"]:
            raise ModelInputError("model inputs changed")
        _remember(self._resolved_cache, build["id"], resolved)
        return resolved

    @staticmethod
    def _write_inputs(directory: Path, markdown: dict[str, str]) -> dict[str, Path]:
        paths: dict[str, Path] = {}
        for module_id, content in markdown.items():
            path = directory / f"{module_id.lower()}.md"
            path.write_text(content, encoding="utf-8")
            paths[module_id] = path
        return paths

    def _calculate(self, build: dict[str, Any], effective: list[dict[str, Any]] | None, deadline: float) -> tuple[Any, Any]:
        self._check_deadline(deadline)
        if effective is not None and len(effective) > MAX_EFFECTIVE_ASSUMPTIONS:
            raise ValueError("MODEL_ASSUMPTION_LIMIT")
        resolved = self._resolved_inputs(build)
        with tempfile.TemporaryDirectory(prefix="caos-model-") as temporary:
            paths = self._write_inputs(Path(temporary), resolved["markdown"])
            model, calculations = self.bundle.calculate(paths, effective_assumptions=effective)
        self._check_deadline(deadline)
        return model, calculations

    def _defaults(self, build: dict[str, Any], deadline: float) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        cached = self._defaults_cache.get(build["id"])
        if cached is None:
            model, calculations = self._calculate(build, None, deadline)
            rows = _assumption_rows(model)
            rows = self._apply_evolution(build["id"], rows)
            cached = (rows, _annual_outputs(calculations))
            _remember(self._defaults_cache, build["id"], cached)
        return copy.deepcopy(cached[0]), copy.deepcopy(cached[1])

    def _apply_evolution(self, build_id: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        evolution = self._evolutions.get(build_id)
        if not evolution:
            return rows
        removed = set(evolution.get("removed", ()))
        changed = set(evolution.get("source_context_changed", ()))
        result = []
        for row in rows:
            if row["assumption_id"] in removed:
                continue
            if row["assumption_id"] in changed:
                row = copy.deepcopy(row)
                row["source_context"]["gap_code"] = "SOURCE_CONTEXT_EVOLVED"
                row["source_context_digest"] = digest(row["source_context"])
            result.append(row)
        return result

    def recover_builds(self) -> int:
        """Worker startup recovery: requeue the builds a predecessor claimed and
        never finished. A hard kill skips run_pending's FAILED fallback, so
        without this a claimed row stayed BUILDING forever (SIM-008)."""
        return self.builds.requeue_building_builds()

    def run_build_for_tests(self, build_id: str) -> dict[str, Any]:
        build = self.builds.get_build(build_id)
        if build is None:
            raise ValueError("MODEL_BUILD_NOT_FOUND")
        if build["status"] == "READY":
            return build
        if not self.builds.update_build(build_id, expected_status=("QUEUED", "FAILED"), status="BUILDING"):
            # Lost the claim: another executor owns this row. Losing is not a
            # failure — never compute, never touch the winner's state.
            return self.builds.get_build(build_id)  # type: ignore[return-value]
        # The CLAIMED row is the authority; the pre-claim read may already be stale.
        build = self.builds.get_build(build_id)
        fingerprint = build["input_fingerprint"]  # type: ignore[index]
        try:
            result, identity = self._compute_build_result(build)
            self._complete(build_id, result, identity, expected_fingerprint=fingerprint)
        except BuildAuthorityChanged:
            pass  # the FAILED record with its typed code is already committed
        except BuildIdentityChanged:
            # Re-pointed mid-flight: the requeued row wins, this calculation dies.
            pass
        except ModelInputError:
            self._fail(build_id, "MODEL_INPUT_INVALID", "The accepted model inputs are invalid.", fingerprint)
        except ModelCalculationTimeout:
            raise
        except ValueError:
            raise
        except Exception:
            self._fail(build_id, "MODEL_CALCULATION_FAILED", "The model calculation did not complete.", fingerprint)
        return self.builds.get_build(build_id)  # type: ignore[return-value]

    def _compute_build_result(
        self, build: dict[str, Any], *, deadline: float | None = None,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        deadline = deadline if deadline is not None else self._new_deadline()
        resolved = self._resolved_inputs(build, deadline=deadline)
        if resolved.get("pathway_effect") is not None:
            base_build = resolved["base_build"]
            return (
                {
                    "payload": copy.deepcopy(resolved["model_payload"]),
                    "payload_digest": digest(resolved["model_payload"]),
                    "qa": copy.deepcopy(base_build["qa"]),
                },
                {
                    "assumptions_digest": base_build["assumptions_digest"],
                    "outputs_digest": base_build["outputs_digest"],
                },
            )
        model, calculations = self._calculate(build, None, deadline)
        serialized = self.bundle.serialize_workbook(model, calculations)
        self._check_deadline(deadline)
        payload = serialized["payload"]
        payload["source_lineage"] = copy.deepcopy(resolved["source_lineage"])
        result = {
            "payload": payload,
            "payload_digest": digest(payload),
            "qa": serialized["qa"],
        }
        rows = self._apply_evolution(build["id"], _assumption_rows(model))
        outputs = _annual_outputs(calculations)
        _remember(self._defaults_cache, build["id"], (rows, outputs))
        identity = {"assumptions_digest": digest(rows), "outputs_digest": digest(outputs)}
        return result, identity

    def valid_build_result_for_tests(self, build_id: str) -> dict[str, Any]:
        build = self.builds.get_build(build_id)
        if build is None:
            raise ValueError("MODEL_BUILD_NOT_FOUND")
        result, _identity = self._compute_build_result(build)
        return result

    def complete_build_for_tests(self, build_id: str, result: dict[str, Any]) -> dict[str, Any]:
        self._validate_result(result)
        self._complete(build_id, copy.deepcopy(result), {})
        return self.builds.get_build(build_id)  # type: ignore[return-value]

    def _pinned_source_ids(self, build: dict[str, Any]) -> list[str]:
        """The source ids a build's snapshot pinned; publication is conditional on
        every one of them still being live at the moment of the write."""
        source_set = self.store.source_set(build.get("source_set_id"))
        if source_set is None or source_set.get("case_id") != build.get("case_id"):
            raise ModelInputError("model source authority is unavailable")
        return list(source_set["source_ids"])

    def _complete(self, build_id: str, result: dict[str, Any], identity: dict[str, Any],
                  *, expected_fingerprint: str | None = None) -> None:
        self._validate_result(result)
        build = self.builds.get_build(build_id)
        if build is None:
            raise BuildIdentityChanged("MODEL_RESULT_INVALID: build is not completable")
        changed = self.builds.update_build(
            build_id, expected_status=("QUEUED", "BUILDING"),
            expected_input_fingerprint=expected_fingerprint,
            expected_live_source_ids=self._pinned_source_ids(build),
            status="READY", payload=result["payload"], payload_digest=result["payload_digest"],
            qa=result["qa"], error=None, completed_at=now_iso(), **identity,
            # Audit lineage rides the same transaction as the READY transition:
            # the build, its snapshot, its run and the digest of the payload
            # that carries the source lineage and any pathway effect.
            audit=lambda conn: self.store._audit(
                conn, "model.build_ready", build.get("created_by") or "system",
                case_id=build["case_id"], build_id=build_id, snapshot_id=build.get("snapshot_id"),
                run_id=build.get("accepted_run_id"), sha256=result["payload_digest"],
            ),
        )
        if changed:
            return
        # Either the row was re-pointed mid-flight (the requeued row wins) or a
        # pinned source was withdrawn while the result was being computed: the
        # second is a typed failure the analyst must see, never a READY model.
        if self.builds.update_build(
            build_id, expected_status=("QUEUED", "BUILDING"),
            expected_input_fingerprint=expected_fingerprint,
            status="FAILED",
            error={"code": "MODEL_AUTHORITY_CHANGED",
                   "detail": "A pinned source was withdrawn before the model could be published."},
        ):
            raise BuildAuthorityChanged("MODEL_AUTHORITY_CHANGED")
        raise BuildIdentityChanged("MODEL_RESULT_INVALID: build is not completable")

    @staticmethod
    def _validate_result(result: dict[str, Any]) -> None:
        def invalid(reason: str) -> ValueError:
            return ValueError(f"MODEL_RESULT_INVALID: {reason}")

        if not isinstance(result, dict) or set(result) != {"payload", "payload_digest", "qa"}:
            raise invalid("result must carry exactly payload, payload_digest, qa")
        payload, qa = result["payload"], result["qa"]
        if not isinstance(qa, dict) or qa.get("status") != "PASS":
            raise invalid("qa status must pass")
        if not isinstance(payload, dict) or not isinstance(payload.get("tabs"), list):
            raise invalid("payload tabs must be a list")
        for tab in payload["tabs"]:
            if not isinstance(tab, dict) or not isinstance(tab.get("cells"), list):
                raise invalid("malformed worksheet tab")
            for cell in tab["cells"]:
                if not isinstance(cell, dict) or not {"address", "row", "column", "value", "value_type"} <= set(cell):
                    raise invalid("incomplete worksheet cell")
                value = cell["value"]
                if isinstance(value, float) and (value != value or value in (float("inf"), float("-inf"))):
                    raise invalid("non-finite worksheet value")
        if result["payload_digest"] != digest(payload):
            raise invalid("payload digest mismatch")

    def _fail(self, build_id: str, code: str, detail: str, expected_fingerprint: str | None = None) -> None:
        self._validate_error(code, detail)
        self.builds.update_build(build_id, expected_status=("QUEUED", "BUILDING"),
                                 expected_input_fingerprint=expected_fingerprint,
                                 status="FAILED", error={"code": code, "detail": detail})

    @staticmethod
    def _validate_error(code: str, detail: str) -> None:
        if code not in BUILD_ERROR_CODES or not isinstance(detail, str) or len(detail) > MAX_ERROR_DETAIL_CHARS:
            raise ValueError("MODEL_ERROR_INVALID: failure records are bounded")

    def fail_build_for_tests(self, build_id: str, *, code: str, detail: str) -> None:
        self._validate_error(code, detail)
        changed = self.builds.update_build(build_id, expected_status=("QUEUED", "BUILDING"),
                                           status="FAILED", error={"code": code, "detail": detail})
        if not changed:
            raise ValueError("MODEL_ERROR_INVALID: build is not failable")

    def insert_ready_build_for_tests(self, case_id: str, *, build_id: str, queued_at: str) -> dict[str, Any]:
        return self.builds.insert_build_row({
            "id": build_id, "case_id": case_id, "status": "READY",
            "input_fingerprint": f"fixture-{build_id}", "queued_at": queued_at,
            "created_by": "test", "calculation_runtime": self.bundle.calculation_runtime,
        })

    # -- deadline -----------------------------------------------------------

    def _new_deadline(self) -> float:
        if self._deadline_exceeded:
            return time.monotonic() - 1.0
        return time.monotonic() + MAX_CALCULATION_SECONDS

    @staticmethod
    def _check_deadline(deadline: float) -> None:
        if deadline - time.monotonic() <= 0:
            raise ModelCalculationTimeout()

    # -- registry, previews, sign-off ---------------------------------------

    def _require_current(self, case_id: str, build_id: str) -> dict[str, Any]:
        build = self.builds.get_build(build_id)
        if build is None or build["case_id"] != case_id or build["status"] != "READY":
            raise ValueError("MODEL_BUILD_NOT_READY")
        current = self.builds.current_build(case_id)
        if current is None or current["id"] != build_id:
            raise ModelBuildStale(current)
        case = self.store.get_case(case_id)
        snapshot = self.engine.runs.get_snapshot(build.get("snapshot_id"))
        try:
            if (
                case is None
                or case.get("accepted_snapshot_id") != build.get("snapshot_id")
                or snapshot is None
                or snapshot.get("case_id") != case_id
            ):
                raise ModelInputError("accepted model authority changed")
            resolved = self._resolve_snapshot(snapshot)
            if (
                resolved.get("input_fingerprint") != build.get("input_fingerprint")
                or resolved.get("source_set", {}).get("id") != build.get("source_set_id")
            ):
                raise ModelInputError("accepted model authority changed")
        except (ModelInputError, ValueError) as exc:
            self._resolved_cache.pop(build_id, None)
            self._defaults_cache.pop(build_id, None)
            raise ValueError("MODEL_BUILD_NOT_READY: accepted model authority is invalid") from exc
        _remember(self._resolved_cache, build_id, resolved)
        return build

    def _validate_registry(self, registry_version: str, registry_digest: str) -> None:
        registry = self.bundle.assumption_registry
        if registry_version != registry["version"] or registry_digest != registry["digest"]:
            raise ValueError("MODEL_REGISTRY_STALE")

    def _parent(self, case_id: str, revision_id: str | None) -> dict[str, Any] | None:
        if revision_id is None:
            return None
        parent = self.builds.get_revision(revision_id)
        if parent is None or parent.get("case_id") != case_id:
            raise ValueError("MODEL_REVISION_NOT_FOUND")
        return parent

    def assumption_registry(self, case_id: str, build_id: str) -> dict[str, Any]:
        build = self._require_current(case_id, build_id)
        defaults, _outputs = self._defaults(build, self._new_deadline())
        registry = copy.deepcopy(self.bundle.assumption_registry)
        return {
            **registry,
            "build_id": build_id,
            "snapshot_id": build["snapshot_id"],
            "input_fingerprint": build["input_fingerprint"],
            "defaults": defaults,
        }

    def preview(self, case_id: str, request: ModelPreviewRequest, *, _deadline: float | None = None) -> dict[str, Any]:
        deadline = _deadline if _deadline is not None else self._new_deadline()
        self._check_deadline(deadline)
        build = self._require_current(case_id, request.build_id)
        self._validate_registry(request.registry_version, request.registry_digest)
        parent = self._parent(case_id, request.parent_revision_id)
        if parent is not None and parent.get("build_id") != request.build_id:
            raise ValueError("MODEL_REVISION_STALE")
        rows = [item.model_dump() for item in request.assumptions]
        defaults, default_outputs = self._defaults(build, deadline)
        model, calculations = self._calculate(build, rows, deadline)
        normalized = _with_default_context(self._apply_evolution(build["id"], _assumption_rows(model)), defaults)
        outputs = _annual_outputs(calculations)
        worksheet = _with_audit_tabs(
            self.bundle.serialize_workbook(model, calculations)["payload"],
            build.get("payload") or {},
        )
        baseline = parent["outputs"] if parent is not None else default_outputs
        envelope = {
            "case_id": case_id,
            "build_id": request.build_id,
            "snapshot_id": build["snapshot_id"],
            "build_input_fingerprint": build["input_fingerprint"],
            "build_payload_digest": build["payload_digest"],
            "registry_version": request.registry_version,
            "registry_digest": request.registry_digest,
            "calculation_contract_version": self.bundle.calculation_runtime["calculation_contract_version"],
            "parent_revision_id": request.parent_revision_id,
            "draft_generation": request.draft_generation,
            "effective_assumptions": normalized,
            "assumptions_digest": digest(normalized),
            "outputs": outputs,
            "outputs_digest": digest(outputs),
            "worksheet": worksheet,
            "deltas": _output_deltas(baseline, outputs),
        }
        envelope["preview_digest"] = digest(envelope)
        return envelope

    def _validate_build_identity(self, case_id: str, build: dict[str, Any], deadline: float) -> None:
        """Sign-off validates the exact current build identity — every pinned
        field recomputed, any mismatch refused as MODEL_REVISION_INVALID."""

        def invalid(field: str) -> ValueError:
            return ValueError(f"MODEL_REVISION_INVALID: build identity mismatch on {field}")

        if build["payload_digest"] != digest(build["payload"]):
            raise invalid("payload_digest")
        if build.get("methodology_build_id") != self.engine.bundle.build_id:
            raise invalid("methodology_build_id")
        if build.get("calculation_runtime") != self.bundle.calculation_runtime:
            raise invalid("calculation_runtime")
        if build["registry_digest"] != self.bundle.assumption_registry["digest"]:
            raise invalid("registry_digest")
        case = self.store.get_case(case_id)
        if case is None or build["snapshot_id"] != case.get("accepted_snapshot_id"):
            raise invalid("snapshot_id")
        snapshot = self.engine.runs.get_snapshot(build["snapshot_id"])
        if snapshot is None:
            raise invalid("snapshot_id")
        try:
            resolved = self._resolve_snapshot(snapshot)
        except (ModelInputError, ValueError) as exc:
            raise invalid("input_fingerprint") from exc
        if resolved["input_fingerprint"] != build["input_fingerprint"]:
            raise invalid("input_fingerprint")
        self._resolved_cache.pop(build["id"], None)
        self._defaults_cache.pop(build["id"], None)
        _remember(self._resolved_cache, build["id"], resolved)
        expected_result, expected_identity = self._compute_build_result(build, deadline=deadline)
        if build["payload"] != expected_result["payload"]:
            raise invalid("payload")
        if build["payload_digest"] != expected_result["payload_digest"]:
            raise invalid("payload_digest")
        if build["qa"] != expected_result["qa"]:
            raise invalid("qa")
        if build["assumptions_digest"] != expected_identity["assumptions_digest"]:
            raise invalid("assumptions_digest")
        if build["outputs_digest"] != expected_identity["outputs_digest"]:
            raise invalid("outputs_digest")

    def sign_off(self, case_id: str, request: ModelSignOffRequest, *, actor: str = "analyst") -> dict[str, Any]:
        deadline = self._new_deadline()
        with self.store.authority_guard(), self._sign_off_lock:
            try:
                build = self._require_current(case_id, request.build_id)
            except ValueError as exc:
                if str(exc).startswith("MODEL_BUILD_NOT_READY: accepted model authority"):
                    raise ValueError("MODEL_REVISION_INVALID: accepted model authority is invalid") from exc
                raise
            self._validate_build_identity(case_id, build, deadline)
            preview = self.preview(case_id, request, _deadline=deadline)
            if preview["preview_digest"] != request.preview_digest:
                raise ValueError("MODEL_PREVIEW_STALE")
            record = {key: preview[key] for key in (
                "case_id", "build_id", "snapshot_id", "build_input_fingerprint", "build_payload_digest",
                "registry_version", "registry_digest", "calculation_contract_version",
                "effective_assumptions", "assumptions_digest", "outputs", "outputs_digest",
                "worksheet", "preview_digest", "parent_revision_id",
            )}
            record["note"] = request.note
            current = self.builds.current_build(case_id)
            if current is None or current["id"] != request.build_id:
                raise ModelBuildStale(current)
            signed = self.builds.sign_off_revision(
                case_id, record, actor, request.expected_head_revision_id, self.store._audit,
            )
        return {**signed, "state": self._revision_state(case_id, signed)}

    def _revision_state(self, case_id: str, revision: dict[str, Any]) -> str:
        current = self.builds.current_build(case_id)
        case = self.store.get_case(case_id)
        accepted_snapshot_id = case.get("accepted_snapshot_id") if case is not None else None
        if (
            current is None
            or accepted_snapshot_id is None
            or current.get("snapshot_id") != accepted_snapshot_id
            or revision.get("snapshot_id") != accepted_snapshot_id
            or revision.get("build_id") != current["id"]
        ):
            return "STALE"
        head = self.builds.head_revision(case_id)
        return "ACTIVE" if head is not None and head["id"] == revision["id"] else "SUPERSEDED"

    def revisions(self, case_id: str) -> list[dict[str, Any]]:
        return [
            {**revision, "state": self._revision_state(case_id, revision)}
            for revision in self.builds.list_revisions(case_id)
        ]

    def revision_export_statuses(self, case_id: str) -> list[dict[str, Any]]:
        safe_fields = ("status", "error", "filename", "sha256", "size")
        statuses = []
        for item in self.builds.list_revision_exports(case_id):
            export = item.get("export") or {"status": "NOT_REQUESTED"}
            statuses.append({
                "revision_id": item["revision_id"],
                "export": {key: export[key] for key in safe_fields if key in export},
            })
        return statuses

    def head_revision(self, case_id: str) -> dict[str, Any] | None:
        head = self.builds.head_revision(case_id)
        if head is None:
            return None
        return {**head, "state": self._revision_state(case_id, head)}

    def publication_guard(self):
        """Serialize filing authority checks with analyst model sign-off."""
        return self._sign_off_lock

    def default_outputs(self, case_id: str, build_id: str) -> dict[str, Any]:
        """The current build's outputs under registry defaults (the fallback
        model authority a Deliverable may pin when no revision is signed)."""
        build = self._require_current(case_id, build_id)
        _defaults_rows, outputs = self._defaults(build, self._new_deadline())
        return outputs

    def validated_build(self, case_id: str, build_id: str) -> dict[str, Any]:
        """Reconstruct and verify a READY build before another service publishes it."""
        try:
            build = self._require_current(case_id, build_id)
        except ValueError as exc:
            if str(exc).startswith(
                "MODEL_BUILD_NOT_READY: accepted model authority"
            ):
                raise ValueError(
                    "MODEL_REVISION_INVALID: accepted model authority is invalid"
                ) from exc
            raise
        self._validate_build_identity(case_id, build, self._new_deadline())
        return build

    def validated_publication_build(
        self,
        case_id: str,
        build_id: str | None = None,
    ) -> dict[str, Any]:
        """Resolve the model a deliverable may publish under current authority.

        Model Builder remains current-snapshot-only. Earnings and covenant
        deliverables are the narrow exception: they may reuse the nearest
        validated Full Credit ancestor while their own accepted analysis stays
        the publication authority.
        """
        try:
            accepted_snapshot = self._accepted_snapshot(case_id)
            if accepted_snapshot is None:
                raise ModelInputError("accepted model authority is unavailable")
            accepted_run = self._validated_snapshot_run(accepted_snapshot)
            if accepted_run["pathway"] in PRIOR_FULL_CREDIT_PUBLICATION_PATHWAYS:
                build, model_authority = (
                    self._validated_prior_full_credit_publication_build(
                        build_id,
                        accepted_snapshot,
                        accepted_run,
                    )
                )
            else:
                candidate = (
                    self.builds.get_build(build_id)
                    if build_id is not None
                    else self.builds.current_build(case_id)
                )
                if candidate is None:
                    raise ModelBuildStale(None)
                build = self._require_current(case_id, candidate["id"])
                self._validate_build_identity(case_id, build, self._new_deadline())
                model_authority = self._publication_model_authority(
                    accepted_snapshot,
                    build,
                    relationship="CURRENT_ACCEPTED_MODEL",
                )
            _defaults_rows, outputs = self._defaults(build, self._new_deadline())
            return {
                "build": build,
                "outputs": outputs,
                "model_authority": model_authority,
            }
        except ModelBuildStale:
            raise
        except ValueError as exc:
            if str(exc).startswith("MODEL_REVISION_INVALID"):
                raise
            raise ValueError(
                "MODEL_REVISION_INVALID: accepted model authority is invalid"
            ) from exc

    # -- transient calculations ---------------------------------------------

    def tornado(self, case_id: str, request: ModelTornadoRequest) -> dict[str, Any]:
        deadline = self._new_deadline()
        build = self._require_current(case_id, request.build_id)
        self._validate_registry(request.registry_version, request.registry_digest)
        parent = self._parent(case_id, request.parent_revision_id)
        if parent is not None and parent.get("build_id") != request.build_id:
            raise ValueError("MODEL_REVISION_STALE")
        if not request.output_period_id.startswith(f"{request.case}::"):
            raise ValueError("MODEL_TORNADO_OUTPUT_INVALID")

        working = [item.model_dump() for item in request.assumptions]
        _model, baseline_calculations = self._calculate(build, working, deadline)
        baseline_outputs = _annual_outputs(baseline_calculations)
        baseline = baseline_outputs.get(request.case, {}).get(request.output_period_id, {}).get(request.output_id)
        if baseline is None:
            raise ValueError("MODEL_TORNADO_OUTPUT_INVALID")

        definitions = {item["assumption_id"]: item for item in self.bundle.assumption_registry["definitions"]}
        ready_indices: dict[str, list[int]] = {}
        for index, row in enumerate(working):
            if row["case"] == request.case and row["status"] == "READY":
                ready_indices.setdefault(row["assumption_id"], []).append(index)
        bars: list[dict[str, Any]] = []
        for choices, label, configured_swing in LEGACY_TORNADO_DRIVERS:
            assumption_id = next((item for item in choices if item in ready_indices), None)
            if assumption_id is None:
                continue
            definition = definitions[assumption_id]
            swing = Decimal(configured_swing) * Decimal(str(request.intensity))
            selected = ready_indices[assumption_id]
            low_rows, high_rows = copy.deepcopy(working), copy.deepcopy(working)
            hard_min, hard_max = Decimal(str(definition["hard_min"])), Decimal(str(definition["hard_max"]))
            for index in selected:
                value = Decimal(str(working[index]["value"]))
                low_rows[index]["value"] = float(max(hard_min, value - swing))
                high_rows[index]["value"] = float(min(hard_max, value + swing))
            _low_model, low_calculations = self._calculate(build, low_rows, deadline)
            _high_model, high_calculations = self._calculate(build, high_rows, deadline)
            low = _annual_outputs(low_calculations)[request.case][request.output_period_id][request.output_id]
            high = _annual_outputs(high_calculations)[request.case][request.output_period_id][request.output_id]
            if low is None or high is None:
                continue
            bars.append({
                "assumption_id": assumption_id,
                "label": label,
                "unit": definition["unit"],
                "swing": format(swing, "f"),
                "low": low,
                "high": high,
            })
        bars.sort(key=lambda bar: (-abs(Decimal(bar["high"]) - Decimal(bar["low"])), bar["assumption_id"]))
        return {
            "case_id": case_id,
            "build_id": request.build_id,
            "parent_revision_id": request.parent_revision_id,
            "registry_version": request.registry_version,
            "registry_digest": request.registry_digest,
            "draft_generation": request.draft_generation,
            "case": request.case,
            "output_period_id": request.output_period_id,
            "output_id": request.output_id,
            "intensity": request.intensity,
            "baseline": baseline,
            "bars": bars,
        }

    def scenario(self, case_id: str, request: ModelScenarioRequest) -> dict[str, Any]:
        deadline = self._new_deadline()
        self._check_deadline(deadline)
        build = self._require_current(case_id, request.build_id)
        self._validate_registry(request.registry_version, request.registry_digest)
        parent = self._parent(case_id, request.base_revision_id)
        if parent is not None and parent.get("build_id") != request.build_id:
            raise ValueError("MODEL_REVISION_STALE")
        defaults, default_outputs = self._defaults(build, deadline)
        baseline_assumptions = copy.deepcopy(parent["effective_assumptions"] if parent else defaults)
        baseline_outputs = parent["outputs"] if parent else default_outputs
        by_key = {(row["assumption_id"], row["case"], row["period_id"]): row for row in baseline_assumptions}
        seen: set[tuple[str, str, str]] = set()
        for shock in request.shocks:
            key = (shock.assumption_id, shock.case, shock.period_id)
            if key in seen or key not in by_key or by_key[key]["status"] != "READY":
                raise ValueError("MODEL_SCENARIO_INVALID")
            seen.add(key)
            by_key[key]["value"] = shock.value
        model, calculations = self._calculate(build, baseline_assumptions, deadline)
        normalized = _with_default_context(self._apply_evolution(build["id"], _assumption_rows(model)), defaults)
        outputs = _annual_outputs(calculations)
        scenario = {
            "case_id": case_id,
            "build_id": request.build_id,
            "base_revision_id": request.base_revision_id,
            "registry_version": request.registry_version,
            "registry_digest": request.registry_digest,
            "draft_generation": request.draft_generation,
            "effective_assumptions": normalized,
            "assumptions_digest": digest(normalized),
            "outputs": outputs,
            "outputs_digest": digest(outputs),
            "deltas": _output_deltas(baseline_outputs, outputs),
        }
        return {
            "draft_generation": request.draft_generation,
            "baseline": baseline_outputs,
            "scenario": scenario,
            "scenario_digest": digest(scenario),
        }

    def one_way(self, case_id: str, request: OneWaySensitivityRequest) -> dict[str, Any]:
        deadline = self._new_deadline()
        self._check_deadline(deadline)
        build = self._require_current(case_id, request.build_id)
        self._validate_registry(request.registry_version, request.registry_digest)
        parent = self._parent(case_id, request.base_revision_id)
        if parent is not None and parent.get("build_id") != request.build_id:
            raise ValueError("MODEL_REVISION_STALE")
        defaults, default_outputs = self._defaults(build, deadline)
        baseline_assumptions = copy.deepcopy(parent["effective_assumptions"] if parent else defaults)
        baseline_outputs = parent["outputs"] if parent else default_outputs
        if not _has_output(baseline_outputs, request.case, request.output_id):
            raise ValueError("MODEL_SENSITIVITY_OUTPUT_INVALID")
        definition = next(
            (item for item in self.bundle.assumption_registry["definitions"] if item["assumption_id"] == request.assumption_id),
            None,
        )
        selected = [
            row for row in baseline_assumptions
            if row["assumption_id"] == request.assumption_id and row["case"] == request.case
            and (request.period_scope == "ALL" or row["period_id"] == request.period_scope)
        ]
        if definition is None or not selected or any(row["status"] != "READY" for row in selected):
            raise ValueError("MODEL_SENSITIVITY_INVALID")
        base_value = Decimal(str(selected[0]["value"]))
        default_range = Decimal(str(definition["sensitivity_default"]["range"]))
        try:
            chosen_step = Decimal(str(request.step)) if request.step is not None else Decimal(str(definition["sensitivity_default"]["step"]))
            lower = Decimal(str(request.minimum)) if request.minimum is not None else base_value - default_range
            upper = Decimal(str(request.maximum)) if request.maximum is not None else base_value + default_range
        except InvalidOperation as exc:
            raise ValueError("MODEL_SENSITIVITY_INVALID") from exc
        hard_min = Decimal(str(definition["hard_min"]))
        hard_max = Decimal(str(definition["hard_max"]))
        if chosen_step <= 0 or lower > upper or lower < hard_min or upper > hard_max:
            raise ValueError("MODEL_SENSITIVITY_INVALID")
        point_count = int((upper - lower) / chosen_step) + 1
        if point_count > MAX_SENSITIVITY_POINTS or lower + chosen_step * (point_count - 1) != upper:
            raise ValueError("MODEL_SENSITIVITY_POINT_LIMIT")
        selected_keys = {(row["assumption_id"], row["case"], row["period_id"]) for row in selected}
        points: list[dict[str, Any]] = []
        breakpoint_record: dict[str, Any] | None = None
        for index in range(point_count):
            value = lower + chosen_step * index
            effective = copy.deepcopy(baseline_assumptions)
            for row in effective:
                if (row["assumption_id"], row["case"], row["period_id"]) in selected_keys:
                    row["value"] = float(value)
            _model, calculations = self._calculate(build, effective, deadline)
            outputs = _annual_outputs(calculations)
            point = {"value": format(value, "f"), "outputs": outputs, "deltas": _output_deltas(baseline_outputs, outputs)}
            points.append(point)
            breaches = [
                breach for breach in outputs.get(f"{request.case}_first_breaches", [])
                if _breach_matches_output(breach, request.output_id)
            ]
            if breakpoint_record is None and breaches:
                breakpoint_record = {"value": point["value"], "threshold": breaches[0]}
        return {
            "case_id": case_id,
            "build_id": request.build_id,
            "base_revision_id": request.base_revision_id,
            "registry_version": request.registry_version,
            "registry_digest": request.registry_digest,
            "assumption_id": request.assumption_id,
            "case": request.case,
            "period_scope": request.period_scope,
            "output_id": request.output_id,
            "draft_generation": request.draft_generation,
            "points": points,
            "breakpoint": breakpoint_record,
        }

    def rebase_preview(self, case_id: str, request: ModelRebasePreviewRequest) -> dict[str, Any]:
        deadline = self._new_deadline()
        source = self._parent(case_id, request.revision_id)
        if source is None:
            raise ValueError("MODEL_REVISION_NOT_FOUND")
        build = self._require_current(case_id, request.build_id)
        new_defaults, _outputs = self._defaults(build, deadline)
        source_by_key = {
            (row["assumption_id"], row["case"], row["period_id"]): row
            for row in source["effective_assumptions"]
        }
        definitions = {item["assumption_id"]: item for item in self.bundle.assumption_registry["definitions"]}
        candidate = copy.deepcopy(new_defaults)
        candidate_by_key = {(row["assumption_id"], row["case"], row["period_id"]): row for row in candidate}
        compatible: list[dict[str, Any]] = []
        changed: list[dict[str, Any]] = []
        invalidated: list[dict[str, Any]] = []
        for key in sorted(set(source_by_key).difference(candidate_by_key)):
            invalidated.append({"identity": list(key), "reason": "ASSUMPTION_NO_LONGER_MAPS"})
        for row in candidate:
            key = (row["assumption_id"], row["case"], row["period_id"])
            prior = source_by_key.get(key)
            definition = definitions.get(row["assumption_id"])
            invalid_reason = None
            if prior is None:
                invalid_reason = "ASSUMPTION_ADDED"
            elif definition is None:
                invalid_reason = "ASSUMPTION_NO_LONGER_MAPS"
            elif prior["unit"] != row["unit"] or prior["status"] != row["status"]:
                invalid_reason = "ASSUMPTION_METADATA_CHANGED"
            elif not isinstance(prior.get("source_context_digest"), str):
                invalid_reason = "ASSUMPTION_CONTEXT_UNAVAILABLE"
            elif prior["status"] == "READY":
                value = Decimal(str(prior["value"]))
                if value < Decimal(str(definition["hard_min"])) or value > Decimal(str(definition["hard_max"])):
                    invalid_reason = "ASSUMPTION_OUTSIDE_NEW_BOUNDS"
            if invalid_reason:
                invalidated.append({"identity": list(key), "reason": invalid_reason})
                continue
            row["value"] = prior["value"]
            row["gap_code"] = prior.get("gap_code", "")
            item = {"identity": list(key), "value": prior["value"]}
            if any(
                prior.get(field) != row.get(field)
                for field in ("default_value", "default_status", "default_gap_code", "source_context_digest")
            ):
                changed.append(item)
            else:
                compatible.append(item)
        preview = None
        if not invalidated:
            registry = self.bundle.assumption_registry
            preview = self.preview(case_id, ModelPreviewRequest.model_validate({
                "build_id": request.build_id,
                "parent_revision_id": None,
                "registry_version": registry["version"],
                "registry_digest": registry["digest"],
                "assumptions": candidate,
                "draft_generation": request.draft_generation,
            }), _deadline=deadline)
        return {
            "case_id": case_id,
            "source_revision_id": request.revision_id,
            "source_build_id": source["build_id"],
            "build_id": request.build_id,
            "draft_generation": request.draft_generation,
            "compatible": compatible,
            "changed": changed,
            "invalidated": invalidated,
            "candidate_assumptions": candidate,
            "preview": preview,
        }

    # -- exports -------------------------------------------------------------

    def queue_export(self, target_id: str, actor: str) -> dict[str, Any]:
        build = self.builds.get_build(target_id)
        if build is not None:
            if (build.get("export") or {}).get("status") == "READY":
                return build
            self.validated_build(build["case_id"], build["id"])
            self.builds.update_build(target_id, export={"status": "QUEUED"})
            return self.builds.get_build(target_id)  # type: ignore[return-value]
        revision = self.builds.get_revision(target_id)
        if revision is None:
            raise ValueError("MODEL_EXPORT_TARGET_NOT_FOUND")
        if (revision.get("export") or {}).get("status") == "READY":
            return revision
        self.validated_build(revision["case_id"], revision["build_id"])
        self.builds.update_revision_export(target_id, {"status": "QUEUED"})
        return self.builds.get_revision(target_id)  # type: ignore[return-value]

    def _publish(self, case_id: str, target_id: str, content: bytes) -> dict[str, Any]:
        import hashlib

        # Hash-addressed atomic publish: O_EXCL + fsync + rename, identity
        # characters checked, and an already-present target verified byte for
        # byte rather than overwritten. A plain write_bytes left a truncated
        # file behind on a crash for `download` to refuse later.
        sha256 = hashlib.sha256(content).hexdigest()
        _path, vault_key, size = publish_hash_addressed_bytes(
            self.vault_dir, ("models", case_id, target_id), "xlsx", content,
            expected_sha256=sha256, max_bytes=MAX_EXPORT_BYTES,
        )
        export = {
            "status": "READY",
            "vault_key": vault_key,
            "filename": f"{target_id}.xlsx",
            "sha256": sha256,
            "size": size,
        }
        export["identity_digest"] = _export_identity_digest(case_id, target_id, export)
        return export

    def _render_bytes(
        self,
        build: dict[str, Any],
        effective: list[dict[str, Any]] | None,
        signed_worksheet: dict[str, Any] | None = None,
    ) -> bytes:
        deadline = self._new_deadline()
        model, calculations = self._calculate(build, effective, deadline)
        with tempfile.TemporaryDirectory(prefix="caos-model-export-") as temporary:
            output = Path(temporary) / "model.xlsx"
            self.bundle.render_workbook(model, calculations, output)
            tabs = (
                signed_worksheet["tabs"]
                if signed_worksheet is not None
                else _audit_tabs(build.get("payload") or {})
            )
            _materialize_workbook_tabs(output, tabs)
            if signed_worksheet is not None:
                _assert_workbook_semantics(output, signed_worksheet)
            return output.read_bytes()

    def run_export_for_tests(self, target_id: str) -> dict[str, Any]:
        build = self.builds.get_build(target_id)
        if build is not None:
            if self._fail_next_export:
                self._fail_next_export = False
                self.builds.update_build(target_id, export={
                    "status": "FAILED", "error": {"code": "MODEL_EXPORT_FAILED", "detail": "The XLSX export did not complete."},
                })
                return self.builds.get_build(target_id)  # type: ignore[return-value]
            build = self.validated_build(build["case_id"], build["id"])
            self._export_input_reads += 1
            export = self._publish(build["case_id"], target_id, self._render_bytes(build, None))
            if not self.builds.update_build(
                target_id, expected_live_source_ids=self._pinned_source_ids(build), export=export,
            ):
                self.builds.update_build(target_id, export={
                    "status": "FAILED",
                    "error": {"code": "MODEL_EXPORT_AUTHORITY_CHANGED",
                              "detail": "A pinned source was withdrawn before the export could be published."},
                })
            return self.builds.get_build(target_id)  # type: ignore[return-value]
        revision = self.builds.get_revision(target_id)
        if revision is None:
            raise ValueError("MODEL_EXPORT_TARGET_NOT_FOUND")
        if (revision.get("export") or {}).get("status") == "READY":
            return revision  # exports are immutable once READY — never re-rendered
        source_build = self.builds.get_build(revision["build_id"])
        if source_build is None or source_build.get("calculation_runtime") != self._current_runtime():
            # §12: the runtime pin is verified before any input is read.
            self.builds.update_revision_export(target_id, {
                "status": "FAILED",
                "error": {"code": "MODEL_REVISION_EXPORT_RUNTIME_UNAVAILABLE",
                          "detail": "The signed revision's pinned calculation runtime is unavailable."},
            })
            return self.builds.get_revision(target_id)  # type: ignore[return-value]
        if self._fail_next_export:
            self._fail_next_export = False
            self.builds.update_revision_export(target_id, {
                "status": "FAILED", "error": {"code": "MODEL_REVISION_EXPORT_FAILED", "detail": "The signed revision XLSX export did not complete."},
            })
            return self.builds.get_revision(target_id)  # type: ignore[return-value]
        source_build = self.validated_build(revision["case_id"], source_build["id"])
        self._export_input_reads += 1
        content = self._render_bytes(
            source_build,
            revision["effective_assumptions"],
            revision["worksheet"],
        )
        export = self._publish(revision["case_id"], target_id, content)
        if not self.builds.update_revision_export(
            target_id, export, expected_live_source_ids=self._pinned_source_ids(source_build),
        ):
            self.builds.update_revision_export(target_id, {
                "status": "FAILED",
                "error": {"code": "MODEL_EXPORT_AUTHORITY_CHANGED",
                          "detail": "A pinned source was withdrawn before the export could be published."},
            })
        return self.builds.get_revision(target_id)  # type: ignore[return-value]

    # The worker's entry points. The `_for_tests` suffix on the underlying
    # methods is historical — the spec suite pinned those names before
    # worker.py existed; the execution path is the production one.
    run_build = run_build_for_tests
    run_export = run_export_for_tests

    def download(self, case_id: str, target_id: str) -> tuple[bytes, str]:
        record = self.builds.get_build(target_id) or self.builds.get_revision(target_id)
        is_revision = record is not None and "revision_number" in record
        if record is None or record.get("case_id") != case_id:
            raise ValueError("MODEL_EXPORT_TARGET_NOT_FOUND")
        export = record.get("export") or {}
        if export.get("status") != "READY":
            raise ValueError("MODEL_EXPORT_NOT_READY")
        integrity_error = (
            "MODEL_REVISION_EXPORT_INTEGRITY_FAILED"
            if is_revision
            else "MODEL_EXPORT_INTEGRITY_FAILED"
        )
        expected_vault_key = (
            f"models/{case_id}/{target_id}/{export.get('sha256')}.xlsx"
        )
        if (
            export.get("vault_key") != expected_vault_key
            or export.get("filename") != f"{target_id}.xlsx"
            or export.get("identity_digest")
            != _export_identity_digest(case_id, target_id, export)
        ):
            raise ValueError(integrity_error)
        # Read through the no-follow descriptor chain: the stored bytes are
        # verified against the recorded digest AND length, and a symlink or a
        # non-regular file where the export should be is refused outright.
        try:
            content = read_verified_vault_bytes(
                self.vault_dir, export["vault_key"], expected_sha256=export["sha256"],
                expected_size=export["size"], max_bytes=MAX_EXPORT_BYTES,
            )
        except (VaultFileUnavailable, VaultFileIntegrityError) as exc:
            raise ValueError(integrity_error) from exc
        return content, export["sha256"]

    # -- test seams ----------------------------------------------------------

    def fail_next_queue_for_tests(self) -> None:
        self._fail_next_queue = True

    def fail_next_dispatch_for_tests(self) -> None:
        self._fail_next_dispatch = True

    def fail_next_export_for_tests(self) -> None:
        self._fail_next_export = True

    def on_queue_for_tests(self, callback: Callable[[str], None]) -> None:
        self._on_queue.append(callback)

    def dispatch_log_for_tests(self) -> list[str]:
        return list(self._dispatch_log)

    def force_readiness_for_tests(self, case_id: str, code: str) -> None:
        self._forced_readiness[case_id] = code

    def validate_bundle_calls_for_tests(self) -> list[dict[str, Any]]:
        return list(self._validate_bundle_calls)

    def set_calculation_runtime_for_tests(self, *, sha256: str | None) -> None:
        self._runtime_sha_override = sha256

    def export_input_reads_for_tests(self) -> int:
        return self._export_input_reads

    def exceed_calculation_deadline_for_tests(self) -> None:
        self._deadline_exceeded = True

    def evolve_registry_for_tests(self, build_id: str, *, removed: tuple[str, ...] = (), source_context_changed: tuple[str, ...] = ()) -> None:
        self._evolutions[build_id] = {"removed": tuple(removed), "source_context_changed": tuple(source_context_changed)}
        self._defaults_cache.pop(build_id, None)

    def tamper_build_identity_for_tests(self, build_id: str, field: str) -> None:
        assert field in {"input_fingerprint", "payload_digest", "registry_digest", "snapshot_id", "assumptions_digest", "outputs_digest"}
        self.builds.update_build(build_id, **{field: "0" * 64})


def _fiscal_year(period_id: str) -> str | None:
    match = _ANNUAL_PERIOD.match(period_id)
    return match.group(1) if match else None


def _forecast_variance(period_updates: list[dict[str, Any]], base_outputs: dict[str, Any]) -> list[dict[str, Any]]:
    """Reported minus forecast for every reported annual period the base model
    forecasts, per case and metric. Both operands are finite by construction;
    a zero forecast leaves the ratio null with a named gap (invariant 7)."""
    rows: list[dict[str, Any]] = []
    for update in period_updates:
        fiscal_year = _fiscal_year(update["period_id"])
        for case in ("BASE", "DOWNSIDE"):
            column = f"{case}::FY{fiscal_year}" if fiscal_year else None
            forecast_values = base_outputs.get(case, {}).get(column) if column else None
            for reported_key, output_id in VARIANCE_METRICS:
                reported = update["reported"].get(reported_key)
                forecast = forecast_values.get(output_id) if isinstance(forecast_values, dict) else None
                row: dict[str, Any] = {
                    "period_id": update["period_id"],
                    "case": case,
                    "metric_id": output_id,
                    "reported_metric": reported_key,
                    "reported": reported,
                    "forecast": forecast,
                    "variance": None,
                    "variance_ratio": None,
                    "gap_code": None,
                    "reported_authority": "REPORTED_ACTUAL",
                    "forecast_authority": "ANALYST_FORECAST",
                }
                if forecast_values is None:
                    row["gap_code"] = "FORECAST_PERIOD_NOT_MODELLED"
                elif reported is None:
                    row["gap_code"] = "ACTUAL_NOT_DISCLOSED"
                elif forecast is None:
                    row["gap_code"] = "FORECAST_NOT_CALCULABLE"
                else:
                    actual, expected = Decimal(reported), Decimal(forecast)
                    if not actual.is_finite() or not expected.is_finite():
                        raise ModelInputError("non-finite forecast variance operand")
                    row["variance"] = format(actual - expected, "f")
                    if expected == 0:
                        row["gap_code"] = "ZERO_FORECAST_DENOMINATOR"
                    else:
                        row["variance_ratio"] = format((actual - expected) / expected, "f")
                rows.append(row)
    return rows


def _covenant_slot(test: dict[str, Any]) -> str | None:
    # ponytail: name-based slot mapping over the one covenant slot the registry
    # declares; a registry-declared test→slot table if the methodology adds one.
    if test.get("test_type") == "max-ratio" and "leverage" in str(test.get("test") or "").casefold():
        return "covenant.max_total_leverage"
    return None


def _covenant_assumption_updates(
    covenant_updates: list[dict[str, Any]], defaults: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """The base registry rows a documentary covenant test would update, as
    proposals: the base value and status stay visible beside the documentary
    threshold, and adoption is the analyst's preview → sign-off."""
    updates: list[dict[str, Any]] = []
    for test in covenant_updates:
        slot = _covenant_slot(test)
        base_rows = [row for row in defaults if row["assumption_id"] == slot] if slot else []
        if not base_rows or test.get("threshold") is None:
            updates.append({
                "test": test.get("test"),
                "assumption_id": slot,
                "gap_code": "COVENANT_THRESHOLD_NOT_DISCLOSED" if base_rows else "UNMAPPED_COVENANT_TEST",
                "treatment": "NO_ASSUMPTION_UPDATE",
                "authority": "DOCUMENTARY_COVENANT_TERMS",
            })
            continue
        for row in base_rows:
            # The pinned engine lets a preview change a READY value, never a
            # status: a slot the accepted CP-2G handoff left UNAVAILABLE can
            # only become READY through a Full Credit re-run that sources it.
            updates.append({
                "test": test.get("test"),
                "assumption_id": slot,
                "case": row["case"],
                "period_id": row["period_id"],
                "unit": row["unit"],
                "base_value": row["value"],
                "base_status": row["status"],
                "base_gap_code": row["gap_code"],
                "proposed_value": test["threshold"],
                "authority": "DOCUMENTARY_COVENANT_TERMS",
                "treatment": (
                    "PROPOSED_FOR_SIGN_OFF" if row["status"] == "READY"
                    else "PROPOSED_REQUIRES_FULL_CREDIT_HANDOFF"
                ),
            })
    return updates


def _latest_reported_period_end(cp1_markdown: str) -> str | None:
    ends = [
        row.get("end_date", "")
        for row in _table_rows(cp1_markdown, "cp1.model_period_register")
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", row.get("end_date", "") or "")
    ]
    return max(ends) if ends else None


def _remember(cache: dict[str, Any], key: str, value: Any) -> Any:
    """Insert, then evict the oldest entries past MAX_CACHED_BUILDS.
    Plain dicts keep insertion order, so the first key is the oldest."""
    cache.pop(key, None)
    cache[key] = value
    while len(cache) > MAX_CACHED_BUILDS:
        del cache[next(iter(cache))]
    return value


# -- pure row/output projections (shape shared with the assumption contract) --


def _assumption_rows(model: Any) -> list[dict[str, Any]]:
    rows = []
    for driver in model.effective_assumptions:
        value = json_value(driver.value) if driver.value is not None else None
        row: dict[str, Any] = {
            "assumption_id": driver.assumption_id,
            "case": driver.case,
            "period_id": driver.period_id,
            "unit": driver.unit,
            "status": driver.status,
            "value": value,
            "gap_code": driver.gap_code,
            "default_value": value,
            "default_status": driver.status,
            "default_gap_code": driver.gap_code,
            "source_context": {
                "authority_module": "CP-2G",
                "gap_code": driver.gap_code,
                "provenance": [
                    {"source_id": item.source_id, "source_locator": item.source_locator, "as_of": item.as_of}
                    for item in driver.provenance
                ],
            },
        }
        row["source_context_digest"] = digest(row["source_context"])
        rows.append(row)
    return sorted(rows, key=lambda row: (row["case"], row["period_id"], row["assumption_id"]))


def _with_default_context(effective: list[dict[str, Any]], defaults: list[dict[str, Any]]) -> list[dict[str, Any]]:
    default_by_key = {(row["assumption_id"], row["case"], row["period_id"]): row for row in defaults}
    if len(default_by_key) != len(defaults):
        raise ValueError("MODEL_ASSUMPTION_DEFAULT_INVALID")
    result = copy.deepcopy(effective)
    for row in result:
        default = default_by_key.get((row["assumption_id"], row["case"], row["period_id"]))
        if default is None:
            raise ValueError("MODEL_ASSUMPTION_DEFAULT_INVALID")
        row.update(
            default_value=default["value"],
            default_status=default["status"],
            default_gap_code=default["gap_code"],
            source_context=copy.deepcopy(default["source_context"]),
            source_context_digest=default["source_context_digest"],
        )
    return result


def _json_number(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        if not value.is_finite():
            raise ModelInputError("non-finite model output")
        return format(value, "f")
    raise ModelInputError("unexpected model output type")


_OUTPUT_VALUE_IDS = (
    "revenue", "adjusted_ebitda_calc", "fcf", "cumulative_fcf", "cash_and_equivalents",
    "accessible_liquidity", "liquidity_headroom", "total_debt_reported", "net_debt",
)
_OUTPUT_METRIC_IDS = (
    "adjusted_ebitda_margin", "total_leverage", "net_leverage", "interest_coverage", "covenant_headroom",
)


def _annual_outputs(calculations: Any) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for case in ("BASE", "DOWNSIDE"):
        periods: dict[str, Any] = {}
        for column in calculations.columns:
            if column.group != case:
                continue
            calculation = calculations.for_column(column.column_id)
            periods[column.column_id] = {
                **{output_id: _json_number(calculation.values.get(output_id)) for output_id in _OUTPUT_VALUE_IDS},
                **{output_id: _json_number(calculation.credit_metrics.get(output_id)) for output_id in _OUTPUT_METRIC_IDS},
            }
        result[case] = periods
        result[f"{case}_first_breaches"] = [
            {
                "period_id": breach.period_id,
                "threshold_id": breach.threshold_id,
                "metric_id": breach.metric_id,
                "limit": _json_number(breach.limit),
                "actual": _json_number(breach.actual),
                "headroom": _json_number(breach.headroom),
            }
            for breach in calculations.first_breaches[case]
        ]
    return result


def _output_deltas(baseline: dict[str, Any], changed: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for case in ("BASE", "DOWNSIDE"):
        result[case] = {}
        for period_id, values in changed.get(case, {}).items():
            base_values = baseline.get(case, {}).get(period_id, {})
            result[case][period_id] = {}
            for output_id, value in values.items():
                base = base_values.get(output_id)
                result[case][period_id][output_id] = (
                    format(Decimal(value) - Decimal(base), "f") if value is not None and base is not None else None
                )
    return result


def _has_output(outputs: dict[str, Any], case: str, output_id: str) -> bool:
    return any(output_id in values for values in outputs.get(case, {}).values() if isinstance(values, dict))


def _breach_matches_output(breach: dict[str, Any], output_id: str) -> bool:
    threshold_outputs = {
        "covenant.max_total_leverage": {"covenant_headroom", "total_leverage"},
        "liquidity.minimum_operating_cash": {"accessible_liquidity", "liquidity_headroom"},
    }
    return breach.get("metric_id") == output_id or output_id in threshold_outputs.get(breach.get("threshold_id"), set())
